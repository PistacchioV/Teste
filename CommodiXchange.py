import win32com.client as win32
from pynput import keyboard
import sys
from tkcalendar import DateEntry
import pythoncom
from CTkScrollableDropdown import *
import win32com.client
import traceback
from typing import List, Dict, Any, Tuple, Union
from datetime import datetime, timedelta
import tkinter as tk
from tkinter.filedialog import askdirectory
from tkinter import ttk
from bs4 import BeautifulSoup
from tkinter.font import Font
import random
import sqlite3
from tkinter import messagebox
import os
import math
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from docx.oxml import OxmlElement
from docx.oxml import parse_xml
from docx.oxml.ns import nsdecls
from PIL import Image, ImageTk
import glob
import getpass
import fitz # PyMuPDF
import numpy as np
import tempfile
from time import sleep
import threading
from tkinter import font

checkbox_states_deals = {}
checkbox_states_fixings = {}
checkbox_states_file = {}
global tabela_termo_cliente, tabela_termo_b2b, tabela_opcao_cliente, tabela_opcao_b2b
global tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b
global tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b
global tabela_boletatermo, tabela_boletaopcao
global deal_option_query_window_active
global deal_query_window_active

# Variáveis globais para armazenar o estado das janelas
deal_query_window_active = False
deal_option_query_window_active = False

script_dir = os.path.dirname(__file__)    
db_path = os.path.join(script_dir, "db", "refData.db")
                       

# Variáveis globais para armazenar os argumentos
window = None
option_menu_trade_date = None
date_entry1_trade_date = None
date_entry2_trade_date = None
option_menu_settlement_date = None
date_entry1_settlement_date = None
date_entry2_settlement_date = None
option_menu_fixing_commodity = None
date_entry1_fixing_commodity = None
date_entry2_fixing_commodity = None
option_menu_fixing_ccy = None
date_entry1_fixing_ccy = None
date_entry2_fixing_ccy = None
counterparty_combobox_termo = None
notional_entry = None
strike_entry = None
athena_id_entry = None
status_option_menu = None
treeviews = None


# Variável global para armazenar as tabelas de feriados
tabela_feriados_global = None
# Definição global das colunas
colunas_anbima = ["DATA", "DIA DA SEMANA", "FERIADO"]
colunas_mercadoria = ["COMMODITY", "DATA", "DIA DA SEMANA", "FERIADO"]
colunas_lme = ["COMMODITY", "DATA", "DIA DA SEMANA", "FERIADO"]
colunas_ice = ["COMMODITY", "DATA", "DIA DA SEMANA", "FERIADO"]
colunas_platts = ["COMMODITY", "DATA", "DIA DA SEMANA", "FERIADO"]
colunas_cbot = ["COMMODITY", "DATA", "DIA DA SEMANA", "FERIADO"]
colunas_nymex = ["COMMODITY", "DATA", "DIA DA SEMANA", "FERIADO"]
colunas_bursa = ["COMMODITY", "DATA", "DIA DA SEMANA", "FERIADO"]
colunas_commodities = [
        "ATIVO SUBJACENTE / RIC", "BOLSA DE NEGOCIACAO", "INDICE VALORIZACAO", "MES VENCIMENTO", "ANO VENCIMENTO", "TIPO",
        "UNIDADE DE NEGOCIACAO", "MOEDA", "FATOR DE CONVERSÃO", "MERCADORIA", "STATUS", "SID", "TIME STAMP"
    ]
 # Columns for deals
termo_deals_columns = [
        "DealName", "TradeDate", "Market", "Type", "Instrument", "Contract", "Strike", "Currency",
        "IntermediateCCY", "TotalNotional", "SettlementDate", "TradingBook", "OtherBook", "ClientValue",
        "SpotFXRate", "FXConvDate", "FixingStartDate", "FixingEndDate", "Counterparty", "Identifier",
        "Index", "Status", "SID"
    ]
opcao_fixings_columns = [ 
    "AthenaID", "Exchange", "Commodity", "Notional", "Number_of_Deates", "DATA_1", "DATA_2", "DATA_3",
    "DATA_4", "DATA_5", "DATA_6", "DATA_7", "DATA_8", "DATA_9", "DATA_10", "DATA_11", "DATA_12", "DATA_13",
    "DATA_14", "DATA_15", "DATA_16", "DATA_17", "DATA_18", "DATA_19", "DATA_20", "DATA_21", "DATA_22",
    "DATA_23", "DATA_24", "DATA_25", "DATA_26", "DATA_27", "DATA_28", "DATA_29", "DATA_30", "DATA_31",
    "DATA_32", "DATA_33", "DATA_34", "DATA_35", "DATA_36", "DATA_37", "DATA_38", "DATA_39", "DATA_40",
    "DATA_41", "DATA_42", "DATA_43", "DATA_44", "DATA_45", "Identifier", "Index", "Status", "SID"
]
termo_fixings_columns = [
    "AthenaID", "Exchange", "Commodity", "Strike_in_BRL", "Number_of_Deals",
    "DATA_1", "DATA_2", "DATA_3", "DATA_4", "DATA_5", "DATA_6", "DATA_7", "DATA_8", "DATA_9", "DATA_10",
    "DATA_11", "DATA_12", "DATA_13", "DATA_14", "DATA_15", "DATA_16", "DATA_17", "DATA_18", "DATA_19", "DATA_20",
    "DATA_21", "DATA_22", "DATA_23", "DATA_24", "DATA_25", "DATA_26", "DATA_27", "DATA_28", "DATA_29", "DATA_30",
    "DATA_31", "DATA_32", "DATA_33", "DATA_34", "DATA_35", "DATA_36", "DATA_37", "DATA_38", "DATA_39", "DATA_40",
    "DATA_41", "DATA_42", "DATA_43", "DATA_44", "DATA_45", "DATA_46", "DATA_47", "DATA_48", "DATA_49", "DATA_50",
    "DATA_51", "DATA_52", "DATA_53", "DATA_54", "DATA_55", "DATA_56", "DATA_57", "DATA_58", "DATA_59", "DATA_60",
    "DATA_61", "DATA_62", "DATA_63", "DATA_64", "DATA_65", "DATA_66", "DATA_67", "DATA_68", "DATA_69", "DATA_70",
    "DATA_71", "DATA_72", "DATA_73", "DATA_74", "DATA_75", "DATA_76", "DATA_77", "DATA_78", "DATA_79", "DATA_80",
    "Identifier", "Index", "Status", "SID"
]   
    # Columns for file
termo_file_columns = [
        "ID_do_Sistema", "ID_Tipo_de_Linha", "Código_operaçao", "Meu_Número", "Lançamento_do_Participante_Conta",
        "Papel_Posiçao_do_participante", "CPF_CNPJ_Cliente_Parte", "Contraparte", "CPF_CNPJ_Cliente_Contraparte",
        "Contrato_Global", "Classe_do_Ativo_Subjacente", "Fonte_Informaçao", "Moeda_de_Referência", "Moeda_Cotada",
        "Cotaçao_para_o_Vencimento", "Valor_Base_Quantidade", "Código_do_Ativo_Subjacente", "Taxa_a_Termo_R_Moeda",
        "Data_de_fixing_do_Ativo_Subjacente", "Data_de_Operaçao", "Data_vencimento", "Boletim", "Tipo_de_Cotaçao",
        "Data_de_Fixing_da_Moeda", "Cross_Rate_na_Avaliaçao", "Fonte_de_Consulta", "Tela_ou_Funçao_de_Consulta",
        "Praça_de_Negociaçao", "Horário_de_Consulta", "Cotaçao_Taxa_de_Câmbio_R_USD", "Cotaçao_Paridade_Moeda_USD_ou_USD_Moeda",
        "Data_de_Avaliaçao", "Código_da_paridade_cross", "Data_de_fixing_da_paridade_cross", "Termo_a_Termo", "Data_de_Fixaçao",
        "Forma_de_Atualizaçao", "Valor_Percentual_Negociado", "Cotaçao_para_fixing", "Atualizar_Valor_Base", "Cotaçao_Inicial",
        "Ajustar_Taxa", "Responsável_pelo_Ajuste_da_Taxa", "Data_Inicial_para_Ajuste_da_Taxa", "Data_Final_para_Ajuste_de_taxa",
        "Limites", "Superior_Paridade", "Inferior_Paridade", "Data_de_Liquidaçao_do_Prêmio", "Prêmio_a_ser_pago_pelo",
        "Valor_do_Prêmio", "Modalidade_de_Liquidaçao", "Prêmio_em_Moeda_Estrangeira", "Data_de_fixing_da_moeda_do_prêmio",
        "Taxa_a_Termo_em_Reais", "Observaçao", "Código_Identificador", "Tipo_Média_Asiático", "Quantidade_de_Datas_de_Verificaçao",
        "Identifier", "Index", "Status", "SID"
    ]
    
    # Columns for base_deals
base_deals_columns = [
        "AthenaID", "DealName", "B3_ID", "B2B_AthenaID", "B2B_B3_ID", "TradeDate", "Instrument", "Status",
        "Maker", "Checker", "Time_Stamp", "Confirmation", "SS_Validation", "Identifier", "Index", "Status", "SID"
    ]
# Define opcao_file_columns
opcao_file_columns = [
    "ID_do_Sistema", "ID_Tipo_de_Linha", "Codigo_da_Operaçao", "Tipo_Contrato", "Código", "Conta_Parte",
    "Conta_Contraparte", "Papel_Parte", "Moeda_Base_Índice_Ações", "Data_Inicio", "Data_de_Vencimento",
    "Preço_de_Exercício", "Casas_Decimais_Preço_Exercicio", "Prêmio_Unitário", "Casas_Decimais_do_Prêmio_Unitário",
    "Valor_Base_em_Moeda_Estrangeira_Quantidade", "Casas_Decimais_do_Valor_Base_em_Moeda_Estrangeira_ou_Quantidade",
    "Cotaçao_Índice_Limite", "Casas_Decimais_da_Cotaçao_Índice_Limite", "Tipo_de_Exercício", "Banco_Liquidante",
    "Modalidade", "Adesao_a_Contrato", "Meu_Número", "Conta_Intermediador", "Comissao_paga_pelo_Titular",
    "Casas_Decimais_da_Comissao_paga_pelo_Titular", "Comissao_paga_pelo_Lançador", "Casas_Decimais_da_Comissao_paga_pelo_Lançador",
    "Cross_Rate_na_Avaliaçao", "Fonte_de_Informaçao", "Cotaçao_para_o_Vencimento", "Boletim", "Horário_do_Boletim",
    "Fonte_de_Consulta", "Outra_Fonte_de_Consulta", "Tela_ou_Funçao_de_Consulta", "Praça_de_Negociaçao", "Horário_de_Consulta",
    "Cotaçao_Taxa_de_Câmbio", "Cotaçao_Paridade", "Data_de_Avaliaçao", "CPF_CNPJ_da_Parte", "CPF_CNPJ_da_Contraparte",
    "Moeda_Cotada", "Barreiras", "Trigger_In", "Casas_Decimais_do_Trigger_In", "Trigger_Out", "Casas_Decimais_do_Trigger_Out",
    "Cesta_de_Garantias_Lançador", "Forma_de_Verificaçao", "Rebate", "Valor_do_Rebate", "Casas_decimais_do_Valor_do_Rebate",
    "Liquidaçao_do_Rebate", "Código_da_Açao_Indice_Internacional", "Ajuste_de_Proventos_pelas",
    "Proteçao_contra_Provento_em_Dinheiro", "Trigger_Proporçao", "Trigger_Forma_de_Disparo", "Trigger_Tipo_de_Disparo",
    "Preço_de_Exercício_em_Reais", "Opçao_Quanto", "Cotaçao_para_Opçao_Quanto", "Casas_decimais_do_Cotaçao_para_Opçao_Quanto",
    "Data_de_Liquidaçao_do_Prêmio", "Mercadoria", "Cotaçao_para_Moeda", "Observaçao", "Média_para_Opçao_Asiática",
    "Data_de_Verificaçao", "Valor_Quantidade_de_Referência", "Casas_Decimais_do_Valor_Quantidade_de_Referência",
    "Data_de_Verificaçao_2", "Valor_Quantidade_de_Referência_2", "Casas_Decimais_do_Valor_Quantidade_de_Referência_2",
    "Identifier", "Index", "Status", "SID"
]

# Define opcao_deals_columns
opcao_deals_columns = [
    "DealName", "TradeDate", "Market", "Type", "Instrument", "Contract", "Strike", "Currency", "IntermediateCCY",
    "TotalNotional", "SettlementDate", "TradingBook", "OtherBook", "ClientValue", "SpotFXRate", "FXConvDate",
    "FixingStartDate", "FixingEndDate", "Counterparty", "Premium", "PremiumPerUnit", "PremiumCCY", "SpotDate",
    "Mnemonico", "Identifier", "Index", "Status", "SID"
]
colunas_comitentes = [
        "CNPJ", "ENTE", "SPN", "ECI", "CLIENTE", "ACCRONYM", "CONTA CETIP", "CGD", "MAPEAMENTO CONFIRMAÇÕES", "BANCO", "AG", "CC", "STATUS", "SID", "TIME STAMP"
    ]
colunas_arquivo_termo = (
    "ID do Sistema", "ID Tipo de Linha", "Código operação", "Meu Número", "Lançamento do Participante (Conta)",
    "Papel (Posição do participante)", "CPF/CNPJ Cliente Parte", "Contraparte", "CPF/CNPJ Cliente Contraparte",
    "Contrato Global", "Classe do Ativo Subjacente", "Fonte Informação", "Moeda de Referência", "Moeda Cotada", 
    "Cotação para o Vencimento", "Valor Base / Quantidade", "Código do Ativo Subjacente", 
    "Taxa a Termo (R$/Moeda)", "Data de fixing do Ativo Subjacente", "Data de Operação", "Data vencimento", 
    "Boletim", "Tipo de Cotação", "Data de Fixing da Moeda", "Cross Rate na Avaliação?", "Fonte de Consulta", 
    "Tela ou Função de Consulta", "Praça de Negociação", "Horário de Consulta", "Cotação - Taxa de Câmbio R$/USD",
    "Cotação - Paridade (Moeda/USD ou USD/ Moeda)", "Data de Avaliação", "Código da paridade cross",
    "Data de fixing da paridade cross", "Termo a Termo", "Data de Fixação", "Forma de Atualização",
    "Valor / Percentual Negociado", "Cotação para fixing", "Atualizar Valor Base?", "Cotação Inicial", "Ajustar Taxa",
    "Responsável pelo Ajuste da Taxa", "Data Inicial para Ajuste da Taxa", "Data Final para Ajuste de taxa", "Limites",
    "Superior (Paridade)", "Inferior (Paridade)", "Data de Liquidação do Prêmio", "Prêmio a ser pago pelo",
    "Valor do Prêmio", "Modalidade de Liquidação", "Prêmio em Moeda Estrangeira", "Data de fixing da moeda do prêmio",
    "Taxa a Termo em Reais", "Observação", "Código Identificador", "Tipo Média Asiático",
    "Quantidade de Datas de Verificação","Identifier", "Index", "Status", "SID"
)
colunas_arquivo_opcao = (
    "ID do Sistema", "ID Tipo de Linha", "Codigo da Operação", "Tipo Contrato", "Código", "Conta Parte", "Conta Contraparte",
    "Papel Parte", "Moeda Base/Índice/Ações", "Data Inicio", "Data de Vencimento", "Preço de Exercício",
    "Casas Decimais Preço Exercicio", "Prêmio Unitário", "Casas Decimais do Prêmio Unitário",
    "Valor Base em Moeda Estrangeira / Quantidade", "Casas Decimais do Valor Base em Moeda Estrangeira ou Quantidade.",
    "Cotação / Índice Limite", "Casas Decimais da Cotação / Índice Limite", "Tipo de Exercício", "Banco Liquidante",
    "Modalidade", "Adesão a Contrato", "Meu Número", "Conta Intermediador", "Comissão paga pelo Titular",
    "Casas Decimais da Comissão paga pelo Titular", "Comissão paga pelo Lançador",
    "Casas Decimais da Comissão paga pelo Lançador", "Cross-Rate na Avaliação", "Fonte de Informação",
    "Cotação para o Vencimento", "Boletim", "Horário do Boletim", "Fonte de Consulta", "Outra Fonte de Consulta",
    "Tela ou Função de Consulta", "Praça de Negociação", "Horário de Consulta", "Cotação – Taxa de Câmbio",
    "Cotação – Paridade", "Data de Avaliação", "CPF / CNPJ da Parte", "CPF / CNPJ da Contraparte", "Moeda Cotada",
    "Barreiras", "Trigger In", "Casas Decimais do Trigger In", "Trigger Out", "Casas Decimais do Trigger Out",
    "Cesta de Garantias - Lançador", "Forma de Verificação", "Rebate", "Valor do Rebate",
    "Casas decimais do Valor do Rebate", "Liquidação do Rebate", "Código da Ação / Indice Internacional",
    "Ajuste de Proventos pelas", "Proteção contra Provento em Dinheiro", "Trigger – Proporção",
    "Trigger – Forma de Disparo", "Trigger – Tipo de Disparo", "Preço de Exercício em Reais", "Opção Quanto",
    "Cotação para Opção Quanto", "Casas decimais do Cotação para Opção Quanto", "Data de Liquidação do Prêmio",
    "Mercadoria", "Cotação para Moeda", "Observação", "Média para Opção Asiática", "Data de Verificação",
    "Valor/Quantidade de Referência.", "Casas Decimais do Valor/Quantidade de Referência.", "Data de Verificação.",
    "Valor/Quantidade de Referência", "Casas Decimais do Valor/Quantidade de Referência","Identifier", "Index", "Status", "SID"
)
colunas_opcao = (
    "DealName", "TradeDate", "Market", "Type", "Instrument", "Contract", "Strike", "Currency",
    "IntermediateCCY", "TotalNotional", "SettlementDate", "TradingBook", "OtherBook", "ClientValue",
    "SpotFXRate", "FXConvDate", "FixingStartDate", "FixingEndDate","Counterparty", "Premium", 
    "PremiumPerUnit", "PremiumCCY", "SpotDate", "Mnemonico","Identifier", "Index", "Status", "SID"
)

colunas_termo = (
    "DealName", "TradeDate", "Market", "Type", "Instrument", "Contract", "Strike", "Currency",
    "IntermediateCCY", "TotalNotional", "SettlementDate", "TradingBook", "OtherBook", "ClientValue",
    "SpotFXRate", "FXConvDate", "FixingStartDate", "FixingEndDate", "Counterparty", "Identifier", "Index", "Status", "SID"
)
# Colunas para fixings de opção
colunas_fixings_opcao = (
    "AthenaID", "Exchange", "Commodity", "Notional", "Number of Dates",
    "DATA 1", "DATA 2", "DATA 3", "DATA 4", "DATA 5", "DATA 6", "DATA 7", "DATA 8", "DATA 9", "DATA 10",
    "DATA 11", "DATA 12", "DATA 13", "DATA 14", "DATA 15", "DATA 16", "DATA 17", "DATA 18", "DATA 19", "DATA 20",
    "DATA 21", "DATA 22", "DATA 23", "DATA 24", "DATA 25", "DATA 26", "DATA 27", "DATA 28", "DATA 29", "DATA 30",
    "DATA 31", "DATA 32", "DATA 33", "DATA 34", "DATA 35", "DATA 36", "DATA 37", "DATA 38", "DATA 39", "DATA 40",
    "DATA 41", "DATA 42", "DATA 43", "DATA 44", "DATA 45", "Identifier", "Index", "Status", "SID"
)

# Colunas para fixings de termo
colunas_fixings_termo = (
    "AthenaID", "Exchange", "Commodity", "Strike in BRL", "Number of Dates",
    "DATA 1", "DATA 2", "DATA 3", "DATA 4", "DATA 5", "DATA 6", "DATA 7", "DATA 8", "DATA 9", "DATA 10",
    "DATA 11", "DATA 12", "DATA 13", "DATA 14", "DATA 15", "DATA 16", "DATA 17", "DATA 18", "DATA 19", "DATA 20",
    "DATA 21", "DATA 22", "DATA 23", "DATA 24", "DATA 25", "DATA 26", "DATA 27", "DATA 28", "DATA 29", "DATA 30",
    "DATA 31", "DATA 32", "DATA 33", "DATA 34", "DATA 35", "DATA 36", "DATA 37", "DATA 38", "DATA 39", "DATA 40",
    "DATA 41", "DATA 42", "DATA 43", "DATA 44", "DATA 45", "DATA 46", "DATA 47", "DATA 48", "DATA 49", "DATA 50",
    "DATA 51", "DATA 52", "DATA 53", "DATA 54", "DATA 55", "DATA 56", "DATA 57", "DATA 58", "DATA 59", "DATA 60",
    "DATA 61", "DATA 62", "DATA 63", "DATA 64", "DATA 65", "DATA 66", "DATA 67", "DATA 68", "DATA 69", "DATA 70",
    "DATA 71", "DATA 72", "DATA 73", "DATA 74", "DATA 75", "DATA 76", "DATA 77", "DATA 78", "DATA 79", "DATA 80",
    "Identifier", "Index", "Status", "SID"
)

colunas_boleta_opcao = (
    "CARTEIRA", "ID do Sistema", "ID Tipo de Linha", "Data de Registro", "Conta Titular", "Nome do Titular",
    "Contrato", "Contrato Cetip", "Meu Número", "Tipo da Operaçao", "Conta Lancador", "Nome do Lancador",
    "Data de Inicio", "Data de Vencimento", "Numero Cetip", "Código SISBACEN da Moeda", "Símbolo da Moeda",
    "Valor da Aplicação", "Valor Base Moeda Estrangeira", "Valor Antecipado", "Preço Unitário da Antecipação",
    "Valor Resgate", "Preço de Exercício de Call", "Preço de Exercício de Put", "Prêmio Unitário de Call",
    "Prêmio Unitário de Put", "Cotação Barreira", "Tipo de Exercício", "Fonte de Informação", "Boletim",
    "Horário do Boletim", "Cotação para Vencimento", "Descrição de Cotação para Vencimento", "Fonte de Consulta",
    "TICKER", "QUANTIDADE", "DATA DE PAGAMENTO DO PRÊMIO", "MÉDIA OPÇÃO ASIÁTICA","Identifier"
)

colunas_boleta_termo = (
    "TIPO DO CONTRATO A TERMO", "CÓDIGO CONTRATO CETIP", "CÓDIGO DA CARTEIRA", "POSIÇÃO DO PARTICIPANTE",
    "CPF/CNPJ CLIENTE PARTE", "CONTRAPARTE", "CPF/CNPJ CLIENTE CONTRAPARTE", "CESTA GARANTIAS CONTRAPARTE",
    "CESTA GARANTIAS PARTE", "VALOR BASE (NOTIONAL)", "DATA DE OPERAÇÃO", "DATA DE REGISTRO", "DATA DE VENCIMENTO",
    "MOEDA", "BOLSA REFERÊNCIA", "COMMODITY", "TIPO (FORMA)", "QUANTIDADE", "UNIDADE DE NEGOCIAÇÃO",
    "PREÇO DA OPERAÇÃO", "PARIDADE PARA LIQUIDAÇÃO", "MÊS E ANO DO VENCIMENTO", "COTAÇÃO PARA AJUSTE/SPOT",
    "TAXA A TERMO (R$/MOEDA)", "TAXA MÉDIA PARA TERMO ASIÁTICO", "FONTE DE INFORMAÇÃO", "COTAÇÃO PARA O VENCIMENTO",
    "TIPO DE AJUSTE", "OBSERVAÇÃO", "Fator de Desconto","Identifier"
)

def highlight_duplicates(treeview, type):
       # Definição das cores
    COLOR_NEW = '#FFFFFF'  # White
    COLOR_PENDING_REVIEW = '#FFC9CA'  # Light Pink
    COLOR_APPROVED = '#CCCCFF'  # Light Purple
    COLOR_GENERATED = '#79BCFF'  # Light Blue
    COLOR_PENDING_MAKER_CHECKER = '#FFCC66'  # Light Orange
    COLOR_CONCLUDED = '#AFFFE4'  # Light Green
    COLOR_DUPLICATE = '#F08080'  # Light Coral

    # Configuração das tags
    treeview.tag_configure('new', background=COLOR_NEW)
    treeview.tag_configure('pending_review', background=COLOR_PENDING_REVIEW)
    treeview.tag_configure('approved', background=COLOR_APPROVED)
    treeview.tag_configure('generated', background=COLOR_GENERATED)
    treeview.tag_configure('pending_maker_checker', background=COLOR_PENDING_MAKER_CHECKER)
    treeview.tag_configure('concluded', background=COLOR_CONCLUDED)
    treeview.tag_configure('duplicate', background=COLOR_DUPLICATE)  

    if type == 'deals':
        # Dicionário para contar ocorrências dos valores no índice 0
        value_count = {}
        # Percorre os itens do Treeview e conta as ocorrências
        for item in treeview.get_children():
            row = treeview.item(item, 'values')
            value = row[0]
            if value in value_count:
                value_count[value].append(item)
            else:
                value_count[value] = [item]

        # Altera a cor de fundo se houver duplicados
        for value, items in value_count.items():
            if len(items) > 1:
                for item in items:
                    treeview.item(item, tags=('duplicate',))
            else:
                for item in items:
                    # Verifica o valor no índice -2 (penúltima coluna)
                    status = treeview.item(item, 'values')[-2] if len(treeview.item(item, 'values')) > 1 else None
                    if status == "New":
                        treeview.item(item, tags=('new',))
                    elif status == "Pending Review":
                        treeview.item(item, tags=('pending_review',))
                    elif status == "Approved":
                        treeview.item(item, tags=('approved',))
                    elif status == "Generated":
                        treeview.item(item, tags=('generated',))
                    elif status == "Pending Maker" or status == "Pending Checker":
                        treeview.item(item, tags=('pending_maker_checker',))
                    elif status == "Concluded":
                        treeview.item(item, tags=('concluded',))
    else:
        for item in treeview.get_children():
            # Verifica o valor no índice -2 (penúltima coluna)
            status = treeview.item(item, 'values')[-2] 
            if status == "New":
                treeview.item(item, tags=('new',))
            elif status == "Pending Review":
                treeview.item(item, tags=('pending_review',))
            elif status == "Approved":
                treeview.item(item, tags=('approved',))
            elif status == "Generated":
                treeview.item(item, tags=('generated',))
            elif status == "Pending Maker" or status == "Pending Checker":
                treeview.item(item, tags=('pending_maker_checker',))
            elif status == "Concluded":
                treeview.item(item, tags=('concluded',))



def ajustar_largura_colunas(treeview, colunas, tabview):
    try:
        font = Font()
        for coluna in colunas:
            # Calcular a largura do título da coluna
            titulo_largura = font.measure(coluna)
            # Calcular a largura máxima do conteúdo da coluna

            conteudo_largura = max(
                (font.measure(treeview.set(item, coluna)) for item in treeview.get_children('')),
                default=0
            )
            # Definir a largura final da coluna
            largura_final = max(titulo_largura, conteudo_largura) + 10 # Adicionar padding para clareza
            treeview.column(coluna, width=largura_final, anchor='center', stretch=0)

    except Exception as e:
        print(f"Erro ao ajustar largura das colunas: {e}")

    #Funções Auxiliares

def lookup(value, lookup_array, return_array):
    try:
        index = lookup_array.index(value)
        return return_array[index]
    except ValueError:
        return None

def strike_formula(market, strike, intermediate_ccy, commodities_ric, commodities_factor,tabview):
    fator_conversao = lookup(market, commodities_ric, commodities_factor)
    if fator_conversao is None:
        # Exibe uma mensagem de aviso se o fator de conversão não for encontrado
        messagebox.showwarning("Atenção", f"Favor cadastrar a commodity {market}!")

        # Seleciona a aba "Cadastro Commodities"
        tabview.set("Commodities")
        return  # Encerra a execução da função, semelhante ao Exit Sub
    
    fator_conversao = fator_conversao.replace(",", ".")  # Substitua a vírgula por um ponto
    fator_conversao = float(fator_conversao)  # Converta para float
    if strike.upper().startswith("TAS"):
        return " " * 20
    
    strike_cleaned = strike.replace(",", "")
    valor = float(strike_cleaned) *  (fator_conversao)            
    int_valor = int(valor)
    valor_str = str(int_valor)
    valor_completo_str = f"{valor:.9f}"
    decimal_index = valor_completo_str.find(".")
    decimal_part = valor_completo_str[decimal_index + 1:decimal_index + 9]
    return "0" * (12 - len(valor_str)) + valor_str + "." + decimal_part[:8] + "0" * (8 - len(decimal_part[:8]))

def strike_confirmation(market, strike, intermediate_ccy, commodities_ric, commodities_factor, tabview):
    fator_conversao = lookup(market, commodities_ric, commodities_factor)
    if fator_conversao is None:
        # Exibe uma mensagem de aviso se o fator de conversão não for encontrado
        messagebox.showwarning("Atenção", f"Favor cadastrar a commodity {market}!")

        # Seleciona a aba "Cadastro Commodities"
        tabview.set("Commodities")
        return  # Encerra a execução da função, semelhante ao Exit Sub
    
    fator_conversao = fator_conversao.replace(",", ".")  # Substitua a vírgula por um ponto
    fator_conversao = float(fator_conversao)  # Converta para float       
    strike_cleaned = strike.replace(",", "")
    valor = float(strike_cleaned) * (fator_conversao if intermediate_ccy != "BRR" else 1)
    
    # Format the value with thousands separator and four decimal places
    valor_completo_str = f"{valor:,.4f}"    
    return valor_completo_str

def pu_formula_opcao(market, premiumperunit, intermediate_ccy, commodities_ric, commodities_factor,tabview):
    fator_conversao = lookup(market, commodities_ric, commodities_factor)
    if fator_conversao is None:        # Exibe uma mensagem de aviso se o fator de conversão não for encontrado
        messagebox.showwarning("Atenção", f"Favor cadastrar a commodity {market}!")    

        # Seleciona a aba "Cadastro Commodities"
        tabview.set("Commodities")
        return  # Encerra a execução da função, semelhante ao Exit Sub
    
    fator_conversao = fator_conversao.replace(",", ".")  # Substitua a vírgula por um ponto
    fator_conversao = float(fator_conversao)  # Converta para float
    premiumperunit_cleaned = premiumperunit.replace(",", "")
    valor = float(premiumperunit_cleaned) * (fator_conversao)
    int_valor = int(valor)
    valor_str = str(int_valor)
    return "0" * (10 - len(valor_str)) + valor_str

def pudecimal_formula_opcao(market, premiumperunit, intermediate_ccy, commodities_ric, commodities_factor,tabview):
    fator_conversao = lookup(market, commodities_ric, commodities_factor)
    if fator_conversao is None:
        # Exibe uma mensagem de aviso se o fator de conversão não for encontrado
        messagebox.showwarning("Atenção", f"Favor cadastrar a commodity {market}!")    

        # Seleciona a aba "Cadastro Commodities"
        # Seleciona a aba "Cadastro Commodities"
        tabview.set("Commodities")
        return  # Encerra a execução da função, semelhante ao Exit Sub
    
    fator_conversao = fator_conversao.replace(",", ".")  # Substitua a vírgula por um ponto
    fator_conversao = float(fator_conversao)  # Converta para float

    premiumperunit_cleaned = premiumperunit.replace(",", "")
    valor = float(premiumperunit_cleaned) * (fator_conversao)
    valor_completo_str = f"{valor:.8f}"
    decimal_index = valor_completo_str.find(".")
    decimal_part = valor_completo_str[decimal_index + 1:decimal_index + 9]
    return decimal_part + "0" * (8 - len(decimal_part))

def strike_formula_opcao(market, strike, intermediate_ccy, commodities_ric, commodities_factor,tabview):
    fator_conversao = lookup(market, commodities_ric, commodities_factor)
    if fator_conversao is None:
        # Exibe uma mensagem de aviso se o fator de conversão não for encontrado
        messagebox.showwarning("Atenção", f"Favor cadastrar a commodity {market}!")    

        # Seleciona a aba "Cadastro Commodities"
        tabview.set("Commodities")
        return  # Encerra a execução da função, semelhante ao Exit Sub
    
    fator_conversao = fator_conversao.replace(",", ".")  # Substitua a vírgula por um ponto
    fator_conversao = float(fator_conversao)  # Converta para float
    strike_cleaned = strike.replace(",", "")
    valor = float(strike_cleaned) * (fator_conversao)
    int_valor = int(valor)
    valor_str = str(int_valor)
    return "0" * (10 - len(valor_str)) + valor_str

def decimal_formula_opcao(market, strike, intermediate_ccy, commodities_ric, commodities_factor,tabview):
    fator_conversao = lookup(market, commodities_ric, commodities_factor)
    if fator_conversao is None:
        # Exibe uma mensagem de aviso se o fator de conversão não for encontrado
        messagebox.showwarning("Atenção", f"Favor cadastrar a commodity {market}!")     

        # Seleciona a aba "Cadastro Commodities"
        tabview.set("Commodities")
        return  # Encerra a execução da função, semelhante ao Exit Sub
    
    fator_conversao = fator_conversao.replace(",", ".")  # Substitua a vírgula por um ponto
    fator_conversao = float(fator_conversao)  # Converta para float
    strike_cleaned = strike.replace(",", "")
    valor = float(strike_cleaned) * (fator_conversao)
    valor_completo_str = f"{valor:.8f}"
    decimal_index = valor_completo_str.find(".")
    decimal_part = valor_completo_str[decimal_index + 1:decimal_index + 9]
    return decimal_part + "0" * (8 - len(decimal_part))

def tas_formula(market, strike, intermediate_ccy, commodities_ric, commodities_factor):
    if strike.upper().startswith("TAS"):
        try:
            strike_cleaned = strike.replace("TAS", "").strip()
            valor = float(strike_cleaned)
            if valor < 0:
                int_valor = int(valor)
                valor_str = str(abs(int_valor))
                valor_completo_str = f"{abs(valor):.8f}"
                decimal_index = valor_completo_str.find(".")
                decimal_part = valor_completo_str[decimal_index + 1:decimal_index + 9]
                return "-" + "0" * (3 - len(valor_str)) + valor_str + "." + decimal_part + "0" * (8 - len(decimal_part))
            else:
                int_valor = int(valor)
                valor_str = str(int_valor)
                valor_completo_str = f"{valor:.8f}"
                decimal_index = valor_completo_str.find(".")
                decimal_part = valor_completo_str[decimal_index + 1:decimal_index + 9]
                return "0" * (4 - len(valor_str)) + valor_str + "." + decimal_part + "0" * (8 - len(decimal_part))
        except Exception:
            return "0" * 12
    else:
        return " " * 12
    
def safe_date_conversion(date_str):
    try:
        return datetime.strptime(date_str, "%d-%b-%Y").strftime("%Y%m%d")
    except (ValueError, TypeError):
        return " " * 8

def safe_date_conversion_confirma(date_str):
    try:
        return datetime.strptime(date_str, "%d-%b-%Y").strftime("%d-%m-%Y")
    except (ValueError, TypeError):
        return " " * 8

def safe_date_conversion_y_m_d(date_str):
    try:
        return datetime.strptime(date_str, "%d-%b-%Y").strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return " " * 8 

def safe_date_conversion_dd_MM(date_str):
    try:
        return datetime.strptime(date_str, "%d-%b-%Y").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return " " * 8
    
def safe_date_conversion_yyyymmaa(date_str):
    try:
        return datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y%m%d")
    except (ValueError, TypeError):
        return " " * 8
    
def data_por_extenso(date_str):
    # Mapeamento dos meses em inglês para português
    meses_portugues = {
        "Jan": "Janeiro", "Feb": "Fevereiro", "Mar": "Março", "Apr": "Abril",
        "May": "Maio", "Jun": "Junho", "Jul": "Julho", "Aug": "Agosto",
        "Sep": "Setembro", "Oct": "Outubro", "Nov": "Novembro", "Dec": "Dezembro"
    }   

    try:
        # Converter a string de data para um objeto datetime
        data = datetime.strptime(date_str, "%d-%b-%Y")
        # Obter o nome do mês em português
        mes_portugues = meses_portugues[data.strftime("%b")]
        # Formatar a data por extenso
        return f"{data.day} de {mes_portugues} de {data.year}"

    except (ValueError, TypeError):

        return "Data inválida"

def ordenar_por(treeview, cols, reverse):
    # Obter os dados das colunas especificadas e as chaves dos itens
    l = [(tuple(treeview.set(k, col) for col in cols), k) for k in treeview.get_children('')]
    
    # Função de chave para ordenação
    def sort_key(t):
        keys = []
        for i, col in enumerate(cols):
            value = t[0][i]
            if col == "SettlementDate":
                try:
                    # Converter a string de data em objeto datetime
                    value = datetime.strptime(value, "%d-%b-%Y")
                except ValueError:
                    value = datetime.min  # Use uma data mínima se a conversão falhar
            keys.append(value)
        return tuple(keys)
    
    # Ordenar a lista usando a chave de ordenação
    l.sort(key=sort_key, reverse=reverse)
    
    # Reordenar os itens no Treeview
    for index, (_, k) in enumerate(l):
        treeview.move(k, '', index)
    
    # Alterar a função de ordenação para a próxima vez que o cabeçalho for clicado
    treeview.heading(cols[0], command=lambda: ordenar_por(treeview, cols, not reverse))

def carregar_counterparty_combobox_opcao():
    try:
        # Tente abrir a conexão com o banco de dados
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()

        # Use DISTINCT para obter valores únicos, filtrar strings vazias e ordenar alfabeticamente
        cursor.execute("""
            SELECT DISTINCT ACCRONYM 
            FROM refData 
            WHERE ACCRONYM <> '' 
            ORDER BY ACCRONYM ASC
        """)

        # Buscar todos os resultados e armazená-los em uma lista
        cntpy_accronym = [row[0] for row in cursor.fetchall()]

    except sqlite3.Error as e:
        # Exibir uma mensagem de erro se ocorrer um problema com o banco de dados
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
        cntpy_accronym = []  # Retornar uma lista vazia em caso de erro

    finally:
        # Fechar a conexão, se foi aberta
        if conn:
            conn.close()

    # Retornar a lista de acrônimos ou usá-la conforme necessário
    return cntpy_accronym


def carregar_counterparty_combobox_termo():
    try:
        # Tente abrir a conexão com o banco de dados
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()

        # Use DISTINCT para obter valores únicos, filtrar strings vazias e ordenar alfabeticamente
        cursor.execute("""
            SELECT DISTINCT ACCRONYM 
            FROM refData 
            WHERE ACCRONYM <> '' 
            ORDER BY ACCRONYM ASC
        """)

        # Buscar todos os resultados e armazená-los em uma lista
        cntpy_accronym = [row[0] for row in cursor.fetchall()]

    except sqlite3.Error as e:
        # Exibir uma mensagem de erro se ocorrer um problema com o banco de dados
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
        cntpy_accronym = []  # Retornar uma lista vazia em caso de erro

    finally:
        # Fechar a conexão, se foi aberta
        if conn:
            conn.close()

    # Retornar a lista de acrônimos ou usá-la conforme necessário
    return cntpy_accronym

def carregar_combobox_bolsa():
    try:
        # Tente abrir a conexão com o banco de dados
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()

        # Use DISTINCT para obter valores únicos, filtrar strings vazias e ordenar alfabeticamente
        cursor.execute("""
            SELECT DISTINCT BolsadeNegociacao 
            FROM COMMODITIES 
            WHERE BolsadeNegociacao <> '' 
            ORDER BY BolsadeNegociacao ASC
        """)
        bolsa_negociacao = [row[0] for row in cursor.fetchall()]

        # Inserir "BOLSA DE NEGOCIAÇÃO" no início da lista
        bolsa_negociacao.insert(0, "BOLSA DE NEGOCIAÇÃO")

    except sqlite3.Error as e:
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
        bolsa_negociacao = ["BOLSA DE NEGOCIAÇÃO"]  # Retornar valor padrão em caso de erro

    finally:
        if conn:
            conn.close()

    return bolsa_negociacao

def carregar_combobox_mercadoria(combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_mercadoria): 
    bolsa = combobox_bolsa_negociacao.get()
    if bolsa == "BOLSA DE NEGOCIÇÃO": 
        mercadoria = ["MERCADORIA"] 
    else:
        try:
            conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT Mercadoria 
                FROM COMMODITIES 
                WHERE BolsadeNegociacao <> '' AND BolsadeNegociacao =  ? AND Mercadoria <> ''
                ORDER BY Mercadoria ASC
            """, (bolsa,))
            mercadoria = [row[0] for row in cursor.fetchall()]        

        except sqlite3.Error as e:
            conn.close()
            messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
            mercadoria = ["MERCADORIA"]

        finally:
            if conn:
                conn.close()

    if len(mercadoria) > 1:
        mercadoria.insert(0, "MERCADORIA") 
        combobox_mercadoria.set("MERCADORIA")
        scrollable_dropdown_mercadoria.configure(values=mercadoria)
    elif len(mercadoria) == 1:
        scrollable_dropdown_mercadoria.configure(values=mercadoria)             
        combobox_mercadoria.set(mercadoria[0])     

def carregar_combobox_tipo(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_tipo): 
    bolsa = combobox_bolsa_negociacao.get()
    mercadoria = combobox_mercadoria.get()    
    if mercadoria == "MERCADORIA": 
        tipo = ["TIPO"] 
    else:
        try:
            conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT Tipo 
                FROM COMMODITIES 
                WHERE BolsadeNegociacao <> '' AND BolsadeNegociacao = ? AND Mercadoria <> '' and Mercadoria = ? and Tipo <> ''
                ORDER BY Tipo ASC
            """, (bolsa, mercadoria,))
            tipo = [row[0] for row in cursor.fetchall()]

        except sqlite3.Error as e:
            conn.close()
            messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
            tipo = ["TIPO"]

        finally:
            if conn:
                conn.close()

    if len(tipo) > 1:
        tipo.insert(0, "TIPO")
        combobox_tipo.set("TIPO")
        scrollable_dropdown_tipo.configure(values=tipo)
    elif len(tipo) == 1:
        scrollable_dropdown_tipo.configure(values=tipo)
        combobox_tipo.set(tipo[0])

def carregar_combobox_unidade(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, scrollable_dropdown_unidade): 
    bolsa = combobox_bolsa_negociacao.get()
    mercadoria = combobox_mercadoria.get()
    tipo = combobox_tipo.get()
    
    if tipo == "TIPO": 
        unidade = ["UNIDADE DE NEGOCIAÇÃO"] 
    else:
        try:
            conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT UnidadedeNegociacao 
                FROM COMMODITIES 
                WHERE BolsadeNegociacao <> '' AND BolsadeNegociacao = ? AND Mercadoria <> '' AND Mercadoria = ? AND Tipo = ? AND Tipo <> '' AND UnidadedeNegociacao <> '' 
                ORDER BY UnidadedeNegociacao ASC
            """, (bolsa, mercadoria, tipo,))
            unidade = [row[0] for row in cursor.fetchall()]

        except sqlite3.Error as e:
            conn.close()
            messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
            unidade = ["UNIDADE DE NEGOCIAÇÃO"]

        finally:
            if conn:
                conn.close()

    if len(unidade) > 1:
        unidade.insert(0, "UNIDADE DE NEGOCIAÇÃO")
        scrollable_dropdown_unidade.configure(values=unidade)
    elif len(unidade) == 1:
        scrollable_dropdown_unidade.configure(values=unidade)        
        combobox_unidade.set(unidade[0])

def carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda): 
    bolsa = combobox_bolsa_negociacao.get()
    mercadoria = combobox_mercadoria.get()
    tipo = combobox_tipo.get()
    unidade = combobox_unidade.get()
    
    if unidade == "UNIDADE DE NEGOCIAÇÃO": 
        moeda = ["MOEDA"] 
    else:
        try:
            conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT Moeda 
                FROM COMMODITIES 
                WHERE BolsadeNegociacao <> '' AND BolsadeNegociacao = ? AND Mercadoria <> '' AND Mercadoria = ? AND Tipo = ? AND Tipo <> '' AND UnidadedeNegociacao <> '' AND UnidadedeNegociacao = ?  AND Moeda <> ''
                ORDER BY Moeda ASC
            """, (bolsa, mercadoria, tipo, unidade,))
            moeda = [row[0] for row in cursor.fetchall()]

        except sqlite3.Error as e:
            conn.close()
            messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
            moeda = ["MOEDA"]

        finally:
            if conn:
                conn.close()

    if len(moeda) > 1:
        moeda.insert(0, "MOEDA")
        scrollable_dropdown_moeda.configure(values=moeda)
    elif len(moeda) == 1:
        scrollable_dropdown_moeda.configure(values=moeda)        
        combobox_moeda.set(moeda[0])

def carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao): 
    bolsa = combobox_bolsa_negociacao.get()
    mercadoria = combobox_mercadoria.get()
    tipo = combobox_tipo.get()
    unidade = combobox_unidade.get()
    moeda = combobox_moeda.get()
    
    if moeda == "MOEDA": 
        conversao = ["FATOR DE CONVERSÃO"] 
    else:
        try:
            conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
            cursor = conn.cursor()

            cursor.execute("""
                SELECT DISTINCT FatordeConversão 
                FROM COMMODITIES 
                WHERE BolsadeNegociacao <> '' AND BolsadeNegociacao = ? AND Mercadoria <> '' AND Mercadoria = ? AND Tipo = ? AND Tipo <> '' AND UnidadedeNegociacao <> '' AND UnidadedeNegociacao = ?  AND Moeda <> '' AND Moeda = ? AND FatordeConversão <> ''
                ORDER BY FatordeConversão ASC
            """, (bolsa, mercadoria, tipo, unidade, moeda,))
            conversao = [row[0] for row in cursor.fetchall()]

        except sqlite3.Error as e:
            conn.close()
            messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
            conversao = ["FATOR DE CONVERSÃO"]

        finally:
            if conn:
                conn.close()

    if len(conversao) > 1:
        conversao.insert(0, "FATOR DE CONVERSÃO")
        scrollable_dropdown_conversao.configure(values=conversao)
    elif len(conversao) == 1:
        scrollable_dropdown_conversao.configure(values=conversao)        
        combobox_conversao.set(conversao[0])
 
def refresh_table_commodities(tree):
    carregar_dados_commodities(tree)
    messagebox.showinfo("Sucesso","Base de Commodities atualizada!")
    
def carregar_dados_commodities(tree):
    # Limpar o Treeview
    tree.delete(*tree.get_children())
    try:
        # Use the existing database connection
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")        
        cursor = conn.cursor()

        # Selecionar todos os dados da tabela Commodities
        cursor.execute("SELECT * FROM Commodities")
        rows = cursor.fetchall()

        # Inserir os dados no Treeview
        for row in rows:
            tree.insert("", "end", values=row)

        # Verificar se a tabela temporária existe e tem dados
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='temp_table_commodities'")
        if cursor.fetchone():
            cursor.execute("SELECT * FROM temp_table_commodities")
            temp_rows = cursor.fetchall()

            # Criar um dicionário para mapear AtivoSubjacenteRIC aos dados da tabela temporária
            temp_data_dict = {row[0]: row for row in temp_rows}

            # Sobrescrever os dados no Treeview com os dados da tabela temporária
            for item in tree.get_children():
                item_values = tree.item(item, "values")
                ativo_subjacente = item_values[0]
                if ativo_subjacente in temp_data_dict:
                    # Atualizar os valores no Treeview
                    tree.item(item, values=temp_data_dict[ativo_subjacente])
    except sqlite3.Error as e:
            conn.close()
            messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
            
    finally:
            if conn:
                conn.close()
    

    # Atualizar a lista de dados original
    global tabela_commodities_data
    tabela_commodities_data = [tree.item(item, "values") for item in tree.get_children()]
    
def rollback_selected_items_comitentes(tree):
    """Revert the last changes made to the selected comitentes."""
    # Use the existing database connection
    
    try:
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()
        selected_items = tree.selection()
        for item in selected_items:
            item_id = tree.item(item, 'values')
            cnpj = item_id[0]
            # Delete the entry from the temporary table
            cursor.execute("DELETE FROM temp_table_comitentes WHERE CNPJ = ?", (cnpj,))        
        conn.commit()        
        carregar_dados_comitentes(tree)
        conn.close()
        messagebox.showinfo("Rollback", "As alterações de comitentes foram revertidas.")
    except sqlite3.Error as e:
        conn.close()
        messagebox.showwarning("Erro de Banco de Dados", f"Ocorreu um erro ao reverter as alterações: {e}")
        
def rollback_selected_items_commodities(tree):
    """Revert the last changes made to the selected items."""
    # Use the existing database connection
    
    try:
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()
        selected_items = tree.selection()
        for item in selected_items:
            item_id = tree.item(item, 'values')
            ativo_subjacente = item_id[0]
            # Delete the entry from the temporary table
            cursor.execute("DELETE FROM temp_table_commodities WHERE AtivoSubjacenteRIC = ?", (ativo_subjacente,))        
        conn.commit()
        carregar_dados_commodities(tree)
        conn.close()
        messagebox.showinfo("Rollback", "As alterações de Commodities foram revertidas.")
    except sqlite3.Error as e:
        conn.close()
        messagebox.showwarning("Erro de Banco de Dados", f"Ocorreu um erro ao reverter as alterações: {e}")

def delete_commodities(tree):  
    resposta = messagebox.askyesnocancel("Confirmação", "Deseja seguir com a atualização?")    
    if resposta == "no":
        return  # Se o usuário clicar em "No", encerra a função
    elif resposta == "cancel":
        rollback_selected_items_commodities(tree)
        return    

    selected_items = tree.selection()
    
    if not selected_items:
        messagebox.showwarning("Attention!", "No item selected.")
        return

    # Use the existing database connection
    
    try:
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")
        
        # Check if the temporary table exists, if not, create it
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS temp_table_commodities AS
            SELECT * FROM Commodities WHERE 0
        """)

        for item in selected_items:
            item_id = tree.item(item, 'values')
            SID = getpass.getuser()
            SID = SID[0].upper() + SID[1:] 
            
            if SID == item_id[11] and item_id[10] == "PENDING INACTIVE":
                messagebox.showwarning("Attention!", "Different SID must perform the Checker")
                return
            
            ativo_subjacente = item_id[0]
            bolsa_negociacao = item_id[1]
            indice_valorizacao = item_id[2]
            mes_vencimento = item_id[3]
            ano_vencimento = item_id[4]
            tipo = item_id[5]
            unidade_negociacao = item_id[6]
            moeda = item_id[7]
            fator_conversao = item_id[8]
            mercadoria = item_id[9]
            status = "PENDING INACTIVE" if item_id[10] == "ACTIVE" else "INACTIVE"
            timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            if item_id[10] == "ACTIVE":
                # Insert the original data into the temporary table
                cursor.execute("""
                    INSERT INTO temp_table_commodities (
                        AtivoSubjacenteRIC, BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo,
                        UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria, Status, SID, timestamp
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (ativo_subjacente, bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo,
                      unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp))
                # Mark the item as pending inactive in the main table
                cursor.execute("""
                    UPDATE Commodities SET
                    Status = ?,
                    SID = ?,
                    timestamp = ?
                    WHERE AtivoSubjacenteRIC = ?
                """, (status, SID, timestamp, ativo_subjacente))
            else:                         
                # Update the main table and delete the entry from the temporary table
                cursor.execute("""
                    UPDATE Commodities SET
                    BolsadeNegociacao = COALESCE(NULLIF(?, ''), BolsadeNegociacao),
                    IndiceValorizacao = COALESCE(NULLIF(?, ''), IndiceValorizacao),
                    MesVencimento = COALESCE(NULLIF(?, ''), MesVencimento),
                    AnoVencimento = COALESCE(NULLIF(?, ''), AnoVencimento),
                    Tipo = COALESCE(NULLIF(?, ''), Tipo),
                    UnidadedeNegociacao = COALESCE(NULLIF(?, ''), UnidadedeNegociacao),
                    Moeda = COALESCE(NULLIF(?, ''), Moeda),
                    FatordeConversão = COALESCE(NULLIF(?, ''), FatordeConversão),
                    Mercadoria = COALESCE(NULLIF(?, ''), Mercadoria),
                    Status = COALESCE(NULLIF(?, ''), Status),
                    SID = COALESCE(NULLIF(?, ''), SID),
                    timestamp = COALESCE(NULLIF(?, ''), timestamp)
                    WHERE AtivoSubjacenteRIC = ?
                """, (bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo, unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp, ativo_subjacente))
                cursor.execute("DELETE FROM temp_table_commodities WHERE AtivoSubjacenteRIC = ?", (ativo_subjacente,))
        
        conn.commit()
        conn.close()
        carregar_dados_commodities(tree)
        messagebox.showinfo("Atualização", "Commodity atualizada com sucesso!")
    except sqlite3.Error as e:
        conn.rollback()
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")

    # Dicionário para armazenar o placeholder_text de cada campo
    placeholders = {
        entry_ativo_subjacente: "ATIVO SUBJACENTE / RIC",        
        entry_indice_valorizacao: "INDICE VALORIZACAO",            
        entry_filtro_commodities: "Ativo Subjacente / RIC",
        entry_filtro_status_commodities: "Status"        
    }

    # Função para limpar um campo de entrada apenas se houver texto do usuário
    def limpar_entry(entry):
        current_text = entry.get()
        placeholder_text = placeholders[entry]
        if current_text and current_text != placeholder_text:  # Verifica se há texto do usuário
            entry.delete(0, tk.END)

    # Limpar os campos de entrada após o cadastro
    limpar_entry(entry_ativo_subjacente)    
    limpar_entry(entry_indice_valorizacao)     
    limpar_entry(entry_filtro_commodities)
    limpar_entry(entry_filtro_status_commodities)
    
def cadastrar_commodities():
    # Perguntar ao usuário se deseja seguir com o cadastro
    resposta = messagebox.askyesno("Confirmação", "Deseja seguir com o cadastro?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função

    # Obter os valores dos campos de entrada
    ativo_subjacente = entry_ativo_subjacente.get()
    bolsa_negociacao = combobox_bolsa_negociacao.get()
    indice_valorizacao = entry_indice_valorizacao.get()
    mes_vencimento_abreviado = combobox_mes_vencimento.get()
    ano_vencimento = combobox_ano_vencimento.get()
    tipo = combobox_tipo.get()
    unidade_negociacao = combobox_unidade.get()
    moeda = combobox_moeda.get()
    fator_conversao = combobox_conversao.get()
    mercadoria = combobox_mercadoria.get()
    status = "PENDING"
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    SID = getpass.getuser()
    SID = SID[0].upper() + SID[1:]

    # Mapeamento de meses abreviados para números
    meses = {
        "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
    }

    # Converter o mês de vencimento para número
    mes_vencimento = meses.get(mes_vencimento_abreviado, 0)  # Retorna 0 se o mês não for encontrado

    # Use the existing database connection
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    try:
        cursor = conn.cursor()

        # Verificar se o ativo já existe na tabela Commodities
        cursor.execute("SELECT * FROM Commodities WHERE AtivoSubjacenteRIC = ?", (ativo_subjacente,))
        existing_record = cursor.fetchone()

        if existing_record:
            # Perguntar ao usuário se deseja seguir com a atualização
            resposta = messagebox.askyesno("Confirmação", "Cadastro já existente. Deseja seguir com a atualização?")
            if not resposta:
                return  # Se o usuário clicar em "No", encerra a função           

            # Verificar se o ativo já existe na tabela temp_table_commodities
            cursor.execute("SELECT * FROM temp_table_commodities WHERE AtivoSubjacenteRIC = ?", (ativo_subjacente,))
            temp_record = cursor.fetchone()

            if temp_record:
                # Atualizar a tabela temp_table_commodities
                cursor.execute("""
                    UPDATE temp_table_commodities SET
                    BolsadeNegociacao = COALESCE(NULLIF(?, ''), BolsadeNegociacao),
                    IndiceValorizacao = COALESCE(NULLIF(?, ''), IndiceValorizacao),
                    MesVencimento = COALESCE(NULLIF(?, ''), MesVencimento),
                    AnoVencimento = COALESCE(NULLIF(?, ''), AnoVencimento),
                    Tipo = COALESCE(NULLIF(?, ''), Tipo),
                    UnidadedeNegociacao = COALESCE(NULLIF(?, ''), UnidadedeNegociacao),
                    Moeda = COALESCE(NULLIF(?, ''), Moeda),
                    FatordeConversão = COALESCE(NULLIF(?, ''), FatordeConversão),
                    Mercadoria = COALESCE(NULLIF(?, ''), Mercadoria),
                    Status = COALESCE(NULLIF(?, ''), Status),
                    SID = COALESCE(NULLIF(?, ''), SID),
                    timestamp = COALESCE(NULLIF(?, ''), timestamp)
                    WHERE AtivoSubjacenteRIC = ?
                """, (bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo, unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp, ativo_subjacente))
            else:
                # Inserir na tabela temp_table_commodities
                cursor.execute("""
                    INSERT INTO temp_table_commodities (AtivoSubjacenteRIC, BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo, UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria, Status, SID, TIMESTAMP)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (ativo_subjacente, bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo, unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp))
            conn.commit()
            messagebox.showinfo("Atualização", "Commodity atualizada com sucesso!")

            # Atualizar a linha existente no Treeview
            for item in tabela_commodities.get_children():
                if tabela_commodities.item(item, "values")[0] == ativo_subjacente:
                    tabela_commodities.item(item, values=(
                        ativo_subjacente,
                        bolsa_negociacao or tabela_commodities.item(item, "values")[1],
                        indice_valorizacao or tabela_commodities.item(item, "values")[2],
                        mes_vencimento or tabela_commodities.item(item, "values")[3],
                        ano_vencimento or tabela_commodities.item(item, "values")[4],
                        tipo or tabela_commodities.item(item, "values")[5],
                        unidade_negociacao or tabela_commodities.item(item, "values")[6],
                        moeda or tabela_commodities.item(item, "values")[7],
                        fator_conversao or tabela_commodities.item(item, "values")[8],
                        mercadoria or tabela_commodities.item(item, "values")[9],
                        status,
                        SID,
                        timestamp
                    ))
                    break
        else:
            # Inserir os dados na tabela Commodities
            cursor.execute("""
                INSERT INTO Commodities (AtivoSubjacenteRIC, BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo, UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria, Status, SID, TIMESTAMP)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (ativo_subjacente, bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo, unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp))

            # Inserir os dados na tabela temp_table_commodities
            cursor.execute("""
                INSERT INTO temp_table_commodities (AtivoSubjacenteRIC, BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo, UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria, Status, SID, TIMESTAMP)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (ativo_subjacente, bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo, unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp))
            conn.commit()
            messagebox.showinfo("Sucesso", "Commodity cadastrada com sucesso!")

            # Inserir os dados na Treeview
            tabela_commodities.insert("", "end", values=(ativo_subjacente, bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo, unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp))

        # Recarregar os dados do banco de dados e atualizar a lista de dados original
        carregar_dados_commodities(tabela_commodities)
        conn.close()

    except sqlite3.Error as e:
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
        

    # Dicionário para armazenar o placeholder_text de cada campo
    placeholders = {
        entry_ativo_subjacente: "ATIVO SUBJACENTE / RIC",       
        entry_indice_valorizacao: "INDICE VALORIZACAO",
    }

    # Função para limpar um campo de entrada apenas se houver texto do usuário
    def limpar_entry(entry):
        current_text = entry.get()
        placeholder_text = placeholders[entry]
        if current_text and current_text != placeholder_text:  # Verifica se há texto do usuário
            entry.delete(0, tk.END)

    # Limpar os campos de entrada após o cadastro
    limpar_entry(entry_ativo_subjacente)    
    limpar_entry(entry_indice_valorizacao)

    # Resetar os comboboxes para seus valores padrão
    combobox_bolsa_negociacao.set("BOLSA DE NEGOCIACAO")
    combobox_mercadoria.set("MERCADORIA")
    combobox_tipo.set("TIPO")
    combobox_unidade.set("UNIDADE DE NEGOCIÇÃO")
    combobox_moeda.set("MOEDA")
    combobox_conversao.set("FATOR DE CONVERSÃO")
    combobox_mes_vencimento.set("MES VENCIMENTO")
    combobox_ano_vencimento.set("ANO VENCIMENTO")

def refresh_table_comitentes(tree):
    carregar_dados_comitentes(tree)
    messagebox.showinfo("Sucesso","Base de comitentes atualizada!")
    
def atualizar_commodities(tree):
    resposta = messagebox.askyesnocancel("Confirmação", "Deseja seguir com a atualização?")
    if resposta == False:
        return
    elif resposta == None:
        rollback_selected_items_commodities(tree)
        return

    selected_items = tree.selection()

    if not selected_items:
        messagebox.showwarning("Attention!", "No item selected.")
        return

    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS temp_table_commodities AS
            SELECT * FROM Commodities WHERE 0
        """)

        for item in selected_items:
            item_id = tree.item(item, 'values')
            SID = getpass.getuser()
            SID = SID[0].upper() + SID[1:]

            if SID == item_id[11] and item_id[10] == "PENDING":
                messagebox.showwarning("Attention!", "Different SID must perform the Checker")
                return

            ativo_subjacente = item_id[0]
            bolsa_negociacao = item_id[1]
            indice_valorizacao = item_id[2]
            mes_vencimento = item_id[3]
            ano_vencimento = item_id[4]
            tipo = item_id[5]
            unidade_negociacao = item_id[6]
            moeda = item_id[7]
            fator_conversao = item_id[8]
            mercadoria = item_id[9]
            status = "PENDING" if item_id[10] == "ACTIVE" else "ACTIVE"
            timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            if item_id[10] == "ACTIVE":
                cursor.execute("""
                    INSERT INTO temp_table_commodities (
                        AtivoSubjacenteRIC, BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo,
                        UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria, Status, SID, timestamp
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (ativo_subjacente, bolsa_negociacao, indice_valorizacao, mes_vencimento, ano_vencimento, tipo,
                      unidade_negociacao, moeda, fator_conversao, mercadoria, status, SID, timestamp))
                carregar_dados_commodities(tree)
            else:
                cursor.execute("""
                    SELECT BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo,
                        UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria
                    FROM temp_table_commodities
                    WHERE AtivoSubjacenteRIC = ?
                """, (ativo_subjacente,))

                result = cursor.fetchone()

                if result:
                    (BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo,
                    UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria) = result

                    cursor.execute("""
                        UPDATE Commodities SET
                        BolsadeNegociacao = ?,
                        IndiceValorizacao = ?,
                        MesVencimento = ?,
                        AnoVencimento = ?,
                        Tipo = ?,
                        UnidadedeNegociacao = ?,
                        Moeda = ?,
                        FatordeConversão = ?,
                        Mercadoria = ?,
                        Status = ?,
                        SID = ?,
                        timestamp = ?
                        WHERE AtivoSubjacenteRIC = ?
                    """, (BolsadeNegociacao, IndiceValorizacao, MesVencimento, AnoVencimento, Tipo,
                        UnidadedeNegociacao, Moeda, FatordeConversão, Mercadoria, status, SID, timestamp, ativo_subjacente))
                cursor.execute("DELETE FROM temp_table_commodities WHERE AtivoSubjacenteRIC = ?", (ativo_subjacente,))

        conn.commit()
        carregar_dados_commodities(tree)
        conn.close()
        messagebox.showinfo("Atualização", "Commodity atualizada com sucesso!")

    except sqlite3.Error as e:
        conn.rollback()
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
    
    # Dicionário para armazenar o placeholder_text de cada campo
    placeholders = {
        entry_ativo_subjacente: "ATIVO SUBJACENTE / RIC",        
        entry_indice_valorizacao: "INDICE VALORIZACAO",
        entry_filtro_commodities: "Ativo Subjacente / RIC",
        entry_filtro_status_commodities: "Status"        
    }

    # Função para limpar um campo de entrada apenas se houver texto do usuário
    def limpar_entry(entry):
        current_text = entry.get()
        placeholder_text = placeholders[entry]
        if current_text and current_text != placeholder_text:  # Verifica se há texto do usuário
            entry.delete(0, tk.END)

    # Limpar os campos de entrada após o cadastro
    limpar_entry(entry_ativo_subjacente)    
    limpar_entry(entry_indice_valorizacao)
    limpar_entry(entry_filtro_commodities)
    limpar_entry(entry_filtro_status_commodities)

def atualizar_comitentes(tree):  
    resposta = messagebox.askyesnocancel("Confirmação", "Deseja seguir com a atualização?")    
    if resposta == False:
        return  # Se o usuário clicar em "No", encerra a função
    elif resposta == None:
        rollback_selected_items_comitentes(tree)
        return    

    selected_items = tree.selection()
    
    if not selected_items:
        messagebox.showwarning("Attention!", "No item selected.")
        return

    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")
        
        # Check if the temporary table exists, if not, create it
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS temp_table_comitentes AS
            SELECT * FROM refData WHERE 0
        """)

        for item in selected_items:
            item_id = tree.item(item, 'values')
            SID = getpass.getuser()
            SID = SID[0].upper() + SID[1:] 
            
            if SID == item_id[13] and item_id[12] == "PENDING":
                messagebox.showwarning("Attention!", "Different SID must perform the Checker")
                return
            
            cnpj = item_id[0]
            ente = item_id[1]
            spn = item_id[2]
            eci = item_id[3]
            cliente  = item_id[4]
            accronym = item_id[5]
            conta_cetip = item_id[6]
            cgd = item_id[7]
            mapeamento_confirmacoes = item_id[8]
            banco = item_id[9]
            ag = item_id[10]
            cc = item_id[11]
            status = "PENDING" if item_id[12] == "ACTIVE" else "ACTIVE"
            timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            if item_id[12] == "ACTIVE":
                # Insert the data from the Treeview into the temporary table
                cursor.execute("""
                    INSERT INTO temp_table_comitentes (
                        CNPJ, ENTE, SPN, ECI, CLIENTE, ACCRONYM, CONTACETIP, CGD, MAPEAMENTOCONFIRMAÇÕES,
                        BANCO, AG, CC, STATUS, SID, TIMESTAMP
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (cnpj, ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes,
                      banco, ag, cc, status, SID, timestamp))
                # Load data from the temporary table into the treeview
                carregar_dados_comitentes(tree)
            else:
                # Update the main table and delete the entry from the temporary table
                cursor.execute("""
                    SELECT CNPJ, ENTE, SPN, ECI, CLIENTE, ACCRONYM, CONTACETIP, CGD, MAPEAMENTOCONFIRMAÇÕES,
                        BANCO, AG, CC
                    FROM temp_table_comitentes
                    WHERE CNPJ = ?
                """, (cnpj,))

                # Obtenha os resultados da consulta
                result = cursor.fetchone()

                if result:
                    # Descompacte os resultados
                    (cnpj, ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes,
                    banco, ag, cc) = result

                    # Atualize a tabela refData com os valores extraídos
                       # Atualize a tabela Commodities com os valores extraídos                 
                    cursor.execute("""
                        UPDATE refData SET
                        ENTE = ?,
                        SPN = ?,
                        ECI = ?,
                        CLIENTE = ?,
                        ACCRONYM = ?,
                        CONTACETIP = ?,
                        CGD = ?,
                        MAPEAMENTOCONFIRMAÇÕES = ?,
                        BANCO = ?,
                        AG = ?,
                        CC = ?,
                        Status = ?,
                        SID = ?,
                        timestamp = ?
                        WHERE CNPJ = ?
                    """, (ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes, banco, ag, cc, status, SID, timestamp, cnpj))      
                cursor.execute("DELETE FROM temp_table_comitentes WHERE CNPJ = ?", (cnpj,))
        conn.commit()
        carregar_dados_comitentes(tree)
        messagebox.showinfo("Atualização", "Comitentes atualizado com sucesso!")
    except sqlite3.Error as e:
        conn.rollback()
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
    finally:
        conn.close()
    # Dicionário para armazenar o placeholder_text de cada campo
    placeholders_comitentes = {
        entry_cnpj: "CNPJ",
        entry_ente: "ENTE",
        entry_spn: "SPN",
        entry_eci: "ECI",
        entry_cliente: "CLIENTE",
        entry_accronym: "ACCRONYM",
        entry_conta_cetip: "CONTA CETIP",
        entry_cgd: "CGD",
        entry_mapeamento_confirmacoes: "MAPEAMENTO CONFIRMAÇÕES",
        entry_banco: "BANCO",
        entry_ag: "AG",
        entry_cc: "CC",
        entry_filtro_comitentes: "Filtrar por Cliente",
        entry_filtro_status_comitentes: "Filtrar por Status"
    }

    # Função para limpar um campo de entrada apenas se houver texto do usuário
    def limpar_entry(entry):
        current_text = entry.get()
        placeholder_text = placeholders_comitentes[entry]
        if current_text and current_text != placeholder_text:  # Verifica se há texto do usuário
            entry.delete(0, tk.END)

    # Limpar os campos de entrada após o cadastro
    limpar_entry(entry_cnpj)
    limpar_entry(entry_ente)
    limpar_entry(entry_spn)
    limpar_entry(entry_eci)
    limpar_entry(entry_cliente)
    limpar_entry(entry_accronym)
    limpar_entry(entry_conta_cetip)
    limpar_entry(entry_cgd)
    limpar_entry(entry_mapeamento_confirmacoes)
    limpar_entry(entry_banco)
    limpar_entry(entry_ag)
    limpar_entry(entry_cc)
    limpar_entry(entry_filtro_comitentes)
    limpar_entry(entry_filtro_status_comitentes)

def delete_comitentes(tree):  
    resposta = messagebox.askyesnocancel("Confirmação", "Deseja seguir com a atualização?")    
    if resposta == "no":
        return  # Se o usuário clicar em "No", encerra a função
    elif resposta == "cancel":
        rollback_selected_items_comitentes(tree)
        return    

    selected_items = tree.selection()
    
    if not selected_items:
        messagebox.showwarning("Attention!", "No item selected.")
        return

    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")
        
        # Check if the temporary table exists, if not, create it
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS temp_table_comitentes AS
            SELECT * FROM refData WHERE 0
        """)

        for item in selected_items:
            item_id = tree.item(item, 'values')
            SID = getpass.getuser()
            SID = SID[0].upper() + SID[1:] 
            
            if SID == item_id[13] and item_id[12] == "PENDING INACTIVE":
                messagebox.showwarning("Attention!", "Different SID must perform the Checker")
                return
            
            cnpj = item_id[0]
            ente = item_id[1]
            spn = item_id[2]
            eci = item_id[3]
            cliente  = item_id[4]
            accronym = item_id[5]
            conta_cetip = item_id[6]
            cgd = item_id[7]
            mapeamento_confirmacoes = item_id[8]
            banco = item_id[9]
            ag = item_id[10]
            cc = item_id[11]
            status = "PENDING INACTIVE" if item_id[12] == "ACTIVE" else "INACTIVE"
            timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            if item_id[12] == "ACTIVE":
                # Insert the data from the Treeview into the temporary table
                cursor.execute("""
                    INSERT INTO temp_table_comitentes (
                        CNPJ, ENTE, SPN, ECI, CLIENTE, ACCRONYM, CONTACETIP, CGD, MAPEAMENTOCONFIRMAÇÕES,
                        BANCO, AG, CC, STATUS, SID, TIMESTAMP
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (cnpj, ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes,
                      banco, ag, cc, status, SID, timestamp))
                # Load data from the temporary table into the treeview
                carregar_dados_comitentes(tree)
            else:
                # Update the main table and delete the entry from the temporary table
                cursor.execute("""
                    UPDATE refData SET
                    ENTE = COALESCE(NULLIF(?, ''), ENTE),
                    SPN = COALESCE(NULLIF(?, ''), SPN),
                    ECI = COALESCE(NULLIF(?, ''), ECI),
                    CLIENTE = COALESCE(NULLIF(?, ''), CLIENTE),
                    ACCRONYM = COALESCE(NULLIF(?, ''), ACCRONYM),
                    CONTACETIP = COALESCE(NULLIF(?, ''), CONTACETIP),
                    CGD = COALESCE(NULLIF(?, ''), CGD),
                    MAPEAMENTOCONFIRMAÇÕES = COALESCE(NULLIF(?, ''), MAPEAMENTOCONFIRMAÇÕES),
                    BANCO = COALESCE(NULLIF(?, ''), BANCO),
                    AG = COALESCE(NULLIF(?, ''), AG),
                    CC = COALESCE(NULLIF(?, ''), CC),
                    Status = COALESCE(NULLIF(?, ''), Status),
                    SID = COALESCE(NULLIF(?, ''), SID),
                    timestamp = COALESCE(NULLIF(?, ''), timestamp)
                    WHERE CNPJ = ?
                """, (ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes, banco, ag, cc, status, SID, timestamp, cnpj))
                cursor.execute("DELETE FROM temp_table_comitentes WHERE CNPJ = ?", (cnpj,))
        
        conn.commit()
        carregar_dados_comitentes(tree)
        messagebox.showinfo("Atualização", "Comitentes atualizado com sucesso!")
    except sqlite3.Error as e:
        conn.rollback()
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")
    finally:
        conn.close()

    # Dicionário para armazenar o placeholder_text de cada campo
    placeholders_comitentes = {
        entry_cnpj: "CNPJ",
        entry_ente: "ENTE",
        entry_spn: "SPN",
        entry_eci: "ECI",
        entry_cliente: "CLIENTE",
        entry_accronym: "ACCRONYM",
        entry_conta_cetip: "CONTA CETIP",
        entry_cgd: "CGD",
        entry_mapeamento_confirmacoes: "MAPEAMENTO CONFIRMAÇÕES",
        entry_banco: "BANCO",
        entry_ag: "AG",
        entry_cc: "CC",
        entry_filtro_comitentes: "Filtrar por Cliente",
        entry_filtro_status_comitentes: "Filtrar por Status"
    }

    # Função para limpar um campo de entrada apenas se houver texto do usuário
    def limpar_entry(entry):
        current_text = entry.get()
        placeholder_text = placeholders_comitentes[entry]
        if current_text and current_text != placeholder_text:  # Verifica se há texto do usuário
            entry.delete(0, tk.END)

    # Limpar os campos de entrada após o cadastro
    limpar_entry(entry_cnpj)
    limpar_entry(entry_ente)
    limpar_entry(entry_spn)
    limpar_entry(entry_eci)
    limpar_entry(entry_cliente)
    limpar_entry(entry_accronym)
    limpar_entry(entry_conta_cetip)
    limpar_entry(entry_cgd)
    limpar_entry(entry_mapeamento_confirmacoes)
    limpar_entry(entry_banco)
    limpar_entry(entry_ag)
    limpar_entry(entry_cc)
    limpar_entry(entry_filtro_comitentes)
    limpar_entry(entry_filtro_status_comitentes)

def carregar_dados_comitentes(tree):
    # Limpar o Treeview
    tree.delete(*tree.get_children())
    
    # Conectar ao banco de dados
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()

    # Executar a consulta para obter os dados de comitentes
    cursor.execute("SELECT CNPJ, ENTE, SPN, ECI, CLIENTE, ACCRONYM, CONTACETIP, CGD, MAPEAMENTOCONFIRMAÇÕES, BANCO, AG, CC, STATUS, SID, TIMESTAMP FROM refData")
    rows = cursor.fetchall()

    # Inserir os dados na Treeview
    for row in rows:
        tree.insert("", "end", values=row)

    # Verificar se a tabela temporária existe e tem dados
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='temp_table_comitentes'")
    if cursor.fetchone():
        cursor.execute("SELECT * FROM temp_table_comitentes")
        temp_rows = cursor.fetchall()

        # Criar um dicionário para mapear CNPJ aos dados da tabela temporária
        temp_data_dict = {row[0]: row for row in temp_rows}

        # Sobrescrever os dados no Treeview com os dados da tabela temporária
        for item in tree.get_children():
            item_values = tree.item(item, "values")
            cnpj = item_values[0]
            if cnpj in temp_data_dict:
                # Atualizar os valores no Treeview
                tree.item(item, values=temp_data_dict[cnpj])

    # Fechar a conexão
    conn.close()
    global tabela_comitentes_data  # Certifique-se de que a variável é global
    # Atualizar a lista de dados original
    tabela_comitentes_data = [tree.item(item, "values") for item in tree.get_children()]

def cadastrar_comitentes():
    # Perguntar ao usuário se deseja seguir com o cadastro
    resposta = messagebox.askyesno("Confirmação", "Deseja seguir com o cadastro?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função
    
    # Obter os valores dos campos de entrada
    cnpj = entry_cnpj.get()
    ente = entry_ente.get()
    spn = entry_spn.get()
    eci = entry_eci.get()
    cliente = entry_cliente.get()
    accronym = entry_accronym.get()
    conta_cetip = entry_conta_cetip.get()
    cgd = entry_cgd.get()
    mapeamento_confirmacoes = entry_mapeamento_confirmacoes.get()
    banco = entry_banco.get()
    ag = entry_ag.get()
    cc = entry_cc.get()
    status = "PENDING"
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    SID = getpass.getuser()
    SID = SID[0].upper() + SID[1:]

    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    try:
        cursor = conn.cursor()

        # Verificar se o CNPJ já existe
        cursor.execute("SELECT * FROM refData WHERE CNPJ = ?", (cnpj,))
        existing_record = cursor.fetchone()

        if existing_record:
            # Perguntar ao usuário se deseja seguir com a atualização
            resposta = messagebox.askyesno("Confirmação", "Cadastro já existente. Deseja seguir com a atualização?")
            if not resposta:
                return  # Se o usuário clicar em "No", encerra a função
            # Atualizar apenas os campos que não estão vazios
            cursor.execute("""
                UPDATE refData SET
                ENTE = COALESCE(NULLIF(?, ''), ENTE),
                SPN = COALESCE(NULLIF(?, ''), SPN),
                ECI = COALESCE(NULLIF(?, ''), ECI),
                CLIENTE = COALESCE(NULLIF(?, ''), CLIENTE),
                ACCRONYM = COALESCE(NULLIF(?, ''), ACCRONYM),
                CONTACETIP = COALESCE(NULLIF(?, ''), CONTACETIP),
                CGD = COALESCE(NULLIF(?, ''), CGD),
                MAPEAMENTOCONFIRMAÇÕES = COALESCE(NULLIF(?, ''), MAPEAMENTOCONFIRMAÇÕES),
                BANCO = COALESCE(NULLIF(?, ''), BANCO),
                AG = COALESCE(NULLIF(?, ''), AG),
                CC = COALESCE(NULLIF(?, ''), CC),
                Status = COALESCE(NULLIF(?, ''), Status),
                SID = COALESCE(NULLIF(?, ''), SID),
                timestamp = COALESCE(NULLIF(?, ''), timestamp)
                WHERE CNPJ = ?
            """, (ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes, banco, ag, cc, cnpj))
            conn.commit()
            messagebox.showinfo("Atualização", "comitentes atualizado com sucesso!")

            # Atualizar a linha existente no Treeview
            for item in tabela_base_comitentes.get_children():
                if tabela_base_comitentes.item(item, "values")[0] == cnpj:
                    tabela_base_comitentes.item(item, values=(
                        cnpj,
                        ente or tabela_base_comitentes.item(item, "values")[1],
                        spn or tabela_base_comitentes.item(item, "values")[2],
                        eci or tabela_base_comitentes.item(item, "values")[3],
                        cliente or tabela_base_comitentes.item(item, "values")[4],
                        accronym or tabela_base_comitentes.item(item, "values")[5],
                        conta_cetip or tabela_base_comitentes.item(item, "values")[6],
                        cgd or tabela_base_comitentes.item(item, "values")[7],
                        mapeamento_confirmacoes or tabela_base_comitentes.item(item, "values")[8],
                        banco or tabela_base_comitentes.item(item, "values")[9],
                        ag or tabela_base_comitentes.item(item, "values")[10],
                        cc or tabela_base_comitentes.item(item, "values")[11],
                        status,
                        SID,
                        timestamp
                    ))
                    break
        else:
            # Inserir os dados na tabela refData
            cursor.execute("""
                INSERT INTO refData (CNPJ, ENTE, SPN, ECI, CLIENTE, ACCRONYM, CONTACETIP, CGD, MAPEAMENTOCONFIRMAÇÕES, BANCO, AG, CC)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (cnpj, ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes, banco, ag, cc, status, SID, timestamp))
            conn.commit()
            messagebox.showinfo("Sucesso", "comitentes cadastrado com sucesso!")

            # Inserir os dados na Treeview
            tabela_base_comitentes.insert("", "end", values=(cnpj, ente, spn, eci, cliente, accronym, conta_cetip, cgd, mapeamento_confirmacoes, banco, ag, cc, status, SID, timestamp))

        # Recarregar os dados do banco de dados e atualizar a lista de dados original
        carregar_dados_comitentes(tabela_base_comitentes)

    except sqlite3.Error as e:
        conn.close()
        messagebox.showerror("Erro de Banco de Dados", f"Ocorreu um erro ao acessar o banco de dados: {e}")

    finally:
        # Fechar a conexão com o banco de dados
        conn.close()

  # Dicionário para armazenar o placeholder_text de cada campo
    placeholders_comitentes = {
        entry_cnpj: "CNPJ",
        entry_ente: "ENTE",
        entry_spn: "SPN",
        entry_eci: "ECI",
        entry_cliente: "CLIENTE",
        entry_accronym: "ACCRONYM",
        entry_conta_cetip: "CONTA CETIP",
        entry_cgd: "CGD",
        entry_mapeamento_confirmacoes: "MAPEAMENTO CONFIRMAÇÕES",
        entry_banco: "BANCO",
        entry_ag: "AG",
        entry_cc: "CC",
        entry_filtro_comitentes: "Filtrar por Cliente",
        entry_filtro_status_comitentes: "Filtrar por Status"
    }

    # Função para limpar um campo de entrada apenas se houver texto do usuário
    def limpar_entry(entry):
        current_text = entry.get()
        placeholder_text = placeholders_comitentes[entry]
        if current_text and current_text != placeholder_text:  # Verifica se há texto do usuário
            entry.delete(0, tk.END)

    # Limpar os campos de entrada após o cadastro
    limpar_entry(entry_cnpj)
    limpar_entry(entry_ente)
    limpar_entry(entry_spn)
    limpar_entry(entry_eci)
    limpar_entry(entry_cliente)
    limpar_entry(entry_accronym)
    limpar_entry(entry_conta_cetip)
    limpar_entry(entry_cgd)
    limpar_entry(entry_mapeamento_confirmacoes)
    limpar_entry(entry_banco)
    limpar_entry(entry_ag)
    limpar_entry(entry_cc)
    limpar_entry(entry_filtro_comitentes)
    limpar_entry(entry_filtro_status_comitentes)

def carregar_dados_pendente_termo(tree):
    # Limpar o Treeview
    tree.delete(*tree.get_children())

    conn = sqlite3.connect("I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\termo_historico.db")
    cursor = conn.cursor()  

    cursor.execute("SELECT * FROM tco_pendente")      
    rows = cursor.fetchall()

    # Inserir os dados na Treeview
    for row in rows:
        tree.insert("", "end", values=row)
    
    conn.close()
    global tabela_termo_pendente_data  # Certifique-se de que a variável é global    
    # Atualizar a lista de dados original    
    tabela_termo_pendente_data = [tree.item(item, "values") for item in tree.get_children()]


def carregar_calendario(tree, query):
    # Conectar ao banco de dados
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()

    # Executar a consulta para obter os dados
    cursor.execute(query)
    rows = cursor.fetchall()

    # Inserir os dados na Treeview
    for row in rows:
        tree.insert("", "end", values=row)

    # Fechar a conexão
    conn.close()  

def excluir_commodities():
    # Obter o valor do campo de entrada
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()

    ativo_subjacente = entry_ativo_subjacente.get()

    # Confirmar a exclusão com o usuário
    resposta = messagebox.askyesno("Confirmação", f"Deseja excluir a commodity '{ativo_subjacente}'?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função

    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()

    # Excluir o registro do banco de dados
    cursor.execute("DELETE FROM Commodities WHERE AtivoSubjacenteRIC = ?", (ativo_subjacente,))
    conn.commit()

    # Fechar a conexão
    conn.close()

    # Excluir a linha correspondente na tabela_commodities
    for item in tabela_commodities.get_children():
        if tabela_commodities.item(item, 'values')[0] == ativo_subjacente:
            tabela_commodities.delete(item)
            break

    messagebox.showinfo("Sucesso", f"Commodity '{ativo_subjacente}' excluída com sucesso!")

def excluir_comitentes():
    # Obter o valor do campo de entrada
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    cnpj = entry_cnpj.get()
    comitante_name = lookup(cnpj, cntpy_taxid, cntpy_name)

    # Confirmar a exclusão com o usuário
    resposta = messagebox.askyesno("Confirmação", f"Deseja excluir o comitentes '{comitante_name}'?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função

    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()

    # Excluir o registro do banco de dados
    cursor.execute("DELETE FROM refData WHERE CNPJ = ?", (cnpj,))
    conn.commit()

    # Fechar a conexão
    conn.close()

    # Excluir a linha correspondente na tabela_base_comitentes
    for item in tabela_base_comitentes.get_children():
        if tabela_base_comitentes.item(item, 'values')[0] == cnpj:
            tabela_base_comitentes.delete(item)
            break

    messagebox.showinfo("Sucesso", f"comitentes com CNPJ '{cnpj}' excluído com sucesso!")
    

# Funções principais
def criar_interface():       
    global combobox_bolsa_negociacao, combobox_mercadoria, combobox_tipo, combobox_unidade, combobox_moeda, combobox_conversao, combobox_mes_vencimento, combobox_ano_vencimento
    global athena_id
    global label_qty_deals_cliente_termo, label_qty_deals_cliente_opcao   
    global aba_inicio
    global tabela_base_comitentes, tabela_commodities 
    global tabela_feriados
    global entry_ativo_subjacente, entry_indice_valorizacao    
    global entry_filtro_status_commodities 
    global entry_cnpj, entry_ente, entry_spn, entry_eci, entry_cliente, entry_filtro_status_comitentes
    global entry_accronym, entry_conta_cetip, entry_cgd, entry_mapeamento_confirmacoes
    global entry_filtro_commodities, entry_filtro_comitentes, entry_banco, entry_ag, entry_cc    
    global janela, toplevel_windows, original_data_commodities, original_data_comitentes
    global switch_cliente_termo
    global switch_b2b_termo
    global switch_cliente_opcao
    global switch_b2b_opcao
    global switch_fixingstermo_cliente
    global switch_fixingstermo_b2b
    global switch_fixingsopcao_cliente
    global switch_fixingsopcao_b2b
    global switch_cliente_arquivotermo
    global switch_b2b_arquivotermo
    global switch_cliente_arquivoopcao
    global switch_b2b_arquivoopcao
    global tabview

    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("dark-blue")   

    janela = ctk.CTk()
    script_dir = os.path.dirname(__file__)    
    icon_path = os.path.join(script_dir, "Icons", "X_icone.ico")
    logo_path = os.path.join(script_dir, "Icons", "fulllogo.png")
    janela.iconbitmap(icon_path)

    janela.title("CommodiXchange")
   # Set the new window size
    janela.geometry("1580x768")   
    
    def on_close_event():
        # Show a warning message box
        if messagebox.askyesno("Close Window", "Are you sure you want to close the window?"):
            janela.destroy()  # Close the window if "Yes" is clicked
        else:
            pass  # Do nothing if "No" is clicked
    janela.protocol("WM_DELETE_WINDOW", on_close_event)
    
    tabview = ctk.CTkTabview(janela, width=188.2421875, height=26, corner_radius=8)
    tabview.pack(expand=True, fill='both')    
    abas_existentes = []    
    
    # Dictionary to store original data for rollback
    original_data_commodities = {}
    original_data_comitentes = {}
    tabela_termo_cliente= None
    tabela_termo_b2b= None
    tabela_arquivotermo_cliente= None
    tabela_arquivotermo_b2b= None
    tabela_fixingstermo_cliente= None
    tabela_fixingstermo_b2b= None
    tabela_opcao_cliente= None
    tabela_opcao_b2b= None
    tabela_arquivoopcao_cliente= None
    tabela_arquivoopcao_b2b= None
    tabela_fixingsopcao_cliente= None
    tabela_fixingsopcao_b2b= None 

    def adicionar_aba(nome_aba):
        if nome_aba not in abas_existentes:
            tabview.add(nome_aba)
            abas_existentes.append(nome_aba)

    aba_inicio = tabview.add("Home")
    frame_inicio = ctk.CTkFrame(aba_inicio, fg_color="#F0F0F0", corner_radius=8)
    frame_inicio.pack(expand=True, fill='both')

    # Configuração do grid para o frame_inicio com 8 colunas e 8 linhas
    for i in range(8):
        frame_inicio.grid_rowconfigure(i, weight=1)
        frame_inicio.grid_columnconfigure(i, weight=1)

    # Inverter o lado do retângulo azul
    canvas = ctk.CTkCanvas(frame_inicio, bg="#1B0447", bd=0, highlightthickness=0, relief="ridge")
    canvas.grid(row=0, column=1, rowspan=8, columnspan=8, sticky="nsew")    
    
    # Defina a fonte usando tkinter.font.Font
    fonte_botao = ctk.CTkFont(family="League Spartan", size=13, weight="bold")    
    
    botao_calendarios_inicio = ctk.CTkButton(frame_inicio, width=122 , height=30, corner_radius=8,fg_color="#5A5368",  bg_color = "#1B0447", text="HOLIDAYS", font= fonte_botao, command=lambda: calendarios_bolsas(tabview, abas_existentes))
    botao_calendarios_inicio.grid(row=0, column=6, padx=10, pady=55, sticky="ne")
   # Carregar e exibir a logo
    logo = Image.open(logo_path)
    logo = logo.resize((245, 87))
    logo_tk = ImageTk.PhotoImage(logo)

    canvas_logo = tk.Canvas(frame_inicio, bg="#F0F0F0", width=245, height=87, bd=0, highlightthickness=0)
    canvas_logo.grid(row=0, column=0, padx=10, pady=20, sticky="nw")
    canvas_logo.create_image(0, 0, anchor=tk.NW, image=logo_tk)
    canvas_logo.image = logo_tk

    tree = []
    
    
    # Configuração da aba de Commodities    
    aba_commodities = tabview.add("Commodities")
    aba_commodities.grid_rowconfigure(0, weight=1)
    aba_commodities.grid_columnconfigure(1, weight=1)

    frame_cadastro_commodities = ctk.CTkFrame(aba_commodities, height=668, width=300)
    frame_cadastro_commodities.grid(row=0, column=0, sticky='ns')

    # Inicializar as entradas
    entry_ativo_subjacente = ctk.CTkEntry(frame_cadastro_commodities, width=300, placeholder_text="ATIVO SUBJACENTE / RIC")
    entry_ativo_subjacente.grid(row=0, column=0, padx=10, pady=3)
    
    bolsa_negociacao = carregar_combobox_bolsa()    
    
    def combobox_bolsa_callback(choice):     
        combobox_bolsa_negociacao.set(choice)          
        carregar_combobox_mercadoria(combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_mercadoria)      
        if combobox_bolsa_negociacao.get() != "BOLSA DE NEGOCIACAO":
            carregar_combobox_mercadoria(combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_mercadoria)
            carregar_combobox_tipo(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_tipo)
            carregar_combobox_unidade(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, scrollable_dropdown_unidade)
            carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)      
      
    def combobox_bolsa_change(event):
        if combobox_bolsa_negociacao.get() != "BOLSA DE NEGOCIACAO":
            carregar_combobox_mercadoria(combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_mercadoria)
            carregar_combobox_tipo(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_tipo)
            carregar_combobox_unidade(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, scrollable_dropdown_unidade)
            carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)
            
    # BLOCO PARA BOLSA
    combobox_bolsa_negociacao = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color ="white", button_color="#5A5368")    
    # Set default value
    combobox_bolsa_negociacao.set("BOLSA DE NEGOCIACAO")
    combobox_bolsa_negociacao.grid(row=1, column=0, padx=10, pady=3, sticky='w')    
    CTkScrollableDropdown(combobox_bolsa_negociacao, values= bolsa_negociacao, justify="left", autocomplete=True, command= combobox_bolsa_callback)
    combobox_bolsa_negociacao.bind("<<ComboboxSelected>>", combobox_bolsa_change)
    
    def combobox_mercadoria_callback(choice):     
        combobox_mercadoria.set(choice)          
        carregar_combobox_tipo(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_tipo)
        if combobox_tipo.get() != "TIPO":
            carregar_combobox_unidade(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, scrollable_dropdown_unidade)
            carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)

    def combobox_mercadoria_change(event):
        if combobox_mercadoria.get() != "MERCADORIA":
            carregar_combobox_tipo(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, scrollable_dropdown_tipo)
            carregar_combobox_unidade(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, scrollable_dropdown_unidade)
            carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)
        
    # BLOCO PARA MERCADORIA
    combobox_mercadoria = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color ="white", button_color="#5A5368")
    # Set default value
    combobox_mercadoria.set("MERCADORIA")
    combobox_mercadoria.grid(row=2, column=0, padx=10, pady=3, sticky='w')
    scrollable_dropdown_mercadoria = CTkScrollableDropdown(combobox_mercadoria, values=["MERCADORIA"], justify="left", autocomplete=True, command= combobox_mercadoria_callback)    
    combobox_mercadoria.bind("<<ComboboxSelected>>", combobox_mercadoria_change)
    
    def combobox_tipo_callback(choice):     
        combobox_tipo.set(choice)          
        carregar_combobox_unidade(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, scrollable_dropdown_unidade)
        if combobox_unidade.get() != "UNIDADE":
            carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)
    def combobox_tipo_change(event):
        if combobox_tipo.get() != "TIPO":
            carregar_combobox_unidade(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, scrollable_dropdown_unidade)
            carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)
      
    # BLOCO PARA TIPO
    combobox_tipo = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color ="white", button_color="#5A5368")
    # Set default value
    combobox_tipo.set("TIPO")
    combobox_tipo.grid(row=3, column=0, padx=10, pady=3, sticky='w')
    scrollable_dropdown_tipo = CTkScrollableDropdown(combobox_tipo, values=["TIPO"], justify="left", autocomplete=True, command= combobox_tipo_callback)        

    # Bind the change event for combobox_tipo
    combobox_tipo.bind("<<ComboboxSelected>>", combobox_tipo_change)

    def combobox_unidade_callback(choice):     
        combobox_unidade.set(choice)          
        carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
        if combobox_moeda.get() != "MOEDA":
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)
    def combobox_unidade_change(event):
        if combobox_unidade.get() != "UNIDADE DE NEGOCIÇÃO":
            carregar_combobox_moeda(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, scrollable_dropdown_moeda)
    

    
    # BLOCO PARA UNIDADE DE NEGOCIACAO
    combobox_unidade = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color ="white", button_color="#5A5368")
    # Set default value
    combobox_unidade.set("UNIDADE DE NEGOCIÇÃO")
    combobox_unidade.grid(row=4, column=0, padx=10, pady=3, sticky='w')
    scrollable_dropdown_unidade = CTkScrollableDropdown(combobox_unidade, values=["UNIDADE DE NEGOCIÇÃO"], justify="left", autocomplete=True, command= combobox_unidade_callback)
    
    # Bind the change event for combobox_unidade
    combobox_unidade.bind("<<ComboboxSelected>>", combobox_unidade_change)
    
    def combobox_moeda_callback(choice):     
        combobox_moeda.set(choice)          
        carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)
        
    def combobox_moeda_change(event):
        if combobox_moeda.get() != "MOEDA":
            carregar_combobox_conversao(combobox_tipo, combobox_bolsa_negociacao, combobox_mercadoria, combobox_unidade, combobox_moeda, combobox_conversao, scrollable_dropdown_conversao)
            
    # BLOCO PARA MOEDA
    combobox_moeda = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color ="white", button_color="#5A5368")
    # Set default value
    combobox_moeda.set("MOEDA")
    combobox_moeda.grid(row=5, column=0, padx=10, pady=3, sticky='w')
    scrollable_dropdown_moeda = CTkScrollableDropdown(combobox_moeda, values=["MOEDA"], justify="left", autocomplete=True, command= combobox_moeda_callback)
    
    # Bind the change event for combobox_unidade
    combobox_moeda.bind("<<ComboboxSelected>>", combobox_moeda_change)
 
    
     # BLOCO PARA FATOR DE CONVERSAO
    combobox_conversao = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color ="white", button_color="#5A5368")
    # Set default value
    combobox_conversao.set("FATOR DE CONVERSÃO")
    combobox_conversao.grid(row=6, column=0, padx=10, pady=3, sticky='w')
    scrollable_dropdown_conversao = CTkScrollableDropdown(combobox_conversao, values=["FATOR DE CONVERSÃO"], justify="left", autocomplete=True)    

    entry_indice_valorizacao = ctk.CTkEntry(frame_cadastro_commodities, width=300, placeholder_text="INDICE VALORIZACAO")
    entry_indice_valorizacao.grid(row=7, column=0, padx=10, pady=3)
    
    # BLOCO PARA MES VENCIMENTO
    combobox_mes_vencimento = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color="white", button_color="#5A5368")
    # Set default value
    
    combobox_mes_vencimento.grid(row=8, column=0, padx=10, pady=3, sticky='w')
    CTkScrollableDropdown(combobox_mes_vencimento, values=["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"], justify="left", autocomplete=True)
    combobox_mes_vencimento.set("MES VENCIMENTO")
    
    # BLOCO PARA ANO VENCIMENTO
    combobox_ano_vencimento = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color="white", button_color="#5A5368")
    # Set default value    
   # Assuming combobox_ano_vencimento is a CTkComboBox or similar widget
    combobox_ano_vencimento = ctk.CTkComboBox(frame_cadastro_commodities, width=300, corner_radius=8, dropdown_fg_color="white", button_color="#5A5368")
    combobox_ano_vencimento.grid(row=9, column=0, padx=10, pady=3, sticky='w')

    # Create a scrollable dropdown with the specified values
    years = [str(year) for year in range(2025, 2041)] + ["2221", "2200", "2201", "2199", "2197", "2196"]
    CTkScrollableDropdown(combobox_ano_vencimento, values=years, justify="left", autocomplete=True)

    # Set the default value for the combobox
    combobox_ano_vencimento.set("ANO VENCIMENTO")

    frame_caracteristicas_commodities = ctk.CTkFrame(aba_commodities, height=668)
    frame_caracteristicas_commodities.grid(row=0, column=1, sticky='nsew')

    scrollbar_x_commodities = ctk.CTkScrollbar(frame_caracteristicas_commodities, orientation='horizontal')
    scrollbar_y_commodities = ctk.CTkScrollbar(frame_caracteristicas_commodities, orientation='vertical')

    tabela_commodities = ttk.Treeview(frame_caracteristicas_commodities, columns=colunas_commodities, show='headings', xscrollcommand=scrollbar_x_commodities.set, yscrollcommand=scrollbar_y_commodities.set)
    for coluna in colunas_commodities:
        tabela_commodities.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_commodities, _col, False))
    tabela_commodities.grid(row=0, column=0, sticky='nsew')

    
    
    scrollbar_x_commodities.configure(command=tabela_commodities.xview, height=25)
    scrollbar_y_commodities.configure(command=tabela_commodities.yview, width=25)
    scrollbar_x_commodities.grid(row=1, column=0, sticky='ew')
    scrollbar_y_commodities.grid(row=0, column=1, sticky='ns')

    frame_caracteristicas_commodities.grid_rowconfigure(0, weight=1)
    frame_caracteristicas_commodities.grid_columnconfigure(0, weight=1)

    carregar_dados_commodities(tabela_commodities)
    ajustar_largura_colunas(tabela_commodities, colunas_commodities, tabview)

    tabela_commodities_data = [tabela_commodities.item(item, "values") for item in tabela_commodities.get_children()]

    frame_filtro_commodities = ctk.CTkFrame(aba_commodities, height=150)
    frame_filtro_commodities.grid(row=1, column=0, columnspan=2, sticky='ew', pady=5)

    botao_cadastrar_commodities = ctk.CTkButton(frame_filtro_commodities, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="REGISTER", font= fonte_botao, command=lambda: cadastrar_commodities())
    botao_cadastrar_commodities.grid(row=0, column=0, padx=10, pady=10, sticky='w')
    botao_refresh_commodities = ctk.CTkButton(frame_filtro_commodities, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="REFRESH", font= fonte_botao, command=lambda:refresh_table_commodities(tabela_commodities))
    botao_refresh_commodities.grid(row=0, column=1, padx=5, pady=10, sticky='w')
    botao_update_commodities = ctk.CTkButton(frame_filtro_commodities, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="UPDATE", font= fonte_botao,command=lambda: atualizar_commodities(tabela_commodities))
    botao_update_commodities.grid(row=0, column=2, padx=5, pady=10, sticky='w')
    botao_excluir_commodities = ctk.CTkButton(frame_filtro_commodities, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", font= fonte_botao,command=lambda: delete_commodities(tabela_commodities))
    botao_excluir_commodities.grid(row=0, column=3, padx=10, pady=10, sticky='w')
    

    entry_filtro_commodities = ctk.CTkEntry(frame_filtro_commodities, width=200, placeholder_text="Ativo Subjacente / RIC")
    entry_filtro_commodities.grid(row=0, column=4, padx=5, sticky='w')

    entry_filtro_type_commodities = ctk.CTkEntry(frame_filtro_commodities, width=200, placeholder_text="Mercadoria")
    entry_filtro_type_commodities.grid(row=0, column=5, padx=5, sticky='w')

    entry_filtro_status_commodities = ctk.CTkEntry(frame_filtro_commodities, width=200, placeholder_text="Status")
    entry_filtro_status_commodities.grid(row=0, column=6, padx=5, sticky='w')

 
    def filtrar_commodities(event):
        global tabela_commodities_data  # Certifique-se de que a variável é global
        filtro = entry_filtro_commodities.get().lower()
        tabela_commodities.delete(*tabela_commodities.get_children())
        for item in tabela_commodities_data:
            if filtro in item[0].lower():
                tabela_commodities.insert("", "end", values=item)    
        
    def filtrar_commodities_mercadoria(event):
        global tabela_commodities_data  # Certifique-se de que a variável é global
        filtro = entry_filtro_type_commodities.get().lower()
        tabela_commodities.delete(*tabela_commodities.get_children())
        for item in tabela_commodities_data:
            if filtro in item[9].lower():
                tabela_commodities.insert("", "end", values=item)

    def filtrar_commodities_status(event):
        global tabela_commodities_data  # Certifique-se de que a variável é global
        filtro = entry_filtro_status_commodities.get().lower()
        tabela_commodities.delete(*tabela_commodities.get_children())
        for item in tabela_commodities_data:
            if filtro in item[10].lower():
                tabela_commodities.insert("", "end", values=item)

    
    entry_filtro_commodities.bind("<KeyRelease>", filtrar_commodities) 
    entry_filtro_type_commodities.bind("<KeyRelease>", filtrar_commodities_mercadoria)
    entry_filtro_status_commodities.bind("<KeyRelease>", filtrar_commodities_status)
    
    # Assuming `tabela_commodities` is your Treeview widget
    vincular_evento_duplo_clique_base(tabela_commodities, colunas_commodities) 
    
    # Configuração da aba de comitentes
    

    aba_comitentes = tabview.add("Counterparty")
    aba_comitentes.grid_rowconfigure(0, weight=1)
    aba_comitentes.grid_columnconfigure(1, weight=1)

    frame_cadastro_comitentes = ctk.CTkFrame(aba_comitentes, height=668, width=250)
    frame_cadastro_comitentes.grid(row=0, column=0, sticky='ns')

    # Inicializar as entradas

    entry_cnpj = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="CNPJ")
    entry_cnpj.grid(row=0, column=0, padx=10, pady=3)

    entry_ente = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="ENTE")
    entry_ente.grid(row=1, column=0, padx=10, pady=3)

    entry_spn = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="SPN")
    entry_spn.grid(row=2, column=0, padx=10, pady=3)

    entry_eci = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="ECI")
    entry_eci.grid(row=3, column=0, padx=10, pady=3)

    entry_cliente = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="CLIENTE")
    entry_cliente.grid(row=4, column=0, padx=10, pady=3)

    entry_accronym = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="ACCRONYM")
    entry_accronym.grid(row=5, column=0, padx=10, pady=3)

    entry_conta_cetip = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="CONTA CETIP")
    entry_conta_cetip.grid(row=6, column=0, padx=10, pady=3)

    entry_cgd = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="CGD")
    entry_cgd.grid(row=7, column=0, padx=10, pady=3)

    entry_mapeamento_confirmacoes = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="MAPEAMENTO CONFIRMAÇÕES")
    entry_mapeamento_confirmacoes.grid(row=8, column=0, padx=10, pady=3)

    entry_banco = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="BANCO")
    entry_banco.grid(row=9, column=0, padx=10, pady=3)

    entry_ag = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="AG")
    entry_ag.grid(row=10, column=0, padx=10, pady=3)

    entry_cc = ctk.CTkEntry(frame_cadastro_comitentes, width=250, placeholder_text="CC")
    entry_cc.grid(row=11, column=0, padx=10, pady=3)

    frame_caracteristicas_comitentes = ctk.CTkFrame(aba_comitentes, height=668)
    frame_caracteristicas_comitentes.grid(row=0, column=1, sticky='nsew')

    scrollbar_x_comitentes = ctk.CTkScrollbar(frame_caracteristicas_comitentes, orientation='horizontal')
    scrollbar_y_comitentes = ctk.CTkScrollbar(frame_caracteristicas_comitentes, orientation='vertical')

    tabela_base_comitentes = ttk.Treeview(frame_caracteristicas_comitentes, columns=colunas_comitentes, show='headings', xscrollcommand=scrollbar_x_comitentes.set, yscrollcommand=scrollbar_y_comitentes.set)
    for coluna in colunas_comitentes:
        tabela_base_comitentes.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_base_comitentes, _col, False))
    tabela_base_comitentes.grid(row=0, column=0, sticky='nsew')
    # Bind the <FocusOut> event to clear the selection
    

    scrollbar_x_comitentes.configure(command=tabela_base_comitentes.xview, height=25)
    scrollbar_y_comitentes.configure(command=tabela_base_comitentes.yview, width=25)
    scrollbar_x_comitentes.grid(row=1, column=0, sticky='ew')
    scrollbar_y_comitentes.grid(row=0, column=1, sticky='ns')

    frame_caracteristicas_comitentes.grid_rowconfigure(0, weight=1)
    frame_caracteristicas_comitentes.grid_columnconfigure(0, weight=1)

    carregar_dados_comitentes(tabela_base_comitentes)
    ajustar_largura_colunas(tabela_base_comitentes, colunas_comitentes, tabview)

    global tabela_comitentes_data
    tabela_comitentes_data = [tabela_base_comitentes.item(item, "values") for item in tabela_base_comitentes.get_children()]

    frame_filtro_comitentes = ctk.CTkFrame(aba_comitentes, height=150)
    frame_filtro_comitentes.grid(row=1, column=0, columnspan=2, sticky='ew', pady=5)

    botao_cadastrar_comitentes = ctk.CTkButton(frame_filtro_comitentes, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="REGISTER", font=fonte_botao, command=cadastrar_comitentes)
    botao_cadastrar_comitentes.grid(row=0, column=0, padx=10, pady=20, sticky='w')
    botao_refresh_comitentes = ctk.CTkButton(frame_filtro_comitentes, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="REFRESH", font= fonte_botao,command=lambda: refresh_table_comitentes(tabela_base_comitentes))
    botao_refresh_comitentes.grid(row=0, column=1, padx=5, pady=10, sticky='w')
    botao_update_comitentes = ctk.CTkButton(frame_filtro_comitentes, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="UPDATE", font= fonte_botao,command=lambda: atualizar_comitentes(tabela_base_comitentes))
    botao_update_comitentes.grid(row=0, column=2, padx=5, pady=10, sticky='w')
    botao_excluir_comitentes = ctk.CTkButton(frame_filtro_comitentes, width=122 , height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", font=fonte_botao, command=lambda: delete_comitentes(tabela_base_comitentes))
    botao_excluir_comitentes.grid(row=0, column=3, padx=10, pady=20, sticky='w')
    

    entry_filtro_comitentes = ctk.CTkEntry(frame_filtro_comitentes, width=250, placeholder_text="Filtrar por Cliente")
    entry_filtro_comitentes.grid(row=0, column=4, padx=5, sticky='w')
    entry_filtro_cnpj = ctk.CTkEntry(frame_filtro_comitentes, width=250, placeholder_text="Filtrar por CNPJ")
    entry_filtro_cnpj.grid(row=0, column=5, padx=5, sticky='w')
    entry_filtro_status_comitentes = ctk.CTkEntry(frame_filtro_comitentes, width=250, placeholder_text="Filtrar por Status")
    entry_filtro_status_comitentes.grid(row=0, column=6, padx=5, sticky='w')



    def filtrar_comitentes_nome(event):
        global tabela_comitentes_data  # Certifique-se de que a variável é global
        filtro = entry_filtro_comitentes.get().lower()
        tabela_base_comitentes.delete(*tabela_base_comitentes.get_children())
        for item in tabela_comitentes_data:
            if filtro in item[4].lower():
                tabela_base_comitentes.insert("", "end", values=item)

  
    entry_filtro_comitentes.bind("<KeyRelease>", filtrar_comitentes_nome)

    def filtrar_comitentes_cnpj(event):
        global tabela_comitentes_data  # Certifique-se de que a variável é global
        filtro = entry_filtro_cnpj.get().lower()
        tabela_base_comitentes.delete(*tabela_base_comitentes.get_children())
        for item in tabela_comitentes_data:
            if filtro in item[0].lower():
                tabela_base_comitentes.insert("", "end", values=item)
  
    entry_filtro_cnpj.bind("<KeyRelease>", filtrar_comitentes_cnpj)

    def filtrar_comitentes_status(event):
        global tabela_comitentes_data  # Certifique-se de que a variável é global
        filtro = entry_filtro_status_comitentes.get().lower()
        tabela_base_comitentes.delete(*tabela_base_comitentes.get_children())
        for item in tabela_comitentes_data:
            if filtro in item[12].lower():
                tabela_base_comitentes.insert("", "end", values=item)
  
    entry_filtro_status_comitentes.bind("<KeyRelease>", filtrar_comitentes_status)
   
    
    vincular_evento_duplo_clique_base(tabela_base_comitentes, colunas_comitentes)   
    monitor_operacoes(tabview, aba_inicio, abas_existentes, tree)
    
    abas_existentes.append("Monitor") 
    botao_importar_inicio = ctk.CTkButton(frame_inicio, width=122 , height=30, corner_radius=8, fg_color="#5A5368",  bg_color = "#1B0447", border_color="#000000", text="IMPORT", font= fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_inicio.grid(row=0, column=6, padx=10, pady=20, sticky="ne")


    # Define o atalho de teclado para abrir a janela de pesquisa
    janela.bind('<Control-Shift-F>', lambda event: open_search_window())
    # Bind the CTRL + A key combination to the select_all function
    janela.bind('<Control-a>', select_all)
    janela.bind('<Control-A>', select_all)
    # Bind the CTRL + P key combination to the open_export_window function
    janela.bind('<Control-p>', lambda event: open_export_window())
    janela.bind('<Control-P>', lambda event: open_export_window())
    # Bind the CTRL + P key combination to the open_export_window function
    janela.bind('<Control-r>', lambda event: approve_window_status())    
    janela.bind('<Control-R>', lambda event: approve_window_status())    
    

    janela.resizable(True, True)
    janela.mainloop()

def generate_mnemonic_cliente(type_value):
    # Parte fixa da string
    prefix = "CHASM"

    # Obter os dois últimos dígitos do ano atual
    year_suffix = str(datetime.now().year)[-2:]

    # Gerar um número aleatório entre 0 e 9
    random_digit = chr(random.randint(48, 57))

    # Função para gerar um caractere aleatório (letra maiúscula ou dígito)
    def random_char():
        if random.randint(0, 1):
            return chr(random.randint(65, 90))  # Letra maiúscula
        else:
            return chr(random.randint(48, 57))  # Dígito

    # Construir a string final
    result = prefix + year_suffix + random_digit + random_char() + random_char() + random_char()
    result = result.replace("e", "f")
    return result

def generate_mnemonic_b2b(type_value):
    # Definir o prefixo com base no valor de type_value
    prefix = "00041" if type_value == "Sell" else "CHASM"

    # Obter os dois últimos dígitos do ano atual
    year_suffix = str(datetime.now().year)[-2:]

    # Gerar um número aleatório entre 0 e 9
    random_digit = chr(random.randint(48, 57))

    # Função para gerar um caractere aleatório (letra maiúscula ou dígito)
    def random_char():
        if random.randint(0, 1):
            return chr(random.randint(65, 90))  # Letra maiúscula
        else:
            return chr(random.randint(48, 57))  # Dígito

    # Construir a string final
    result = prefix + year_suffix + random_digit + random_char() + random_char() + random_char()
    result = result.replace("e", "f")
    return result

def networkdays(start_date, end_date, tabela_anbima=None):
    """
    Calcula o número de dias úteis entre duas datas, excluindo fins de semana e, opcionalmente, feriados da tabela Anbima.

    :param start_date: Data inicial no formato 'DD-MMM-YYYY'
    :param end_date: Data final no formato 'DD-MMM-YYYY'
    :param tabela_anbima: Tabela opcional de feriados Anbima
    :return: Número de dias úteis
    """
    # Converter strings de data para objetos datetime usando o formato 'DD-MMM-YYYY'
    start_date = datetime.strptime(start_date, '%d-%b-%Y')
    end_date = datetime.strptime(end_date, '%d-%b-%Y')

    # Extrair feriados da tabela Anbima, se fornecida
    holidays = []
    if tabela_anbima is not None:
        for item in tabela_anbima.get_children():
            values = tabela_anbima.item(item, 'values')
            holiday_date = datetime.strptime(values[0], '%d/%m/%Y')  # Supondo que a data está na primeira coluna
            holidays.append(holiday_date)

    # Converter feriados para o formato necessário
    holidays = [holiday.strftime('%Y-%m-%d') for holiday in holidays]

    workdays = np.busday_count(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'), holidays=holidays)     
    return workdays



def next_workday_opcao(start_date, workdays):    
    # Converte a string de data para um objeto datetime, se necessário    
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, "%d/%m/%Y")
    # Converte workdays para int padrão do Python
    workdays = int(workdays)    
    daystoadd = 0
    next_day = start_date
    while daystoadd < workdays:
        next_day += timedelta(days=1) 
        while next_day.weekday() >= 5: # 5 = sábado, 6 = domingo             
            next_day += timedelta(days=1)    
        daystoadd += 1   
    return next_day

def next_workday(start_date, holidays, workdays):
    # Converte workdays para int padrão do Python
    workdays = int(workdays)
    # Adiciona dias até encontrar um dia útil
    next_day = start_date + timedelta(days=workdays)
    while next_day.weekday() >= 5 or next_day in holidays:  # 5 = sábado, 6 = domingo
        next_day += timedelta(days=workdays)
    return next_day



def custom_workday_ex(start_date, exchange_value, commodity_value, tabela_feriados, fixing_end_date):
    nome_bolsa = {
        "MDE-BURSA MALAYSIA": "BURSA",
        "BLOOMBERG": "PLATTS",
        "LME": "LME",
        "ICE (NYBOT)": "ICE",
        "NYMEX": "NYMEX",
        "CBOT": "CBOT"
    }.get(exchange_value)

    if not nome_bolsa:
        messagebox.showwarning("Atenção", f"Favor cadastrar o calendário da bolsa {exchange_value}!")
        return None

    holidays = []
    if nome_bolsa in tabela_feriados:
        for feriado_item in tabela_feriados[nome_bolsa].get_children():
            feriado_values = tabela_feriados[nome_bolsa].item(feriado_item, 'values')
            typecommodity_value = feriado_values[0]            
            holidays_value = datetime.strptime(feriado_values[1],"%d/%m/%Y") #).strftime("%d/%m/%Y")            
            if commodity_value in typecommodity_value:
                holidays.append(holidays_value)

    # Defina o valor de workdays conforme necessário
    workdays = 1  # Exemplo: pode ser um valor fixo ou calculado

    final_date = next_workday(start_date, holidays, workdays)    
    if final_date > fixing_end_date:
        return ""
    else:
        return final_date.strftime("%d/%m/%Y")
    
def calendarios_bolsas(tabview, abas_existentes):
    global tabela_feriados_global
    global tabela_anbima, tabela_ice, tabela_nymex, tabela_bursa, tabela_cbot, tabela_platts, tabela_lme
    global tabview_calendarios

    if "Holidays" in abas_existentes:
        # Se a aba já existe, retorne as tabelas armazenadas globalmente
        return tabela_feriados_global

    abas_existentes.append("Holidays")

    aba_calendarios = tabview.add("Holidays")
    tabview_calendarios = ctk.CTkTabview(aba_calendarios)
    tabview_calendarios.pack(expand=True, fill='both')

    
    

    style = ttk.Style()
    style.theme_use("default")
    style.configure(
        "Treeview",        
        borderwidth=0
    )

    tabela_feriados = {}  # Inicializa o dicionário para armazenar as tabelas

    # Bloco para ANBIMA
    nome_sub_aba = "ANBIMA"
    tabview_calendarios.add(nome_sub_aba)
    sub_aba = tabview_calendarios.tab(nome_sub_aba)

    frame_calendario_anbima = ctk.CTkFrame(sub_aba)
    frame_calendario_anbima.pack(expand=True, fill='both')

    frame_scrollbary_anbima = ctk.CTkFrame(frame_calendario_anbima, width=4)
    frame_scrollbary_anbima.pack(fill='y', side='right')

    colunas = colunas_anbima
    query = "SELECT Data, DiadaSemana, FeriadoAnmbima FROM feriados_ANBIMA"
    data_col_idx = 0

    scrollbar_x_anbima = ctk.CTkScrollbar(frame_calendario_anbima, orientation='horizontal')
    scrollbar_y_anbima = ctk.CTkScrollbar(frame_scrollbary_anbima, orientation='vertical')

    tabela_anbima = ttk.Treeview(frame_calendario_anbima, columns=colunas, show='headings', xscrollcommand=scrollbar_x_anbima.set, yscrollcommand=scrollbar_y_anbima.set)
    for coluna in colunas:
        tabela_anbima.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_anbima, _col, False))
    tabela_anbima.pack(expand=True, fill='both')

    scrollbar_x_anbima.configure(command=tabela_anbima.xview, height=25)
    scrollbar_y_anbima.configure(command=tabela_anbima.yview, width=25)
    scrollbar_x_anbima.pack(side='bottom', fill='x')
    scrollbar_y_anbima.pack(side='right', fill='y')

    carregar_calendario(tabela_anbima, query)
    ajustar_largura_colunas(tabela_anbima, colunas, tabview)

    tabela_feriados[nome_sub_aba] = tabela_anbima  # Armazenar a tabela no dicionário

    dados_originais_anbima = [tabela_anbima.item(item, "values") for item in tabela_anbima.get_children()]

    frame_filtro_anbima = ctk.CTkFrame(frame_calendario_anbima)
    frame_filtro_anbima = ctk.CTkFrame(frame_calendario_anbima)
    frame_filtro_anbima.pack(fill='x', pady=5)

    entry_filtro_data_anbima = ctk.CTkEntry(frame_filtro_anbima, width=100, placeholder_text="Filtrar por Data")
    entry_filtro_data_anbima.pack(side='left', padx=10)

    entry_filtro_dia_anbima = ctk.CTkEntry(frame_filtro_anbima, width=100, placeholder_text="Filtrar por Dia")
    entry_filtro_dia_anbima.pack(side='left', padx=10)

    entry_filtro_feriado_anbima = ctk.CTkEntry(frame_filtro_anbima, width=120, placeholder_text="Filtrar por Feriado")
    entry_filtro_feriado_anbima.pack(side='left', padx=10)

    def filtrar_data_anbima(event):
        filtro = entry_filtro_data_anbima.get().lower()
        tabela_anbima.delete(*tabela_anbima.get_children())
        for item in dados_originais_anbima:
            if filtro in item[0].lower():
                tabela_anbima.insert("", "end", values=item)

    def filtrar_dia_anbima(event):
        filtro = entry_filtro_dia_anbima.get().lower()
        tabela_anbima.delete(*tabela_anbima.get_children())
        for item in dados_originais_anbima:
            if filtro in item[1].lower():
                tabela_anbima.insert("", "end", values=item)

    def filtrar_feriado_anbima(event):
        filtro = entry_filtro_feriado_anbima.get().lower()
        tabela_anbima.delete(*tabela_anbima.get_children())
        for item in dados_originais_anbima:
            if filtro in item[2].lower():
                tabela_anbima.insert("", "end", values=item)

    entry_filtro_data_anbima.bind("<KeyRelease>", filtrar_data_anbima)
    entry_filtro_dia_anbima.bind("<KeyRelease>", filtrar_dia_anbima)
    entry_filtro_feriado_anbima.bind("<KeyRelease>", filtrar_feriado_anbima)

    # Bloco para ICE
    nome_sub_aba = "ICE"
    tabview_calendarios.add(nome_sub_aba)
    sub_aba = tabview_calendarios.tab(nome_sub_aba)

    frame_calendario_ice = ctk.CTkFrame(sub_aba)
    frame_calendario_ice.pack(expand=True, fill='both')

    frame_scrollbary_ice = ctk.CTkFrame(frame_calendario_ice, width=4)
    frame_scrollbary_ice.pack(fill='y', side='right')

    colunas = colunas_mercadoria
    feriado_coluna = f"Feriado{nome_sub_aba}US"
    query = f"SELECT Commodity, Data, DiadaSemana, {feriado_coluna} FROM feriados_{nome_sub_aba.upper()}"
    data_col_idx = 1

    scrollbar_x_ice = ctk.CTkScrollbar(frame_calendario_ice, orientation='horizontal')
    scrollbar_y_ice = ctk.CTkScrollbar(frame_scrollbary_ice, orientation='vertical')

    tabela_ice = ttk.Treeview(frame_calendario_ice, columns=colunas, show='headings', xscrollcommand=scrollbar_x_ice.set, yscrollcommand=scrollbar_y_ice.set)
    for coluna in colunas:
        tabela_ice.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_ice, _col, False))
    tabela_ice.pack(expand=True, fill='both')

    scrollbar_x_ice.configure(command=tabela_ice.xview, height=25)
    scrollbar_y_ice.configure(command=tabela_ice.yview, width=25)
    scrollbar_x_ice.pack(side='bottom', fill='x')
    scrollbar_y_ice.pack(side='right', fill='y')

    carregar_calendario(tabela_ice, query)
    ajustar_largura_colunas(tabela_ice, colunas, tabview)

    tabela_feriados[nome_sub_aba] = tabela_ice  # Armazenar a tabela no dicionário

    dados_originais_ice = [tabela_ice.item(item, "values") for item in tabela_ice.get_children()]

    frame_filtro_ice = ctk.CTkFrame(frame_calendario_ice)
    frame_filtro_ice.pack(fill='x', pady=5)

    entry_filtro_comm_ice = ctk.CTkEntry(frame_filtro_ice, width=135, placeholder_text="Filtrar por Commodity")
    entry_filtro_comm_ice.pack(side='left', padx=10)

    entry_filtro_data_ice = ctk.CTkEntry(frame_filtro_ice, width=100, placeholder_text="Filtrar por Data")
    entry_filtro_data_ice.pack(side='left', padx=10)

    entry_filtro_dia_ice = ctk.CTkEntry(frame_filtro_ice, width=100, placeholder_text="Filtrar por Dia")
    entry_filtro_dia_ice.pack(side='left', padx=10)

    entry_filtro_feriado_ice = ctk.CTkEntry(frame_filtro_ice, width=120, placeholder_text="Filtrar por Feriado")
    entry_filtro_feriado_ice.pack(side='left', padx=10)

    def filtrar_comm_ice(event):
        filtro = entry_filtro_comm_ice.get().lower()
        tabela_ice.delete(*tabela_ice.get_children())
        for item in dados_originais_ice:
            if filtro in item[0].lower():
                tabela_ice.insert("", "end", values=item)

    def filtrar_dia_ice(event):
        filtro = entry_filtro_dia_ice.get().lower()
        tabela_ice.delete(*tabela_ice.get_children())
        for item in dados_originais_ice:
            if filtro in item[2].lower():
                tabela_ice.insert("", "end", values=item)

    def filtrar_data_ice(event):
        filtro = entry_filtro_data_ice.get().lower()
        tabela_ice.delete(*tabela_ice.get_children())
        for item in dados_originais_ice:
            if filtro in item[1].lower():
                tabela_ice.insert("", "end", values=item)

    def filtrar_feriado_ice(event):
        filtro = entry_filtro_feriado_ice.get().lower()
        tabela_ice.delete(*tabela_ice.get_children())
        for item in dados_originais_ice:
            if filtro in item[3].lower():
                tabela_ice.insert("", "end", values=item)

    entry_filtro_comm_ice.bind("<KeyRelease>", filtrar_comm_ice)
    entry_filtro_data_ice.bind("<KeyRelease>", filtrar_data_ice)
    entry_filtro_dia_ice.bind("<KeyRelease>", filtrar_dia_ice)
    entry_filtro_feriado_ice.bind("<KeyRelease>", filtrar_feriado_ice)

    # Bloco para NYMEX
    nome_sub_aba = "NYMEX"
    tabview_calendarios.add(nome_sub_aba)
    sub_aba = tabview_calendarios.tab(nome_sub_aba)

    frame_calendario_nymex = ctk.CTkFrame(sub_aba)
    frame_calendario_nymex.pack(expand=True, fill='both')

    frame_scrollbary_nymex = ctk.CTkFrame(frame_calendario_nymex, width=4)
    frame_scrollbary_nymex.pack(fill='y', side='right')

    colunas = colunas_mercadoria
    feriado_coluna = f"Feriado{nome_sub_aba}"
    query = f"SELECT Commodity, Data, DiadaSemana, {feriado_coluna} FROM feriados_{nome_sub_aba.upper()}"
    data_col_idx = 1

    scrollbar_x_nymex = ctk.CTkScrollbar(frame_calendario_nymex, orientation='horizontal')
    scrollbar_y_nymex = ctk.CTkScrollbar(frame_scrollbary_nymex, orientation='vertical')

    tabela_nymex = ttk.Treeview(frame_calendario_nymex, columns=colunas, show='headings', xscrollcommand=scrollbar_x_nymex.set, yscrollcommand=scrollbar_y_nymex.set)
    for coluna in colunas:
        tabela_nymex.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_nymex, _col, False))
    tabela_nymex.pack(expand=True, fill='both')

    scrollbar_x_nymex.configure(command=tabela_nymex.xview, height=25)
    scrollbar_y_nymex.configure(command=tabela_nymex.yview, width=25)
    scrollbar_x_nymex.pack(side='bottom', fill='x')
    scrollbar_y_nymex.pack(side='right', fill='y')

    carregar_calendario(tabela_nymex, query)
    ajustar_largura_colunas(tabela_nymex, colunas, tabview)

    tabela_feriados[nome_sub_aba] = tabela_nymex  # Armazenar a tabela no dicionário

    dados_originais_nymex = [tabela_nymex.item(item, "values") for item in tabela_nymex.get_children()]

    frame_filtro_nymex = ctk.CTkFrame(frame_calendario_nymex)
    frame_filtro_nymex.pack(fill='x', pady=5)

    entry_filtro_comm_nymex = ctk.CTkEntry(frame_filtro_nymex, width=135, placeholder_text="Filtrar por Commodity")
    entry_filtro_comm_nymex.pack(side='left', padx=10)

    entry_filtro_data_nymex = ctk.CTkEntry(frame_filtro_nymex, width=100, placeholder_text="Filtrar por Data")
    entry_filtro_data_nymex.pack(side='left', padx=10)

    entry_filtro_dia_nymex = ctk.CTkEntry(frame_filtro_nymex, width=100, placeholder_text="Filtrar por Dia")
    entry_filtro_dia_nymex.pack(side='left', padx=10)

    entry_filtro_feriado_nymex = ctk.CTkEntry(frame_filtro_nymex, width=120, placeholder_text="Filtrar por Feriado")
    entry_filtro_feriado_nymex.pack(side='left', padx=10)

    def filtrar_comm_nymex(event):
        filtro = entry_filtro_comm_nymex.get().lower()
        tabela_nymex.delete(*tabela_nymex.get_children())
        for item in dados_originais_nymex:
            if filtro in item[0].lower():
                tabela_nymex.insert("", "end", values=item)

    def filtrar_dia_nymex(event):
        filtro = entry_filtro_dia_nymex.get().lower()
        tabela_nymex.delete(*tabela_nymex.get_children())
        for item in dados_originais_nymex:
            if filtro in item[2].lower():
                tabela_nymex.insert("", "end", values=item)

    def filtrar_data_nymex(event):
        filtro = entry_filtro_data_nymex.get().lower()
        tabela_nymex.delete(*tabela_nymex.get_children())
        for item in dados_originais_nymex:
            if filtro in item[1].lower():
                tabela_nymex.insert("", "end", values=item)
    def filtrar_feriado_nymex(event):
        filtro = entry_filtro_feriado_nymex.get().lower()
        tabela_nymex.delete(*tabela_nymex.get_children())
        for item in dados_originais_nymex:
            if filtro in item[3].lower():
                tabela_nymex.insert("", "end", values=item)

    entry_filtro_feriado_nymex.bind("<KeyRelease>", filtrar_feriado_nymex)
    entry_filtro_comm_nymex.bind("<KeyRelease>", filtrar_comm_nymex)
    entry_filtro_data_nymex.bind("<KeyRelease>", filtrar_data_nymex)
    entry_filtro_dia_nymex.bind("<KeyRelease>", filtrar_dia_nymex)

    # Bloco para BURSA
    nome_sub_aba = "BURSA"
    tabview_calendarios.add(nome_sub_aba)
    sub_aba = tabview_calendarios.tab(nome_sub_aba)

    frame_calendario_bursa = ctk.CTkFrame(sub_aba)
    frame_calendario_bursa.pack(expand=True, fill='both')

    frame_scrollbary_bursa = ctk.CTkFrame(frame_calendario_bursa, width=4)
    frame_scrollbary_bursa.pack(fill='y', side='right')

    colunas = colunas_mercadoria
    feriado_coluna = f"Feriado{nome_sub_aba}"
    query = f"SELECT Commodity, Data, DiadaSemana, {feriado_coluna} FROM feriados_{nome_sub_aba.upper()}"
    data_col_idx = 1

    scrollbar_x_bursa = ctk.CTkScrollbar(frame_calendario_bursa, orientation='horizontal')
    scrollbar_y_bursa = ctk.CTkScrollbar(frame_scrollbary_bursa, orientation='vertical')

    tabela_bursa = ttk.Treeview(frame_calendario_bursa, columns=colunas, show='headings', xscrollcommand=scrollbar_x_bursa.set, yscrollcommand=scrollbar_y_bursa.set)
    for coluna in colunas:
        tabela_bursa.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_bursa, _col, False))
    tabela_bursa.pack(expand=True, fill='both')

    scrollbar_x_bursa.configure(command=tabela_bursa.xview, height=25)
    scrollbar_y_bursa.configure(command=tabela_bursa.yview, width=25)
    scrollbar_x_bursa.pack(side='bottom', fill='x')
    scrollbar_y_bursa.pack(side='right', fill='y')

    carregar_calendario(tabela_bursa, query)
    ajustar_largura_colunas(tabela_bursa, colunas, tabview)

    tabela_feriados[nome_sub_aba] = tabela_bursa  # Armazenar a tabela no dicionário

    dados_originais_bursa = [tabela_bursa.item(item, "values") for item in tabela_bursa.get_children()]

    frame_filtro_bursa = ctk.CTkFrame(frame_calendario_bursa)
    frame_filtro_bursa.pack(fill='x', pady=5)

    entry_filtro_comm_bursa = ctk.CTkEntry(frame_filtro_bursa, width=135, placeholder_text="Filtrar por Commodity")
    entry_filtro_comm_bursa.pack(side='left', padx=10)

    entry_filtro_data_bursa = ctk.CTkEntry(frame_filtro_bursa, width=100, placeholder_text="Filtrar por Data")
    entry_filtro_data_bursa.pack(side='left', padx=10)

    entry_filtro_dia_bursa = ctk.CTkEntry(frame_filtro_bursa, width=100, placeholder_text="Filtrar por Dia")
    entry_filtro_dia_bursa.pack(side='left', padx=10)

    entry_filtro_feriado_bursa = ctk.CTkEntry(frame_filtro_bursa, width=120, placeholder_text="Filtrar por Feriado")
    entry_filtro_feriado_bursa.pack(side='left', padx=10)

    def filtrar_feriado_bursa(event):
        filtro = entry_filtro_feriado_bursa.get().lower()
        tabela_bursa.delete(*tabela_bursa.get_children())
        for item in dados_originais_bursa:
            if filtro in item[3].lower():
                tabela_bursa.insert("", "end", values=item)

    def filtrar_comm_bursa(event):
        filtro = entry_filtro_comm_bursa.get().lower()
        tabela_bursa.delete(*tabela_bursa.get_children())
        for item in dados_originais_bursa:
            if filtro in item[0].lower():
                tabela_bursa.insert("", "end", values=item)

    def filtrar_dia_bursa(event):
        filtro = entry_filtro_dia_bursa.get().lower()
        tabela_bursa.delete(*tabela_bursa.get_children())
        for item in dados_originais_bursa:
            if filtro in item[2].lower():
                tabela_bursa.insert("", "end", values=item)

    def filtrar_data_bursa(event):
        filtro = entry_filtro_data_bursa.get().lower()
        tabela_bursa.delete(*tabela_bursa.get_children())
        for item in dados_originais_bursa:
            if filtro in item[1].lower():
                tabela_bursa.insert("", "end", values=item)

    entry_filtro_comm_bursa.bind("<KeyRelease>", filtrar_comm_bursa)
    entry_filtro_data_bursa.bind("<KeyRelease>", filtrar_data_bursa)
    entry_filtro_dia_bursa.bind("<KeyRelease>", filtrar_dia_bursa)
    entry_filtro_feriado_bursa.bind("<KeyRelease>", filtrar_feriado_bursa)

    # Bloco para CBOT
    nome_sub_aba = "CBOT"
    tabview_calendarios.add(nome_sub_aba)
    sub_aba = tabview_calendarios.tab(nome_sub_aba)

    frame_calendario_cbot = ctk.CTkFrame(sub_aba)
    frame_calendario_cbot.pack(expand=True, fill='both')

    frame_scrollbary_cbot = ctk.CTkFrame(frame_calendario_cbot, width=4)
    frame_scrollbary_cbot.pack(fill='y', side='right')

    colunas = colunas_mercadoria
    feriado_coluna = f"Feriado{nome_sub_aba}"
    query = f"SELECT Commodity, Data, DiadaSemana, {feriado_coluna} FROM feriados_{nome_sub_aba.upper()}"
    data_col_idx = 1

    scrollbar_x_cbot = ctk.CTkScrollbar(frame_calendario_cbot, orientation='horizontal')
    scrollbar_y_cbot = ctk.CTkScrollbar(frame_scrollbary_cbot, orientation='vertical')

    tabela_cbot = ttk.Treeview(frame_calendario_cbot, columns=colunas, show='headings', xscrollcommand=scrollbar_x_cbot.set, yscrollcommand=scrollbar_y_cbot.set)
    for coluna in colunas:
                tabela_cbot.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_cbot, _col, False))
    tabela_cbot.pack(expand=True, fill='both')

    scrollbar_x_cbot.configure(command=tabela_cbot.xview, height=25)
    scrollbar_y_cbot.configure(command=tabela_cbot.yview, width=25)
    scrollbar_x_cbot.pack(side='bottom', fill='x')
    scrollbar_y_cbot.pack(side='right', fill='y')

    carregar_calendario(tabela_cbot, query)
    ajustar_largura_colunas(tabela_cbot, colunas, tabview)

    tabela_feriados[nome_sub_aba] = tabela_cbot  # Armazenar a tabela no dicionário

    dados_originais_cbot = [tabela_cbot.item(item, "values") for item in tabela_cbot.get_children()]

    frame_filtro_cbot = ctk.CTkFrame(frame_calendario_cbot)
    frame_filtro_cbot.pack(fill='x', pady=5)

    entry_filtro_comm_cbot = ctk.CTkEntry(frame_filtro_cbot, width=135, placeholder_text="Filtrar por Commodity")
    entry_filtro_comm_cbot.pack(side='left', padx=10)

    entry_filtro_data_cbot = ctk.CTkEntry(frame_filtro_cbot, width=100, placeholder_text="Filtrar por Data")
    entry_filtro_data_cbot.pack(side='left', padx=10)

    entry_filtro_dia_cbot = ctk.CTkEntry(frame_filtro_cbot, width=100, placeholder_text="Filtrar por Dia")
    entry_filtro_dia_cbot.pack(side='left', padx=10)

    entry_filtro_feriado_cbot = ctk.CTkEntry(frame_filtro_cbot, width=120, placeholder_text="Filtrar por Feriado")
    entry_filtro_feriado_cbot.pack(side='left', padx=10)

    def filtrar_feriado_cbot(event):
        filtro = entry_filtro_feriado_cbot.get().lower()
        tabela_cbot.delete(*tabela_cbot.get_children())
        for item in dados_originais_cbot:
            if filtro in item[3].lower():
                tabela_cbot.insert("", "end", values=item)

    def filtrar_comm_cbot(event):
        filtro = entry_filtro_comm_cbot.get().lower()
        tabela_cbot.delete(*tabela_cbot.get_children())
        for item in dados_originais_cbot:
            if filtro in item[0].lower():
                tabela_cbot.insert("", "end", values=item)

    def filtrar_dia_cbot(event):
        filtro = entry_filtro_dia_cbot.get().lower()
        tabela_cbot.delete(*tabela_cbot.get_children())
        for item in dados_originais_cbot:
            if filtro in item[2].lower():
                tabela_cbot.insert("", "end", values=item)

    def filtrar_data_cbot(event):
        filtro = entry_filtro_data_cbot.get().lower()
        tabela_cbot.delete(*tabela_cbot.get_children())
        for item in dados_originais_cbot:
            if filtro in item[1].lower():
                tabela_cbot.insert("", "end", values=item)

    entry_filtro_comm_cbot.bind("<KeyRelease>", filtrar_comm_cbot)
    entry_filtro_data_cbot.bind("<KeyRelease>", filtrar_data_cbot)
    entry_filtro_dia_cbot.bind("<KeyRelease>", filtrar_dia_cbot)
    entry_filtro_feriado_cbot.bind("<KeyRelease>", filtrar_feriado_cbot)

    # Bloco para PLATTS
    nome_sub_aba = "PLATTS"
    tabview_calendarios.add(nome_sub_aba)
    sub_aba = tabview_calendarios.tab(nome_sub_aba)

    frame_calendario_platts = ctk.CTkFrame(sub_aba)
    frame_calendario_platts.pack(expand=True, fill='both')

    frame_scrollbary_platts = ctk.CTkFrame(frame_calendario_platts, width=4)
    frame_scrollbary_platts.pack(fill='y', side='right')

    colunas = colunas_mercadoria
    feriado_coluna = f"Feriado{nome_sub_aba}"
    query = f"SELECT Commodity, Data, DiadaSemana, {feriado_coluna} FROM feriados_{nome_sub_aba.upper()}"
    data_col_idx = 1

    scrollbar_x_platts = ctk.CTkScrollbar(frame_calendario_platts, orientation='horizontal')
    scrollbar_y_platts = ctk.CTkScrollbar(frame_scrollbary_platts, orientation='vertical')

    tabela_platts = ttk.Treeview(frame_calendario_platts, columns=colunas, show='headings', xscrollcommand=scrollbar_x_platts.set, yscrollcommand=scrollbar_y_platts.set)
    for coluna in colunas:
        tabela_platts.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_platts, _col, False))
    tabela_platts.pack(expand=True, fill='both')

    scrollbar_x_platts.configure(command=tabela_platts.xview, height=25)
    scrollbar_y_platts.configure(command=tabela_platts.yview, width=25)
    scrollbar_x_platts.pack(side='bottom', fill='x')
    scrollbar_y_platts.pack(side='right', fill='y')

    carregar_calendario(tabela_platts, query)
    ajustar_largura_colunas(tabela_platts, colunas, tabview)

    tabela_feriados[nome_sub_aba] = tabela_platts  # Armazenar a tabela no dicionário

    dados_originais_platts = [tabela_platts.item(item, "values") for item in tabela_platts.get_children()]

    frame_filtro_platts = ctk.CTkFrame(frame_calendario_platts)
    frame_filtro_platts.pack(fill='x', pady=5)

    entry_filtro_comm_platts = ctk.CTkEntry(frame_filtro_platts, width=135, placeholder_text="Filtrar por Commodity")
    entry_filtro_comm_platts.pack(side='left', padx=10)

    entry_filtro_data_platts = ctk.CTkEntry(frame_filtro_platts, width=100, placeholder_text="Filtrar por Data")
    entry_filtro_data_platts.pack(side='left', padx=10)

    entry_filtro_dia_platts = ctk.CTkEntry(frame_filtro_platts, width=100, placeholder_text="Filtrar por Dia")
    entry_filtro_dia_platts.pack(side='left', padx=10)

    entry_filtro_feriado_platts = ctk.CTkEntry(frame_filtro_platts, width=120, placeholder_text="Filtrar por Feriado")
    entry_filtro_feriado_platts.pack(side='left', padx=10)

    def filtrar_feriado_platts(event):
        filtro = entry_filtro_feriado_platts.get().lower()
        tabela_platts.delete(*tabela_platts.get_children())
        for item in dados_originais_platts:
            if filtro in item[3].lower():
                tabela_platts.insert("", "end", values=item)

    def filtrar_comm_platts(event):
        filtro = entry_filtro_comm_platts.get().lower()
        tabela_platts.delete(*tabela_platts.get_children())
        for item in dados_originais_platts:
            if filtro in item[0].lower():
                tabela_platts.insert("", "end", values=item)

    def filtrar_dia_platts(event):
        filtro = entry_filtro_dia_platts.get().lower()
        tabela_platts.delete(*tabela_platts.get_children())
        for item in dados_originais_platts:
            if filtro in item[2].lower():
                tabela_platts.insert("", "end", values=item)

    def filtrar_data_platts(event):
        filtro = entry_filtro_data_platts.get().lower()
        tabela_platts.delete(*tabela_platts.get_children())
        for item in dados_originais_platts:
            if filtro in item[1].lower():
                tabela_platts.insert("", "end", values=item)

    entry_filtro_comm_platts.bind("<KeyRelease>", filtrar_comm_platts)
    entry_filtro_data_platts.bind("<KeyRelease>", filtrar_data_platts)
    entry_filtro_dia_platts.bind("<KeyRelease>", filtrar_dia_platts)
    entry_filtro_feriado_platts.bind("<KeyRelease>", filtrar_feriado_platts)

    # Bloco para LME
    nome_sub_aba = "LME"
    tabview_calendarios.add(nome_sub_aba)
    sub_aba = tabview_calendarios.tab(nome_sub_aba)

    frame_calendario_lme = ctk.CTkFrame(sub_aba)
    frame_calendario_lme.pack(expand=True, fill='both')

    frame_scrollbary_lme = ctk.CTkFrame(frame_calendario_lme, width=4)
    frame_scrollbary_lme.pack(fill='y', side='right')

    colunas = colunas_mercadoria
    feriado_coluna = f"Feriado{nome_sub_aba}"
    query = f"SELECT Commodity, Data, DiadaSemana, {feriado_coluna} FROM feriados_{nome_sub_aba.upper()}"
    data_col_idx = 1

    scrollbar_x_lme = ctk.CTkScrollbar(frame_calendario_lme, orientation='horizontal')
    scrollbar_y_lme = ctk.CTkScrollbar(frame_scrollbary_lme, orientation='vertical')

    tabela_lme = ttk.Treeview(frame_calendario_lme, columns=colunas, show='headings', xscrollcommand=scrollbar_x_lme.set, yscrollcommand=scrollbar_y_lme.set)
    for coluna in colunas:
        tabela_lme.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_lme, _col, False))
    tabela_lme.pack(expand=True, fill='both')

    scrollbar_x_lme.configure(command=tabela_lme.xview, height=25)
    scrollbar_y_lme.configure(command=tabela_lme.yview, width=25)
    scrollbar_x_lme.pack(side='bottom', fill='x')
    scrollbar_y_lme.pack(side='right', fill='y')

    carregar_calendario(tabela_lme, query)
    ajustar_largura_colunas(tabela_lme, colunas, tabview)

    tabela_feriados[nome_sub_aba] = tabela_lme  # Armazenar a tabela no dicionário

    dados_originais_lme = [tabela_lme.item(item, "values") for item in tabela_lme.get_children()]

    frame_filtro_lme = ctk.CTkFrame(frame_calendario_lme)
    frame_filtro_lme.pack(fill='x', pady=5)

    entry_filtro_comm_lme = ctk.CTkEntry(frame_filtro_lme, width=135, placeholder_text="Filtrar por Commodity")
    entry_filtro_comm_lme.pack(side='left', padx=10)

    entry_filtro_data_lme = ctk.CTkEntry(frame_filtro_lme, width=100, placeholder_text="Filtrar por Data")
    entry_filtro_data_lme.pack(side='left', padx=10)

    entry_filtro_dia_lme = ctk.CTkEntry(frame_filtro_lme, width=100, placeholder_text="Filtrar por Dia")
    entry_filtro_dia_lme.pack(side='left', padx=10)

    entry_filtro_feriado_lme = ctk.CTkEntry(frame_filtro_lme, width=120, placeholder_text="Filtrar por Feriado")
    entry_filtro_feriado_lme.pack(side='left', padx=10)

    def filtrar_feriado_lme(event):
        filtro = entry_filtro_feriado_lme.get().lower()
        tabela_lme.delete(*tabela_lme.get_children())
        for item in dados_originais_lme:
            if filtro in item[3].lower():
                tabela_lme.insert("", "end", values=item)
    

    def filtrar_comm_lme(event):
            filtro = filtrar_comm_lme.get().lower()
            tabela_lme.delete(*tabela_lme.get_children())
            for item in dados_originais_lme:
                if filtro in item[0].lower():
                    tabela_lme.insert("", "end", values=item) 

    def filtrar_dia_lme(event):
        filtro = entry_filtro_dia_lme.get().lower()
        tabela_lme.delete(*tabela_lme.get_children())
        for item in dados_originais_lme:
            if filtro in item[2].lower():
                tabela_lme.insert("", "end", values=item)  

    def filtrar_data_lme(event):
        filtro = entry_filtro_data_lme.get().lower()
        tabela_lme.delete(*tabela_lme.get_children())
        for item in dados_originais_lme:
            if filtro in item[1].lower():
                tabela_lme.insert("", "end", values=item)   

    entry_filtro_comm_lme.bind("<KeyRelease>", filtrar_comm_lme)
    entry_filtro_data_lme.bind("<KeyRelease>", filtrar_data_lme)
    entry_filtro_dia_lme.bind("<KeyRelease>", filtrar_dia_lme)
    entry_filtro_feriado_lme.bind("<KeyRelease>", filtrar_feriado_lme)

    # Armazena as tabelas globalmente para acesso futuro
    tabela_feriados_global = tabela_feriados

    return tabela_feriados

    

def chamar_preencher_fixings(tabview, abas_existentes):
    preencher_fixings(tabview, abas_existentes)
    messagebox.showinfo("Sucesso","Datas de Fixings preenchidas!")



def preencher_fixings(tabview, abas_existentes):
    # Extrair dados necessários das tabelas
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()       
    if "Holidays" not in abas_existentes:
        abas_existentes.append("Holidays")
    else:
        tabela_fixingsopcao_cliente.delete(*tabela_fixingsopcao_cliente.get_children())
        tabela_fixingsopcao_b2b.delete(*tabela_fixingsopcao_b2b.get_children())
        tabela_fixingstermo_cliente.delete(*tabela_fixingstermo_cliente.get_children())
        tabela_fixingstermo_b2b.delete(*tabela_fixingstermo_b2b.get_children())
        
    tabela_feriados = calendarios_bolsas(tabview, abas_existentes)
    entry_filtro_commodities.delete(0, tk.END)
    # Processar dados para tabela_opcao_cliente
    for item in tabela_opcao_cliente.get_children():
        values_m = tabela_opcao_cliente.item(item, 'values')
        fixing_start_date = datetime.strptime(values_m[16].strip(), "%d-%b-%Y")
        fixing_end_date = datetime.strptime(values_m[17].strip(), "%d-%b-%Y")
        if fixing_start_date != fixing_end_date:
            market = values_m[2]
            bolsa_commodity = lookup(market, commodities_ric, commodities_exchange)
            commodity_type = lookup(market, commodities_ric, commodities_mercadoria)

            linha = [
                values_m[0],  # ATHENA ID
                bolsa_commodity,  # BOLSA
                commodity_type,  # COMMODITY
                values_m[9].replace("-", ""),  # NOTIONAL
                "",  # Placeholder para a quantidade de datas
                safe_date_conversion_dd_MM(values_m[16]),
                custom_workday_ex(fixing_start_date, bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date)
            ]
            new_item_id = tabela_fixingsopcao_cliente.insert("", "end", values=linha)

            # Atualizar valores nas colunas subsequentes para a linha recém-inserida
            values = list(tabela_fixingsopcao_cliente.item(new_item_id, 'values'))

            # Certifique-se de que a lista tem o tamanho necessário
            while len(values) < 54:
                values.append("")

            for i in range(6, 54):
                if i < len(values) and i-1 < len(values):
                    values[i] = custom_workday_ex(datetime.strptime(values[i-1], "%d/%m/%Y"), bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date) if values[i-1] != "" else ""

            # Contar a quantidade de campos não vazios no range 7 ao 50
            quantidade_datas = sum(1 for v in values[5:50] if v != "")
            values[4] = quantidade_datas  # Atualizar a coluna 5 com a contagem
            values[50] = values_m[-4]
            values[51] = values_m[-3]
            values[52] = values_m[-2]
            values[53] = values_m[-1]
            tabela_fixingsopcao_cliente.item(new_item_id, values=values)
            ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)

    # Processar dados para tabela_opcao_b2b
    for item in tabela_opcao_b2b.get_children():
        values_m = tabela_opcao_b2b.item(item, 'values')
        fixing_start_date = datetime.strptime(values_m[16].strip(), "%d-%b-%Y")
        fixing_end_date = datetime.strptime(values_m[17].strip(), "%d-%b-%Y")

        if fixing_start_date != fixing_end_date:
            market = values_m[2]
            bolsa_commodity = lookup(market, commodities_ric, commodities_exchange)
            commodity_type = lookup(market, commodities_ric, commodities_mercadoria)

            linha = [
                values_m[0],  # ATHENA ID
                bolsa_commodity,  # BOLSA
                commodity_type,  # COMMODITY
                values_m[9].replace("-", ""),  # NOTIONAL
                "",  # Placeholder para a quantidade de datas
                safe_date_conversion_dd_MM(values_m[16]),
                custom_workday_ex(fixing_start_date, bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date)
            ]
            new_item_id = tabela_fixingsopcao_b2b.insert("", "end", values=linha)

            # Atualizar valores nas colunas subsequentes para a linha recém-inserida
            values = list(tabela_fixingsopcao_b2b.item(new_item_id, 'values'))
            
            # Certifique-se de que a lista tem o tamanho necessário
            while len(values) < 54:
                values.append("")

            for i in range(6, 54):
                if i < len(values) and i-1 < len(values):
                    values[i] = custom_workday_ex(datetime.strptime(values[i-1], "%d/%m/%Y"), bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date) if values[i-1] != "" else ""

            # Contar a quantidade de campos não vazios no range 7 ao 50
            quantidade_datas = sum(1 for v in values[5:50] if v != "")
            values[4] = quantidade_datas  # Atualizar a coluna 5 com a contagem
            values[50] = values_m[-4]
            values[51] = values_m[-3]
            values[52] = values_m[-2]
            values[53] = values_m[-1]
            tabela_fixingsopcao_b2b.item(new_item_id, values=values)
            ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)

    vincular_evento_duplo_clique(tabela_fixingstermo_cliente, colunas_fixings_termo)
    vincular_evento_duplo_clique(tabela_fixingstermo_b2b, colunas_fixings_termo)
    vincular_evento_duplo_clique(tabela_fixingsopcao_b2b, colunas_fixings_opcao)
    vincular_evento_duplo_clique(tabela_fixingsopcao_cliente, colunas_fixings_opcao)

    # Processar dados para tabela_termo_cliente
    for item in tabela_termo_cliente.get_children():
        values_m = tabela_termo_cliente.item(item, 'values')
        fixing_start_date = datetime.strptime(values_m[16].strip(), "%d-%b-%Y")
        fixing_end_date = datetime.strptime(values_m[17].strip(), "%d-%b-%Y")

        if fixing_start_date != fixing_end_date:
            market = values_m[2]
            bolsa_commodity = lookup(market, commodities_ric, commodities_exchange)
            commodity_type = lookup(market, commodities_ric, commodities_mercadoria)

            linha = [
                values_m[0],  # ATHENA ID
                bolsa_commodity,  # BOLSA
                commodity_type,  # COMMODITY
                "SIM" if values_m[8] == "BRR" else "NÃO",  # TAXA BRR
                "",  # Placeholder para a quantidade de datas
                safe_date_conversion_dd_MM(values_m[16]),
                custom_workday_ex(fixing_start_date, bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date)
            ]
            new_item_id = tabela_fixingstermo_cliente.insert("", "end", values=linha)

            # Atualizar valores nas colunas subsequentes para a linha recém-inserida
            values = list(tabela_fixingstermo_cliente.item(new_item_id, 'values'))
            
            # Certifique-se de que a lista tem o tamanho necessário
            while len(values) < 89:
                values.append("")

            for i in range(6, 89):
                if i < len(values) and i-1 < len(values):
                    values[i] = custom_workday_ex(datetime.strptime(values[i-1], "%d/%m/%Y"), bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date) if values[i-1] != "" else ""

            # Contar a quantidade de campos não vazios no range 7 ao 50
            quantidade_datas = sum(1 for v in values[5:85] if v != "")
            values[4] = quantidade_datas  # Atualizar a coluna 5 com a contagem
            values[85] = values_m[-4]
            values[86] = values_m[-3]
            values[87] = values_m[-2]
            values[88] = values_m[-1]

            tabela_fixingstermo_cliente.item(new_item_id, values=values)
            ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)

    # Processar dados para tabela_termo_b2b
    for item in tabela_termo_b2b.get_children():
        values_m = tabela_termo_b2b.item(item, 'values')
        fixing_start_date = datetime.strptime(values_m[16].strip(), "%d-%b-%Y")
        fixing_end_date = datetime.strptime(values_m[17].strip(), "%d-%b-%Y")

        if fixing_start_date != fixing_end_date:
            market = values_m[2]
            bolsa_commodity = lookup(market, commodities_ric, commodities_exchange)
            commodity_type = lookup(market, commodities_ric, commodities_mercadoria)

            linha = [
                values_m[0],  # ATHENA ID
                bolsa_commodity,  # BOLSA
                commodity_type,  # COMMODITY
                "SIM" if values_m[8] == "BRR" else "NÃO",  # TAXA BRR
                "",  # Placeholder para a quantidade de datas
                safe_date_conversion_dd_MM(values_m[16]),
                custom_workday_ex(fixing_start_date, bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date)
            ]
            new_item_id = tabela_fixingstermo_b2b.insert("", "end", values=linha)

            # Atualizar valores nas colunas subsequentes para a linha recém-inserida
            values = list(tabela_fixingstermo_b2b.item(new_item_id, 'values'))
            
            # Certifique-se de que a lista tem o tamanho necessário
            while len(values) < 89:
                values.append("")

            for i in range(6, 89):
                if i < len(values) and i-1 < len(values):
                    values[i] = custom_workday_ex(datetime.strptime(values[i-1], "%d/%m/%Y"), bolsa_commodity, commodity_type, tabela_feriados, fixing_end_date) if values[i-1] != "" else ""

            # Contar a quantidade de campos não vazios no range 7 ao 50
            quantidade_datas = sum(1 for v in values[5:86] if v != "")
            values[4] = quantidade_datas  # Atualizar a coluna 5 com a contagem
            values[85] = values_m[-4]
            values[86] = values_m[-3]
            values[87] = values_m[-2]
            values[88] = values_m[-1]

            tabela_fixingstermo_b2b.item(new_item_id, values=values)
            ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo') 
def filtrar_commodities_export(tabela, filtro, coluna_index, tabela_commodities_data):
    tabela.delete(*tabela.get_children())
    for item in tabela_commodities_data:
        if filtro in item[coluna_index].lower():
            tabela.insert("", "end", values=item)

def ajustar_colunas_sheet(sheet, headers):
# Ajusta a largura das colunas fixings_cliente
    for col_idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)), start=1):
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2.5)  # Ajuste para espaçamento
        sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width 

def gerar_excel_termo(abas_existentes, downloads_path, cliente, accronym, tabela_termo_cliente, tabela_termo_b2b, tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, mercadoria, identifiers_termo, index):
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    treeview_arquivotermo_cliente = None
    treeview_arquivotermo_b2b = None
    
    if "Arquivo B3" in abas_existentes:
        treeview_arquivotermo_cliente = tabela_arquivotermo_cliente
        treeview_arquivotermo_b2b = tabela_arquivotermo_b2b

    # Inicializa o workbook
    workbook = Workbook()
    # Remove a planilha padrão criada automaticamente
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    caracteristicas_commodities_sheet = workbook.create_sheet(title="Características Commodities")
    # Obter os cabeçalhos das colunas da tabela de commodities
    headers_caracteristicas_commodities = [tabela_commodities.heading(col)["text"] for col in tabela_commodities["columns"]]
    caracteristicas_commodities_sheet.append(headers_caracteristicas_commodities) 
    
    # Verifica se a planilha já existe, caso contrário, cria uma nova                  
    if f"Cliente_{mercadoria}" not in workbook.sheetnames:
        sheet_cliente = workbook.create_sheet(title=f"Cliente_{mercadoria}")
        headers_cliente = [tabela_termo_cliente.heading(col)["text"] for col in tabela_termo_cliente["columns"]]
        sheet_cliente.append(headers_cliente)

    # Adiciona todas as linhas
    for row_id in tabela_termo_cliente.get_children():
        row = tabela_termo_cliente.item(row_id)["values"]
        if str(row[-4]) in identifiers_termo:
                sheet_cliente.append(row)    
    
    if f"B2B_{mercadoria}" not in workbook.sheetnames:
        sheet_b2b = workbook.create_sheet(title=f"B2B_{mercadoria}")
        headers_b2b = [tabela_termo_b2b.heading(col)["text"] for col in tabela_termo_b2b["columns"]]
        sheet_b2b.append(headers_b2b) 
    
    for row_id in tabela_termo_b2b.get_children():
        row = tabela_termo_b2b.item(row_id)["values"]
        if str(row[-4]) in identifiers_termo:
                sheet_b2b.append(row)

    ajustar_colunas_sheet(sheet_cliente, headers_cliente)
    ajustar_colunas_sheet(sheet_b2b, headers_b2b)
        
    # Se a aba "Arquivo B3" existir, preenche as planilhas B3File
    if "Arquivo B3" in abas_existentes:
        if f"Cliente_B3_{mercadoria}" not in workbook.sheetnames:
            b3file_cliente_sheet = workbook.create_sheet(title=(f"Cliente_B3File_{mercadoria}"[:30] if len(f"Cliente_B3File_{mercadoria}") > 30 else f"Cliente_B3File_{mercadoria}"))            
            headers_b3file_cliente = [treeview_arquivotermo_cliente.heading(col)["text"] for col in treeview_arquivotermo_cliente["columns"]]
            b3file_cliente_sheet.append(headers_b3file_cliente)
        
        if f"B2B_B3_{mercadoria}" not in workbook.sheetnames:
            b3file_b2b_sheet = workbook.create_sheet(title=(f"B2B_B3File_{mercadoria}"[:30] if len(f"B2B_B3File_{mercadoria}") > 30 else f"B2B_B3File_{mercadoria}"))            
            headers_b3file_b2b = [treeview_arquivotermo_b2b.heading(col)["text"] for col in treeview_arquivotermo_b2b["columns"]]
            b3file_b2b_sheet.append(headers_b3file_b2b)        

        
        for file_row in treeview_arquivotermo_cliente.get_children():
            file_values = treeview_arquivotermo_cliente.item(file_row, 'values')            
            if file_values[-4] in identifiers_termo and lookup(file_values[16].strip(), commodities_ric, commodities_mercadoria) == mercadoria and file_values[-3] == str(index):
                b3file_cliente_sheet.append(file_values)
            elif file_values[-4] in identifiers_termo and file_values[1] == '2':
                b3file_cliente_sheet.append(file_values)
            
        for file_row in treeview_arquivotermo_b2b.get_children():
            file_values = treeview_arquivotermo_b2b.item(file_row, 'values')
            if file_values[-4] in identifiers_termo and lookup(file_values[16].strip(), commodities_ric, commodities_mercadoria) == mercadoria and file_values[-3] == str(index):
                b3file_b2b_sheet.append(file_values)
            elif file_values[-4] in identifiers_termo and file_values[1] == '2':
                b3file_b2b_sheet.append(file_values)
        
        # Ajusta a largura das colunas Arquivo B3
        ajustar_colunas_sheet(b3file_cliente_sheet, headers_b3file_cliente)
        ajustar_colunas_sheet(b3file_b2b_sheet, headers_b3file_b2b)
                    
        # Verifica se existe alguma linha para o acrônimo atual onde os valores nas colunas de índices 16 e 17 são diferentes                
    criar_fixings = any(row[16] != row[17] and row[18] == accronym and lookup(row[2], commodities_ric, commodities_mercadoria) == mercadoria for row_id in tabela_termo_cliente.get_children() if (row := tabela_termo_cliente.item(row_id)["values"]))
    if criar_fixings:
        if f"Fixings_Cliente_{mercadoria}" not in workbook.sheetnames:
            fixings_cliente_sheet = workbook.create_sheet(title=f"Fixings_Cliente_{mercadoria}")
            headers_fixings_cliente = [tabela_fixingstermo_cliente.heading(col)["text"] for col in tabela_fixingstermo_cliente["columns"]]
            fixings_cliente_sheet.append(headers_fixings_cliente)
        
        if f"Fixings_B2B_{mercadoria}" not in workbook.sheetnames:
            fixings_b2b_sheet = workbook.create_sheet(title=f"Fixings_B2B_{mercadoria}")
            headers_fixings_b2b = [tabela_fixingstermo_b2b.heading(col)["text"] for col in tabela_fixingstermo_b2b["columns"]]
            fixings_b2b_sheet.append(headers_fixings_b2b)             
        
        # Preenche as planilhas de fixings
        for fix_row in tabela_fixingstermo_cliente.get_children():
            fix_values = tabela_fixingstermo_cliente.item(fix_row, 'values')
            if fix_values[-4] in identifiers_termo and fix_values[2] == mercadoria:
                fixings_cliente_sheet.append(fix_values)

        for fix_row in tabela_fixingstermo_b2b.get_children():
            fix_values = tabela_fixingstermo_b2b.item(fix_row, 'values')
            if fix_values[-4] in identifiers_termo and fix_values[2] == mercadoria:
                fixings_b2b_sheet.append(fix_values)

        # Ajusta a largura das colunas fixings
        ajustar_colunas_sheet(fixings_cliente_sheet, headers_fixings_cliente)
        ajustar_colunas_sheet(fixings_b2b_sheet, headers_fixings_b2b)                   
               
    #Sheet Características                    
    for row_id in tabela_termo_cliente.get_children():
        row = tabela_termo_cliente.item(row_id)["values"]
        if str(row[-4]) in identifiers_termo:
            # Aplicar o filtro na entrada de texto
            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            entry_filtro_commodities.insert(0, row[2])   # Inserir o tipo de commodity no campo de entrada            
            # Chamar a função de filtragem
            tabela_commodities_data = [tabela_commodities.item(item, "values") for item in tabela_commodities.get_children()]
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data)                 

            # Extrair os dados filtrados e adicioná-los à planilha
            for item_id in tabela_commodities.get_children():
                item_values = tabela_commodities.item(item_id)["values"]                        
                caracteristicas_commodities_sheet.append(item_values)

            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data)            

    #Ajustar a largura das colunas para melhor visualização                
    ajustar_colunas_sheet(caracteristicas_commodities_sheet, headers_caracteristicas_commodities)
    

    # Define o nome do arquivo com base no tipo de dados, acrônimo e tipo de mercadoria
    base_filename = "Termo"
    max_length = 30
    max_accronym_length = 10

    def truncate_string(s, max_len):
        return s if len(s) <= max_len else s[:max_len]

    # Trunca o acrônimo se necessário
    truncated_accronym = truncate_string(accronym, max_accronym_length)

    if index == 1:
        # Calcula o espaço disponível para mercadoria
        available_length = max_length - len(base_filename) - len(".xlsx") - 2 - len(truncated_accronym)  # 2 underscores
        truncated_mercadoria = truncate_string(mercadoria, available_length)
        
        filename = f"{base_filename}_{truncated_accronym}_{truncated_mercadoria}.xlsx"
    else:
        # Calcula o espaço disponível para mercadoria e índice
        available_length = max_length - len(base_filename) - len(".xlsx") - 3 - len(truncated_accronym) - len(str(index))  # 3 underscores and index length
        truncated_mercadoria = truncate_string(mercadoria, available_length)
        
        filename = f"{base_filename}_{truncated_accronym}_{truncated_mercadoria}_#{index}.xlsx"

    
    workbook.save(os.path.join(downloads_path, filename))   
    
        
def export_to_excel_termo(abas_existentes, tabela_termo_cliente, tabela_termo_b2b, tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, switch_cliente_termo):
    # Perguntar ao usuário se deseja seguir com a exportação
    resposta = messagebox.askyesno("Export to Excel - Termo", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função
    
    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"

    # Data atual
    for item in tabela_termo_cliente.get_children():
        data_atual_str = tabela_termo_cliente.item(item)["values"][1]  # Ajuste conforme necessário

        # Converter a string para um objeto datetime
        try:
            data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Ajuste o formato conforme necessário
        except ValueError as e:
            messagebox.showerror("Erro", f"Formato de data inválido: {data_atual_str}")
            continue

    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    # Caminho completo para o diretório
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")

    # Criar o diretório se não existir
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)

    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()       

    if "Arquivo B3" in abas_existentes:
        treeview_arquivotermo_cliente = tabela_arquivotermo_cliente
        treeview_arquivotermo_b2b = tabela_arquivotermo_b2b

    # Verifica se o switch está na posição "Per Client"
    if switch_cliente_termo.get() == "on":   
        # Extrai os acrônimos únicos da tabela "Cliente"         
        accronyms = []
        mercadorias = []
        indexes = []
        for row_id_comm in tabela_termo_cliente.get_children():
            row_comm = tabela_termo_cliente.item(row_id_comm)["values"]                    
            mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)            
            accronym = row_comm[18]
            index = row_comm[-3]            
            if mercadoria not in mercadorias:
                mercadorias.append(mercadoria)
            if accronym not in accronyms:    
                accronyms.append(accronym)    
            if index not in indexes:
               indexes.append(index)    

        for index in indexes:
            for mercadoria in mercadorias:   
                # Itera sobre cada acrônimo único
                for accronym in accronyms:               
                    cliente = lookup(accronym, cntpy_acronym, cntpy_name) 
                    
                    identifiers_termo = []

                    for row_id_tco in tabela_termo_cliente.get_children():
                        row_comm_tco = tabela_termo_cliente.item(row_id_tco)["values"]                    
                        identifier_termo = str(row_comm_tco[-4])                        
                        index_tco = row_comm_tco[-3]                        
                        if lookup(row_comm_tco[2], commodities_ric, commodities_mercadoria) == mercadoria and row_comm_tco[18] == accronym and index_tco == index:
                            identifiers_termo.append(identifier_termo)

                    if identifiers_termo:
                        # Salva o arquivo Excel no diretório de downloads
                        if index == 1:                           
                            downloads_path = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f"{cliente}", "TERMO", f"{mercadoria}")
                            if not os.path.exists(downloads_path):
                                os.makedirs(downloads_path)
                        else:
                            downloads_path = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f"{cliente}", "TERMO", f"#{index} {mercadoria}")
                            if not os.path.exists(downloads_path):
                                os.makedirs(downloads_path)
                        #downloads_path = ensure_directory_exists(downloads_path, diretorio_raiz, ano, mes2, mes, dia, cliente, "TERMO", f"{mercadoria}") 
                        gerar_excel_termo(abas_existentes, downloads_path, cliente, accronym, tabela_termo_cliente, tabela_termo_b2b, tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, mercadoria, identifiers_termo, index)
                        
        messagebox.showinfo("Sucesso", "Excel por Cliente extraído com sucesso!")
    else:        
        # Se o checkbox não estiver marcado, gera um único arquivo com todos os dados
        workbook = Workbook()        
        # Adiciona a primeira planilha para o Treeview "Cliente"
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        sheet_cliente = workbook.create_sheet(title="Cliente", index=0)

        
        headers_cliente = [tabela_termo_cliente.heading(col)["text"] for col in tabela_termo_cliente["columns"]]
        sheet_cliente.append(headers_cliente)

        # Adiciona todas as linhas
        for row_id in tabela_termo_cliente.get_children():
            row = tabela_termo_cliente.item(row_id)["values"]
            sheet_cliente.append(row) 

            # Adiciona a segunda planilha para o Treeview "B2B"
        sheet_b2b = workbook.create_sheet(title="B2B", index=1)        
        headers_b2b = [tabela_termo_b2b.heading(col)["text"] for col in tabela_termo_b2b["columns"]]
        sheet_b2b.append(headers_b2b)

        # Adiciona todas as linhas
        for row_id in tabela_termo_b2b.get_children():
            row = tabela_termo_b2b.item(row_id)["values"]
            sheet_b2b.append(row)         
             
        # Adiciona planilhas de fixings se houver dados
        if any(tabela_fixingstermo_cliente.get_children()):
            fixings_cliente_sheet = workbook.create_sheet(title="Cliente_Fixings", index=2)
            headers_fixings_cliente = [tabela_fixingstermo_cliente.heading(col)["text"] for col in tabela_fixingstermo_cliente["columns"]]
            fixings_cliente_sheet.append(headers_fixings_cliente)
            for fix_row in tabela_fixingstermo_cliente.get_children():
                fix_values = tabela_fixingstermo_cliente.item(fix_row, 'values')
                fixings_cliente_sheet.append(list(fix_values))   
            fixings_b2b_sheet = workbook.create_sheet(title="B2B_Fixings", index=3)
            headers_fixings_b2b = [tabela_fixingstermo_b2b.heading(col)["text"] for col in tabela_fixingstermo_b2b["columns"]]
            fixings_b2b_sheet.append(headers_fixings_b2b)
            for fix_row in tabela_fixingstermo_b2b.get_children():
                fix_values = tabela_fixingstermo_b2b.item(fix_row, 'values')
                fixings_b2b_sheet.append(list(fix_values))

                 # Ajusta a largura das colunas fixings
            ajustar_colunas_sheet(fixings_cliente_sheet, headers_fixings_cliente)
            ajustar_colunas_sheet(fixings_b2b_sheet, headers_fixings_b2b)

        # Adiciona planilhas de arquivoB3 se houver dados
        if "Arquivo B3" in abas_existentes:
            caracteristicas_commodities_sheet = workbook.create_sheet(title="Características Commodities", index=6)
            b3file_cliente_sheet = workbook.create_sheet(title="Cliente_B3", index=4)
            b3file_b2b_sheet = workbook.create_sheet(title=(f"B2B_B3File_{mercadoria}"[:30] if len(f"B2B_B3File_{mercadoria}") > 30 else f"B2B_B3File_{mercadoria}"), index=5)            
            headers_b3file_cliente = [treeview_arquivotermo_cliente.heading(col)["text"] for col in treeview_arquivotermo_cliente["columns"]]
            b3file_cliente_sheet.append(headers_b3file_cliente)
            headers_b3file_b2b = [treeview_arquivotermo_b2b.heading(col)["text"] for col in treeview_arquivotermo_b2b["columns"]]
            b3file_b2b_sheet.append(headers_b3file_b2b)

            for file_row in treeview_arquivotermo_cliente.get_children():
                file_values = treeview_arquivotermo_cliente.item(file_row, 'values')                
                b3file_cliente_sheet.append(file_values)

            for file_row in treeview_arquivotermo_b2b.get_children():
                file_values = treeview_arquivotermo_b2b.item(file_row, 'values')                
                b3file_b2b_sheet.append(file_values)
            # Ajusta a largura das colunas Arquivo B3
            ajustar_colunas_sheet(b3file_cliente_sheet, headers_b3file_cliente)
            ajustar_colunas_sheet(b3file_b2b_sheet, headers_b3file_b2b)        
        else:
            caracteristicas_commodities_sheet = workbook.create_sheet(title="Caracteristicas_Comm", index=4)
     
         # Obter os cabeçalhos das colunas da tabela de commodities
        headers_caracteristicas_commodities = [tabela_commodities.heading(col)["text"] for col in tabela_commodities["columns"]]
        caracteristicas_commodities_sheet.append(headers_caracteristicas_commodities)
        tipos_mercadoria = set(tabela_termo_cliente.item(row_id)["values"][2] for row_id in tabela_termo_cliente.get_children())
        for tipo in tipos_mercadoria:
            # Aplicar o filtro na entrada de texto
            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            entry_filtro_commodities.insert(0, tipo)   # Inserir o tipo de commodity no campo de entrada            
            # Chamar a função de filtragem
            tabela_commodities_data = [tabela_commodities.item(item, "values") for item in tabela_commodities.get_children()]
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data)                 

            # Extrair os dados filtrados e adicioná-los à planilha
            for item_id in tabela_commodities.get_children():
                item_values = tabela_commodities.item(item_id)["values"]
                caracteristicas_commodities_sheet.append(item_values)

            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data) 

       
        
        ajustar_colunas_sheet(caracteristicas_commodities_sheet, headers_caracteristicas_commodities)
        

        # Define o nome do arquivo com base no tipo de dados
        base_filename = "Termo_Data_All"
        filename = f"{base_filename}.xlsx"

        # Salva o arquivo Excel no diretório de downloads        
        workbook.save(os.path.join(caminho_completo, filename))
        messagebox.showinfo("Sucesso","Excel extraído com sucesso!")


def gerar_excel_opcao(abas_existentes, downloads_path, cliente, accronym, tabela_opcao_cliente, tabela_opcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b, mercadoria, identifiers_opcao, index):
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    treeview_arquivoopcao_cliente = None
    treeview_arquivoopcao_b2b = None
    
    if "Arquivo B3" in abas_existentes:
        treeview_arquivoopcao_cliente = tabela_arquivoopcao_cliente
        treeview_arquivoopcao_b2b = tabela_arquivoopcao_b2b

    # Inicializa o workbook
    workbook = Workbook()
    # Remove a planilha padrão criada automaticamente
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    caracteristicas_commodities_sheet = workbook.create_sheet(title="Características Commodities")
    # Obter os cabeçalhos das colunas da tabela de commodities
    headers_caracteristicas_commodities = [tabela_commodities.heading(col)["text"] for col in tabela_commodities["columns"]]
    caracteristicas_commodities_sheet.append(headers_caracteristicas_commodities) 
    
    # Verifica se a planilha já existe, caso contrário, cria uma nova                  
    if f"Cliente_{mercadoria}" not in workbook.sheetnames:
        sheet_cliente = workbook.create_sheet(title=f"Cliente_{mercadoria}")
        headers_cliente = [tabela_opcao_cliente.heading(col)["text"] for col in tabela_opcao_cliente["columns"]]
        sheet_cliente.append(headers_cliente)

    # Adiciona todas as linhas
    for row_id in tabela_opcao_cliente.get_children():
        row = tabela_opcao_cliente.item(row_id)["values"]
        if str(row[24]) in identifiers_opcao:
                sheet_cliente.append(row)    
    
    if f"B2B_{mercadoria}" not in workbook.sheetnames:
        sheet_b2b = workbook.create_sheet(title=f"B2B_{mercadoria}")
        headers_b2b = [tabela_opcao_b2b.heading(col)["text"] for col in tabela_opcao_b2b["columns"]]
        sheet_b2b.append(headers_b2b) 
    
    for row_id in tabela_opcao_b2b.get_children():
        row = tabela_opcao_b2b.item(row_id)["values"]
        if str(row[24]) in identifiers_opcao:
                sheet_b2b.append(row)

    ajustar_colunas_sheet(sheet_cliente, headers_cliente)
    ajustar_colunas_sheet(sheet_b2b, headers_b2b)
        
    # Se a aba "Arquivo B3" existir, preenche as planilhas B3File
    if "Arquivo B3" in abas_existentes:
        if f"Cliente_B3_{mercadoria}" not in workbook.sheetnames:
            b3file_cliente_sheet = workbook.create_sheet(title=f"Cliente_B3_{mercadoria}")
            headers_b3file_cliente = [treeview_arquivoopcao_cliente.heading(col)["text"] for col in treeview_arquivoopcao_cliente["columns"]]
            b3file_cliente_sheet.append(headers_b3file_cliente)
        
        if f"B2B_B3_{mercadoria}" not in workbook.sheetnames:
            b3file_b2b_sheet = workbook.create_sheet(title=f"B2B_B3_{mercadoria}")
            headers_b3file_b2b = [treeview_arquivoopcao_b2b.heading(col)["text"] for col in treeview_arquivoopcao_b2b["columns"]]
            b3file_b2b_sheet.append(headers_b3file_b2b)

        for file_row in treeview_arquivoopcao_cliente.get_children():
            file_values = treeview_arquivoopcao_cliente.item(file_row, 'values')            
            if file_values[77] in identifiers_opcao and lookup(file_values[67].strip(), commodities_ric, commodities_mercadoria) == mercadoria and file_values[78] == str(index):
                b3file_cliente_sheet.append(file_values)
            #elif file_values[77] in identifiers_opcao and file_values[1] == '2':
            #   b3file_cliente_sheet.append(file_values)
            
        for file_row in treeview_arquivoopcao_b2b.get_children():
            file_values = treeview_arquivoopcao_b2b.item(file_row, 'values')
            if file_values[77] in identifiers_opcao and lookup(file_values[67].strip(), commodities_ric, commodities_mercadoria) == mercadoria and file_values[78] == str(index):
                b3file_b2b_sheet.append(file_values)
            #elif file_values[77] in identifiers_opcao and file_values[1] == '2':
            #    b3file_b2b_sheet.append(file_values)
        
        # Ajusta a largura das colunas Arquivo B3
        ajustar_colunas_sheet(b3file_cliente_sheet, headers_b3file_cliente)
        ajustar_colunas_sheet(b3file_b2b_sheet, headers_b3file_b2b)
                    
        # Verifica se existe alguma linha para o acrônimo atual onde os valores nas colunas de índices 16 e 17 são diferentes                
    criar_fixings = any(row[16] != row[17] and row[18] == accronym and lookup(row[2], commodities_ric, commodities_mercadoria) == mercadoria for row_id in tabela_opcao_cliente.get_children() if (row := tabela_opcao_cliente.item(row_id)["values"]))
    if criar_fixings:
        if f"Fixings_Cliente_{mercadoria}" not in workbook.sheetnames:
            fixings_cliente_sheet = workbook.create_sheet(title=f"Fixings_Cliente_{mercadoria}")
            headers_fixings_cliente = [tabela_fixingsopcao_cliente.heading(col)["text"] for col in tabela_fixingsopcao_cliente["columns"]]
            fixings_cliente_sheet.append(headers_fixings_cliente)
        
        if f"Fixings_B2B_{mercadoria}" not in workbook.sheetnames:
            fixings_b2b_sheet = workbook.create_sheet(title=f"Fixings_B2B_{mercadoria}")
            headers_fixings_b2b = [tabela_fixingsopcao_b2b.heading(col)["text"] for col in tabela_fixingsopcao_b2b["columns"]]
            fixings_b2b_sheet.append(headers_fixings_b2b)             
        
        # Preenche as planilhas de fixings
        for fix_row in tabela_fixingsopcao_cliente.get_children():
            fix_values = tabela_fixingsopcao_cliente.item(fix_row, 'values')
            if fix_values[50] in identifiers_opcao and fix_values[2] == mercadoria:
                fixings_cliente_sheet.append(fix_values)

        for fix_row in tabela_fixingsopcao_b2b.get_children():
            fix_values = tabela_fixingsopcao_b2b.item(fix_row, 'values')
            if fix_values[50] in identifiers_opcao and fix_values[2] == mercadoria:
                fixings_b2b_sheet.append(fix_values)

        # Ajusta a largura das colunas fixings
        ajustar_colunas_sheet(fixings_cliente_sheet, headers_fixings_cliente)
        ajustar_colunas_sheet(fixings_b2b_sheet, headers_fixings_b2b)                   
               
    #Sheet Características                    
    for row_id in tabela_opcao_cliente.get_children():
        row = tabela_opcao_cliente.item(row_id)["values"]
        if str(row[-4]) in identifiers_opcao:
            # Aplicar o filtro na entrada de texto
            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            entry_filtro_commodities.insert(0, row[2])   # Inserir o tipo de commodity no campo de entrada            
            # Chamar a função de filtragem
            tabela_commodities_data = [tabela_commodities.item(item, "values") for item in tabela_commodities.get_children()]
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data)                 

            # Extrair os dados filtrados e adicioná-los à planilha
            for item_id in tabela_commodities.get_children():
                item_values = tabela_commodities.item(item_id)["values"]                        
                caracteristicas_commodities_sheet.append(item_values)

            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data)            

    #Ajustar a largura das colunas para melhor visualização                
    ajustar_colunas_sheet(caracteristicas_commodities_sheet, headers_caracteristicas_commodities)
    

    # Define o nome do arquivo com base no tipo de dados, acrônimo e tipo de mercadoria
    base_filename = "Opção"        
    max_length = 31
    max_accronym_length = 10

    def truncate_string(s, max_len):
        return s if len(s) <= max_len else s[:max_len]

    # Trunca o acrônimo se necessário
    truncated_accronym = truncate_string(accronym, max_accronym_length)

    if index == 1:
        # Calcula o espaço disponível para mercadoria
        available_length = max_length - len(base_filename) - len(".xlsx") - 2 - len(truncated_accronym)  # 2 underscores
        truncated_mercadoria = truncate_string(mercadoria, available_length)
        
        filename = f"{base_filename}_{truncated_accronym}_{truncated_mercadoria}.xlsx"
    else:
        # Calcula o espaço disponível para mercadoria e índice
        available_length = max_length - len(base_filename) - len(".xlsx") - 3 - len(truncated_accronym) - len(str(index))  # 3 underscores and index length
        truncated_mercadoria = truncate_string(mercadoria, available_length)
        
        filename = f"{base_filename}_{truncated_accronym}_{truncated_mercadoria}_#{index}.xlsx"


    
    workbook.save(os.path.join(downloads_path, filename))          



def export_to_excel_opcao(abas_existentes, tabela_opcao_cliente, tabela_opcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b, switch_cliente_opcao):
    # Perguntar ao usuário se deseja seguir com a exportação
    resposta = messagebox.askyesno("Export to Excel - Opção", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função
    
    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"

    # Data atual
    for item in tabela_opcao_cliente.get_children():
        data_atual_str = tabela_opcao_cliente.item(item)["values"][1]  # Ajuste conforme necessário

        # Converter a string para um objeto datetime
        try:
            data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Ajuste o formato conforme necessário
        except ValueError as e:
            messagebox.showerror("Erro", f"Formato de data inválido: {data_atual_str}")
            continue

    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    # Caminho completo para o diretório
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")

    # Criar o diretório se não existir
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)

    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()       

    if "Arquivo B3" in abas_existentes:
        treeview_arquivoopcao_cliente = tabela_arquivoopcao_cliente
        treeview_arquivoopcao_b2b = tabela_arquivoopcao_b2b

    # Verifica se o switch está na posição "Per Client"
    if switch_cliente_opcao.get() == "on":   
        # Extrai os acrônimos únicos da tabela "Cliente"         
        accronyms = []
        mercadorias = []
        indexes = []
        for row_id_comm in tabela_opcao_cliente.get_children():
            row_comm = tabela_opcao_cliente.item(row_id_comm)["values"]                    
            mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)            
            accronym = row_comm[18]
            index = row_comm[-3]            
            if mercadoria not in mercadorias:
                mercadorias.append(mercadoria)
            if accronym not in accronyms:    
                accronyms.append(accronym)    
            if index not in indexes:
               indexes.append(index)    

        for index in indexes:
            for mercadoria in mercadorias:   
                # Itera sobre cada acrônimo único
                for accronym in accronyms:               
                    cliente = lookup(accronym, cntpy_acronym, cntpy_name) 
                    
                    identifiers_opcao = []

                    for row_id_tco in tabela_opcao_cliente.get_children():
                        row_comm_tco = tabela_opcao_cliente.item(row_id_tco)["values"]                    
                        identifier_opcao = str(row_comm_tco[-4])                        
                        index_tco = row_comm_tco[-3]                        
                        if lookup(row_comm_tco[2], commodities_ric, commodities_mercadoria) == mercadoria and row_comm_tco[18] == accronym and index_tco == index:
                            identifiers_opcao.append(identifier_opcao)

                    if identifiers_opcao:
                        # Salva o arquivo Excel no diretório de downloads
                        if index == 1:                           
                            downloads_path = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f"{cliente}", "OPÇÃO", f"{mercadoria}")
                            if not os.path.exists(downloads_path):
                                os.makedirs(downloads_path)
                        else:
                            downloads_path = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f"{cliente}", "OPÇÃO", f"#{index} {mercadoria}")
                            if not os.path.exists(downloads_path):
                                os.makedirs(downloads_path)
                        #downloads_path = ensure_directory_exists(downloads_path, diretorio_raiz, ano, mes2, mes, dia, cliente, "opcao", f"{mercadoria}") 
                        gerar_excel_opcao(abas_existentes, downloads_path, cliente, accronym, tabela_opcao_cliente, tabela_opcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b, mercadoria, identifiers_opcao, index)
                        
        messagebox.showinfo("Sucesso", "Excel por Cliente extraído com sucesso!")
    else:        
        # Se o checkbox não estiver marcado, gera um único arquivo com todos os dados
        workbook = Workbook()        
        # Adiciona a primeira planilha para o Treeview "Cliente"
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        sheet_cliente = workbook.create_sheet(title="Cliente", index=0)

        
        headers_cliente = [tabela_opcao_cliente.heading(col)["text"] for col in tabela_opcao_cliente["columns"]]
        sheet_cliente.append(headers_cliente)

        # Adiciona todas as linhas
        for row_id in tabela_opcao_cliente.get_children():
            row = tabela_opcao_cliente.item(row_id)["values"]
            sheet_cliente.append(row) 

            # Adiciona a segunda planilha para o Treeview "B2B"
        sheet_b2b = workbook.create_sheet(title="B2B", index=1)        
        headers_b2b = [tabela_opcao_b2b.heading(col)["text"] for col in tabela_opcao_b2b["columns"]]
        sheet_b2b.append(headers_b2b)

        # Adiciona todas as linhas
        for row_id in tabela_opcao_b2b.get_children():
            row = tabela_opcao_b2b.item(row_id)["values"]
            sheet_b2b.append(row)         
             
        # Adiciona planilhas de fixings se houver dados
        if any(tabela_fixingsopcao_cliente.get_children()):
            fixings_cliente_sheet = workbook.create_sheet(title="Cliente_Fixings", index=2)
            headers_fixings_cliente = [tabela_fixingsopcao_cliente.heading(col)["text"] for col in tabela_fixingsopcao_cliente["columns"]]
            fixings_cliente_sheet.append(headers_fixings_cliente)
            for fix_row in tabela_fixingsopcao_cliente.get_children():
                fix_values = tabela_fixingsopcao_cliente.item(fix_row, 'values')
                fixings_cliente_sheet.append(list(fix_values))   
            fixings_b2b_sheet = workbook.create_sheet(title="B2B_Fixings", index=3)
            headers_fixings_b2b = [tabela_fixingsopcao_b2b.heading(col)["text"] for col in tabela_fixingsopcao_b2b["columns"]]
            fixings_b2b_sheet.append(headers_fixings_b2b)
            for fix_row in tabela_fixingsopcao_b2b.get_children():
                fix_values = tabela_fixingsopcao_b2b.item(fix_row, 'values')
                fixings_b2b_sheet.append(list(fix_values))

                 # Ajusta a largura das colunas fixings
            ajustar_colunas_sheet(fixings_cliente_sheet, headers_fixings_cliente)
            ajustar_colunas_sheet(fixings_b2b_sheet, headers_fixings_b2b)

        # Adiciona planilhas de arquivoB3 se houver dados
        if "Arquivo B3" in abas_existentes:
            caracteristicas_commodities_sheet = workbook.create_sheet(title="Características Commodities", index=6)
            b3file_cliente_sheet = workbook.create_sheet(title="Cliente_B3", index=4)
            b3file_b2b_sheet = workbook.create_sheet(title="B2B_B3", index=5)
            headers_b3file_cliente = [treeview_arquivoopcao_cliente.heading(col)["text"] for col in treeview_arquivoopcao_cliente["columns"]]
            b3file_cliente_sheet.append(headers_b3file_cliente)
            headers_b3file_b2b = [treeview_arquivoopcao_b2b.heading(col)["text"] for col in treeview_arquivoopcao_b2b["columns"]]
            b3file_b2b_sheet.append(headers_b3file_b2b)

            for file_row in treeview_arquivoopcao_cliente.get_children():
                file_values = treeview_arquivoopcao_cliente.item(file_row, 'values')                
                b3file_cliente_sheet.append(file_values)

            for file_row in treeview_arquivoopcao_b2b.get_children():
                file_values = treeview_arquivoopcao_b2b.item(file_row, 'values')                
                b3file_b2b_sheet.append(file_values)
            # Ajusta a largura das colunas Arquivo B3
            ajustar_colunas_sheet(b3file_cliente_sheet, headers_b3file_cliente)
            ajustar_colunas_sheet(b3file_b2b_sheet, headers_b3file_b2b)        
        else:
            caracteristicas_commodities_sheet = workbook.create_sheet(title="Características Commodities", index=4)
     
         # Obter os cabeçalhos das colunas da tabela de commodities
        headers_caracteristicas_commodities = [tabela_commodities.heading(col)["text"] for col in tabela_commodities["columns"]]
        caracteristicas_commodities_sheet.append(headers_caracteristicas_commodities)
        tipos_mercadoria = set(tabela_opcao_cliente.item(row_id)["values"][2] for row_id in tabela_opcao_cliente.get_children())
        for tipo in tipos_mercadoria:
            # Aplicar o filtro na entrada de texto
            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            entry_filtro_commodities.insert(0, tipo)   # Inserir o tipo de commodity no campo de entrada            
            # Chamar a função de filtragem
            tabela_commodities_data = [tabela_commodities.item(item, "values") for item in tabela_commodities.get_children()]
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data)                 

            # Extrair os dados filtrados e adicioná-los à planilha
            for item_id in tabela_commodities.get_children():
                item_values = tabela_commodities.item(item_id)["values"]
                caracteristicas_commodities_sheet.append(item_values)

            entry_filtro_commodities.delete(0, 'end')  # Limpar o campo de entrada
            filtro = entry_filtro_commodities.get().lower()
            filtrar_commodities_export(tabela_commodities, filtro, 0, tabela_commodities_data)        
        
        ajustar_colunas_sheet(caracteristicas_commodities_sheet, headers_caracteristicas_commodities)
        

        # Define o nome do arquivo com base no tipo de dados
        base_filename = "opcao_Data_All"
        filename = f"{base_filename}.xlsx"

        # Salva o arquivo Excel no diretório de downloads        
        workbook.save(os.path.join(caminho_completo, filename))
        messagebox.showinfo("Sucesso","Excel extraído com sucesso!")



def delete_selected_item(treeview, label_qty_deals):
    # Obtém os itens selecionados
    selected_items = treeview.selection()
    
    if selected_items:
        # Deleta todos os itens selecionados
        for item in selected_items:
            treeview.delete(item)
        
        # Chama a função para destacar duplicatas, se necessário
        highlight_duplicates(treeview, '')
        # Atualiza o label com a quantidade de deals                        
        qty_deals = number_of_deals(treeview)
        label_qty_deals.configure(text=str(qty_deals))


    else:
        # Mostra uma mensagem de aviso se nenhum item estiver selecionado
        messagebox.showwarning("Warning", "No item selected.")

def approve_selected_item(treeview_deals, treeview_file):
    selected_items_deals = treeview_deals.selection()
    indentifiers = []
    if selected_items_deals:
        # Deleta todos os itens selecionados
        for item in selected_items_deals:
            row_data_deals = list(treeview_deals.item(item, 'values')) # Converte a tupla em uma lista
            identifier = str(row_data_deals[-4])
            status = row_data_deals[-2]
            if status == "Pending Review":
                row_data_deals[-2] = "Approved"   
                indentifiers.append(identifier)        
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
        
        
        for item in treeview_file.get_children():
            row_data_file = list(treeview_file.item(item, 'values')) # Converte a tupla em uma lista            
            status = row_data_deals[-2]
            if status == "Pending Review" and str(row_data_deals[-4]) in indentifiers:
                row_data_file[-2] = "Approved" 
                  
            treeview_file.item(item, values=row_data_deals)  # Atualiza o Treeview

    
def create_color_legend(frame):
    # Definição das cores
    COLOR_NEW = '#FFFFFF'  # White
    COLOR_PENDING_REVIEW = '#FFC9CA'  # Light Pink
    COLOR_APPROVED = '#CCCCFF'  # Light Purple
    COLOR_GENERATED = '#79BCFF'  # Light Blue
    COLOR_PENDING_MAKER_CHECKER = '#FFCC66'  # Light Orange
    COLOR_CONCLUDED = '#AFFFE4'  # Light Green
    COLOR_DUPLICATE = '#F08080'  # Light Coral
    BORDER_COLOR = '#000000'  # Cor da borda (preto)
    BORDER_WIDTH = 2  # Largura da borda

    font_legend = ctk.CTkFont(family="League Spartan", size=12, weight="bold")   

    # Criação das legendas com borda
    def create_bordered_label(text, color):
        bordered_frame = ctk.CTkFrame(frame, fg_color=BORDER_COLOR, width=100, height=30)
        bordered_frame.pack_propagate(False)  # Impede que o frame se ajuste ao tamanho do label
        label = ctk.CTkLabel(bordered_frame, text=text, fg_color=color, font=font_legend)
        label.pack(expand=True, fill='both', padx=BORDER_WIDTH, pady=BORDER_WIDTH)
        return bordered_frame

    legend_new = create_bordered_label("New", COLOR_NEW)
    legend_pending_review = create_bordered_label("Pending Review", COLOR_PENDING_REVIEW)
    legend_approved = create_bordered_label("Approved", COLOR_APPROVED)
    legend_generated = create_bordered_label("Generated", COLOR_GENERATED)
    legend_pending_maker_checker = create_bordered_label("Maker/Checker", COLOR_PENDING_MAKER_CHECKER)
    legend_concluded = create_bordered_label("Concluded", COLOR_CONCLUDED)
    legend_duplicate = create_bordered_label("Duplicated", COLOR_DUPLICATE)

    # Posicionamento das legendas    
    legend_concluded.pack(side='bottom', padx=5, pady=1)
    legend_pending_maker_checker.pack(side='bottom', padx=5, pady=1)
    legend_generated.pack(side='bottom', padx=5, pady=1)
    legend_approved.pack(side='bottom', padx=5, pady=1)
    legend_pending_review.pack(side='bottom', padx=5, pady=1)
    legend_new.pack(side='bottom', padx=5, pady=1)
    legend_duplicate.pack(side='bottom', padx=5, pady=1)
    

def number_of_deals(treeview):
    # Conta a quantidade de linhas no Treeview
    return len(treeview.get_children())

def monitor_operacoes(tabview, aba_inicio, abas_existentes, tree):
    global sub_notebook_monitor, sub_sub_notebook_operacoes, sub_sub_sub_notebook_termo, sub_sub_sub_notebook_opcao
    global sub_sub_notebook_fixings, sub_sub_sub_notebook_termo_fixings, sub_sub_sub_notebook_opcao_fixings
    # Adicionar a aba "Monitor de Operações" ao tabview
    tabview.insert(1, "Monitor")
    aba_monitor_operacoes = tabview.tab("Monitor")

    sub_notebook_monitor = ctk.CTkTabview(aba_monitor_operacoes)
    sub_notebook_monitor.pack(expand=True, fill='both')

    # Configuração para a aba "Operações"
    sub_notebook_monitor.add("Operações")
    aba_operacoes = sub_notebook_monitor.tab("Operações")

    sub_sub_notebook_operacoes = ctk.CTkTabview(aba_operacoes)
    sub_sub_notebook_operacoes.pack(expand=True, fill='both')

    # Configuração para a sub-aba "Termo"
    sub_sub_notebook_operacoes.add("Termo")
    aba_termo = sub_sub_notebook_operacoes.tab("Termo")

    sub_sub_sub_notebook_termo = ctk.CTkTabview(aba_termo)
    sub_sub_sub_notebook_termo.pack(expand=True, fill='both')

    # Configuração para a sub-sub-aba "Cliente" em "Termo"
    sub_sub_sub_notebook_termo.add("Cliente")
    aba_cliente_termo = sub_sub_sub_notebook_termo.tab("Cliente")
    
    frame_query_cliente_termo = ctk.CTkFrame(aba_cliente_termo, height=1, fg_color="#D3D3D3")
    frame_query_cliente_termo.pack(fill='x')   
        
    frame_approval_cliente_termo = ctk.CTkFrame(aba_cliente_termo, width=220, fg_color="#D3D3D3")
    frame_approval_cliente_termo.pack(fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_cliente_termo)

   
    frame_treeview_cliente_termo = ctk.CTkFrame(aba_cliente_termo)
    frame_treeview_cliente_termo.pack(expand=True, fill='both', side='left')  

    
    frame_scrollbary_cliente_termo = ctk.CTkFrame(frame_treeview_cliente_termo, width=4, fg_color="#D3D3D3")
    frame_scrollbary_cliente_termo.pack(fill='y', side='right')

    scrollbar_x_cliente_termo = ctk.CTkScrollbar(frame_treeview_cliente_termo, orientation='horizontal')
    scrollbar_y_cliente_termo = ctk.CTkScrollbar(frame_scrollbary_cliente_termo, orientation='vertical')

    global tabela_termo_cliente
    tabela_termo_cliente = ttk.Treeview(frame_treeview_cliente_termo, columns=colunas_termo, show='headings', xscrollcommand=scrollbar_x_cliente_termo.set, yscrollcommand=scrollbar_y_cliente_termo.set)
    tabela_termo_cliente.pack(expand=True, fill='both')
    scrollbar_x_cliente_termo.configure(command=tabela_termo_cliente.xview, height=25)
    scrollbar_y_cliente_termo.configure(command=tabela_termo_cliente.yview, width=25)
    
    # Configurar cabeçalhos das colunas
    for coluna in colunas_termo:
        tabela_termo_cliente.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_termo_cliente, _col, False))
    vincular_evento_duplo_clique_status(tabela_termo_cliente, colunas_termo)

    frame_botoes_cliente_termo = ctk.CTkFrame(frame_treeview_cliente_termo, height=150)
    frame_botoes_cliente_termo.pack(fill='x', side='bottom')

    
    # Configuração para a sub-sub-aba "B2B" em "Termo"
    sub_sub_sub_notebook_termo.add("B2B")
    aba_termo_b2b = sub_sub_sub_notebook_termo.tab("B2B")
    
    frame_query_b2b_termo = ctk.CTkFrame(aba_termo_b2b, height=1, fg_color="#D3D3D3")
    frame_query_b2b_termo.pack(fill='x')   

    frame_approval_b2b_termo = ctk.CTkFrame(aba_termo_b2b, width=220, fg_color="#D3D3D3")
    frame_approval_b2b_termo.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_b2b_termo)

    frame_treeview_termo_b2b = ctk.CTkFrame(aba_termo_b2b, fg_color="#D3D3D3")
    frame_treeview_termo_b2b.pack(expand=True, fill='both', side='left')    

    frame_scrollbary_termo_b2b = ctk.CTkFrame(frame_treeview_termo_b2b, width=4)
    frame_scrollbary_termo_b2b.pack(fill='y', side='right')

    scrollbar_x_termo_b2b = ctk.CTkScrollbar(frame_treeview_termo_b2b, orientation='horizontal')
    scrollbar_y_termo_b2b = ctk.CTkScrollbar(frame_scrollbary_termo_b2b, orientation='vertical')

    global tabela_termo_b2b
    tabela_termo_b2b = ttk.Treeview(frame_treeview_termo_b2b, columns=colunas_termo, show='headings', xscrollcommand=scrollbar_x_termo_b2b.set, yscrollcommand=scrollbar_y_termo_b2b.set)
    tabela_termo_b2b.pack(expand=True, fill='both')
    scrollbar_x_termo_b2b.configure(command=tabela_termo_b2b.xview, height=25)
    scrollbar_y_termo_b2b.configure(command=tabela_termo_b2b.yview, width=25)
    

    # Configurar cabeçalhos das colunas
    for coluna in colunas_termo:
        tabela_termo_b2b.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_termo_b2b, _col, False))
    vincular_evento_duplo_clique_status(tabela_termo_b2b, colunas_termo)
    frame_botoes_termo_b2b = ctk.CTkFrame(frame_treeview_termo_b2b, height=150)
    frame_botoes_termo_b2b.pack(fill='x', side='bottom')



    # Configuração para a sub-aba "Opção"
    sub_sub_notebook_operacoes.add("Opção")
    aba_opcao = sub_sub_notebook_operacoes.tab("Opção")

    sub_sub_sub_notebook_opcao = ctk.CTkTabview(aba_opcao)
    sub_sub_sub_notebook_opcao.pack(expand=True, fill='both')

    # Configuração para a sub-sub-aba "Cliente" em "Opção"
    sub_sub_sub_notebook_opcao.add("Cliente")
    aba_cliente_opcao = sub_sub_sub_notebook_opcao.tab("Cliente")
    
    
    frame_query_cliente_opcao = ctk.CTkFrame(aba_cliente_opcao, height=1, fg_color="#D3D3D3")
    frame_query_cliente_opcao.pack(fill='x')   

    frame_approval_cliente_opcao = ctk.CTkFrame(aba_cliente_opcao, width=220, fg_color="#D3D3D3")
    frame_approval_cliente_opcao.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_cliente_opcao)


    frame_treeview_cliente_opcao = ctk.CTkFrame(aba_cliente_opcao)
    frame_treeview_cliente_opcao.pack(expand=True, fill='both', side='left')

    frame_scrollbary_cliente_opcao = ctk.CTkFrame(frame_treeview_cliente_opcao, width=4)
    frame_scrollbary_cliente_opcao.pack(fill='y', side='right')

    scrollbar_x_cliente_opcao = ctk.CTkScrollbar(frame_treeview_cliente_opcao, orientation='horizontal')
    scrollbar_y_cliente_opcao = ctk.CTkScrollbar(frame_scrollbary_cliente_opcao, orientation='vertical')

    global tabela_opcao_cliente
    tabela_opcao_cliente = ttk.Treeview(frame_treeview_cliente_opcao, columns=colunas_opcao, show='headings', xscrollcommand=scrollbar_x_cliente_opcao.set, yscrollcommand=scrollbar_y_cliente_opcao.set)
    tabela_opcao_cliente.pack(expand=True, fill='both')
    scrollbar_x_cliente_opcao.configure(command=tabela_opcao_cliente.xview, height=25)
    scrollbar_y_cliente_opcao.configure(command=tabela_opcao_cliente.yview, width=25)
    
    

    # Configurar cabeçalhos das colunas
    for coluna in colunas_opcao:
        tabela_opcao_cliente.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_opcao_cliente, _col, False))
    vincular_evento_duplo_clique_status(tabela_opcao_cliente, colunas_opcao)
    frame_botoes_cliente_opcao = ctk.CTkFrame(frame_treeview_cliente_opcao)
    frame_botoes_cliente_opcao.pack(fill='x', side='bottom')

    

    # Configuração para a sub-sub-aba "B2B" em "Opção"
    sub_sub_sub_notebook_opcao.add("B2B")
    aba_b2b_opcao = sub_sub_sub_notebook_opcao.tab("B2B")
    
    frame_query_b2b_opcao = ctk.CTkFrame(aba_b2b_opcao, height=1, fg_color="#D3D3D3")
    frame_query_b2b_opcao.pack(fill='x')  
    
    frame_approval_b2b_opcao = ctk.CTkFrame(aba_b2b_opcao, width=220, fg_color="#D3D3D3")
    frame_approval_b2b_opcao.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_b2b_opcao)

    frame_treeview_b2b_opcao = ctk.CTkFrame(aba_b2b_opcao)
    frame_treeview_b2b_opcao.pack(expand=True, fill='both', side='left')

    frame_scrollbary_b2b_opcao = ctk.CTkFrame(frame_treeview_b2b_opcao, width=4)
    frame_scrollbary_b2b_opcao.pack(fill='y', side='right')

    scrollbar_x_b2b_opcao = ctk.CTkScrollbar(frame_treeview_b2b_opcao, orientation='horizontal')
    scrollbar_y_b2b_opcao = ctk.CTkScrollbar(frame_scrollbary_b2b_opcao, orientation='vertical')

    global tabela_opcao_b2b
    tabela_opcao_b2b = ttk.Treeview(frame_treeview_b2b_opcao, columns=colunas_opcao, show='headings', xscrollcommand=scrollbar_x_b2b_opcao.set, yscrollcommand=scrollbar_y_b2b_opcao.set)
    tabela_opcao_b2b.pack(expand=True, fill='both')
    scrollbar_x_b2b_opcao.configure(command=tabela_opcao_b2b.xview, height=25)
    scrollbar_y_b2b_opcao.configure(command=tabela_opcao_b2b.yview, width=25)
    
    # Configurar cabeçalhos das colunas
    for coluna in colunas_opcao:
        tabela_opcao_b2b.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_opcao_b2b, _col, False))
    vincular_evento_duplo_clique_status(tabela_opcao_b2b, colunas_opcao)
    frame_botoes_b2b_opcao = ctk.CTkFrame(frame_treeview_b2b_opcao, height=150)
    frame_botoes_b2b_opcao.pack(fill='x', side='bottom')
    

    # Configuração para a aba "Fixings"
    sub_notebook_monitor.add("Fixings")
    aba_fixings = sub_notebook_monitor.tab("Fixings")

    sub_sub_notebook_fixings = ctk.CTkTabview(aba_fixings)
    sub_sub_notebook_fixings.pack(expand=True, fill='both')

    # Configuração para a sub-aba "Termo" em "Fixings"
    sub_sub_notebook_fixings.add("Termo")
    aba_termo_fixings = sub_sub_notebook_fixings.tab("Termo")

    sub_sub_sub_notebook_termo_fixings = ctk.CTkTabview(aba_termo_fixings)
    sub_sub_sub_notebook_termo_fixings.pack(expand=True, fill='both')

    # Configuração para a sub-sub-aba "Cliente" em "Termo" de "Fixings"
    sub_sub_sub_notebook_termo_fixings.add("Cliente")
    aba_fixingstermo_cliente = sub_sub_sub_notebook_termo_fixings.tab("Cliente")
    
    frame_approval_cliente_fixingstermo = ctk.CTkFrame(aba_fixingstermo_cliente, width=220, fg_color="#D3D3D3")
    frame_approval_cliente_fixingstermo.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_cliente_fixingstermo)

    frame_treeview_fixingstermo_cliente = ctk.CTkFrame(aba_fixingstermo_cliente)
    frame_treeview_fixingstermo_cliente.pack(expand=True, fill='both', side='left')

    frame_scrollbary_fixingstermo_cliente = ctk.CTkFrame(frame_treeview_fixingstermo_cliente, width=4)
    frame_scrollbary_fixingstermo_cliente.pack(fill='y', side='right')

    scrollbar_x_fixingstermo_cliente = ctk.CTkScrollbar(frame_treeview_fixingstermo_cliente, orientation='horizontal')
    scrollbar_y_fixingstermo_cliente = ctk.CTkScrollbar(frame_scrollbary_fixingstermo_cliente, orientation='vertical')

    global tabela_fixingstermo_cliente
    tabela_fixingstermo_cliente = ttk.Treeview(frame_treeview_fixingstermo_cliente, columns=colunas_fixings_termo, show='headings', xscrollcommand=scrollbar_x_fixingstermo_cliente.set, yscrollcommand=scrollbar_y_fixingstermo_cliente.set)
    tabela_fixingstermo_cliente.pack(expand=True, fill='both')
    scrollbar_x_fixingstermo_cliente.configure(command=tabela_fixingstermo_cliente.xview, height=25)
    scrollbar_y_fixingstermo_cliente.configure(command=tabela_fixingstermo_cliente.yview, width=25)
    
    # Configurar cabeçalhos das colunas
    for coluna in colunas_fixings_termo:
        tabela_fixingstermo_cliente.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_fixingstermo_cliente, _col, False))
    vincular_evento_duplo_clique_status(tabela_fixingstermo_cliente, colunas_fixings_termo)
    frame_botoes_fixingstermo_cliente = ctk.CTkFrame(frame_treeview_fixingstermo_cliente, height=150)
    frame_botoes_fixingstermo_cliente.pack(fill='x', side='bottom')

    

    # Configuração para a sub-sub-aba "B2B" em "Termo" de "Fixings"
    sub_sub_sub_notebook_termo_fixings.add("B2B")
    aba_fixingstermo_b2b = sub_sub_sub_notebook_termo_fixings.tab("B2B")
    
    frame_approval_b2b_fixingstermo = ctk.CTkFrame(aba_fixingstermo_b2b, width=220, fg_color="#D3D3D3")
    frame_approval_b2b_fixingstermo.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_b2b_fixingstermo)

    frame_treeview_fixingstermo_b2b  = ctk.CTkFrame(aba_fixingstermo_b2b )
    frame_treeview_fixingstermo_b2b .pack(expand=True, fill='both', side='left')

    frame_scrollbary_fixingstermo_b2b  = ctk.CTkFrame(frame_treeview_fixingstermo_b2b , width=4)
    frame_scrollbary_fixingstermo_b2b .pack(fill='y', side='right')

    scrollbar_x_fixingstermo_b2b  = ctk.CTkScrollbar(frame_treeview_fixingstermo_b2b , orientation='horizontal')
    scrollbar_y_fixingstermo_b2b  = ctk.CTkScrollbar(frame_scrollbary_fixingstermo_b2b , orientation='vertical')

    global tabela_fixingstermo_b2b 
    tabela_fixingstermo_b2b  = ttk.Treeview(frame_treeview_fixingstermo_b2b , columns=colunas_fixings_termo, show='headings', xscrollcommand=scrollbar_x_fixingstermo_b2b .set, yscrollcommand=scrollbar_y_fixingstermo_b2b .set)
    tabela_fixingstermo_b2b .pack(expand=True, fill='both')
    scrollbar_x_fixingstermo_b2b .configure(command=tabela_fixingstermo_b2b.xview, height=25)
    scrollbar_y_fixingstermo_b2b .configure(command=tabela_fixingstermo_b2b.yview, width=25)
    

    # Configurar cabeçalhos das colunas
    for coluna in colunas_fixings_termo:
        tabela_fixingstermo_b2b .heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_fixingstermo_b2b, _col, False))
    vincular_evento_duplo_clique_status(tabela_fixingstermo_b2b, colunas_fixings_termo)
    frame_botoes_fixingstermo_b2b  = ctk.CTkFrame(frame_treeview_fixingstermo_b2b, height=150)
    frame_botoes_fixingstermo_b2b .pack(fill='x', side='bottom')   

    # Configuração para a sub-aba "Opção" em "Fixings"
    sub_sub_notebook_fixings.add("Opção")
    aba_opcao_fixings = sub_sub_notebook_fixings.tab("Opção")

    sub_sub_sub_notebook_opcao_fixings = ctk.CTkTabview(aba_opcao_fixings)
    sub_sub_sub_notebook_opcao_fixings.pack(expand=True, fill='both')

    # Configuração para a sub-sub-aba "Cliente" em "Opção" de "Fixings"
    sub_sub_sub_notebook_opcao_fixings.add("Cliente")
    aba_fixingsopcao_cliente = sub_sub_sub_notebook_opcao_fixings.tab("Cliente")
    
    frame_approval_cliente_fixingsopcao = ctk.CTkFrame(aba_fixingsopcao_cliente, width=220, fg_color="#D3D3D3")
    frame_approval_cliente_fixingsopcao.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_cliente_fixingsopcao)   
    

    frame_treeview_fixingsopcao_cliente = ctk.CTkFrame(aba_fixingsopcao_cliente)
    frame_treeview_fixingsopcao_cliente.pack(expand=True, fill='both', side='left')

    frame_scrollbary_fixingsopcao_cliente = ctk.CTkFrame(frame_treeview_fixingsopcao_cliente, width=4)
    frame_scrollbary_fixingsopcao_cliente.pack(fill='y', side='right')

    scrollbar_x_fixingsopcao_cliente = ctk.CTkScrollbar(frame_treeview_fixingsopcao_cliente, orientation='horizontal')
    scrollbar_y_fixingsopcao_cliente = ctk.CTkScrollbar(frame_scrollbary_fixingsopcao_cliente, orientation='vertical')

    global tabela_fixingsopcao_cliente
    tabela_fixingsopcao_cliente = ttk.Treeview(frame_treeview_fixingsopcao_cliente, columns=colunas_fixings_opcao, show='headings', xscrollcommand=scrollbar_x_fixingsopcao_cliente.set, yscrollcommand=scrollbar_y_fixingsopcao_cliente.set)
    tabela_fixingsopcao_cliente.pack(expand=True, fill='both')
    scrollbar_x_fixingsopcao_cliente.configure(command=tabela_fixingsopcao_cliente.xview, height=25)
    scrollbar_y_fixingsopcao_cliente.configure(command=tabela_fixingsopcao_cliente.yview, width=25)
    

    # Configurar cabeçalhos das colunas
    for coluna in colunas_fixings_opcao:
        tabela_fixingsopcao_cliente.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_fixingsopcao_cliente, _col, False))
    vincular_evento_duplo_clique_status(tabela_fixingsopcao_cliente, colunas_fixings_opcao)
    frame_botoes_fixingsopcao_cliente = ctk.CTkFrame(frame_treeview_fixingsopcao_cliente, height=150)
    frame_botoes_fixingsopcao_cliente.pack(fill='x', side='bottom')

   

    # Configuração para a sub-sub-aba "B2B" em "Opção" de "Fixings"
    sub_sub_sub_notebook_opcao_fixings.add("B2B")
    aba_fixingsopcao_b2b = sub_sub_sub_notebook_opcao_fixings.tab("B2B")
    
    frame_approval_b2b_fixingsopcao = ctk.CTkFrame(aba_fixingsopcao_b2b, width=220, fg_color="#D3D3D3")
    frame_approval_b2b_fixingsopcao.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_b2b_fixingsopcao)

    frame_treeview_fixingsopcao_b2b = ctk.CTkFrame(aba_fixingsopcao_b2b)
    frame_treeview_fixingsopcao_b2b.pack(expand=True, fill='both', side='left')

    frame_scrollbary_fixingsopcao_b2b = ctk.CTkFrame(frame_treeview_fixingsopcao_b2b, width=4)
    frame_scrollbary_fixingsopcao_b2b.pack(fill='y', side='right')

    scrollbar_x_fixingsopcao_b2b = ctk.CTkScrollbar(frame_treeview_fixingsopcao_b2b, orientation='horizontal')
    scrollbar_y_fixingsopcao_b2b = ctk.CTkScrollbar(frame_scrollbary_fixingsopcao_b2b, orientation='vertical')

    global tabela_fixingsopcao_b2b
    tabela_fixingsopcao_b2b = ttk.Treeview(frame_treeview_fixingsopcao_b2b, columns=colunas_fixings_opcao, show='headings', xscrollcommand=scrollbar_x_fixingsopcao_b2b.set, yscrollcommand=scrollbar_y_fixingsopcao_b2b.set)
    tabela_fixingsopcao_b2b.pack(expand=True, fill='both')
    scrollbar_x_fixingsopcao_b2b.configure(command=tabela_fixingsopcao_b2b.xview, height=25)
    scrollbar_y_fixingsopcao_b2b.configure(command=tabela_fixingsopcao_b2b.yview, width=25)
    

    # Configurar cabeçalhos das colunas
    for coluna in colunas_fixings_opcao:
        tabela_fixingsopcao_b2b.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_fixingsopcao_b2b, _col, False))
    vincular_evento_duplo_clique_status(tabela_fixingsopcao_b2b, colunas_fixings_opcao)
    
    frame_botoes_fixingsopcao_b2b = ctk.CTkFrame(frame_treeview_fixingsopcao_b2b, height=150)
    frame_botoes_fixingsopcao_b2b.pack(fill='x', side='bottom')

    # Botoes Monitor Operações Termo Cliente
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")    
    

    treeviews = []

   

    botao_review_termo_cliente = ctk.CTkButton(frame_approval_cliente_termo, width=100, height=26, corner_radius=8, fg_color="#5A5368", text="REVIEW", font=fonte_botao, command=lambda: approve_window_status)
    botao_review_termo_cliente.pack(padx=5, pady=5)   
    
  
    botao_importar_termo_cliente = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_termo_cliente.pack(side='left', padx=1, pady=10)
    
    botao_email_termo_cliente = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=20, fg_color="#5A5368", text="EA", font=fonte_botao, command=lambda: email_if(tabela_termo_cliente))
    botao_email_termo_cliente.pack(side='left', padx=0, pady=10)

    botao_limpar_termo_cliente = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_termo(label_qty_deals_cliente_termo))
    botao_limpar_termo_cliente.pack(side='left', padx=1, pady=10)
    
    botao_populardatas_termo_cliente = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_termo_cliente.pack(side='left', padx=1, pady=10)

    botao_popular_cliente_termo = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_cliente_termo.pack(side='left', padx=1, pady=10)

    botao_delete_termo_cliente = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_termo_cliente, label_qty_deals_cliente_termo))
    botao_delete_termo_cliente.pack(side='left', padx=1, pady=10)
    
    botao_excel_cliente_termo = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=20, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_termo(abas_existentes, tabela_termo_cliente, tabela_termo_b2b, tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, switch_cliente_termo))
    botao_excel_cliente_termo.pack(side='left', padx=0, pady=10)

    botao_popularboleta_cliente_termo = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_cliente_termo.pack(side='left', padx=1, pady=10)   

    botao_confirmation_cliente_termo = ctk.CTkButton(frame_botoes_cliente_termo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_cliente_termo.pack(side='left', padx=1, pady=10)

    switch_var_cliente_termo = ctk.StringVar(value="on")  

    def alterna_switch_cliente_termo():        
        switch_cliente_termo.configure(text="Per Client" if switch_var_cliente_termo.get() == "on" else "All")  

    switch_cliente_termo = ctk.CTkSwitch(frame_botoes_cliente_termo, text="Per Client", variable=switch_var_cliente_termo, onvalue="on", offvalue="off", command=lambda: alterna_switch_cliente_termo())
    switch_cliente_termo.pack(side='left', padx=1, pady=10)   
    global label_qty_deals_cliente_termo
    label_deals_cliente_termo = ctk.CTkLabel(frame_botoes_cliente_termo, width=50, height=26, corner_radius=8, fg_color="#D3D3D3", text="Deals:", font=fonte_botao)
    label_deals_cliente_termo.pack(side='left', padx=5, pady=10)  
    label_qty_deals_cliente_termo = ctk.CTkLabel(frame_botoes_cliente_termo, width=80, height=26, corner_radius=8, fg_color="#D3D3D3", text="", font=fonte_botao)
    label_qty_deals_cliente_termo.pack(side='left', padx=1, pady=10)
    

    scrollbar_x_cliente_termo.pack(side='bottom', fill='x')
    scrollbar_y_cliente_termo.pack(side='right', fill='y')

     # Botoes Monitor Operações Termo B2b
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")
    botao_review_termo_b2b = ctk.CTkButton(frame_approval_b2b_termo, width=100, height=26, corner_radius=8, fg_color="#5A5368", text="REVIEW", font=fonte_botao, command=lambda: approve_window_status)
    botao_review_termo_b2b.pack(padx=5, pady=5)   

    botao_importar_termo_b2b = ctk.CTkButton(frame_botoes_termo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_termo_b2b.pack(side='left', padx=1, pady=10)

    botao_limpar_termo_b2b = ctk.CTkButton(frame_botoes_termo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_termo(label_qty_deals_b2b_termo))
    botao_limpar_termo_b2b.pack(side='left', padx=1, pady=10)

    botao_populardatas_termo_b2b = ctk.CTkButton(frame_botoes_termo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_termo_b2b.pack(side='left', padx=1, pady=10)

    botao_popular_termo_b2b = ctk.CTkButton(frame_botoes_termo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_termo_b2b.pack(side='left', padx=1, pady=10)

    botao_delete_termo_b2b = ctk.CTkButton(frame_botoes_termo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_termo_b2b, label_qty_deals_b2b_termo))
    botao_delete_termo_b2b.pack(side='left', padx=1, pady=10)

    botao_popularboleta_b2b_termo = ctk.CTkButton(frame_botoes_termo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview, abas_existentes))
    botao_popularboleta_b2b_termo.pack(side='left', padx=1, pady=10)

    botao_confirmation_termo_b2b = ctk.CTkButton(frame_botoes_termo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_termo_b2b.pack(side='left', padx=1, pady=10)    

    switch_var_termo_b2b = ctk.StringVar(value="on")

    def alterna_switch_termo_b2b(): 
        switch_termo_b2b.configure(text="Per Client" if switch_var_termo_b2b.get() == "on" else "All")

    switch_termo_b2b = ctk.CTkSwitch(frame_botoes_termo_b2b, text="Per Client", variable=switch_var_termo_b2b, onvalue="on", offvalue="off", command=lambda: alterna_switch_termo_b2b())
    switch_termo_b2b.pack(side='left', padx=1, pady=10)

    global label_qty_deals_b2b_termo
    label_deals_b2b_termo = ctk.CTkLabel(frame_botoes_termo_b2b, width=50, height=26, corner_radius=8, fg_color="#D3D3D3", text="Deals:", font=fonte_botao)
    label_deals_b2b_termo.pack(side='left', padx=5, pady=10)  
    label_qty_deals_b2b_termo = ctk.CTkLabel(frame_botoes_termo_b2b, width=80, height=26, corner_radius=8, fg_color="#D3D3D3", text="", font=fonte_botao)
    label_qty_deals_b2b_termo.pack(side='left', padx=1, pady=10)

    scrollbar_x_termo_b2b.pack(side='bottom', fill='x')
    scrollbar_y_termo_b2b.pack(side='right', fill='y')

     # Botoes Monitor Operações Opção Cliente
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")    

    botao_review_opcao_cliente = ctk.CTkButton(frame_approval_cliente_opcao, width=100, height=26, corner_radius=8, fg_color="#5A5368", text="REVIEW", font=fonte_botao, command=lambda: approve_window_status)
    botao_review_opcao_cliente.pack(padx=5, pady=5)    

    botao_email_premio_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="PREMIUM ADVICE", font=fonte_botao, command=lambda: email_Premio(tabela_opcao_cliente))
    botao_email_premio_opcao.pack(side='left', padx=1, pady=10)
    
    #botao_email_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=20, fg_color="#5A5368", text="E-MAIL", font=fonte_botao, command=lambda: validation_email_opcao(tabela_opcao_cliente))
    #botao_email_cliente_opcao.pack(side='left', padx=0, pady=10)
    
    botao_importar_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IIMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_cliente_opcao.pack(side='left', padx=1, pady=10)

    botao_limpar_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_opcao())
    botao_limpar_cliente_opcao.pack(side='left', padx=1, pady=10)

    botao_populardatas_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_cliente_opcao.pack(side='left', padx=1, pady=10)

    botao_popular_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_cliente_opcao.pack(side='left', padx=1, pady=10)
    
    botao_excel_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=20, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_opcao(abas_existentes, tabela_opcao_cliente, tabela_opcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b, switch_cliente_opcao))
    botao_excel_cliente_opcao.pack(side='left', padx=0, pady=10)
    
    botao_delete_opcao_cliente = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_opcao_cliente, label_qty_deals_cliente_opcao))
    botao_delete_opcao_cliente.pack(side='left', padx=1, pady=10)  

    botao_popularboleta_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview, abas_existentes))
    botao_popularboleta_cliente_opcao.pack(side='left', padx=1, pady=10)        

    botao_confirmation_cliente_opcao = ctk.CTkButton(frame_botoes_cliente_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_opcao(tabview))
    botao_confirmation_cliente_opcao.pack(side='left', padx=1, pady=10)

    switch_var_cliente_opcao = ctk.StringVar(value="on")

    def alterna_switch_cliente_opcao(): 
        switch_cliente_opcao.configure(text="Per Client" if switch_var_cliente_termo.get() == "on" else "All")

    switch_cliente_opcao = ctk.CTkSwitch(frame_botoes_cliente_opcao, text="Per Client", variable=switch_var_cliente_opcao, onvalue="on", offvalue="off", command=lambda: alterna_switch_cliente_opcao())
    switch_cliente_opcao.pack(side='left', padx=1, pady=10)

    global label_qty_deals_cliente_opcao
    label_deals_cliente_opcao = ctk.CTkLabel(frame_botoes_cliente_opcao, width=50, height=26, corner_radius=8, fg_color="#D3D3D3", text="Deals:", font=fonte_botao)
    label_deals_cliente_opcao.pack(side='left', padx=5, pady=10)  
    label_qty_deals_cliente_opcao = ctk.CTkLabel(frame_botoes_cliente_opcao, width=80, height=26, corner_radius=8, fg_color="#D3D3D3", text="", font=fonte_botao)
    label_qty_deals_cliente_opcao.pack(side='left', padx=1, pady=10)
    


    scrollbar_x_cliente_opcao.pack(side='bottom', fill='x')
    scrollbar_y_cliente_opcao.pack(side='right', fill='y')

    # Botoes Monitor Operações Opção B2b
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")   
    
    botao_review_opcao_b2b = ctk.CTkButton(frame_approval_b2b_opcao, width=100, height=26, corner_radius=8, fg_color="#5A5368", text="REVIEW", font=fonte_botao, command=lambda: approve_window_status)
    botao_review_opcao_b2b.pack(padx=5, pady=5)   
   
    botao_importar_b2b_opcao = ctk.CTkButton(frame_botoes_b2b_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_b2b_opcao.pack(side='left', padx=1, pady=10)

    botao_limpar_b2b_opcao = ctk.CTkButton(frame_botoes_b2b_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda: limpar_dados_opcao(label_qty_deals_b2b_opcao))
    botao_limpar_b2b_opcao.pack(side='left', padx=1, pady=10)

    botao_populardatas_b2b_opcao = ctk.CTkButton(frame_botoes_b2b_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_b2b_opcao.pack(side='left', padx=1, pady=10)

    botao_popular_b2b_opcao = ctk.CTkButton(frame_botoes_b2b_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_b2b_opcao.pack(side='left', padx=1, pady=10)    

    botao_delete_opcao_b2b = ctk.CTkButton(frame_botoes_b2b_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_opcao_b2b, label_qty_deals_b2b_opcao))
    botao_delete_opcao_b2b.pack(side='left', padx=1, pady=10)

    botao_popularboleta_b2b_opcao = ctk.CTkButton(frame_botoes_b2b_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview, abas_existentes))
    botao_popularboleta_b2b_opcao.pack(side='left', padx=1, pady=10)

    botao_confirmation_b2b_opcao = ctk.CTkButton(frame_botoes_b2b_opcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_opcao(tabview))
    botao_confirmation_b2b_opcao.pack(side='left', padx=1, pady=10)

    switch_var_b2b_opcao = ctk.StringVar(value="on")

    def alterna_switch_b2b_opcao(): 
        switch_b2b_opcao.configure(text="Per Client" if switch_var_b2b_opcao.get() == "on" else "All")

    switch_b2b_opcao = ctk.CTkSwitch(frame_botoes_b2b_opcao, text="Per Client", variable=switch_var_b2b_opcao, onvalue="on", offvalue="off", command=lambda: alterna_switch_b2b_opcao())
    switch_b2b_opcao.pack(side='left', padx=1, pady=10)

    global label_qty_deals_b2b_opcao
    label_deals_b2b_opcao = ctk.CTkLabel(frame_botoes_b2b_opcao, width=50, height=26, corner_radius=8, fg_color="#D3D3D3", text="Deals:", font=fonte_botao)
    label_deals_b2b_opcao.pack(side='left', padx=5, pady=10)  
    label_qty_deals_b2b_opcao = ctk.CTkLabel(frame_botoes_b2b_opcao, width=80, height=26, corner_radius=8, fg_color="#D3D3D3", text="", font=fonte_botao)
    label_qty_deals_b2b_opcao.pack(side='left', padx=1, pady=10)

    scrollbar_x_b2b_opcao.pack(side='bottom', fill='x')
    scrollbar_y_b2b_opcao.pack(side='right', fill='y')

    # Botoes Fixings Termo Cliente

    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    botao_importar_fixingstermo_cliente = ctk.CTkButton(frame_botoes_fixingstermo_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_fixingstermo_cliente.pack(side='left', padx=1, pady=10)

    botao_limpar_fixingstermo_cliente = ctk.CTkButton(frame_botoes_fixingstermo_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda: limpar_dados_fixingstermo())
    botao_limpar_fixingstermo_cliente.pack(side='left', padx=1, pady=10)

    botao_populardatas_fixingstermo_cliente = ctk.CTkButton(frame_botoes_fixingstermo_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_fixingstermo_cliente.pack(side='left', padx=1, pady=10)

    botao_popular_cliente_fixingstermo = ctk.CTkButton(frame_botoes_fixingstermo_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_cliente_fixingstermo.pack(side='left', padx=1, pady=10)

    botao_delete_cliente_fixingtermo = ctk.CTkButton(frame_botoes_fixingstermo_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_fixingstermo_cliente, []))
    botao_delete_cliente_fixingtermo.pack(side='left', padx=1, pady=10)

    botao_popularboleta_cliente_fixingstermo = ctk.CTkButton(frame_botoes_fixingstermo_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_cliente_fixingstermo.pack(side='left', padx=1, pady=10)    

    botao_confirmation_cliente_fixingstermo = ctk.CTkButton(frame_botoes_fixingstermo_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_cliente_fixingstermo.pack(side='left', padx=1, pady=10)

    switch_var_fixingstermo_cliente = ctk.StringVar(value="on")

    def alterna_switch_fixingstermo_cliente(): 
        switch_fixingstermo_cliente.configure(text="Per Client" if switch_var_fixingstermo_cliente.get() == "on" else "All")

    switch_fixingstermo_cliente = ctk.CTkSwitch(frame_botoes_fixingstermo_cliente, text="Per Client", variable=switch_var_fixingstermo_cliente, onvalue="on", offvalue="off", command=lambda: alterna_switch_fixingstermo_cliente())
    switch_fixingstermo_cliente.pack(side='left', padx=1, pady=10)

    scrollbar_x_fixingstermo_cliente.pack(side='bottom', fill='x')
    scrollbar_y_fixingstermo_cliente.pack(side='right', fill='y')

      # Botoes Fixings Termo B2b

    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    botao_importar_fixingstermo_b2b = ctk.CTkButton(frame_botoes_fixingstermo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_fixingstermo_b2b.pack(side='left', padx=1, pady=10)

    botao_limpar_fixingstermo_b2b = ctk.CTkButton(frame_botoes_fixingstermo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda: limpar_dados_fixingstermo())
    botao_limpar_fixingstermo_b2b.pack(side='left', padx=1, pady=10)

    botao_populardatas_fixingstermo_b2b = ctk.CTkButton(frame_botoes_fixingstermo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_fixingstermo_b2b.pack(side='left', padx=1, pady=10)

    botao_popular_b2b_fixingstermo = ctk.CTkButton(frame_botoes_fixingstermo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_b2b_fixingstermo.pack(side='left', padx=1, pady=10)

    botao_delete_b2b_fixingtermo = ctk.CTkButton(frame_botoes_fixingstermo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_fixingstermo_b2b, []))
    botao_delete_b2b_fixingtermo.pack(side='left', padx=1, pady=10) 

    botao_popularboleta_b2b_fixingstermo = ctk.CTkButton(frame_botoes_fixingstermo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_b2b_fixingstermo.pack(side='left', padx=1, pady=10)     

    botao_confirmation_b2b_fixingstermo = ctk.CTkButton(frame_botoes_fixingstermo_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_b2b_fixingstermo.pack(side='left', padx=1, pady=10)

    switch_var_fixingstermo_b2b = ctk.StringVar(value="on")

    def alterna_switch_fixingstermo_b2b(): 
        switch_fixingstermo_b2b.configure(text="Per Client" if switch_var_fixingstermo_b2b.get() == "on" else "All")

    switch_fixingstermo_b2b = ctk.CTkSwitch(frame_botoes_fixingstermo_b2b, text="Per Client", variable=switch_var_fixingstermo_b2b, onvalue="on", offvalue="off", command=lambda: alterna_switch_fixingstermo_b2b())
    switch_fixingstermo_b2b.pack(side='left', padx=1, pady=10)

    scrollbar_x_fixingstermo_b2b.pack(side='bottom', fill='x')
    scrollbar_y_fixingstermo_b2b.pack(side='right', fill='y')

    # Botoes Fixings Opção Cliente
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")    
 

    botao_importar_fixingsopcao_cliente = ctk.CTkButton(frame_botoes_fixingsopcao_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_fixingsopcao_cliente.pack(side='left', padx=1, pady=10)

    botao_limpar_fixingsopcao_cliente = ctk.CTkButton(frame_botoes_fixingsopcao_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_fixingsopcao())
    botao_limpar_fixingsopcao_cliente.pack(side='left', padx=1, pady=10)

    botao_populardatas_fixingsopcao_cliente = ctk.CTkButton(frame_botoes_fixingsopcao_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_fixingsopcao_cliente.pack(side='left', padx=1, pady=10)

    botao_popular_cliente_fixingsopcao = ctk.CTkButton(frame_botoes_fixingsopcao_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_cliente_fixingsopcao.pack(side='left', padx=1, pady=10)

    botao_delete_cliente_fixingsopcao = ctk.CTkButton(frame_botoes_fixingsopcao_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_fixingsopcao_cliente, []))
    botao_delete_cliente_fixingsopcao.pack(side='left', padx=1, pady=10)

    botao_popularboleta_cliente_fixingsopcao = ctk.CTkButton(frame_botoes_fixingsopcao_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_cliente_fixingsopcao.pack(side='left', padx=1, pady=10)       

    botao_confirmation_fixingsopcao_cliente = ctk.CTkButton(frame_botoes_fixingsopcao_cliente, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_opcao(tabview))
    botao_confirmation_fixingsopcao_cliente.pack(side='left', padx=1, pady=10)

    switch_var_fixingsopcao_cliente = ctk.StringVar(value="on")

    def alterna_switch_fixingsopcao_cliente(): 
        switch_fixingsopcao_cliente.configure(text="Per Client" if switch_var_fixingsopcao_cliente.get() == "on" else "All")

    switch_fixingsopcao_cliente = ctk.CTkSwitch(frame_botoes_fixingsopcao_cliente, text="Per Client", variable=switch_var_fixingsopcao_cliente, onvalue="on", offvalue="off", command=lambda: alterna_switch_fixingsopcao_cliente())
    switch_fixingsopcao_cliente.pack(side='left', padx=1, pady=10)

    scrollbar_x_fixingsopcao_cliente.pack(side='bottom', fill='x')
    scrollbar_y_fixingsopcao_cliente.pack(side='right', fill='y')    
    
    # Botoes Fixings Opção B2b
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")  

    botao_importar_fixingsopcao_b2b = ctk.CTkButton(frame_botoes_fixingsopcao_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_fixingsopcao_b2b.pack(side='left', padx=1, pady=10)

    botao_limpar_fixingsopcao_b2b = ctk.CTkButton(frame_botoes_fixingsopcao_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_fixingsopcao())
    botao_limpar_fixingsopcao_b2b.pack(side='left', padx=1, pady=10)

    botao_populardatas_fixingsopcao_b2b = ctk.CTkButton(frame_botoes_fixingsopcao_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_fixingsopcao_b2b.pack(side='left', padx=1, pady=10)

    botao_popular_b2b_fixingsopcao = ctk.CTkButton(frame_botoes_fixingsopcao_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_b2b_fixingsopcao.pack(side='left', padx=1, pady=10)

    botao_delete_b2b_fixingsopcao = ctk.CTkButton(frame_botoes_fixingsopcao_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_fixingsopcao_b2b, []))
    botao_delete_b2b_fixingsopcao.pack(side='left', padx=1, pady=10) 

    botao_popularboleta_b2b_fixingsopcao = ctk.CTkButton(frame_botoes_fixingsopcao_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_b2b_fixingsopcao.pack(side='left', padx=1, pady=10)    

    botao_confirmation_fixingsopcao_b2b = ctk.CTkButton(frame_botoes_fixingsopcao_b2b, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_opcao(tabview))
    botao_confirmation_fixingsopcao_b2b.pack(side='left', padx=1, pady=10)

    switch_var_fixingsopcao_b2b = ctk.StringVar(value="on")

    def alterna_switch_fixingsopcao_b2b(): 
        switch_fixingsopcao_b2b.configure(text="Per Client" if switch_var_fixingsopcao_b2b.get() == "on" else "All")

    switch_fixingsopcao_b2b = ctk.CTkSwitch(frame_botoes_fixingsopcao_b2b, text="Per Client", variable=switch_var_fixingsopcao_b2b, onvalue="on", offvalue="off", command=lambda: alterna_switch_fixingsopcao_b2b())
    switch_fixingsopcao_b2b.pack(side='left', padx=1, pady=10)

    scrollbar_x_fixingsopcao_b2b.pack(side='bottom', fill='x')
    scrollbar_y_fixingsopcao_b2b.pack(side='right', fill='y')
    arquivo_b3(tabview,abas_existentes)   
    calendarios_bolsas(tabview, abas_existentes)
    
    treeviews = {
    'tabela_termo_cliente': tabela_termo_cliente,
    'tabela_termo_b2b': tabela_termo_b2b,
    'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
    'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
    'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
    'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b,
    'tabela_opcao_cliente': tabela_opcao_cliente,
    'tabela_opcao_b2b': tabela_opcao_b2b,
    'tabela_arquivoopcao_cliente': tabela_arquivoopcao_cliente,
    'tabela_arquivoopcao_b2b': tabela_arquivoopcao_b2b,
    'tabela_fixingsopcao_cliente': tabela_fixingsopcao_cliente,
    'tabela_fixingsopcao_b2b': tabela_fixingsopcao_b2b
    }
    
    # Query Termo Cliente
    botao_query_termo_cliente = tk.Button(frame_query_cliente_termo, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_query_window)
    botao_query_termo_cliente.pack(side='left', padx=2, pady=0)  
    global entry_query_termo_cliente
    entry_query_termo_cliente = ctk.CTkEntry(frame_query_cliente_termo, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_termo_cliente.pack(side='left', padx=2, pady=0)  
    botao_load_termo_cliente = tk.Button(frame_query_cliente_termo, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry(entry_query_termo_cliente, treeviews))                               
    botao_load_termo_cliente.pack(side='left', padx=2, pady=0)  
    # Query Termo B2B
    botao_query_termo_b2b = tk.Button(frame_query_b2b_termo, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_query_window)
    botao_query_termo_b2b.pack(side='left', padx=2, pady=0)  
    global entry_query_termo_b2b
    entry_query_termo_b2b = ctk.CTkEntry(frame_query_b2b_termo, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_termo_b2b.pack(side='left', padx=2, pady=0)  
    botao_load_termo_b2b = tk.Button(frame_query_b2b_termo, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry(entry_query_termo_b2b, treeviews))                               
    botao_load_termo_b2b.pack(side='left', padx=2, pady=0)  
    # Query Opção Cliente
    botao_query_opcao_cliente = tk.Button(frame_query_cliente_opcao, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_option_query_window)
    botao_query_opcao_cliente.pack(side='left', padx=2, pady=0)  
    global entry_query_opcao_cliente
    entry_query_opcao_cliente = ctk.CTkEntry(frame_query_cliente_opcao, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_opcao_cliente.pack(side='left', padx=2, pady=0)  
    botao_load_opcao_cliente = tk.Button(frame_query_cliente_opcao, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry_option(entry_query_opcao_cliente, treeviews))                               
    botao_load_opcao_cliente.pack(side='left', padx=2, pady=0)  
    # Query Opção B2B
    botao_query_opcao_b2b = tk.Button(frame_query_b2b_opcao, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_option_query_window)
    botao_query_opcao_b2b.pack(side='left', padx=2, pady=0)  
    global entry_query_opcao_b2b
    entry_query_opcao_b2b = ctk.CTkEntry(frame_query_b2b_opcao, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_opcao_b2b.pack(side='left', padx=2, pady=0)  
    botao_load_opcao_b2b = tk.Button(frame_query_b2b_opcao, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry_option(entry_query_opcao_b2b, treeviews))                               
    botao_load_opcao_b2b.pack(side='left', padx=2, pady=0)  
    
    
    
    ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)
    
  
            
            
# CTRL + A event, select all lines for the Active Tabview + Treeview
def select_all(event):
    # Determine the active tab at each level
    main_tab = tabview.get()
    if main_tab == "Monitor":
        sub_tab = sub_notebook_monitor.get()
        if sub_tab == "Operações":
            sub_sub_tab = sub_sub_notebook_operacoes.get()
            if sub_sub_tab == "Termo":
                sub_sub_sub_tab = sub_sub_sub_notebook_termo.get()
                if sub_sub_sub_tab == "Cliente":
                    for item in tabela_termo_cliente.get_children():
                        tabela_termo_cliente.selection_add(item)
                elif sub_sub_sub_tab == "B2B":
                    for item in tabela_termo_b2b.get_children():
                        tabela_termo_b2b.selection_add(item)
            elif sub_sub_tab == "Opção":
                sub_sub_sub_tab = sub_sub_sub_notebook_opcao.get()
                if sub_sub_sub_tab == "Cliente":
                    for item in tabela_opcao_cliente.get_children():
                        tabela_opcao_cliente.selection_add(item)
                elif sub_sub_sub_tab == "B2B":
                    for item in tabela_opcao_b2b.get_children():
                        tabela_opcao_b2b.selection_add(item)
        elif sub_tab == "Fixings":
            sub_sub_tab = sub_sub_notebook_fixings.get()
            if sub_sub_tab == "Termo":
                sub_sub_sub_tab = sub_sub_sub_notebook_termo_fixings.get()
                if sub_sub_sub_tab == "Cliente":
                    for item in tabela_fixingstermo_cliente.get_children():
                        tabela_fixingstermo_cliente.selection_add(item)
                elif sub_sub_sub_tab == "B2B":
                    for item in tabela_fixingstermo_b2b.get_children():
                        tabela_fixingstermo_b2b.selection_add(item)
            elif sub_sub_tab == "Opção":
                sub_sub_sub_tab = sub_sub_sub_notebook_opcao_fixings.get()
                if sub_sub_sub_tab == "Cliente":
                    for item in tabela_fixingsopcao_cliente.get_children():
                        tabela_fixingsopcao_cliente.selection_add(item)
                elif sub_sub_sub_tab == "B2B":
                    for item in tabela_fixingsopcao_b2b.get_children():
                        tabela_fixingsopcao_b2b.selection_add(item)
    elif main_tab == "Arquivo B3":
        sub_tab = sub_notebook_arquivo_b3.get()
        if sub_tab == "Termo":
            sub_sub_tab = sub_sub_notebook_arquivotermo.get()
            if sub_sub_tab == "Cliente":
                for item in tabela_arquivotermo_cliente.get_children():
                    tabela_arquivotermo_cliente.selection_add(item)
            elif sub_sub_tab == "B2B":
                for item in tabela_arquivotermo_b2b.get_children():
                    tabela_arquivotermo_b2b.selection_add(item)
        elif sub_tab == "Opção":
            sub_sub_tab = sub_sub_notebook_arquivoopcao.get()
            if sub_sub_tab == "Cliente":
                for item in tabela_arquivoopcao_cliente.get_children():
                    tabela_arquivoopcao_cliente.selection_add(item)
            elif sub_sub_tab == "B2B":
                for item in tabela_arquivoopcao_b2b.get_children():
                    tabela_arquivoopcao_b2b.selection_add(item)
    elif main_tab == "Boleta Dinâmica - Intrag":
        sub_tab = sub_notebook_boleta_dinamica.get()
        if sub_tab == "Termo":
            for item in tabela_boletatermo.get_children():
                tabela_boletatermo.selection_add(item)
        elif sub_tab == "Opção":
            for item in tabela_boletaopcao.get_children():
                tabela_boletaopcao.selection_add(item)
    elif main_tab == "Commodities":
        for item in tabela_commodities.get_children():
            tabela_commodities.selection_add(item)
    elif main_tab == "Counterparty":
        for item in tabela_base_comitentes.get_children():
            tabela_base_comitentes.selection_add(item)
    elif main_tab == "Holidays":
        sub_tab = tabview_calendarios.get()
        if sub_tab == "ANBIMA":
            for item in tabela_anbima.get_children():
                tabela_anbima.selection_add(item)
        elif sub_tab == "ICE":
            for item in tabela_ice.get_children():
                tabela_ice.selection_add(item)
        elif sub_tab == "NYMEX":
            for item in tabela_nymex.get_children():
                tabela_nymex.selection_add(item)
        elif sub_tab == "BURSA":
            for item in tabela_bursa.get_children():
                tabela_bursa.selection_add(item)
        elif sub_tab == "CBOT":
            for item in tabela_cbot.get_children():
                tabela_cbot.selection_add(item)
        elif sub_tab == "PLATTS":
            for item in tabela_platts.get_children():
                tabela_platts.selection_add(item)
        elif sub_tab == "LME":
            for item in tabela_lme.get_children():
                tabela_lme.selection_add(item)
                
def approve_window_status():
    # Determine the active tab at each level
    main_tab = tabview.get()
    if main_tab == "Monitor":
        sub_tab = sub_notebook_monitor.get()
        if sub_tab == "Operações":
            sub_sub_tab = sub_sub_notebook_operacoes.get()
            if sub_sub_tab == "Termo":
                sub_sub_sub_tab = sub_sub_sub_notebook_termo.get()
                if sub_sub_sub_tab == "Cliente":
                    create_list_status("Termo Cliente")
                elif sub_sub_sub_tab == "B2B":
                    create_list_status("Termo B2B")
            elif sub_sub_tab == "Opção":
                sub_sub_sub_tab = sub_sub_sub_notebook_opcao.get()
                if sub_sub_sub_tab == "Cliente":
                    create_list_status("Opção Cliente")
                elif sub_sub_sub_tab == "B2B":
                    create_list_status("Opção B2B")
        elif sub_tab == "Fixings":
            sub_sub_tab = sub_sub_notebook_fixings.get()
            if sub_sub_tab == "Termo":
                sub_sub_sub_tab = sub_sub_sub_notebook_termo_fixings.get()
                if sub_sub_sub_tab == "Cliente":
                    create_list_status("Fixings Termo Cliente")
                elif sub_sub_sub_tab == "B2B":
                    create_list_status("Fixings Termo B2B")
            elif sub_sub_tab == "Opção":
                sub_sub_sub_tab = sub_sub_sub_notebook_opcao_fixings.get()
                if sub_sub_sub_tab == "Cliente":
                    create_list_status("Fixings Opção Cliente")
                elif sub_sub_sub_tab == "B2B":
                    create_list_status("Fixings Opção B2B")
    elif main_tab == "Arquivo B3":
        sub_tab = sub_notebook_arquivo_b3.get()
        if sub_tab == "Termo":
            sub_sub_tab = sub_sub_notebook_arquivotermo.get()
            if sub_sub_tab == "Cliente":
                create_list_status("Arquivo Termo Cliente")
            elif sub_sub_tab == "B2B":
                create_list_status("Arquivo Termo B2B")
        elif sub_tab == "Opção":
            sub_sub_tab = sub_sub_notebook_arquivoopcao.get()
            if sub_sub_tab == "Cliente":
                create_list_status("Arquivo Opção Cliente")
            elif sub_sub_tab == "B2B":
                create_list_status("Arquivo Opção B2B")

                
def create_list_status(tab_name):
    # Determine which treeview to use based on the tab name
    if "Termo Cliente" in tab_name:
        treeview = tabela_termo_cliente
    elif "Termo B2B" in tab_name:
        treeview = tabela_termo_b2b
    elif "Opção Cliente" in tab_name:
        treeview = tabela_opcao_cliente
    elif "Opção B2B" in tab_name:
        treeview = tabela_opcao_b2b
    elif "Arquivo Termo Cliente" in tab_name:
        treeview = tabela_arquivotermo_cliente
    elif "Arquivo Termo B2B" in tab_name:
        treeview = tabela_arquivotermo_b2b
    elif "Arquivo Opção Cliente" in tab_name:
        treeview = tabela_arquivoopcao_cliente
    elif "Arquivo Opção B2B" in tab_name:
        treeview = tabela_arquivoopcao_b2b
    elif "Fixings Termo Cliente" in tab_name:
        treeview = tabela_fixingstermo_cliente
    elif "Fixings Termo B2B" in tab_name:
        treeview = tabela_fixingstermo_b2b
    elif "Fixings Opção Cliente" in tab_name:
        treeview = tabela_fixingsopcao_cliente
    elif "Fixings Opção B2B" in tab_name:
        treeview = tabela_fixingsopcao_b2b
    else:
        return  # No valid tab name, exit function

    if "Termo" in tab_name:
        treeview_deals = tabela_termo_cliente
        treeview_deals_b2b = tabela_termo_b2b
        treeview_file = tabela_arquivotermo_cliente
        treeview_file_b2b = tabela_arquivotermo_b2b
        treeview_fixings = tabela_fixingstermo_cliente
        treeview_fixings_b2b = tabela_fixingstermo_b2b
        instrument_deals = "termo_base_deals"
        instrument_file = "termo_base_file"
        instrument_fixings = "termo_base_fixings"
        columns_deals = termo_deals_columns
        columns_file = termo_file_columns
        columns_fixings = termo_fixings_columns
        
    elif "Opção" in tab_name:
        treeview_deals = tabela_opcao_cliente
        treeview_deals_b2b = tabela_opcao_b2b
        treeview_file = tabela_arquivoopcao_cliente
        treeview_file_b2b = tabela_arquivoopcao_b2b
        treeview_fixings = tabela_fixingsopcao_cliente
        treeview_fixings_b2b = tabela_fixingsopcao_b2b
        instrument_deals = "opcao_base_deals"
        instrument_file = "opcao_base_file"
        instrument_fixings = "opcao_base_fixings"
        columns_deals = opcao_deals_columns
        columns_file = opcao_file_columns
        columns_fixings = opcao_fixings_columns      
        
    identifiers_pending = set()
    identifiers_maker_checker = set()
    identifiers_approved = set()
    identifiers_concluded = set()
    
    # Export data from the active treeview
    selected_items = treeview.selection()
    if selected_items:
        for item in selected_items:
            row_data = treeview.item(item, 'values')
            identifier = str(row_data[-4])
            if row_data[-2] == "Pending Review":
                identifiers_pending.add(identifier)
            elif row_data[-2] in {"Generated", "Pending Maker", "Pending Checker"}:
                identifiers_maker_checker.add(identifier)
            elif row_data[-2] == "Approved":
                identifiers_approved.add(identifier)
            elif row_data[-2] == "Concluded":
                identifiers_concluded.add(identifier)

    # Convert sets back to lists if needed
    identifiers_pending = list(identifiers_pending)
    identifiers_maker_checker = list(identifiers_maker_checker)
    identifiers_approved = list(identifiers_approved)
    identifiers_concluded = list(identifiers_concluded)

    #If theres is any item with Pending Review Status
    if identifiers_pending:
        create_approve_status_window(tab_name, columns_deals,  columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, identifiers_pending, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b)
    #If theres is any item with Generated, Pending Maker, Pendin Checker Status
    if identifiers_maker_checker:
        create_maker_checker_window(tab_name, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, identifiers_maker_checker, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b)
        
    #If theres is any item with Pending Review Status 
    if identifiers_approved:
        create_edit_window(tab_name, columns_deals,  columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, identifiers_approved, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b)
    
    if identifiers_concluded:
        # Assuming the first selected item is used for details
        selected_item = selected_items[0]
        row_data = treeview.item(selected_item, 'values')
        deal_name = row_data[0]  # Assuming DealName is the first column
        identifier = row_data[-4]
        if 'termo' in instrument_deals:            
            # Call termo_details_deals with the deal_name and identifier
            termo_details_deals(deal_name, identifier)
        else:
            # Call opcao_details_deals with the deal_name and identifier
            opcao_details_deals(deal_name, identifier)
        
def termo_details_deals(deal_name, identifier):
    global janela
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
   
    # Create a Toplevel window
    toplevel = tk.Toplevel(janela)    
    toplevel.title(f"{deal_name}")  
    toplevel.geometry("1350x500")    

    # Configure grid to expand with window
    toplevel.grid_rowconfigure(0, weight=1)
    toplevel.grid_rowconfigure(1, weight=1)
    toplevel.grid_columnconfigure(0, weight=1)
    toplevel.grid_columnconfigure(1, weight=1)

    # Create a scrollable frame for the left side
    left_frame = tk.Frame(toplevel, bg="#FFFFFF")
    left_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

    left_canvas = tk.Canvas(left_frame, bg="#FFFFFF")
    left_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=left_canvas.yview)
    left_scrollable_frame = tk.Frame(left_canvas, bg="#FFFFFF")

    left_scrollable_frame.bind(
        "<Configure>",
        lambda e: left_canvas.configure(
            scrollregion=left_canvas.bbox("all")
        )
    )

    left_canvas.create_window((0, 0), window=left_scrollable_frame, anchor="nw")
    left_canvas.configure(yscrollcommand=left_scrollbar.set)

    left_canvas.pack(side="left", fill="both", expand=True)
    left_scrollbar.pack(side="right", fill="y")

    # Add labels to the left scrollable frame
    left_labels = [
        "DealName", "TradeDate", "Market", "Type", "Instrument", "Contract", "Strike", "Currency",
        "IntermediateCCY", "TotalNotional", "SettlementDate", "TradingBook", "OtherBook", "ClientValue",
        "SpotFXRate", "FXConvDate", "FixingStartDate", "FixingEndDate", "Counterparty", "Identifier",
        "Index", "Status", "SID"
    ]

    # Fetch data from termo_base_deals
    cursor.execute('SELECT * FROM termo_base_deals WHERE DealName = ?', (deal_name,))
    termo_base_deals_data = cursor.fetchone()

    left_label_widgets = []
    for i, label in enumerate(left_labels):
        label_bg = "#FFFFFF" if label == "SettlementDate" else "#FFE67D"
        value_bg = "#FFFFFF"
        
        tk.Label(left_scrollable_frame, text=label, anchor="w", bg=label_bg, width=15).grid(row=i, column=0, sticky="w", padx=5)
        value_label = tk.Label(left_scrollable_frame, text=termo_base_deals_data[i] if termo_base_deals_data else "Example", anchor="w", bg=value_bg, width=15)
        value_label.grid(row=i, column=1, sticky="w", padx=5)
        left_label_widgets.append(value_label)

        # Add a button next to "SettlementDate"
        if label == "SettlementDate":
            tk.Button(left_scrollable_frame, text="Customize Period", font=("Arial", 8, "bold")).grid(row=i, column=2, padx=2)

    # Create a scrollable frame for the right side
    right_frame = tk.Frame(toplevel, bg="#FFFFFF")
    right_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

    right_canvas = tk.Canvas(right_frame, bg="#FFFFFF")
    right_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=right_canvas.yview)
    right_scrollable_frame = tk.Frame(right_canvas, bg="#FFFFFF")

    right_scrollable_frame.bind(
        "<Configure>",
        lambda e: right_canvas.configure(
            scrollregion=right_canvas.bbox("all")
        )
    )

    right_canvas.create_window((0, 0), window=right_scrollable_frame, anchor="nw")
    right_canvas.configure(yscrollcommand=right_scrollbar.set)

    right_canvas.pack(side="left", fill="both", expand=True)
    right_scrollbar.pack(side="right", fill="y")

    # Add labels to the right scrollable frame
    right_labels = [
        "ID do Sistema", "ID Tipo de Linha", "Código operação", "Meu Número", "Lançamento do Participante (Conta)",
        "Papel (Posição do participante)", "CPF/CNPJ Cliente Parte", "Contraparte", "CPF/CNPJ Cliente Contraparte",
        "Contrato Global", "Classe do Ativo Subjacente", "Fonte Informação", "Moeda de Referência", "Moeda Cotada",
        "Cotação para o Vencimento", "Valor Base / Quantidade", "Código do Ativo Subjacente", "Taxa a Termo (R$/Moeda)",
        "Data de fixing do Ativo Subjacente", "Data de Operação", "Data vencimento", "Boletim", "Tipo de Cotação",
        "Data de Fixing da Moeda", "Cross Rate na Avaliação?", "Fonte de Consulta", "Tela ou Função de Consulta",
        "Praça de Negociação", "Horário de Consulta", "Cotação - Taxa de Câmbio R$/USD", "Cotação - Paridade (Moeda/USD ou USD/ Moeda)",
        "Data de Avaliação", "Código da paridade cross", "Data de fixing da paridade cross", "Termo a Termo", "Data de Fixação",
        "Forma de Atualização", "Valor / Percentual Negociado", "Cotação para fixing", "Atualizar Valor Base?", "Cotação Inicial",
        "Ajustar Taxa", "Responsável pelo Ajuste da Taxa", "Data Inicial para Ajuste da Taxa", "Data Final para Ajuste de taxa",
        "Limites", "Superior (Paridade)", "Inferior (Paridade)", "Data de Liquidação do Prêmio", "Prêmio a ser pago pelo",
        "Valor do Prêmio", "Modalidade de Liquidação", "Prêmio em Moeda Estrangeira", "Data de fixing da moeda do prêmio",
        "Taxa a Termo em Reais", "Observação", "Código Identificador", "Tipo Média Asiático", "Quantidade de Datas de Verificação",
        "Identifier", "Index", "Status", "SID"
    ]

    # Fetch data from base_termo_file
    cursor.execute('SELECT * FROM termo_base_file WHERE Identifier = ? AND Contraparte <> "00041007"', (identifier,))
    base_termo_file_data = cursor.fetchone()

    right_label_widgets = []
    for i, label in enumerate(right_labels):
        tk.Label(right_scrollable_frame, text=label, anchor="w", bg="#E8E8E8", width=30).grid(row=i, column=0, sticky="w", padx=5)
        value_label = tk.Label(right_scrollable_frame, text=base_termo_file_data[i] if base_termo_file_data else "Example", anchor="w", bg="#FFFFFF", width=30)
        value_label.grid(row=i, column=1, sticky="w", padx=5)
        right_label_widgets.append(value_label)

    # Create a scrollable frame for the bottom side
    bottom_frame = tk.Frame(toplevel, width=1350)
    bottom_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5)

    bottom_canvas = tk.Canvas(bottom_frame)
    bottom_scrollbar = ttk.Scrollbar(bottom_frame, orient="horizontal", command=bottom_canvas.xview)
    bottom_canvas.configure(xscrollcommand=bottom_scrollbar.set)

    tree_frame = tk.Frame(bottom_canvas)
    bottom_canvas.create_window((0, 0), window=tree_frame, anchor="nw")
    
    # Configure grid to expand with window
    bottom_frame.grid_rowconfigure(0, weight=1)
    bottom_frame.grid_columnconfigure(0, weight=1)
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)

    # Create a Treeview in the bottom frame
    columns = [
        "Trade Date", "Counterparty", 'AthenaID', 'B3 ID', 'B2B AthenaID', 'B2B B3 ID', 'Instrument', 'Status', 
        'Maker', 'Checker', 'Time_Stamp', 'Confirmation', 'SS_Validation', 'Identifier', 'Index'
    ]

    style = ttk.Style()
    style.configure("Treeview.Heading", font=("Arial", 8, "bold"), background="#E8E8E8", foreground="black", relief="raised")

    tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=font.Font().measure(col) + 20)

    # Fetch data from base_deals where AthenaID matches deal_name
    cursor.execute('SELECT * FROM base_deals WHERE AthenaID = ? OR B2B_AthenaID = ?', (deal_name, deal_name))
    base_deals_data = cursor.fetchall()

    # Insert data into the Treeview
    for data in base_deals_data:
        tree.insert("", "end", values=data)

    tree.pack(expand=True, fill="both")
    bottom_canvas.pack(side="top", fill="both", expand=True)
    bottom_scrollbar.pack(side="bottom", fill="x")

    # Update the scrollregion after packing
    tree_frame.update_idletasks()
    bottom_canvas.configure(scrollregion=bottom_canvas.bbox("all"))
    
    conn.close()
       
def opcao_details_deals(deal_name, identifier):
    global janela
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
   
    # Create a Toplevel window
    toplevel = tk.Toplevel(janela)    
    toplevel.title(f"{deal_name}")   
    toplevel.geometry("1350x500")   

    # Configure grid to expand with window
    toplevel.grid_rowconfigure(0, weight=1)
    toplevel.grid_rowconfigure(1, weight=1)
    toplevel.grid_columnconfigure(0, weight=1)
    toplevel.grid_columnconfigure(1, weight=1)

    # Create a scrollable frame for the left side
    left_frame = tk.Frame(toplevel, bg="#FFFFFF")
    left_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

    left_canvas = tk.Canvas(left_frame, bg="#FFFFFF")
    left_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=left_canvas.yview)
    left_scrollable_frame = tk.Frame(left_canvas, bg="#FFFFFF")

    left_scrollable_frame.bind(
        "<Configure>",
        lambda e: left_canvas.configure(
            scrollregion=left_canvas.bbox("all")
        )
    )

    left_canvas.create_window((0, 0), window=left_scrollable_frame, anchor="nw")
    left_canvas.configure(yscrollcommand=left_scrollbar.set)

    left_canvas.pack(side="left", fill="both", expand=True)
    left_scrollbar.pack(side="right", fill="y")
    # Add labels to the left scrollable frame
    left_labels = [
    "DealName", "TradeDate", "Market", "Type", "Instrument", "Contract", "Strike", "Currency",
    "IntermediateCCY", "TotalNotional", "SettlementDate", "TradingBook", "OtherBook", "ClientValue",
    "SpotFXRate", "FXConvDate", "FixingStartDate", "FixingEndDate","Counterparty", "Premium", 
    "PremiumPerUnit", "PremiumCCY", "SpotDate", "Mnemonico","Identifier", "Index", "Status", "SID"
    ]

    # Fetch data from termo_base_deals
    cursor.execute('SELECT * FROM opcao_base_deals WHERE DealName = ?', (deal_name,))
    termo_base_deals_data = cursor.fetchone()

    left_label_widgets = []
    for i, label in enumerate(left_labels):
        label_bg = "#FFFFFF" if label == "SettlementDate" else "#FFE67D"
        value_bg = "#FFFFFF"
        
        tk.Label(left_scrollable_frame, text=label, anchor="w", bg=label_bg, width=15).grid(row=i, column=0, sticky="w", padx=5)
        value_label = tk.Label(left_scrollable_frame, text=termo_base_deals_data[i] if termo_base_deals_data else "Example", anchor="w", bg=value_bg, width=15)
        value_label.grid(row=i, column=1, sticky="w", padx=5)
        left_label_widgets.append(value_label)

        # Add a button next to "SettlementDate"
        if label == "SettlementDate":
            tk.Button(left_scrollable_frame, text="Customize Period", font=("Arial", 8, "bold")).grid(row=i, column=2, padx=2)

    # Create a scrollable frame for the right side
    right_frame = tk.Frame(toplevel, bg="#FFFFFF")
    right_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

    right_canvas = tk.Canvas(right_frame, bg="#FFFFFF")
    right_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=right_canvas.yview)
    right_scrollable_frame = tk.Frame(right_canvas, bg="#FFFFFF")

    right_scrollable_frame.bind(
        "<Configure>",
        lambda e: right_canvas.configure(
            scrollregion=right_canvas.bbox("all")
        )
    )

    right_canvas.create_window((0, 0), window=right_scrollable_frame, anchor="nw")
    right_canvas.configure(yscrollcommand=right_scrollbar.set)

    right_canvas.pack(side="left", fill="both", expand=True)
    right_scrollbar.pack(side="right", fill="y")

    # Add labels to the right scrollable frame
    right_labels = [
        "ID do Sistema", "ID Tipo de Linha", "Codigo da Operação", "Tipo Contrato", "Código", "Conta Parte", "Conta Contraparte",
        "Papel Parte", "Moeda Base/Índice/Ações", "Data Inicio", "Data de Vencimento", "Preço de Exercício",
        "Casas Decimais Preço Exercicio", "Prêmio Unitário", "Casas Decimais do Prêmio Unitário",
        "Valor Base em Moeda Estrangeira / Quantidade", "Casas Decimais do Valor Base em Moeda Estrangeira ou Quantidade.",
        "Cotação / Índice Limite", "Casas Decimais da Cotação / Índice Limite", "Tipo de Exercício", "Banco Liquidante",
        "Modalidade", "Adesão a Contrato", "Meu Número", "Conta Intermediador", "Comissão paga pelo Titular",
        "Casas Decimais da Comissão paga pelo Titular", "Comissão paga pelo Lançador",
        "Casas Decimais da Comissão paga pelo Lançador", "Cross-Rate na Avaliação", "Fonte de Informação",
        "Cotação para o Vencimento", "Boletim", "Horário do Boletim", "Fonte de Consulta", "Outra Fonte de Consulta",
        "Tela ou Função de Consulta", "Praça de Negociação", "Horário de Consulta", "Cotação – Taxa de Câmbio",
        "Cotação – Paridade", "Data de Avaliação", "CPF / CNPJ da Parte", "CPF / CNPJ da Contraparte", "Moeda Cotada",
        "Barreiras", "Trigger In", "Casas Decimais do Trigger In", "Trigger Out", "Casas Decimais do Trigger Out",
        "Cesta de Garantias - Lançador", "Forma de Verificação", "Rebate", "Valor do Rebate",
        "Casas decimais do Valor do Rebate", "Liquidação do Rebate", "Código da Ação / Indice Internacional",
        "Ajuste de Proventos pelas", "Proteção contra Provento em Dinheiro", "Trigger – Proporção",
        "Trigger – Forma de Disparo", "Trigger – Tipo de Disparo", "Preço de Exercício em Reais", "Opção Quanto",
        "Cotação para Opção Quanto", "Casas decimais do Cotação para Opção Quanto", "Data de Liquidação do Prêmio",
        "Mercadoria", "Cotação para Moeda", "Observação", "Média para Opção Asiática", "Data de Verificação",
        "Valor/Quantidade de Referência.", "Casas Decimais do Valor/Quantidade de Referência.", "Data de Verificação.",
        "Valor/Quantidade de Referência", "Casas Decimais do Valor/Quantidade de Referência","Identifier", "Index", "Status", "SID"
    ]

    # Fetch data from base_termo_file
    cursor.execute('SELECT * FROM opcao_base_file WHERE Identifier = ? AND Contraparte <> "00041007"', (identifier,))
    base_termo_file_data = cursor.fetchone()

    right_label_widgets = []
    for i, label in enumerate(right_labels):
        tk.Label(right_scrollable_frame, text=label, anchor="w", bg="#E8E8E8", width=30).grid(row=i, column=0, sticky="w", padx=5)
        value_label = tk.Label(right_scrollable_frame, text=base_termo_file_data[i] if base_termo_file_data else "Example", anchor="w", bg="#FFFFFF", width=30)
        value_label.grid(row=i, column=1, sticky="w", padx=5)
        right_label_widgets.append(value_label)

    # Create a scrollable frame for the bottom side
    bottom_frame = tk.Frame(toplevel, width=1350)
    bottom_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=5)

    bottom_canvas = tk.Canvas(bottom_frame)
    bottom_scrollbar = ttk.Scrollbar(bottom_frame, orient="horizontal", command=bottom_canvas.xview)
    bottom_canvas.configure(xscrollcommand=bottom_scrollbar.set)

    tree_frame = tk.Frame(bottom_canvas)
    bottom_canvas.create_window((0, 0), window=tree_frame, anchor="nw")
    
    # Configure grid to expand with window
    bottom_frame.grid_rowconfigure(0, weight=1)
    bottom_frame.grid_columnconfigure(0, weight=1)
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)

    # Create a Treeview in the bottom frame
    columns = [
        "Trade Date", "Counterparty", 'AthenaID', 'B3 ID', 'B2B AthenaID', 'B2B B3 ID', 'Instrument', 'Status', 
        'Maker', 'Checker', 'Time_Stamp', 'Confirmation', 'SS_Validation', 'Identifier', 'Index'
    ]

    style = ttk.Style()
    style.configure("Treeview.Heading", font=("Arial", 8, "bold"), background="#E8E8E8", foreground="black", relief="raised")

    tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=font.Font().measure(col) + 20)

    # Fetch data from base_deals where AthenaID matches deal_name
    cursor.execute('SELECT * FROM base_deals WHERE AthenaID = ? OR B2B_AthenaID = ?', (deal_name, deal_name))
    base_deals_data = cursor.fetchall()

    # Insert data into the Treeview
    for data in base_deals_data:
        tree.insert("", "end", values=data)

    tree.pack(expand=True, fill="both")
    bottom_canvas.pack(side="top", fill="both", expand=True)
    bottom_scrollbar.pack(side="bottom", fill="x")

    # Update the scrollregion after packing
    tree_frame.update_idletasks()
    bottom_canvas.configure(scrollregion=bottom_canvas.bbox("all"))  
    conn.close()    

def extrair_dados_deals(treeview_deals, treeview_deals_b2b):
    deals_cliente	 = []
    td_cliente	 = []
    market_client	 = []
    type_cliente	 = []
    instrument_cliente	 = []
    strike_cliente	 = []
    intCCY_cliente	 = []
    tn_cliente	 = []
    sd_cliente	 = []
    fxd_cliente	 = []
    fsd_cliente	 = []
    fed_cliente	 = []
    accronym	 = []
    identifier_cliente	 = []
    status_cliente	 = []
    sid_cliente	 = []

    
    for item in treeview_deals.get_children():
        values = treeview_deals.item(item, 'values')
        deals_cliente.append(values[0])
        td_cliente.append(values[1])
        market_client.append(values[2])
        type_cliente.append(values[3])
        instrument_cliente.append(values[4])
        strike_cliente.append(values[6])
        intCCY_cliente.append(values[8])
        tn_cliente.append(values[9])
        sd_cliente.append(values[10])
        fxd_cliente.append(values[15])
        fsd_cliente.append(values[16])
        fed_cliente.append(values[17])
        accronym.append(values[18])
        identifier_cliente.append(values[-4])
        status_cliente.append(values[-2])
        sid_cliente.append(values[-1])

        
    deals_b2b = []
    td_b2b = []
    market_client = []
    type_b2b = []
    instrument_b2b = []
    strike_b2b = []
    intCCY_b2b = []
    tn_b2b = []
    sd_b2b = []
    fxd_b2b	= []
    fsd_b2b	= []
    fed_b2b	= []
    accronym_b2b = []
    identifier_b2b = []
    status_b2b = []
    sid_b2b = []

   
    
    for item in treeview_deals_b2b.get_children():
        values = treeview_deals_b2b.item(item, 'values')
        deals_b2b.append(values[0])
        td_b2b.append(values[1])
        market_client.append(values[2])
        type_b2b.append(values[3])
        instrument_b2b.append(values[4])
        strike_b2b.append(values[6])
        intCCY_b2b.append(values[8])
        tn_b2b.append(values[9])
        sd_b2b.append(values[10])
        fxd_b2b.append(values[15])
        fsd_b2b.append(values[16])
        fed_b2b.append(values[17])
        accronym_b2b.append(values[18])
        identifier_b2b.append(values[-4])
        status_b2b.append(values[-2])
        sid_b2b.append(values[-1])

        
    return deals_cliente, td_cliente, market_client, type_cliente, instrument_cliente, strike_cliente, intCCY_cliente, tn_cliente, sd_cliente, fxd_cliente, fsd_cliente, fed_cliente, accronym, identifier_cliente, status_cliente, sid_cliente, deals_b2b, td_b2b, market_client, type_b2b, instrument_b2b, strike_b2b, intCCY_b2b, tn_b2b, sd_b2b, fxd_b2b, fsd_b2b, fed_b2b, accronym_b2b, identifier_b2b, status_b2b, sid_b2b



def lookup_approve(identifier, identifier_list, value_list):
    try:
        index = identifier_list.index(identifier)
        return value_list[index]
    except ValueError:
        # Handle the case where the identifier is not found
        return None  # or some default value
    
def update_base_deals(conn, row_data, table_name, columns, treeview_name):
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    try:
        cursor = conn.cursor()
        #cursor.execute("BEGIN TRANSACTION;")
        
        # Prepare the SQL update statement with double quotes around column names
        set_clause = ", ".join([f'"{col}" = ?' for col in columns[:-2]])  # Exclude Identifier and Index from SET clause
        sql_update = f"""
        UPDATE "{table_name}"
        SET {set_clause}
        WHERE "Identifier" = ? AND "Index" = ?
        """     
        
        # Map treeview data to base_deals columns
        if treeview_name == "treeview_deals":
            mapped_data = [
                row_data[0], row_data[0], row_data[0], row_data[0], row_data[1], row_data[4], row_data[21],
                "", "", datetime.now() if row_data[21] == "Concluded" else "", "", "", row_data[19], row_data[20]
            ]
        elif treeview_name == "treeview_deals_b2b":
            mapped_data = [
                row_data[0], row_data[0], row_data[0], row_data[0], row_data[1], row_data[4], row_data[21],
                "", "", datetime.now() if row_data[21] == "Concluded" else "", "", "", row_data[19], row_data[20]
            ]
        
        # Execute the update statement
        cursor.execute(sql_update, (*mapped_data[:-2], mapped_data[-2], mapped_data[-1]))  # Exclude Identifier and Index from mapped_data for SET clause
        
        conn.commit()
        conn.close()
    except Exception as e:
        conn.close()
        messagebox.showerror("Database Error", f"An error occurred: {e}")

   
def create_approve_status_window(tab_name, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, identifiers_pending, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b):
    deals_cliente, td_cliente, market_client, type_cliente, instrument_cliente, strike_cliente, intCCY_cliente, tn_cliente, sd_cliente, fxd_cliente, fsd_cliente, fed_cliente, accronym, identifier_cliente, status_cliente, sid_cliente, deals_b2b, td_b2b, market_client, type_b2b, instrument_b2b, strike_b2b, intCCY_b2b, tn_b2b, sd_b2b, fxd_b2b, fsd_b2b, fed_b2b, accronym_b2b, identifier_b2b, status_b2b, sid_b2b = extrair_dados_deals(treeview_deals, treeview_deals_b2b)
    global janela
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")
    review_window = ctk.CTkToplevel(janela)
    review_window.title("Review Queue")
    review_window.geometry("1300x600")
    review_window.lift()
    review_window.focus_set()
    review_window.grab_set()

    # Font for Buttons and Labels
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")
    fonte_label = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    frame_export = ctk.CTkScrollableFrame(review_window)
    frame_export.pack(fill='both', expand=True)

    # Create frames for each section
    frame_checkboxes = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_checkboxes.pack(side='left', fill='both', expand=True, padx=10, pady=8)
    
    label = ctk.CTkLabel(frame_checkboxes, text="Select", font=fonte_label)
    label.pack(pady=5, padx=5)

    frame_labels_cliente = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_cliente.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_cliente = ctk.CTkLabel(frame_labels_cliente, text="Client", font=fonte_label)
    label_cliente.pack(pady=5, padx=5)

    frame_labels_b2b = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_b2b.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_b2b = ctk.CTkLabel(frame_labels_b2b, text="B2B", font=fonte_label)
    label_b2b.pack(pady=5, padx=5)
    
    frame_labels_status = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_status.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_status = ctk.CTkLabel(frame_labels_status, text="Status", font=fonte_label)
    label_status.pack(pady=5, padx=5)

    # Dictionary to store checkbox states
    checkbox_states = {}

    # Create checkboxes and labels based on identifiers
    for idx, identifier in enumerate(identifiers_pending):
        accronym_value = lookup_approve(identifier, identifier_cliente, accronym)
        deal_cliente_value = lookup_approve(identifier, identifier_cliente, deals_cliente)
        deal_b2b_value = lookup_approve(identifier, identifier_b2b, deals_b2b)
        status_cliente_value = lookup_approve(identifier, identifier_cliente, status_cliente)
        status_b2b_value = lookup_approve(identifier, identifier_cliente, status_b2b)
        
        var = ctk.BooleanVar(value=True)
        checkbox_states[identifier] = var  # Store the state of the checkbox
        checkbox = ctk.CTkCheckBox(frame_checkboxes, text=f"{accronym_value if accronym_value else 'Unknown'}", checkbox_height=18, checkbox_width=18, height=20, variable=var)
        checkbox.pack(pady=5, padx=5)

        label_cliente = ctk.CTkLabel(frame_labels_cliente, text=f"{deal_cliente_value if deal_cliente_value else 'Unknown'}", anchor='center', height=20)
        label_cliente.pack(pady=5, padx=5)

        label_b2b = ctk.CTkLabel(frame_labels_b2b, text=f"{deal_b2b_value if deal_b2b_value else 'Unknown'}", anchor='center', height=20)
        label_b2b.pack(pady=5, padx=5)

        label_status = ctk.CTkLabel(frame_labels_status, text=f"{status_cliente_value if status_cliente_value == status_b2b_value else 'Unknown'}", anchor='center', height=20)
        label_status.pack(pady=5, padx=5)

    # Create a frame for the buttons
    frame_buttons = ctk.CTkFrame(review_window)
    frame_buttons.pack(fill='x', pady=5)

    # Create the APPROVE button
    button_approve = ctk.CTkButton(frame_buttons, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="APPROVE", font=fonte_botao, command=lambda: approve_action(review_window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_fixings, treeview_fixings_b2b, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b))
    button_approve.pack(side='left', pady=5, padx=5)

    # Create the REJECT button
    button_reject = ctk.CTkButton(frame_buttons, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="REJECT", font=fonte_botao, command=lambda: reject_action(review_window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_fixings, treeview_fixings_b2b, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b))
    button_reject.pack(side='left', pady=5, padx=2)
    
def approve_action(window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_fixings, treeview_fixings_b2b, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b):
    # Filter identifiers based on checkbox state
    identifiers_to_approve = [identifier for identifier, var in checkbox_states.items() if var.get()]     
    SID = getpass.getuser()
    SID = SID[0].upper() + SID[1:]
    
    # Lists to accumulate row data for batch update
    rows_to_update_deals = []
    rows_to_update_deals_b2b = []
    rows_to_update_fixings = []
    rows_to_update_fixings_b2b = []
    rows_to_update_file = []
    rows_to_update_file_b2b = []

    # Lists to accumulate data for database update
    DealNames = []
    Markets = []
    Types = []
    Strikes = []
    IntermediatesCCY = []
    TotalNotionals = []
    SettlementDates = []
    FxConvDates = []
    FirstFixingDates = []
    LastFixingDates = []
    SIDS = []
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    TradeDates_update = []
    Counterparties = []
    Counterparties_update = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []
    Statuses_update = []
    Status_update = "Approved"

    # Variable to track if an error has occurred
    error_occurred = False

    # Update the Client Treeview Deals   
    for item in treeview_deals.get_children():        
        row_data_deals = list(treeview_deals.item(item, 'values'))
        identifier_approve = str(row_data_deals[-4])
        if row_data_deals[-2] == "Pending Review" and identifier_approve in identifiers_to_approve and row_data_deals[-1] != SID:
            Instruments.append(row_data_deals[4])
            TradeDates_update.append(row_data_deals[1])
            TradeDates.append(row_data_deals[1])
            Counterparties.append(row_data_deals[18])
            Counterparties_update.append(row_data_deals[18])
            AthenaIDs.append(row_data_deals[0])      
            DealNames.append(row_data_deals[0])      
            Identifiers.append(row_data_deals[-4])
            Indexes.append(row_data_deals[-3])
            Markets.append(row_data_deals[2])
            Types.append(row_data_deals[3])
            Strikes.append(row_data_deals[6])
            IntermediatesCCY.append(row_data_deals[8])
            TotalNotionals.append(row_data_deals[9])
            SettlementDates.append(row_data_deals[10])
            FxConvDates.append(row_data_deals[15])
            FirstFixingDates.append(row_data_deals[16])
            LastFixingDates.append(row_data_deals[17])            
            Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))     
            row_data_deals[-2] = "Approved"
            row_data_deals[-1] = SID
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
            SIDS.append(row_data_deals[-1])
            Statuses.append(row_data_deals[-2])
            Statuses_update.append(row_data_deals[-2])
            rows_to_update_deals.append(row_data_deals)  # Acumula para atualização em lote
        elif row_data_deals[-2] == "Pending Review" and row_data_deals[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview Deals
    for item in treeview_deals_b2b.get_children():
        row_data_deals_b2b = list(treeview_deals_b2b.item(item, 'values'))
        identifier_approve = str(row_data_deals_b2b[-4])            
        if row_data_deals_b2b[-2] == "Pending Review" and identifier_approve in identifiers_to_approve and row_data_deals_b2b[-1] != SID:
            B2B_AthenaIDs.append(row_data_deals_b2b[0])
            Instruments.append(row_data_deals_b2b[4])
            TradeDates_update.append(row_data_deals_b2b[1])
            Counterparties_update.append(row_data_deals_b2b[18])
            AthenaIDs.append(row_data_deals_b2b[0])      
            DealNames.append(row_data_deals_b2b[0])                  
            Indexes.append(row_data_deals_b2b[-3])
            Markets.append(row_data_deals_b2b[2])
            Types.append(row_data_deals_b2b[3])
            Strikes.append(row_data_deals_b2b[6])
            IntermediatesCCY.append(row_data_deals_b2b[8])
            TotalNotionals.append(row_data_deals_b2b[9])
            SettlementDates.append(row_data_deals_b2b[10])
            FxConvDates.append(row_data_deals_b2b[15])
            FirstFixingDates.append(row_data_deals_b2b[16])
            LastFixingDates.append(row_data_deals_b2b[17])            
            row_data_deals_b2b[-2] = "Approved"
            row_data_deals_b2b[-1] = SID
            treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
            SIDS.append(row_data_deals_b2b[-1])
            rows_to_update_deals_b2b.append(row_data_deals_b2b)  # Acumula para atualização em lote
            Statuses_update.append(row_data_deals_b2b[-2])
        elif row_data_deals_b2b[-2] == "Pending Review" and row_data_deals_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the Client Treeview Fixings
    for item in treeview_fixings.get_children():
        row_data_fixings = list(treeview_fixings.item(item, 'values'))
        identifier_approve = str(row_data_fixings[-4])
        if row_data_fixings[-2] == "Pending Review" and identifier_approve in identifiers_to_approve and row_data_fixings[-1] != SID:
            row_data_fixings[-2] = "Approved"
            row_data_fixings[-1] = SID
            treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
            rows_to_update_fixings.append(row_data_fixings)  # Acumula para atualização em lote
        elif row_data_fixings[-2] == "Pending Review" and row_data_fixings[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break
        
    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview Fixings
    for item in treeview_fixings_b2b.get_children():
        row_data_fixings_b2b = list(treeview_fixings_b2b.item(item, 'values'))
        identifier_approve = str(row_data_fixings_b2b[-4])
        if row_data_fixings_b2b[-2] == "Pending Review" and identifier_approve in identifiers_to_approve and row_data_fixings_b2b[-1] != SID:
            row_data_fixings_b2b[-2] = "Approved"
            row_data_fixings_b2b[-1] = SID
            treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview
            rows_to_update_fixings_b2b.append(row_data_fixings_b2b)  # Acumula para atualização em lote
        elif row_data_fixings_b2b[-2] == "Pending Review" and row_data_fixings_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the Client Treeview File       
    for item in treeview_file.get_children():
        row_data_file = list(treeview_file.item(item, 'values'))
        identifier_approve = str(row_data_file[-4])            
        if row_data_file[-2] == "Pending Review" and identifier_approve in identifiers_to_approve and row_data_file[-1] != SID:
            row_data_file[-2] = "Approved"
            row_data_file[-1] = SID
            treeview_file.item(item, values=row_data_file)  # Atualiza o Treeview
            rows_to_update_file.append(row_data_file)  # Acumula para atualização em lote
        elif row_data_file[-2] == "Pending Review" and row_data_file[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview File
    for item in treeview_file_b2b.get_children():
        row_data_file_b2b = list(treeview_file_b2b.item(item, 'values'))
        identifier_approve = str(row_data_file_b2b[-4])            
        if row_data_file_b2b[-2] == "Pending Review" and identifier_approve in identifiers_to_approve and row_data_file_b2b[-1] != SID:
            row_data_file_b2b[-2] = "Approved"
            row_data_file_b2b[-1] = SID
            treeview_file_b2b.item(item, values=row_data_file_b2b)  # Atualiza o Treeview
            rows_to_update_file_b2b.append(row_data_file_b2b)  # Acumula para atualização em lote
        elif row_data_file_b2b[-2] == "Pending Review" and row_data_file_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Prepare updates list
    data_and_tables = [
        (rows_to_update_deals, instrument_deals),
        (rows_to_update_deals_b2b, instrument_deals),
        (rows_to_update_fixings, instrument_fixings),
        (rows_to_update_fixings_b2b, instrument_fixings),
        (rows_to_update_file, instrument_file),
        (rows_to_update_file_b2b, instrument_file)
    ]

    insert_or_update_all_tables(data_and_tables)
      
   # Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += [""] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
    ajustar_largura_colunas(tabela_arquivoopcao_cliente, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivoopcao_b2b, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_cliente, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_b2b, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)       
    highlight_duplicates(tabela_opcao_cliente, 'deals')
    highlight_duplicates(tabela_opcao_b2b, 'deals')
    highlight_duplicates(tabela_termo_cliente, 'deals')
    highlight_duplicates(tabela_termo_b2b, 'deals')
    highlight_duplicates(tabela_arquivoopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivoopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo')     
    # Close the review window
    window.destroy()
    status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update)
    
def reject_action(window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_fixings, treeview_fixings_b2b, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b):
    # Get database connection    
    
    # Filter identifiers based on checkbox state
    identifiers_to_reject = [identifier for identifier, var in checkbox_states.items() if var.get()]
    SID = getpass.getuser()
    SID = SID[0].upper() + SID[1:]

    # Lists to accumulate row data for batch update
    rows_to_update_deals = []
    rows_to_update_deals_b2b = []
    rows_to_update_fixings = []
    rows_to_update_fixings_b2b = []
    rows_to_update_file = []
    rows_to_update_file_b2b = []

    # Lists to accumulate data for database update
    DealNames = []
    Markets = []
    Types = []
    Strikes = []
    IntermediatesCCY = []
    TotalNotionals = []
    SettlementDates = []
    FxConvDates = []
    FirstFixingDates = []
    LastFixingDates = []
    SIDS = []
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    TradeDates_update = []
    Counterparties = []
    Counterparties_update = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []
    Statuses_update = []
    Status_update = "New"

    # Variable to track if an error has occurred
    error_occurred = False

    # Update the Client Treeview Deals   
    for item in treeview_deals.get_children():        
        row_data_deals = list(treeview_deals.item(item, 'values'))
        identifier_reject = str(row_data_deals[-4])
        if row_data_deals[-2] == "Pending Review" and identifier_reject in identifiers_to_reject and row_data_deals[-1] != SID:
            Instruments.append(row_data_deals[4])
            TradeDates_update.append(row_data_deals[1])
            TradeDates.append(row_data_deals[1])
            Counterparties.append(row_data_deals[18])
            Counterparties_update.append(row_data_deals[18])
            AthenaIDs.append(row_data_deals[0])      
            DealNames.append(row_data_deals[0])      
            Identifiers.append(row_data_deals[-4])
            Indexes.append(row_data_deals[-3])
            Markets.append(row_data_deals[2])
            Types.append(row_data_deals[3])
            Strikes.append(row_data_deals[6])
            IntermediatesCCY.append(row_data_deals[8])
            TotalNotionals.append(row_data_deals[9])
            SettlementDates.append(row_data_deals[10])
            FxConvDates.append(row_data_deals[15])
            FirstFixingDates.append(row_data_deals[16])
            LastFixingDates.append(row_data_deals[17])            
            Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))     
            row_data_deals[-2] = "New"
            row_data_deals[-1] = SID
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
            SIDS.append(row_data_deals[-1])
            Statuses.append(row_data_deals[-2])
            Statuses_update.append(row_data_deals[-2])
            rows_to_update_deals.append(row_data_deals)  # Acumula para atualização em lote
        elif row_data_deals[-2] == "Pending Review" and row_data_deals[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview Deals
    for item in treeview_deals_b2b.get_children():
        row_data_deals_b2b = list(treeview_deals_b2b.item(item, 'values'))
        identifier_reject = str(row_data_deals_b2b[-4])            
        if row_data_deals_b2b[-2] == "Pending Review" and identifier_reject in identifiers_to_reject and row_data_deals_b2b[-1] != SID:
            B2B_AthenaIDs.append(row_data_deals_b2b[0])
            Instruments.append(row_data_deals_b2b[4])
            TradeDates_update.append(row_data_deals_b2b[1])
            Counterparties_update.append(row_data_deals_b2b[18])
            AthenaIDs.append(row_data_deals_b2b[0])      
            DealNames.append(row_data_deals_b2b[0])                  
            Indexes.append(row_data_deals_b2b[-3])
            Markets.append(row_data_deals_b2b[2])
            Types.append(row_data_deals_b2b[3])
            Strikes.append(row_data_deals_b2b[6])
            IntermediatesCCY.append(row_data_deals_b2b[8])
            TotalNotionals.append(row_data_deals_b2b[9])
            SettlementDates.append(row_data_deals_b2b[10])
            FxConvDates.append(row_data_deals_b2b[15])
            FirstFixingDates.append(row_data_deals_b2b[16])
            LastFixingDates.append(row_data_deals_b2b[17])            
            row_data_deals_b2b[-2] = "New"
            row_data_deals_b2b[-1] = SID            
            treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
            SIDS.append(row_data_deals_b2b[-1])
            rows_to_update_deals_b2b.append(row_data_deals_b2b)  # Acumula para atualização em lote
            Statuses_update.append(row_data_deals_b2b[-2])
        elif row_data_deals_b2b[-2] == "Pending Review" and row_data_deals_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the Client Treeview Fixings
    for item in treeview_fixings.get_children():
        row_data_fixings = list(treeview_fixings.item(item, 'values'))
        identifier_reject = str(row_data_fixings[-4])
        if row_data_fixings[-2] == "Pending Review" and identifier_reject in identifiers_to_reject and row_data_fixings[-1] != SID:
            row_data_fixings[-2] = "New"
            row_data_fixings[-1] = SID
            treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
            rows_to_update_fixings.append(row_data_fixings)  # Acumula para atualização em lote
        elif row_data_fixings[-2] == "Pending Review" and row_data_fixings[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return
        
    # Update the B2b Treeview Fixings
    for item in treeview_fixings_b2b.get_children():
        row_data_fixings_b2b = list(treeview_fixings_b2b.item(item, 'values'))
        identifier_reject = str(row_data_fixings_b2b[-4])
        if row_data_fixings_b2b[-2] == "Pending Review" and identifier_reject in identifiers_to_reject and row_data_fixings_b2b[-1] != SID:
            row_data_fixings_b2b[-2] = "New"
            row_data_fixings_b2b[-1] = SID
            treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview
            rows_to_update_fixings_b2b.append(row_data_fixings_b2b)  # Acumula para atualização em lote
        elif row_data_fixings_b2b[-2] == "Pending Review" and row_data_fixings_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return
        
    # Update the Client Treeview File       
    for item in treeview_file.get_children():
        row_data_file = list(treeview_file.item(item, 'values'))
        identifier_reject = str(row_data_file[-4])            
        if row_data_file[-2] == "Pending Review" and identifier_reject in identifiers_to_reject and row_data_file[-1] != SID:           
            rows_to_update_file.append(row_data_file)  # Acumula para atualização em lote
            treeview_file.delete(item)            
        elif row_data_file[-2] == "Pending Review" and row_data_file[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview File
    for item in treeview_file_b2b.get_children():
        row_data_file_b2b = list(treeview_file_b2b.item(item, 'values'))
        identifier_reject = str(row_data_file_b2b[-4])            
        if row_data_file_b2b[-2] == "Pending Review" and identifier_reject in identifiers_to_reject and row_data_file_b2b[-1] != SID:           
            rows_to_update_file_b2b.append(row_data_file_b2b)  # Acumula para atualização em lote
            treeview_file_b2b.delete(item)
        elif row_data_file_b2b[-2] == "Pending Review" and row_data_file_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Prepare updates list
    data_and_tables = [
        (rows_to_update_deals, instrument_deals),
        (rows_to_update_deals_b2b, instrument_deals),
        (rows_to_update_fixings, instrument_fixings),
        (rows_to_update_fixings_b2b, instrument_fixings),
        (rows_to_update_file, instrument_file),
        (rows_to_update_file_b2b, instrument_file)
    ]

    insert_or_update_all_tables(data_and_tables)
    
    # Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += [""] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
    
    
    ajustar_largura_colunas(tabela_arquivoopcao_cliente, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivoopcao_b2b, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_cliente, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_b2b, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)       
    highlight_duplicates(tabela_opcao_cliente, 'deals')
    highlight_duplicates(tabela_opcao_b2b, 'deals')
    highlight_duplicates(tabela_termo_cliente, 'deals')
    highlight_duplicates(tabela_termo_b2b, 'deals')
    highlight_duplicates(tabela_arquivoopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivoopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo')   
    
    # Close the review window
    window.destroy()
    
    # Send status change email
    status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update)

def create_maker_checker_window(tab_name, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, identifiers_maker_checker, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b):
    global janela
    deals_cliente, td_cliente, market_client, type_cliente, instrument_cliente, strike_cliente, intCCY_cliente, tn_cliente, sd_cliente, fxd_cliente, fsd_cliente, fed_cliente, accronym, identifier_cliente, status_cliente, sid_cliente, deals_b2b, td_b2b, market_client, type_b2b, instrument_b2b, strike_b2b, intCCY_b2b, tn_b2b, sd_b2b, fxd_b2b, fsd_b2b, fed_b2b, accronym_b2b, identifier_b2b, status_b2b, sid_b2b = extrair_dados_deals(treeview_deals, treeview_deals_b2b)
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")
    review_window = ctk.CTkToplevel(janela)
    review_window.geometry("1300x600")
    review_window.title("Maker and Checker Queue")
    review_window.lift()
    review_window.focus_set()
    review_window.grab_set()

    # Font for Buttons and Labels
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")
    fonte_label = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    frame_export = ctk.CTkScrollableFrame(review_window)
    frame_export.pack(fill='both', expand=True)

    # Create frames for each section
    frame_checkboxes = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_checkboxes.pack(side='left', fill='both', expand=True, padx=10, pady=8)
    
    label = ctk.CTkLabel(frame_checkboxes, text="Select", font=fonte_label)
    label.pack(pady=5, padx=5)

    frame_labels_cliente = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_cliente.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_cliente = ctk.CTkLabel(frame_labels_cliente, text="Client", font=fonte_label)
    label_cliente.pack(pady=5, padx=5)

    frame_labels_b2b = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_b2b.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_b2b = ctk.CTkLabel(frame_labels_b2b, text="B2B", font=fonte_label)
    label_b2b.pack(pady=5, padx=5)
    
    frame_labels_status = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_status.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_status = ctk.CTkLabel(frame_labels_status, text="Status", font=fonte_label)
    label_status.pack(pady=5, padx=5)

    # Dictionary to store checkbox states
    checkbox_states = {}

    # Create checkboxes and labels based on identifiers
    for idx, identifier in enumerate(identifiers_maker_checker):
        accronym_value = lookup_approve(identifier, identifier_cliente, accronym)
        deal_cliente_value = lookup_approve(identifier, identifier_cliente, deals_cliente)
        deal_b2b_value = lookup_approve(identifier, identifier_b2b, deals_b2b)
        status_cliente_value = lookup_approve(identifier, identifier_cliente, status_cliente)
        status_b2b_value = lookup_approve(identifier, identifier_cliente, status_b2b)
        
        var = ctk.BooleanVar(value=True)
        checkbox_states[identifier] = var  # Store the state of the checkbox
        checkbox = ctk.CTkCheckBox(frame_checkboxes, text=f"{accronym_value if accronym_value else 'Unknown'}", checkbox_height=18, checkbox_width=18, height=20, variable=var)
        checkbox.pack(pady=5, padx=5)

        label_cliente = ctk.CTkLabel(frame_labels_cliente, text=f"{deal_cliente_value if deal_cliente_value else 'Unknown'}", anchor='center', height=20)
        label_cliente.pack(pady=5, padx=5)

        label_b2b = ctk.CTkLabel(frame_labels_b2b, text=f"{deal_b2b_value if deal_b2b_value else 'Unknown'}", anchor='center', height=20)
        label_b2b.pack(pady=5, padx=5)

        label_status = ctk.CTkLabel(frame_labels_status, text=f"{status_cliente_value if status_cliente_value == status_b2b_value else 'Unknown'}", anchor='center', height=20)
        label_status.pack(pady=5, padx=5)

    # Create a frame for the buttons
    frame_buttons = ctk.CTkFrame(review_window)
    frame_buttons.pack(fill='x', pady=5)

    # Create the MAKER button
    button_maker = ctk.CTkButton(frame_buttons, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="MAKER", font=fonte_botao, command=lambda: maker_action(review_window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b))
    button_maker.pack(side='left', pady=5, padx=5)

    # Create the CHECKER button
    button_checker = ctk.CTkButton(frame_buttons, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CHECKER", font=fonte_botao, command=lambda: checker_action(review_window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b))
    button_checker.pack(side='left', pady=5, padx=2)
    
def update_database_batch(updates):
    """
    Atualiza o banco de dados em lote para múltiplas tabelas.

    :param conn: Conexão com o banco de dados.
    :param updates: Lista de tuplas, onde cada tupla contém (rows_data, table_name).
    """
    try:
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")

        for rows_data, table_name in updates:
            sql_update = f"""
            UPDATE "{table_name}"
            SET "Status" = ?
            WHERE "Identifier" = ? 
            """
            
            # Execute the update statement for each row
            for row_data in rows_data:
                # row_data should be a tuple like (new_status, identifier)
                cursor.executemany(sql_update, rows_data)
                

        conn.commit()
        conn.close()
    except Exception as e:
        conn.close()
        messagebox.showerror("Database Error", f"An error occurred: {e}")
        
def maker_action(window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b):
    # Get database connection
    
    # Filter identifiers based on checkbox state
    identifiers_to_make = [identifier for identifier, var in checkbox_states.items() if var.get()]
    SID = getpass.getuser()
    SID = SID[0].upper() + SID[1:]
    
    # Lists to accumulate row data for batch update
    rows_to_update_deals = []
    rows_to_update_deals_b2b = []
    rows_to_update_fixings = []
    rows_to_update_fixings_b2b = []
    rows_to_update_file = []
    rows_to_update_file_b2b = []

    # Lists to accumulate data for database update
    DealNames = []
    Markets = []
    Types = []
    Strikes = []
    IntermediatesCCY = []
    TotalNotionals = []
    SettlementDates = []
    FxConvDates = []
    FirstFixingDates = []
    LastFixingDates = []
    SIDS = []
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    TradeDates_update = []
    Counterparties = []
    Counterparties_update = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []
    Statuses_update = []
    Status_update = ""

    # Variable to track if an error has occurred
    error_occurred = False

    # Update the Client Treeview Deals
    for item in treeview_deals.get_children():
        row_data_deals = list(treeview_deals.item(item, 'values'))
        identifier_maker = str(row_data_deals[-4])
        if row_data_deals[-2] == "Pending Maker" and row_data_deals[-1] != SID and identifier_maker in identifiers_to_make:
            Checkers.append(row_data_deals[-1])
            Makers.append(SID)
            Instruments.append(row_data_deals[4])
            TradeDates_update.append(row_data_deals[1])
            TradeDates.append(row_data_deals[1])
            Counterparties.append(row_data_deals[18])
            Counterparties_update.append(row_data_deals[18])
            AthenaIDs.append(row_data_deals[0])      
            DealNames.append(row_data_deals[0])      
            Identifiers.append(row_data_deals[-4])
            Indexes.append(row_data_deals[-3])
            Markets.append(row_data_deals[2])
            Types.append(row_data_deals[3])
            Strikes.append(row_data_deals[6])
            IntermediatesCCY.append(row_data_deals[8])
            TotalNotionals.append(row_data_deals[9])
            SettlementDates.append(row_data_deals[10])
            FxConvDates.append(row_data_deals[15])
            FirstFixingDates.append(row_data_deals[16])
            LastFixingDates.append(row_data_deals[17])            
            Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))     
            row_data_deals[-2] = "Concluded"
            Status_update = "Concluded"
            row_data_deals[-1] = SID
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
            SIDS.append(row_data_deals[-1])
            Statuses.append(row_data_deals[-2])
            Statuses_update.append(row_data_deals[-2])
            rows_to_update_deals.append(row_data_deals)  # Acumula para atualização em lote
        elif row_data_deals[-2] == "Generated" and identifier_maker in identifiers_to_make:            
            Makers.append(SID)
            Instruments.append(row_data_deals[4])
            TradeDates_update.append(row_data_deals[1])
            TradeDates.append(row_data_deals[1])
            Counterparties.append(row_data_deals[18])
            Counterparties_update.append(row_data_deals[18])
            AthenaIDs.append(row_data_deals[0])          
            DealNames.append(row_data_deals[0])      
            Identifiers.append(row_data_deals[-4])
            Indexes.append(row_data_deals[-3])
            Markets.append(row_data_deals[2])
            Types.append(row_data_deals[3])
            Strikes.append(row_data_deals[6])
            IntermediatesCCY.append(row_data_deals[8])
            TotalNotionals.append(row_data_deals[9])
            SettlementDates.append(row_data_deals[10])
            FxConvDates.append(row_data_deals[15])
            FirstFixingDates.append(row_data_deals[16])
            LastFixingDates.append(row_data_deals[17])            
            Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))   
            row_data_deals[-2] = "Pending Checker"
            Status_update = "Pending Checker"
            row_data_deals[-1] = SID
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
            SIDS.append(row_data_deals[-1])
            Statuses.append(row_data_deals[-2])
            Statuses_update.append(row_data_deals[-2])
            rows_to_update_deals.append(row_data_deals)  # Acumula para atualização em lote
        elif (row_data_deals[-2] == "Pending Maker" or row_data_deals[-2] == "Pending Checker") and row_data_deals[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Maker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the B2b Treeview Deals
    for item in treeview_deals_b2b.get_children():
        row_data_deals_b2b = list(treeview_deals_b2b.item(item, 'values'))
        identifier_maker = str(row_data_deals_b2b[-4])
        if row_data_deals_b2b[-2] == "Pending Maker" and row_data_deals_b2b[-1] != SID and identifier_maker in identifiers_to_make:
            B2B_AthenaIDs.append(row_data_deals_b2b[0])
            Instruments.append(row_data_deals_b2b[4])
            TradeDates_update.append(row_data_deals_b2b[1])
            Counterparties_update.append(row_data_deals_b2b[18])
            AthenaIDs.append(row_data_deals_b2b[0])      
            DealNames.append(row_data_deals_b2b[0])                  
            Indexes.append(row_data_deals_b2b[-3])
            Markets.append(row_data_deals_b2b[2])
            Types.append(row_data_deals_b2b[3])
            Strikes.append(row_data_deals_b2b[6])
            IntermediatesCCY.append(row_data_deals_b2b[8])
            TotalNotionals.append(row_data_deals_b2b[9])
            SettlementDates.append(row_data_deals_b2b[10])
            FxConvDates.append(row_data_deals_b2b[15])
            FirstFixingDates.append(row_data_deals_b2b[16])
            LastFixingDates.append(row_data_deals_b2b[17])            
            row_data_deals_b2b[-2] = "Concluded"
            row_data_deals_b2b[-1] = SID
            treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
            SIDS.append(row_data_deals_b2b[-1])
            rows_to_update_deals_b2b.append(row_data_deals_b2b)  # Acumula para atualização em lote
            Statuses_update.append(row_data_deals_b2b[-2])
        elif row_data_deals_b2b[-2] == "Generated" and identifier_maker in identifiers_to_make:
            B2B_AthenaIDs.append(row_data_deals_b2b[0])
            Instruments.append(row_data_deals_b2b[4])
            TradeDates_update.append(row_data_deals_b2b[1])
            Counterparties_update.append(row_data_deals_b2b[18])
            AthenaIDs.append(row_data_deals_b2b[0])      
            DealNames.append(row_data_deals_b2b[0])                  
            Indexes.append(row_data_deals_b2b[-3])
            Markets.append(row_data_deals_b2b[2])
            Types.append(row_data_deals_b2b[3])
            Strikes.append(row_data_deals_b2b[6])
            IntermediatesCCY.append(row_data_deals_b2b[8])
            TotalNotionals.append(row_data_deals_b2b[9])
            SettlementDates.append(row_data_deals_b2b[10])
            FxConvDates.append(row_data_deals_b2b[15])
            FirstFixingDates.append(row_data_deals_b2b[16])
            LastFixingDates.append(row_data_deals_b2b[17])            
            row_data_deals_b2b[-2] = "Pending Checker"
            row_data_deals_b2b[-1] = SID
            treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
            SIDS.append(row_data_deals_b2b[-1])
            rows_to_update_deals_b2b.append(row_data_deals_b2b)  # Acumula para atualização em lote
            Statuses_update.append(row_data_deals_b2b[-2])
        elif (row_data_deals_b2b[-2] == "Pending Maker" or row_data_deals_b2b[-2] == "Pending Checker") and row_data_deals_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Maker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the Client Treeview Fixings
    for item in treeview_fixings.get_children():
        row_data_fixings = list(treeview_fixings.item(item, 'values'))
        identifier_maker = str(row_data_fixings[-4])
        if row_data_fixings[-2] == "Pending Maker" and row_data_fixings[-1] != SID and identifier_maker in identifiers_to_make:
            row_data_fixings[-2] = "Concluded"
            row_data_fixings[-1] = SID
            treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
            rows_to_update_fixings.append(row_data_fixings)  # Acumula para atualização em lote
        elif row_data_fixings[-2] == "Generated" and identifier_maker in identifiers_to_make:
            row_data_fixings[-2] = "Pending Checker"
            row_data_fixings[-1] = SID
            treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
            rows_to_update_fixings.append(row_data_fixings)  # Acumula para atualização em lote
        elif (row_data_fixings[-2] == "Pending Maker" or row_data_fixings[-2] == "Pending Checker") and row_data_fixings[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Maker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the B2b Treeview Fixings
    for item in treeview_fixings_b2b.get_children():
        row_data_fixings_b2b = list(treeview_fixings_b2b.item(item, 'values'))
        identifier_maker = str(row_data_fixings_b2b[-4])
        if row_data_fixings_b2b[-2] == "Pending Maker" and row_data_fixings_b2b[-1] != SID and identifier_maker in identifiers_to_make:
            row_data_fixings_b2b[-2] = "Concluded"
            row_data_fixings_b2b[-1] = SID
            treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview
            rows_to_update_fixings_b2b.append(row_data_fixings_b2b)  # Acumula para atualização em lote
        elif row_data_fixings_b2b[-2] == "Generated" and identifier_maker in identifiers_to_make:
            row_data_fixings_b2b[-2] = "Pending Checker"
            row_data_fixings_b2b[-1] = SID
            treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview
            rows_to_update_fixings_b2b.append(row_data_fixings_b2b)  # Acumula para atualização em lote
        elif row_data_fixings_b2b[-2] == "Pending Maker" and row_data_fixings_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Maker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the Client Treeview File
    for item in treeview_file.get_children():
        row_data_file = list(treeview_file.item(item, 'values'))
        identifier_maker = str(row_data_file[-4])        
        if row_data_file[-2] == "Pending Maker" and row_data_file[-1] != SID and identifier_maker in identifiers_to_make:
            row_data_file[-2] = "Concluded"
            row_data_file[-1] = SID
            treeview_file.item(item, values=row_data_file)  # Atualiza o Treeview
            rows_to_update_file.append(row_data_file)  # Acumula para atualização em lote
        elif row_data_file[-2] == "Generated" and identifier_maker in identifiers_to_make:
            row_data_file[-2] = "Pending Checker"
            row_data_file[-1] = SID
            treeview_file.item(item, values=row_data_file)  # Atualiza o Treeview
            rows_to_update_file.append(row_data_file)  # Acumula para atualização em lote
        elif (row_data_file[-2] == "Pending Maker" or row_data_file[-2] == "Pending Checker") and row_data_file[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Maker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the B2b Treeview File
    for item in treeview_file_b2b.get_children():
        row_data_file_b2b = list(treeview_file_b2b.item(item, 'values'))
        identifier_maker = str(row_data_file_b2b[-4])
        if row_data_file_b2b[-2] == "Pending Maker" and row_data_file_b2b[-1] != SID and identifier_maker in identifiers_to_make:
            row_data_file_b2b[-2] = "Concluded"
            row_data_file_b2b[-1] = SID
            treeview_file_b2b.item(item, values=row_data_file_b2b)  # Atualiza o Treeview
            rows_to_update_file_b2b.append(row_data_file_b2b)  # Acumula para atualização em lote
        elif row_data_file_b2b[-2] == "Generated" and identifier_maker in identifiers_to_make:
            row_data_file_b2b[-2] = "Pending Checker"
            row_data_file_b2b[-1] = SID
            treeview_file_b2b.item(item, values=row_data_file_b2b)  # Atualiza o Treeview
            rows_to_update_file_b2b.append(row_data_file_b2b)  # Acumula para atualização em lote
        elif (row_data_file_b2b[-2] == "Pending Maker" or row_data_file_b2b[-2] == "Pending Checker") and row_data_file_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Maker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Prepare updates list
    data_and_tables = [
        (rows_to_update_deals, instrument_deals),
        (rows_to_update_deals_b2b, instrument_deals),
        (rows_to_update_fixings, instrument_fixings),
        (rows_to_update_fixings_b2b, instrument_fixings),
        (rows_to_update_file, instrument_file),
        (rows_to_update_file_b2b, instrument_file)
    ]

    insert_or_update_all_tables(data_and_tables)
    
    # Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += [""] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
        
    ajustar_largura_colunas(tabela_arquivoopcao_cliente, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivoopcao_b2b, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_cliente, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_b2b, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)       
    highlight_duplicates(tabela_opcao_cliente, 'deals')
    highlight_duplicates(tabela_opcao_b2b, 'deals')
    highlight_duplicates(tabela_termo_cliente, 'deals')
    highlight_duplicates(tabela_termo_b2b, 'deals')
    highlight_duplicates(tabela_arquivoopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivoopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo') 
    
    # Close the review window
    window.destroy()
    
    # Send status change email
    status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update)
                                 
def checker_action(window, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b):
    # Get database connection
    
    # Filter identifiers based on checkbox state
    identifiers_to_check = [identifier for identifier, var in checkbox_states.items() if var.get()]
    SID = getpass.getuser()
    SID = SID[0].upper() + SID[1:]
    
    # Lists to accumulate row data for batch update
    rows_to_update_deals = []
    rows_to_update_deals_b2b = []
    rows_to_update_fixings = []
    rows_to_update_fixings_b2b = []
    rows_to_update_file = []
    rows_to_update_file_b2b = []

    # Lists to accumulate data for database update
    DealNames = []
    Markets = []
    Types = []
    Strikes = []
    IntermediatesCCY = []
    TotalNotionals = []
    SettlementDates = []
    FxConvDates = []
    FirstFixingDates = []
    LastFixingDates = []
    SIDS = []
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    TradeDates_update = []
    Counterparties = []
    Counterparties_update = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []
    Statuses_update = []
    Status_update = ""

    # Variable to track if an error has occurred
    error_occurred = False

    # Update the Client Treeview Deals
    for item in treeview_deals.get_children():
        row_data_deals = list(treeview_deals.item(item, 'values'))
        identifier_checker = str(row_data_deals[-4])
        if row_data_deals[-2] == "Pending Checker" and row_data_deals[-1] != SID and identifier_checker in identifiers_to_check:
            Makers.append(row_data_deals[-1])
            Checkers.append(SID)
            Instruments.append(row_data_deals[4])
            TradeDates_update.append(row_data_deals[1])
            TradeDates.append(row_data_deals[1])
            Counterparties.append(row_data_deals[18])
            Counterparties_update.append(row_data_deals[18])
            AthenaIDs.append(row_data_deals[0])      
            DealNames.append(row_data_deals[0])      
            Identifiers.append(row_data_deals[-4])
            Indexes.append(row_data_deals[-3])
            Markets.append(row_data_deals[2])
            Types.append(row_data_deals[3])
            Strikes.append(row_data_deals[6])
            IntermediatesCCY.append(row_data_deals[8])
            TotalNotionals.append(row_data_deals[9])
            SettlementDates.append(row_data_deals[10])
            FxConvDates.append(row_data_deals[15])
            FirstFixingDates.append(row_data_deals[16])
            LastFixingDates.append(row_data_deals[17])            
            Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))     
            row_data_deals[-2] = "Concluded"
            Status_update = "Concluded"
            row_data_deals[-1] = SID
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
            SIDS.append(row_data_deals[-1])
            Statuses.append(row_data_deals[-2])
            Statuses_update.append(row_data_deals[-2])
            rows_to_update_deals.append(row_data_deals)  # Acumula para atualização em lote
        elif row_data_deals[-2] == "Generated" and identifier_checker in identifiers_to_check:            
            Checkers.append(SID)
            Instruments.append(row_data_deals[4])
            TradeDates_update.append(row_data_deals[1])
            TradeDates.append(row_data_deals[1])
            Counterparties.append(row_data_deals[18])
            Counterparties_update.append(row_data_deals[18])
            AthenaIDs.append(row_data_deals[0])          
            DealNames.append(row_data_deals[0])      
            Identifiers.append(row_data_deals[-4])
            Indexes.append(row_data_deals[-3])
            Markets.append(row_data_deals[2])
            Types.append(row_data_deals[3])
            Strikes.append(row_data_deals[6])
            IntermediatesCCY.append(row_data_deals[8])
            TotalNotionals.append(row_data_deals[9])
            SettlementDates.append(row_data_deals[10])
            FxConvDates.append(row_data_deals[15])
            FirstFixingDates.append(row_data_deals[16])
            LastFixingDates.append(row_data_deals[17])            
            Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))            
            row_data_deals[-2] = "Pending Maker"
            Status_update = "Pending Maker"
            row_data_deals[-1] = SID
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
            SIDS.append(row_data_deals[-1])
            Statuses.append(row_data_deals[-2])
            Statuses_update.append(row_data_deals[-2])
            rows_to_update_deals.append(row_data_deals)  # Acumula para atualização em lote
        elif (row_data_deals[-2] == "Pending Maker" or row_data_deals[-2] == "Pending Checker") and row_data_deals[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the B2b Treeview Deals
    for item in treeview_deals_b2b.get_children():
        row_data_deals_b2b = list(treeview_deals_b2b.item(item, 'values'))
        identifier_checker = str(row_data_deals_b2b[-4])
        if row_data_deals_b2b[-2] == "Pending Checker" and row_data_deals_b2b[-1] != SID and identifier_checker in identifiers_to_check:
            B2B_AthenaIDs.append(row_data_deals_b2b[0])
            Instruments.append(row_data_deals_b2b[4])
            TradeDates_update.append(row_data_deals_b2b[1])
            Counterparties_update.append(row_data_deals_b2b[18])
            AthenaIDs.append(row_data_deals_b2b[0])      
            DealNames.append(row_data_deals_b2b[0])                  
            Indexes.append(row_data_deals_b2b[-3])
            Markets.append(row_data_deals_b2b[2])
            Types.append(row_data_deals_b2b[3])
            Strikes.append(row_data_deals_b2b[6])
            IntermediatesCCY.append(row_data_deals_b2b[8])
            TotalNotionals.append(row_data_deals_b2b[9])
            SettlementDates.append(row_data_deals_b2b[10])
            FxConvDates.append(row_data_deals_b2b[15])
            FirstFixingDates.append(row_data_deals_b2b[16])
            LastFixingDates.append(row_data_deals_b2b[17])            
            row_data_deals_b2b[-2] = "Concluded"
            row_data_deals_b2b[-1] = SID
            treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
            SIDS.append(row_data_deals_b2b[-1])
            rows_to_update_deals_b2b.append(row_data_deals_b2b)  # Acumula para atualização em lote
            Statuses_update.append(row_data_deals_b2b[-2])
        elif row_data_deals_b2b[-2] == "Generated" and identifier_checker in identifiers_to_check:
            B2B_AthenaIDs.append(row_data_deals_b2b[0])
            Instruments.append(row_data_deals_b2b[4])
            TradeDates_update.append(row_data_deals_b2b[1])
            Counterparties_update.append(row_data_deals_b2b[18])
            AthenaIDs.append(row_data_deals_b2b[0])      
            DealNames.append(row_data_deals_b2b[0])                  
            Indexes.append(row_data_deals_b2b[-3])
            Markets.append(row_data_deals_b2b[2])
            Types.append(row_data_deals_b2b[3])
            Strikes.append(row_data_deals_b2b[6])
            IntermediatesCCY.append(row_data_deals_b2b[8])
            TotalNotionals.append(row_data_deals_b2b[9])
            SettlementDates.append(row_data_deals_b2b[10])
            FxConvDates.append(row_data_deals_b2b[15])
            FirstFixingDates.append(row_data_deals_b2b[16])
            LastFixingDates.append(row_data_deals_b2b[17])            
            row_data_deals_b2b[-2] = "Pending Maker"
            row_data_deals_b2b[-1] = SID
            treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
            SIDS.append(row_data_deals_b2b[-1])
            rows_to_update_deals_b2b.append(row_data_deals_b2b)  # Acumula para atualização em lote
            Statuses_update.append(row_data_deals_b2b[-2])
        elif (row_data_deals_b2b[-2] == "Pending Maker" or row_data_deals_b2b[-2] == "Pending Checker") and row_data_deals_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the Client Treeview Fixings
    for item in treeview_fixings.get_children():
        row_data_fixings = list(treeview_fixings.item(item, 'values'))
        identifier_checker = str(row_data_fixings[-4])
        if row_data_fixings[-2] == "Pending Checker" and row_data_fixings[-1] != SID and identifier_checker in identifiers_to_check:
            row_data_fixings[-2] = "Concluded"
            row_data_fixings[-1] = SID
            treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
            rows_to_update_fixings.append(row_data_fixings)  # Acumula para atualização em lote
        elif row_data_fixings[-2] == "Generated" and identifier_checker in identifiers_to_check:
            row_data_fixings[-2] = "Pending Maker"
            row_data_fixings[-1] = SID
            treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
            rows_to_update_fixings.append(row_data_fixings)  # Acumula para atualização em lote
        elif (row_data_fixings[-2] == "Pending Maker" or row_data_fixings[-2] == "Pending Checker") and row_data_fixings[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the B2b Treeview Fixings
    for item in treeview_fixings_b2b.get_children():
        row_data_fixings_b2b = list(treeview_fixings_b2b.item(item, 'values'))
        identifier_checker = str(row_data_fixings_b2b[-4])
        if row_data_fixings_b2b[-2] == "Pending Checker" and row_data_fixings_b2b[-1] != SID and identifier_checker in identifiers_to_check:
            row_data_fixings_b2b[-2] = "Concluded"
            row_data_fixings_b2b[-1] = SID
            treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview
            rows_to_update_fixings_b2b.append(row_data_fixings_b2b)  # Acumula para atualização em lote
        elif row_data_fixings_b2b[-2] == "Generated" and identifier_checker in identifiers_to_check:
            row_data_fixings_b2b[-2] = "Pending Maker"
            row_data_fixings_b2b[-1] = SID
            treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview
            rows_to_update_fixings_b2b.append(row_data_fixings_b2b)  # Acumula para atualização em lote
        elif (row_data_fixings_b2b[-2] == "Pending Maker" or row_data_fixings_b2b[-2] == "Pending Checker") and row_data_fixings_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the Client Treeview File
    for item in treeview_file.get_children():
        row_data_file = list(treeview_file.item(item, 'values'))
        identifier_checker = str(row_data_file[-4])
        if row_data_file[-2] == "Pending Checker" and row_data_file[-1] != SID and identifier_checker in identifiers_to_check:
            row_data_file[-2] = "Concluded"
            row_data_file[-1] = SID
            treeview_file.item(item, values=row_data_file)  # Atualiza o Treeview
            rows_to_update_file.append(row_data_file)  # Acumula para atualização em lote
        elif row_data_file[-2] == "Generated" and identifier_checker in identifiers_to_check:
            row_data_file[-2] = "Pending Maker"
            row_data_file[-1] = SID
            treeview_file.item(item, values=row_data_file)  # Atualiza o Treeview
            rows_to_update_file.append(row_data_file)  # Acumula para atualização em lote
        elif (row_data_file[-2] == "Pending Maker" or row_data_file[-2] == "Pending Checker") and row_data_file[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Update the B2b Treeview File
    for item in treeview_file_b2b.get_children():
        row_data_file_b2b = list(treeview_file_b2b.item(item, 'values'))
        identifier_checker = str(row_data_file_b2b[-4])
        if row_data_file_b2b[-2] == "Pending Checker" and row_data_file_b2b[-1] != SID and identifier_checker in identifiers_to_check:
            row_data_file_b2b[-2] = "Concluded"
            row_data_file_b2b[-1] = SID
            treeview_file_b2b.item(item, values=row_data_file_b2b)  # Atualiza o Treeview
            rows_to_update_file_b2b.append(row_data_file_b2b)  # Acumula para atualização em lote
        elif row_data_file_b2b[-2] == "Generated" and identifier_checker in identifiers_to_check:
            row_data_file_b2b[-2] = "Pending Maker"
            row_data_file_b2b[-1] = SID
            treeview_file_b2b.item(item, values=row_data_file_b2b)
            # Atualiza o Treeview
            rows_to_update_file_b2b.append(row_data_file_b2b)  # Acumula para atualização em lote
        elif (row_data_file_b2b[-2] == "Pending Maker" or row_data_file_b2b[-2] == "Pending Checker") and row_data_file_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Checker")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Prepare updates list
    data_and_tables = [
        (rows_to_update_deals, instrument_deals),
        (rows_to_update_deals_b2b, instrument_deals),
        (rows_to_update_fixings, instrument_fixings),
        (rows_to_update_fixings_b2b, instrument_fixings),
        (rows_to_update_file, instrument_file),
        (rows_to_update_file_b2b, instrument_file)
    ]
    
    insert_or_update_all_tables(data_and_tables)
    
    # Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += [""] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
       
    ajustar_largura_colunas(tabela_arquivoopcao_cliente, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivoopcao_b2b, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_cliente, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_b2b, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)       
    highlight_duplicates(tabela_opcao_cliente, 'deals')
    highlight_duplicates(tabela_opcao_b2b, 'deals')
    highlight_duplicates(tabela_termo_cliente, 'deals')
    highlight_duplicates(tabela_termo_b2b, 'deals')
    highlight_duplicates(tabela_arquivoopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivoopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo') 
    
    # Close the review window
    window.destroy()
    
    # Send status change email
    status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update)
def insert_or_update_base_deals(TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs, Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations, SS_Validations, Identifiers, Indexes):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    try:
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")
        
        # Iterate over the lists and update the database for each set of values
        for i in range(len(AthenaIDs)):
            # Check if the record already exists in the base_deals table
            cursor.execute("SELECT * FROM base_deals WHERE AthenaID = ? OR B2B_AthenaID = ?", (AthenaIDs[i], B2B_AthenaIDs[i]))
            existing_record = cursor.fetchone()

            if existing_record:
                # Update the existing record in the base_deals table
                cursor.execute("""
                    UPDATE base_deals SET
                    TradeDate = COALESCE(NULLIF(?, ''), TradeDate),
                    Counterparty = COALESCE(NULLIF(?, ''), Counterparty),                
                    B3_ID = COALESCE(NULLIF(?, ''), B3_ID),
                    B2B_AthenaID = COALESCE(NULLIF(?, ''), B2B_AthenaID),
                    B2B_B3_ID = COALESCE(NULLIF(?, ''), B2B_B3_ID),
                    Instrument = COALESCE(NULLIF(?, ''), Instrument),
                    Status = COALESCE(NULLIF(?, ''), Status),
                    Maker = COALESCE(NULLIF(?, ''), Maker),
                    Checker = COALESCE(NULLIF(?, ''), Checker),
                    Time_Stamp = COALESCE(NULLIF(?, ''), Time_Stamp),
                    Confirmation = COALESCE(NULLIF(?, ''), Confirmation),
                    SS_Validation = COALESCE(NULLIF(?, ''), SS_Validation),
                    "Identifier" = COALESCE(NULLIF(?, ''), "Identifier"),
                    "Index" = COALESCE(NULLIF(?, ''), "Index")              
                    WHERE AthenaID = ? OR B2B_AthenaID = ?       
                """, (TradeDates[i], Counterparties[i], B3_IDs[i], B2B_AthenaIDs[i], B2B_B3_IDs[i], Instruments[i], Statuses[i], Makers[i], Checkers[i], Time_Stamps[i], Confirmations[i], SS_Validations[i], Identifiers[i], Indexes[i], AthenaIDs[i], B2B_AthenaIDs[i]))
            else:
                # Insert a new record into the base_deals table
                cursor.execute("""
                    INSERT INTO base_deals (TradeDate, Counterparty, AthenaID, B3_ID, B2B_AthenaID, B2B_B3_ID, Instrument, Status, Maker, Checker, Time_Stamp, Confirmation, SS_Validation, "Identifier", "Index")
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (TradeDates[i], Counterparties[i], AthenaIDs[i], B3_IDs[i], B2B_AthenaIDs[i], B2B_B3_IDs[i], Instruments[i], Statuses[i], Makers[i], Checkers[i], Time_Stamps[i], Confirmations[i], SS_Validations[i], Identifiers[i], Indexes[i]))

        # Commit the transaction after all updates/inserts
        conn.commit()
    except sqlite3.Error as e:
        # Rollback the transaction in case of an error
        conn.rollback()
        conn.close()
        # Display an error message if there is a problem with the database
        messagebox.showerror("Database Error", f"An error occurred while accessing the database: {e}")
    finally:
        # Close the database connection
        conn.close()


                                                                                 
def create_edit_window(tab_name, columns_deals, columns_file, columns_fixings, instrument_deals, instrument_file, instrument_fixings, identifiers_approved, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b):
    deals_cliente, td_cliente, market_client, type_cliente, instrument_cliente, strike_cliente, intCCY_cliente, tn_cliente, sd_cliente, fxd_cliente, fsd_cliente, fed_cliente, accronym, identifier_cliente, status_cliente, sid_cliente, deals_b2b, td_b2b, market_client, type_b2b, instrument_b2b, strike_b2b, intCCY_b2b, tn_b2b, sd_b2b, fxd_b2b, fsd_b2b, fed_b2b, accronym_b2b, identifier_b2b, status_b2b, sid_b2b = extrair_dados_deals(treeview_deals, treeview_deals_b2b)
    global janela
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")
    review_window = ctk.CTkToplevel(janela)
    review_window.geometry("1300x600")
    review_window.title("Edit Queue")
    review_window.lift()
    review_window.focus_set()
    review_window.grab_set()

    # Font for Buttons and Labels
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")
    fonte_label = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    frame_export = ctk.CTkScrollableFrame(review_window)
    frame_export.pack(fill='both', expand=True)

    # Create frames for each section
    frame_checkboxes = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_checkboxes.pack(side='left', fill='both', expand=True, padx=10, pady=8)
    
    label = ctk.CTkLabel(frame_checkboxes, text="Select", font=fonte_label)
    label.pack(pady=5, padx=5)

    frame_labels_cliente = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_cliente.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_cliente = ctk.CTkLabel(frame_labels_cliente, text="Client", font=fonte_label)
    label_cliente.pack(pady=5, padx=5)

    frame_labels_b2b = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_b2b.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_b2b = ctk.CTkLabel(frame_labels_b2b, text="B2B", font=fonte_label)
    label_b2b.pack(pady=5, padx=5)
    
    frame_labels_status = ctk.CTkFrame(frame_export, border_width=1, border_color='black')
    frame_labels_status.pack(side='left', fill='both', expand=True, padx=5, pady=5)
    
    label_status = ctk.CTkLabel(frame_labels_status, text="Status", font=fonte_label)
    label_status.pack(pady=5, padx=5)

    # Dictionary to store checkbox states
    checkbox_states = {}

    # Create checkboxes and labels based on identifiers
    for idx, identifier in enumerate(identifiers_approved):
        accronym_value = lookup_approve(identifier, identifier_cliente, accronym)
        deal_cliente_value = lookup_approve(identifier, identifier_cliente, deals_cliente)
        deal_b2b_value = lookup_approve(identifier, identifier_b2b, deals_b2b)
        status_cliente_value = lookup_approve(identifier, identifier_cliente, status_cliente)
        status_b2b_value = lookup_approve(identifier, identifier_cliente, status_b2b)
        
        var = ctk.BooleanVar(value=True)
        checkbox_states[identifier] = var  # Store the state of the checkbox
        checkbox = ctk.CTkCheckBox(frame_checkboxes, text=f"{accronym_value if accronym_value else 'Unknown'}", checkbox_height=18, checkbox_width=18, height=20, variable=var)
        checkbox.pack(pady=5, padx=5)

        label_cliente = ctk.CTkLabel(frame_labels_cliente, text=f"{deal_cliente_value if deal_cliente_value else 'Unknown'}", anchor='center', height=20)
        label_cliente.pack(pady=5, padx=5)

        label_b2b = ctk.CTkLabel(frame_labels_b2b, text=f"{deal_b2b_value if deal_b2b_value else 'Unknown'}", anchor='center', height=20)
        label_b2b.pack(pady=5, padx=5)

        label_status = ctk.CTkLabel(frame_labels_status, text=f"{status_cliente_value if status_cliente_value == status_b2b_value else 'Unknown'}", anchor='center', height=20)
        label_status.pack(pady=5, padx=5)

    # Create a frame for the buttons
    frame_buttons = ctk.CTkFrame(review_window)
    frame_buttons.pack(fill='x', pady=5)

    # Create the EDIT button
    button_approve = ctk.CTkButton(frame_buttons, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EDIT", font=fonte_botao, command=lambda: edit_action(review_window, columns_deals, columns_fixings, columns_file, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b))
    button_approve.pack(side='left', pady=5, padx=5)
    
    
def edit_action(window, columns_deals, columns_fixings, columns_file, instrument_deals, instrument_file, instrument_fixings, checkbox_states, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b):
    # Filter identifiers based on checkbox state
    identifiers_to_edit = [identifier for identifier, var in checkbox_states.items() if var.get()]
    SID = getpass.getuser()
    SID = SID[0].upper() + SID[1:]
    
    # Lists to accumulate row data for batch update
    rows_to_update_deals = []
    rows_to_update_deals_b2b = []
    rows_to_update_fixings = []
    rows_to_update_fixings_b2b = []
    rows_to_update_file = []
    rows_to_update_file_b2b = []

    # Lists to accumulate data for database update
    DealNames = []
    Markets = []
    Types = []
    Strikes = []
    IntermediatesCCY = []
    TotalNotionals = []
    SettlementDates = []
    FxConvDates = []
    FirstFixingDates = []
    LastFixingDates = []
    SIDS = []
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    TradeDates_update = []
    Counterparties = []
    Counterparties_update = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []
    Statuses_update = []
    Status_update = "New"

    # Variable to track if an error has occurred
    error_occurred = False

    # Update the Client Treeview Deals   
    for item in treeview_deals.get_children():        
        row_data_deals = list(treeview_deals.item(item, 'values'))
        identifier_edit = str(row_data_deals[-4])
        if row_data_deals[-2] == "Approved" and identifier_edit in identifiers_to_edit and row_data_deals[-1] != SID:
            Instruments.append(row_data_deals[4])
            TradeDates_update.append(row_data_deals[1])
            TradeDates.append(row_data_deals[1])
            Counterparties.append(row_data_deals[18])
            Counterparties_update.append(row_data_deals[18])
            AthenaIDs.append(row_data_deals[0])      
            DealNames.append(row_data_deals[0])      
            Identifiers.append(row_data_deals[-4])
            Indexes.append(row_data_deals[-3])
            Markets.append(row_data_deals[2])
            Types.append(row_data_deals[3])
            Strikes.append(row_data_deals[6])
            IntermediatesCCY.append(row_data_deals[8])
            TotalNotionals.append(row_data_deals[9])
            SettlementDates.append(row_data_deals[10])
            FxConvDates.append(row_data_deals[15])
            FirstFixingDates.append(row_data_deals[16])
            LastFixingDates.append(row_data_deals[17])            
            Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))     
            row_data_deals[-2] = "New"
            row_data_deals[-1] = SID
            treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
            SIDS.append(row_data_deals[-1])
            Statuses.append(row_data_deals[-2])
            Statuses_update.append(row_data_deals[-2])
            rows_to_update_deals.append(row_data_deals)  # Acumula para atualização em lote
        elif row_data_deals[-2] == "Approved" and row_data_deals[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Edit")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview Deals
    for item in treeview_deals_b2b.get_children():
        row_data_deals_b2b = list(treeview_deals_b2b.item(item, 'values'))
        identifier_edit = str(row_data_deals_b2b[-4])            
        if row_data_deals_b2b[-2] == "Approved" and identifier_edit in identifiers_to_edit and row_data_deals_b2b[-1] != SID:
            B2B_AthenaIDs.append(row_data_deals_b2b[0])
            Instruments.append(row_data_deals_b2b[4])
            TradeDates_update.append(row_data_deals_b2b[1])
            Counterparties_update.append(row_data_deals_b2b[18])
            AthenaIDs.append(row_data_deals_b2b[0])      
            DealNames.append(row_data_deals_b2b[0])                  
            Indexes.append(row_data_deals_b2b[-3])
            Markets.append(row_data_deals_b2b[2])
            Types.append(row_data_deals_b2b[3])
            Strikes.append(row_data_deals_b2b[6])
            IntermediatesCCY.append(row_data_deals_b2b[8])
            TotalNotionals.append(row_data_deals_b2b[9])
            SettlementDates.append(row_data_deals_b2b[10])
            FxConvDates.append(row_data_deals_b2b[15])
            FirstFixingDates.append(row_data_deals_b2b[16])
            LastFixingDates.append(row_data_deals_b2b[17])            
            row_data_deals_b2b[-2] = "New"
            row_data_deals_b2b[-1] = SID
            treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
            SIDS.append(row_data_deals_b2b[-1])
            rows_to_update_deals_b2b.append(row_data_deals_b2b)  # Acumula para atualização em lote
            Statuses_update.append(row_data_deals_b2b[-2])
        elif row_data_deals_b2b[-2] == "Approved" and row_data_deals_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Edit")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the Client Treeview Fixings
    for item in treeview_fixings.get_children():
        row_data_fixings = list(treeview_fixings.item(item, 'values'))
        identifier_edit = str(row_data_fixings[-4])
        if row_data_fixings[-2] == "Approved" and identifier_edit in identifiers_to_edit and row_data_fixings[-1] != SID:
            row_data_fixings[-2] = "New"
            row_data_fixings[-1] = SID
            treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
            rows_to_update_fixings.append(row_data_fixings)  # Acumula para atualização em lote
        elif row_data_fixings[-2] == "Approved" and row_data_fixings[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Edit")
            error_occurred = True
            break
        
    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview Fixings
    for item in treeview_fixings_b2b.get_children():
        row_data_fixings_b2b = list(treeview_fixings_b2b.item(item, 'values'))
        identifier_edit = str(row_data_fixings_b2b[-4])
        if row_data_fixings_b2b[-2] == "Approved" and identifier_edit in identifiers_to_edit and row_data_fixings_b2b[-1] != SID:
            row_data_fixings_b2b[-2] = "New"
            row_data_fixings_b2b[-1] = SID
            treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview
            rows_to_update_fixings_b2b.append(row_data_fixings_b2b)  # Acumula para atualização em lote
        elif row_data_fixings_b2b[-2] == "Approved" and row_data_fixings_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Edit")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the Client Treeview File       
    for item in treeview_file.get_children():
        row_data_file = list(treeview_file.item(item, 'values'))
        identifier_edit = str(row_data_file[-4])            
        if row_data_file[-2] == "Approved" and identifier_edit in identifiers_to_edit and row_data_file[-1] != SID:
            row_data_file[-2] = "New"
            row_data_file[-1] = SID
            treeview_file.item(item, values=row_data_file)  # Atualiza o Treeview
            rows_to_update_file.append(row_data_file)  # Acumula para atualização em lote
        elif row_data_file[-2] == "Approved" and row_data_file[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Edit")
            error_occurred = True
            break
    
    if error_occurred:
        window.destroy()
        return
    
    # Update the B2b Treeview File
    for item in treeview_file_b2b.get_children():
        row_data_file_b2b = list(treeview_file_b2b.item(item, 'values'))
        identifier_edit = str(row_data_file_b2b[-4])            
        if row_data_file_b2b[-2] == "Approved" and identifier_edit in identifiers_to_edit and row_data_file_b2b[-1] != SID:
            row_data_file_b2b[-2] = "New"
            row_data_file_b2b[-1] = SID
            treeview_file_b2b.item(item, values=row_data_file_b2b)  # Atualiza o Treeview
            rows_to_update_file_b2b.append(row_data_file_b2b)  # Acumula para atualização em lote
        elif row_data_file_b2b[-2] == "Approved" and row_data_file_b2b[-1] == SID:
            messagebox.showwarning("Attention!", "Different SID must perform the Edit")
            error_occurred = True
            break

    if error_occurred:
        window.destroy()
        return

    # Prepare updates list
    data_and_tables = [
        (rows_to_update_deals, instrument_deals),
        (rows_to_update_deals_b2b, instrument_deals),
        (rows_to_update_fixings, instrument_fixings),
        (rows_to_update_fixings_b2b, instrument_fixings),
        (rows_to_update_file, instrument_file),
        (rows_to_update_file_b2b, instrument_file)
    ]

    insert_or_update_all_tables(data_and_tables)
      
    # Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += [""] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
    ajustar_largura_colunas(tabela_arquivoopcao_cliente, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivoopcao_b2b, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_cliente, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_b2b, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)       
    highlight_duplicates(tabela_opcao_cliente, 'deals')
    highlight_duplicates(tabela_opcao_b2b, 'deals')
    highlight_duplicates(tabela_termo_cliente, 'deals')
    highlight_duplicates(tabela_termo_b2b, 'deals')
    highlight_duplicates(tabela_arquivoopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivoopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_cliente, 'arquivo')
    highlight_duplicates(tabela_arquivotermo_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
    highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo') 
    # Close the review window
    window.destroy()
    
    # Send status change email
    status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update)
    
def status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update):
    # Construct the email body in HTML
    email_body = f"""
    <html>
    <body style="font-family: 'Calibri'; font-size: 11pt;">               
    <p>The status of the following operation(s) has been changed to {Status_update}:</p>
    
    <table style="font-family: 'Arial'; font-size: 10pt; border-collapse: collapse; width: auto; border: 1px solid black; text-align: center;">
        <tr style="font-weight: bold; border: 1px solid black;">
            <td style="border: 1px solid black;">DealName</td>
            <td style="border: 1px solid black;">TradeDate</td>
            <td style="border: 1px solid black;">Market</td>
            <td style="border: 1px solid black;">Type</td>
            <td style="border: 1px solid black;">Instrument</td>
            <td style="border: 1px solid black;">Strike</td>
            <td style="border: 1px solid black;">IntermediateCCY</td>
            <td style="border: 1px solid black;">TotalNotional</td>
            <td style="border: 1px solid black;">SettlementDate</td>
            <td style="border: 1px solid black;">FXConvDate</td>
            <td style="border: 1px solid black;">FixingStartDate</td>
            <td style="border: 1px solid black;">FixingEndDate</td>
            <td style="border: 1px solid black;">Counterparty</td>
            <td style="border: 1px solid black;">Status</td>
            <td style="border: 1px solid black;">SID</td>
        </tr>      
    """    

    # Iterate over the items to populate the table
    for i in range(len(DealNames)):
        email_body += f"""
        <tr style="border: 1px solid black;">
            <td style="border: 1px solid black;">{DealNames[i]}</td>
            <td style="border: 1px solid black;">{TradeDates_update[i]}</td>
            <td style="border: 1px solid black;">{Markets[i]}</td>
            <td style="border: 1px solid black;">{Types[i]}</td>
            <td style="border: 1px solid black;">{Instruments[i]}</td>
            <td style="border: 1px solid black;">{Strikes[i]}</td>
            <td style="border: 1px solid black;">{IntermediatesCCY[i]}</td>
            <td style="border: 1px solid black;">{TotalNotionals[i]}</td>
            <td style="border: 1px solid black;">{SettlementDates[i]}</td>
            <td style="border: 1px solid black;">{FxConvDates[i]}</td>
            <td style="border: 1px solid black;">{FirstFixingDates[i]}</td>
            <td style="border: 1px solid black;">{LastFixingDates[i]}</td>
            <td style="border: 1px solid black;">{Counterparties_update[i]}</td>
            <td style="border: 1px solid black;">{Statuses_update[i]}</td>
            <td style="border: 1px solid black;">{SIDS[i]}</td>                        
        </tr>        
        """
    
    # Add footer
    email_body += """
    </table>
    <p>Best regards,</p>
    <p>Banco J.P. Morgan S.A. | Av. Brigadeiro Faria Lima, 3729 - 15º andar - São Paulo - SP | T: 55 11 4950 6717 | F: 55 11 4950 3557 |<br>
    brsp_otc_derivatives_ops@jpmorgan.com | jpmorgan.com | Ouvidoria JPMorgan:  Tel.: 0800 – 7700847 / E-mail: ouvidoria.jp.morgan@jpmorgan.com</p>
    <p>JPMC Internal Use Only</p>
    </body>
    </html>
    """
    
    subject = f"Status Changed - CommodiXchange - {Status_update} - {datetime.today().strftime('%d/%m/%Y')}"

    # Create and send the email in Outlook
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = "OTeamC"  # Replace with the actual recipient
    mail.Subject = subject
    mail.HTMLBody = email_body  # Use HTMLBody for HTML content
    mail.Send()  # Correct method to send the email
    
def open_export_window():
    # Determine the active tab at each level
    main_tab = tabview.get()
    if main_tab == "Monitor":
        sub_tab = sub_notebook_monitor.get()
        if sub_tab == "Operações":
            sub_sub_tab = sub_sub_notebook_operacoes.get()
            if sub_sub_tab == "Termo":
                sub_sub_sub_tab = sub_sub_sub_notebook_termo.get()
                if sub_sub_sub_tab == "Cliente":
                    create_export_window("Termo Cliente", colunas_termo, colunas_fixings_termo, colunas_arquivo_termo)
                elif sub_sub_sub_tab == "B2B":
                    create_export_window("Termo B2B", colunas_termo, colunas_fixings_termo, colunas_arquivo_termo)
            elif sub_sub_tab == "Opção":
                sub_sub_sub_tab = sub_sub_sub_notebook_opcao.get()
                if sub_sub_sub_tab == "Cliente":
                    create_export_window("Opção Cliente", colunas_opcao, colunas_fixings_opcao, colunas_arquivo_opcao)
                elif sub_sub_sub_tab == "B2B":
                    create_export_window("Opção B2B", colunas_opcao, colunas_fixings_opcao, colunas_arquivo_opcao)
        elif sub_tab == "Fixings":
            sub_sub_tab = sub_sub_notebook_fixings.get()
            if sub_sub_tab == "Termo":
                sub_sub_sub_tab = sub_sub_sub_notebook_termo_fixings.get()
                if sub_sub_sub_tab == "Cliente":
                    create_export_window("Fixings Termo Cliente", colunas_termo, colunas_fixings_termo, colunas_arquivo_termo)
                elif sub_sub_sub_tab == "B2B":
                    create_export_window("Fixings Termo B2B", colunas_termo, colunas_fixings_termo, colunas_arquivo_termo)
            elif sub_sub_tab == "Opção":
                sub_sub_sub_tab = sub_sub_sub_notebook_opcao_fixings.get()
                if sub_sub_sub_tab == "Cliente":
                    create_export_window("Fixings Opção Cliente", colunas_opcao, colunas_fixings_opcao, colunas_arquivo_opcao)
                elif sub_sub_sub_tab == "B2B":
                    create_export_window("Fixings Opção B2B", colunas_opcao, colunas_fixings_opcao, colunas_arquivo_opcao)
    elif main_tab == "Arquivo B3":
        sub_tab = sub_notebook_arquivo_b3.get()
        if sub_tab == "Termo":
            sub_sub_tab = sub_sub_notebook_arquivotermo.get()
            if sub_sub_tab == "Cliente":
                create_export_window("Arquivo Termo Cliente", colunas_termo, colunas_fixings_termo, colunas_arquivo_termo)
            elif sub_sub_tab == "B2B":
                create_export_window("Arquivo Termo B2B", colunas_termo, colunas_fixings_termo, colunas_arquivo_termo)
        elif sub_tab == "Opção":
            sub_sub_tab = sub_sub_notebook_arquivoopcao.get()
            if sub_sub_tab == "Cliente":
                create_export_window("Arquivo Opção Cliente", colunas_opcao, colunas_fixings_opcao, colunas_arquivo_opcao)
            elif sub_sub_tab == "B2B":
                create_export_window("Arquivo Opção B2B", colunas_opcao, colunas_fixings_opcao, colunas_arquivo_opcao)
    elif main_tab == "Boleta Dinâmica - Intrag":
        sub_tab = sub_notebook_boleta_dinamica.get()
        if sub_tab == "Termo":
            create_export_window("Boleta Termo", colunas_termo, colunas_fixings_termo, colunas_arquivo_termo)
        elif sub_tab == "Opção":
            create_export_window("Boleta Opção", colunas_opcao, colunas_fixings_opcao, colunas_arquivo_opcao)
    elif main_tab == "Commodities":
        create_export_window("Commodities", colunas_commodities, [], [])
    elif main_tab == "Counterparty":
        create_export_window("Counterparty", colunas_comitentes, [], [])
    elif main_tab == "Holidays":
        sub_tab = tabview_calendarios.get()
        if sub_tab == "ANBIMA":
            create_export_window("Holidays ANBIMA", colunas_anbima, [], [])
        elif sub_tab == "ICE":
            create_export_window("Holidays ICE", colunas_ice, [], [])
        elif sub_tab == "NYMEX":
            create_export_window("Holidays NYMEX", colunas_nymex, [], [])
        elif sub_tab == "BURSA":
            create_export_window("Holidays BURSA", colunas_bursa, [], [])
        elif sub_tab == "CBOT":
            create_export_window("Holidays CBOT", colunas_cbot, [], [])
        elif sub_tab == "PLATTS":
            create_export_window("Holidays PLATTS", colunas_platts, [], [])
        elif sub_tab == "LME":
            create_export_window("Holidays LME", colunas_lme, [], [])

def create_export_window(tab_name, colunas_deals, colunas_fixings, colunas_file):
    global janela
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")
    export_window = ctk.CTkToplevel(janela)    
    export_window.title("Dump to Excel")
    export_window.lift()
    export_window.focus_set()
    export_window.grab_set()

    # Font Buttons
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    frame_export = ctk.CTkFrame(export_window)
    frame_export.grid(row=0, column=0, sticky='nsew')  

    # Configure grid to ensure frames are side by side
    frame_export.grid_rowconfigure(1, weight=1)
    frame_export.grid_columnconfigure((0, 1, 2), weight=1)  # Ensure all columns have equal weight

    frame_title = ctk.CTkFrame(frame_export)
    frame_title.grid(row=0, column=0, columnspan=3, pady=5, sticky='n')   

    label = ctk.CTkLabel(frame_title, text=tab_name)
    label.grid(row=0, column=0, pady=5, padx=5, sticky='n')   

    var_b2b = ctk.BooleanVar(value=True)    
    checkbox_b2b = ctk.CTkCheckBox(frame_title, text="Include B2B", variable=var_b2b)
    checkbox_b2b.grid(row=0, column=1, pady=5, padx=5, sticky='n')   

    # Create frames for each section
    var_deals = ctk.BooleanVar(value=True)    
    checkbox_deals = ctk.CTkCheckBox(frame_title, text="Deals", variable=var_deals, command=lambda: toggle_checkboxes(frame_deals_info, var_deals.get()))
    checkbox_deals.grid(row=0, column=2, pady=2)
    frame_deals_info = ctk.CTkScrollableFrame(frame_export)
    frame_deals_info.grid(row=1, column=0, pady=8, padx=5, sticky='nsew')

    var_file = ctk.BooleanVar(value=True)    
    checkbox_file = ctk.CTkCheckBox(frame_title, text="File", variable=var_file, command=lambda: toggle_checkboxes(frame_file_info, var_file.get()))
    checkbox_file.grid(row=0, column=3, pady=2)
    frame_file_info = ctk.CTkScrollableFrame(frame_export)
    frame_file_info.grid(row=1, column=1, pady=8, padx=5, sticky='nsew')

    var_fixings = ctk.BooleanVar(value=True)    
    checkbox_fixings = ctk.CTkCheckBox(frame_title, text="Fixings", variable=var_fixings, command=lambda: toggle_checkboxes(frame_fixing_info, var_fixings.get()))
    checkbox_fixings.grid(row=0, column=4, pady=2)
    frame_fixing_info = ctk.CTkScrollableFrame(frame_export)
    frame_fixing_info.grid(row=1, column=2, pady=8, padx=5, sticky='nsew')

    # Determine which treeview to use based on the tab name
    if "Termo Cliente" in tab_name:
        treeview = tabela_termo_cliente
    elif "Termo B2B" in tab_name:
        treeview = tabela_termo_b2b
    elif "Opção Cliente" in tab_name:
        treeview = tabela_opcao_cliente
    elif "Opção B2B" in tab_name:
        treeview = tabela_opcao_b2b
    elif "Arquivo Termo Cliente" in tab_name:
        treeview = tabela_arquivotermo_cliente
    elif "Arquivo Termo B2B" in tab_name:
        treeview = tabela_arquivotermo_b2b
    elif "Arquivo Opção Cliente" in tab_name:
        treeview = tabela_arquivoopcao_cliente
    elif "Arquivo Opção B2B" in tab_name:
        treeview = tabela_arquivoopcao_b2b
    elif "Fixings Termo Cliente" in tab_name:
        treeview = tabela_fixingstermo_cliente
    elif "Fixings Termo B2B" in tab_name:
        treeview = tabela_fixingstermo_b2b
    elif "Fixings Opção Cliente" in tab_name:
        treeview = tabela_fixingsopcao_cliente
    elif "Fixings Opção B2B" in tab_name:
        treeview = tabela_fixingsopcao_b2b
    else:
        return  # No valid tab name, exit function

    if "Termo" in tab_name:
        treeview_deals = tabela_termo_cliente
        treeview_deals_b2b = tabela_termo_b2b
        treeview_file = tabela_arquivotermo_cliente
        treeview_file_b2b = tabela_arquivotermo_b2b
        treeview_fixings = tabela_fixingstermo_cliente
        treeview_fixings_b2b = tabela_fixingstermo_b2b
        
    elif "Opção" in tab_name:
        treeview_deals = tabela_opcao_cliente
        treeview_deals_b2b = tabela_opcao_b2b
        treeview_file = tabela_arquivoopcao_cliente
        treeview_file_b2b = tabela_arquivoopcao_b2b
        treeview_fixings = tabela_fixingsopcao_cliente
        treeview_fixings_b2b = tabela_fixingsopcao_b2b
        
    identifiers = set()
    # Export data from the active treeview
    selected_items = treeview.selection()
    if selected_items:
        for item in selected_items:
            row_data = treeview.item(item, 'values')
            identifier = str(row_data[-4])
            identifiers.add(identifier)

    identifiers = list(identifiers)

    # Create checkboxes for each column in the respective frames
    create_checkboxes(frame_deals_info, colunas_deals, checkbox_states_deals)
    create_checkboxes(frame_file_info, colunas_file, checkbox_states_file)
    create_checkboxes(frame_fixing_info, colunas_fixings, checkbox_states_fixings)

    # Create a frame for the export button
    frame_botao = ctk.CTkFrame(export_window)
    frame_botao.grid(row=2, column=0, columnspan=3, pady=5,sticky='w')

    # Create the export button
    botao_export = ctk.CTkButton(frame_botao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EXPORT", font=fonte_botao, command=lambda: export_data(export_window, tab_name, var_b2b.get(), colunas_deals, colunas_fixings, colunas_file, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b, var_deals.get(), var_fixings.get(), var_file.get(), identifiers))
    botao_export.grid(row=0, column=0, pady=5,  padx=2, sticky='w')

def create_checkboxes(frame, colunas, checkbox_states):
    for coluna in colunas:
        var = ctk.BooleanVar(value=True)
        checkbox_states[coluna] = var
        checkbox = ctk.CTkCheckBox(frame, text=coluna, variable=var)
        checkbox.pack(anchor='w')

def toggle_checkboxes(frame, state):
    for child in frame.winfo_children():
        if isinstance(child, ctk.CTkCheckBox):
            child.deselect() if not state else child.select()

def ajustar_colunas_sheet_export(sheet, headers):
    # Adjust column widths based on content
    for col_idx, col in enumerate(sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)), start=1):
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 6)  # Adjust for spacing
        sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

def export_data(window, tab_name, include_b2b, colunas_deals, colunas_fixings, colunas_file, selected_items, treeview, treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, treeview_fixings, treeview_fixings_b2b, var_deals, var_fixings, var_file, identifiers):
    # Gather selected columns based on checkbox states
    selected_deals_columns = [col for col in checkbox_states_deals if checkbox_states_deals[col].get()]
    selected_fixings_columns = [col for col in checkbox_states_fixings if checkbox_states_fixings[col].get()]
    selected_file_columns = [col for col in checkbox_states_file if checkbox_states_file[col].get()]

    # Create a new Excel workbook
    workbook = Workbook()
    # Remove the default sheet created automatically
    default_sheet = workbook.active
    workbook.remove(default_sheet)    
   
    # Deals Info
    if selected_items and var_deals and treeview_deals.get_children():
        deals_sheet = workbook.create_sheet(title="Deals Info")
        deals_sheet.append(selected_deals_columns)
        for item in treeview_deals.get_children():
            row_data_deals = treeview_deals.item(item, 'values')
            filtered_data_deals = [row_data_deals[colunas_deals.index(col)] for col in selected_deals_columns if col in colunas_deals and colunas_deals.index(col) < len(row_data_deals)]
            if str(row_data_deals[-4]) in identifiers:
                deals_sheet.append(filtered_data_deals)        
        if include_b2b and treeview_deals_b2b.get_children():
            for item in treeview_deals_b2b.get_children():
                row_data_deals_b2b = treeview_deals_b2b.item(item, 'values')
                filtered_data_deals_b2b = [row_data_deals_b2b[colunas_deals.index(col)] for col in selected_deals_columns if col in colunas_deals and colunas_deals.index(col) < len(row_data_deals)]
                if str(row_data_deals_b2b[-4]) in identifiers:
                    deals_sheet.append(filtered_data_deals_b2b)
        ajustar_colunas_sheet_export(deals_sheet, selected_deals_columns)         


    # File Info
    if selected_items and var_file and treeview_file.get_children() :
        file_sheet = workbook.create_sheet(title="File Info")
        file_sheet.append(selected_file_columns)
        for item in treeview_file.get_children():
            row_data_file = treeview_file.item(item, 'values')
            filtered_data_file = [row_data_file[colunas_file.index(col)] for col in selected_file_columns if col in colunas_file and colunas_file.index(col) < len(row_data_file)]
            if str(row_data_file[-4]) in identifiers:
                file_sheet.append(filtered_data_file)
        if include_b2b and treeview_file_b2b.get_children():
            for item in treeview_file_b2b.get_children():
                row_data_file_b2b = treeview_file_b2b.item(item, 'values')
                filtered_data_file_b2b = [row_data_file_b2b[colunas_file.index(col)] for col in selected_file_columns if col in colunas_file and colunas_file.index(col) < len(row_data_file)]
                if str(row_data_file_b2b[-4]) in identifiers:
                    file_sheet.append(filtered_data_file_b2b)                
        ajustar_colunas_sheet_export(file_sheet, selected_file_columns)

    # Fixings Info
    indentifiers_fixings = []
    indentifiers_fixings_b2b = []
    if selected_items and var_fixings and treeview_fixings.get_children():    
        for item in treeview_fixings.get_children():
            row_data_fixings = treeview_fixings.item(item, 'values')
            filtered_data_fixings = [row_data_fixings[colunas_fixings.index(col)] for col in selected_fixings_columns if col in colunas_fixings and colunas_fixings.index(col) < len(row_data_fixings)]
            if str(row_data_fixings[-4]) in identifiers:
                indentifiers_fixings.append(str(row_data_fixings[-4]))                           
        if include_b2b and treeview_fixings_b2b.get_children():
            for item in treeview_fixings_b2b.get_children():
                row_data_fixings_b2b = treeview_fixings_b2b.item(item, 'values')
                filtered_data_fixings_b2b = [row_data_fixings[colunas_fixings.index(col)] for col in selected_fixings_columns if col in colunas_fixings and colunas_fixings.index(col) < len(row_data_fixings)]
                if str(row_data_fixings_b2b[-4]) in identifiers:
                    indentifiers_fixings_b2b.append(str(row_data_fixings_b2b[-4]))                    
        if indentifiers_fixings or indentifiers_fixings_b2b:
            fixings_sheet = workbook.create_sheet(title="Fixings Info")
            fixings_sheet.append(selected_fixings_columns)
            fixings_sheet.append(filtered_data_fixings)
            fixings_sheet.append(filtered_data_fixings_b2b)
            ajustar_colunas_sheet_export(fixings_sheet, selected_fixings_columns)

    
                
    # Save the workbook to a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    workbook.save(temp_file.name)
    temp_file.close()

    # Open the Excel file
    os.startfile(temp_file.name)

    # Close the export window
    window.destroy()


def open_search_window():
    global janela
    # Cria um search_janelalevel para a janela de consulta     
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")
    # Cria uma nova janela search_janelalevel
    search_janela = ctk.CTkToplevel(janela)    
    search_janela.title("CommodiXChange Global Search")
    search_janela.iconbitmap(os.path.join(r"I:\Confirmation\Derivativos\Movimento\Liquidações do Dia\X_icone.ico"))

    # Traz a janela search_janelalevel para o primeiro plano
    search_janela.lift()
    search_janela.focus_set()
    search_janela.grab_set()  # Opcional: impede interação com a janela principal até que a search_janelalevel seja fechada   

    # Define o tamanho da janela
    search_janela.geometry("600x300")

    # Menu suspenso para "Search what?"
    search_label = ctk.CTkLabel(search_janela, text="Search what?")
    search_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

    search_options = ["All", "Swap", "Option"]
    search_combobox = ctk.CTkComboBox(search_janela, values=search_options)
    search_combobox.set("All")
    search_combobox.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    # Variáveis para as caixas de seleção
    include_dead_deals_var = tk.BooleanVar(value=True)
    economic_info_var = tk.BooleanVar(value=True)

    # Caixas de seleção
    include_dead_deals = ctk.CTkCheckBox(search_janela, text="Include dead deals", variable=include_dead_deals_var)
    include_dead_deals.grid(row=0, column=2, padx=5, pady=5, sticky="w")

    economic_info = ctk.CTkCheckBox(search_janela, text="Economic Info only", variable=economic_info_var)
    economic_info.grid(row=0, column=3, padx=5, pady=5, sticky="w")

    # Área de texto para resultados usando CTkTextbox
    results_textbox = ctk.CTkTextbox(search_janela, height=10, width=50)
    results_textbox.grid(row=1, column=0, columnspan=7, padx=5, pady=5, sticky="nsew")
    
    treeviews = {
        'tabela_termo_cliente': tabela_termo_cliente,
        'tabela_termo_b2b': tabela_termo_b2b,
        'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
        'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
        'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
        'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b,
        'tabela_opcao_cliente': tabela_opcao_cliente,
        'tabela_opcao_b2b': tabela_opcao_b2b,
        'tabela_arquivoopcao_cliente': tabela_arquivoopcao_cliente,
        'tabela_arquivoopcao_b2b': tabela_arquivoopcao_b2b,
        'tabela_fixingsopcao_cliente': tabela_fixingsopcao_cliente,
        'tabela_fixingsopcao_b2b': tabela_fixingsopcao_b2b        
    }

    # Botões de ação 
    fonte_botao = ctk.CTkFont(family="League Spartan", size=13, weight="bold")      
    # Modifique a chamada do botão "Find" para usar a nova função
    find_button = ctk.CTkButton(
        search_janela, text="Find", width=100, height=26, corner_radius=8, fg_color="#5A5368", font=fonte_botao,
        command=lambda: on_run_athenaid_search_click(search_janela, search_combobox, results_textbox, treeviews))
            
              # Dicionário de treeviews que você deve definir
       
    find_button.grid(row=2, column=0, padx=5, pady=5, sticky="w")
    find_button.grid(row=2, column=0, padx=5, pady=5, sticky="w")

    cancel_button = ctk.CTkButton(search_janela, width=100, height=26, corner_radius=8, fg_color="#5A5368", font=fonte_botao, text="Cancel", command=search_janela.destroy)
    cancel_button.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    # Área de texto para resultados usando CTkTextbox
    results_textbox = ctk.CTkTextbox(search_janela, height=10, width=50)
    results_textbox.grid(row=1, column=0, columnspan=7, padx=5, pady=5, sticky="nsew")
    
    

    # Configuração de redimensionamento
    search_janela.grid_rowconfigure(1, weight=1)
    search_janela.grid_columnconfigure(6, weight=1)


        
def execute_load_query_sqlite( query_text: str = None) -> Dict[str, List[List[Any]]]:   
    
    """
    Executa uma query no banco de dados SQLite com base nos filtros fornecidos
    e distribui os resultados em múltiplas tabelas, incluindo dados relacionados de outras tabelas.

    Args:
        conn: Caminho para o arquivo do banco de dados SQLite
        trade_date_filter: Dicionário com tipo de filtro e valor(es) para Trade Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        settlement_date_filter: Dicionário com tipo de filtro e valor(es) para Settlement Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        fixing_commodity_filter: Dicionário com tipo de filtro e valor para Fixing Commodity
            Formato: {'type': 'is', 'value': str}
        fixing_ccy_filter: Dicionário com tipo de filtro e valor para Fixing CCY
            Formato: {'type': 'is', 'value': str}
        counterparty_filter: Valor para filtrar Counterparty
        notional_filter: Valor para filtrar Notional
        strike_filter: Valor para filtrar Strike
        athena_id_filter: Valor para filtrar Athena ID
        status_filter: Tipo de filtro para Status ('Pending', 'OK', 'All')

    Returns:
        Dicionário contendo seis listas:
        - 'tabela_termo_cliente': Registros de termo_base_deals onde Counterparty != 'Lawton'
        - 'tabela_termo_b2b': Registros de termo_base_deals onde Counterparty == 'Lawton'
        - 'tabela_arquivotermo_cliente': Registros de termo_base_file relacionados com tabela_termo_cliente
        - 'tabela_arquivotermo_b2b': Registros de termo_base_file relacionados com tabela_termo_b2b
        - 'tabela_fixingstermo_cliente': Registros de termo_base_fixings relacionados com tabela_termo_cliente
        - 'tabela_fixingstermo_b2b': Registros de termo_base_fixings relacionados com tabela_termo_b2b
    """
    # Separar a consulta SQL e os parâmetros        
 
    
    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
    
    # Construir a consulta SQL base    
    params = []    
    query, params = query_text.split(" , ")
    params = eval(params)  # Converte a string de parâmetros de volta para uma lista/tupla
    
    # Executar a consulta
    cursor.execute(query, params)
    filtered_data = cursor.fetchall()
    
    # Distribuir dados filtrados para as tabelas principais
    tabela_termo_cliente = []
    tabela_termo_b2b = []
    
    # Coletar identificadores para busca em tabelas relacionadas
    cliente_identifiers = []
    b2b_identifiers = []
    id_cliente = []
    id_b2b = []
    contraparte = "00041007"
    for row in filtered_data:
        # Obter o valor do campo "Identifier" (índice -4)
        identifier = row[-4] if len(row) >= 4 else None
        id = row[0]
            
        if row[18] == 'Lawton':
            tabela_termo_b2b.append(row)
            if identifier:
                b2b_identifiers.append(identifier)
                id_b2b.append(id)
        else:
            tabela_termo_cliente.append(row)
            if identifier:
                cliente_identifiers.append(identifier)
                id_cliente.append(id)
    
    # Buscar dados relacionados na tabela termo_base_file
    tabela_arquivotermo_cliente = []
    tabela_arquivotermo_b2b = []
    
    if cliente_identifiers:
        placeholders = ', '.join(['?' for _ in cliente_identifiers])        
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND Contraparte <> ?"
        cursor.execute(query, cliente_identifiers + [contraparte])
        tabela_arquivotermo_cliente = cursor.fetchall()
    
    if b2b_identifiers:
        placeholders = ', '.join(['?' for _ in b2b_identifiers])
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND (Contraparte = ? OR Observaçao = ?)"
        cursor.execute(query, b2b_identifiers + [contraparte, contraparte])
        tabela_arquivotermo_b2b = cursor.fetchall()
    
    # Buscar dados relacionados na tabela termo_base_fixings
    tabela_fixingstermo_cliente = []
    tabela_fixingstermo_b2b = []
    
    if cliente_identifiers:
        placeholders_identifiers = ', '.join(['?' for _ in cliente_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_cliente])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, cliente_identifiers + id_cliente)
        tabela_fixingstermo_cliente = cursor.fetchall()
    
    if b2b_identifiers:        
        placeholders_identifiers = ', '.join(['?' for _ in b2b_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_b2b])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"        
        cursor.execute(query, b2b_identifiers + id_b2b)
        tabela_fixingstermo_b2b = cursor.fetchall()
    
    # Fechar a conexão com o banco de dados
    conn.commit()
    conn.close()
    
    # Retornar todas as tabelas em um dicionário
    return {
        'tabela_termo_cliente': tabela_termo_cliente,
        'tabela_termo_b2b': tabela_termo_b2b,
        'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
        'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
        'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
        'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b
    }

def create_deal_option_query_window():
    global janela    
    global deal_query_window_active
    global deal_option_query_window_active
    deal_query_window_active = False
    deal_option_query_window_active = True
    # Cria um Toplevel para a janela de consulta     
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")

    query_janela = ctk.CTkToplevel(janela)    
    query_janela.title("Deal Query")
    query_janela.iconbitmap(os.path.join(r"I:\Confirmation\Derivativos\Movimento\Liquidações do Dia\X_icone.ico"))

    # Traz a janela Toplevel para o primeiro plano
    query_janela.lift()
    query_janela.focus_set()
    query_janela.grab_set()  # Opcional: impede interação com a janela principal até que a Toplevel seja fechada

    # Frame para os filtros de consulta
    frame_filters = ctk.CTkFrame(query_janela)
    frame_filters.pack(side="top", fill="both", expand=True, padx=10, pady=5)

    # Variáveis para armazenar as seleções dos OptionMenus
    trade_date_option = tk.StringVar(value="is")
    settlement_date_option = tk.StringVar(value="is")
    fixing_commodity_option = tk.StringVar(value="is")
    fixing_ccy_option = tk.StringVar(value="is")
    status_option = tk.StringVar(value="All")  # Variável para o novo OptionMenu
    
    # Inicialização das entradas de data
    date_entry1_trade_date = None
    date_entry2_trade_date = None 
    date_entry1_settlement_date = None
    date_entry2_settlement_date = None
    date_entry1_fixing_commodity = None
    date_entry2_fixing_commodity = None
    date_entry1_fixing_ccy = None
    date_entry2_fixing_ccy = None

    # Função para limpar todas as entradas e comboboxes
    def clear_entries():
        for widget in frame_filters.winfo_children():
            if isinstance(widget, ctk.CTkEntry) or isinstance(widget, ttk.Combobox):
                widget.delete(0, tk.END)
            elif isinstance(widget, DateEntry):
                widget.delete(0, tk.END)  # Limpa a entrada de data
            elif isinstance(widget, ctk.CTkOptionMenu):
                widget.set('')  # Reseta o OptionMenu

    # Função para fechar a janela Toplevel
    def cancel():
        query_janela.destroy()

    # Função para atualizar as entradas de data para Trade Date
    def update_trade_date_entries(option):
        nonlocal option_menu_trade_date, date_entry1_trade_date, date_entry2_trade_date
        for widget in frame_filters.grid_slaves(row=0, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=0, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=0, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=0, column=4):
            widget.destroy()

        option_menu_trade_date = ctk.CTkOptionMenu(frame_filters, variable=trade_date_option, values=["is", "after", "before", "is between"], width=50, command=update_trade_date_entries)
        option_menu_trade_date.grid(row=0, column=1, padx=5, pady=5)

        date_entry1_trade_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_trade_date.grid(row=0, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=0, column=3, padx=5, pady=5)
            date_entry2_trade_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_trade_date.grid(row=0, column=4, padx=5, pady=5)

    # Função para atualizar as entradas de data para Settlement Date
    def update_settlement_date_entries(option):
        nonlocal option_menu_settlement_date, date_entry1_settlement_date, date_entry2_settlement_date
        for widget in frame_filters.grid_slaves(row=1, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=1, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=1, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=1, column=4):
            widget.destroy()

        option_menu_settlement_date = ctk.CTkOptionMenu(frame_filters, variable=settlement_date_option, values=["is", "after", "before", "is between"], width=50, command=update_settlement_date_entries)
        option_menu_settlement_date.grid(row=1, column=1, padx=5, pady=5)

        date_entry1_settlement_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_settlement_date.delete(0, "end")
        date_entry1_settlement_date.grid(row=1, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=1, column=3, padx=5, pady=5)
            date_entry2_settlement_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_settlement_date.delete(0, "end")
            date_entry2_settlement_date.grid(row=1, column=4, padx=5, pady=5)

    # Função para atualizar as entradas de data para Fixing Commodity
    def update_fixing_commodity_entries(option):
        nonlocal option_menu_fixing_commodity, date_entry1_fixing_commodity, date_entry2_fixing_commodity
        for widget in frame_filters.grid_slaves(row=2, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=2, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=2, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=2, column=4):
            widget.destroy()

        option_menu_fixing_commodity = ctk.CTkOptionMenu(frame_filters, variable=fixing_commodity_option, values=["is", "after", "before", "is between"], width=50, command=update_fixing_commodity_entries)
        option_menu_fixing_commodity.grid(row=2, column=1, padx=5, pady=5)

        date_entry1_fixing_commodity = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_fixing_commodity.delete(0, "end")
        date_entry1_fixing_commodity.grid(row=2, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=2, column=3, padx=5, pady=5)
            date_entry2_fixing_commodity = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_fixing_commodity.delete(0, "end")
            date_entry2_fixing_commodity.grid(row=2, column=4, padx=5, pady=5)

    # Função para atualizar as entradas de data para Fixing CCY
    def update_fixing_ccy_entries(option):
        nonlocal option_menu_fixing_ccy, date_entry1_fixing_ccy, date_entry2_fixing_ccy
        for widget in frame_filters.grid_slaves(row=3, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=3, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=3, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=3, column=4):
            widget.destroy()

        option_menu_fixing_ccy = ctk.CTkOptionMenu(frame_filters, variable=fixing_ccy_option, values=["is", "after", "before", "is between"], width=50, command=update_fixing_ccy_entries)
        option_menu_fixing_ccy.grid(row=3, column=1, padx=5, pady=5)

        date_entry1_fixing_ccy = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_fixing_ccy.delete(0, "end")
        date_entry1_fixing_ccy.grid(row=3, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=3, column=3, padx=5, pady=5)
            date_entry2_fixing_ccy = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_fixing_ccy.delete(0, "end")
            date_entry2_fixing_ccy.grid(row=3, column=4, padx=5, pady=5)

    # Filtro específico para "Trade Date"
    label_trade_date = ctk.CTkLabel(frame_filters, text="Trade Date")
    label_trade_date.grid(row=0, column=0, padx=5, pady=5)

    # Menu de opções para "Trade Date"
    option_menu_trade_date = ctk.CTkOptionMenu(frame_filters, variable=trade_date_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_trade_date_entries)
    option_menu_trade_date.grid(row=0, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_trade_date_entries("is")

    # Filtro específico para "Settlement Date"
    label_settlement_date = ctk.CTkLabel(frame_filters, text="Settlement Date")
    label_settlement_date.grid(row=1, column=0, padx=5, pady=5)

    # Menu de opções para "Settlement Date"
    option_menu_settlement_date = ctk.CTkOptionMenu(frame_filters, variable=settlement_date_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_settlement_date_entries)
    option_menu_settlement_date.grid(row=1, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_settlement_date_entries("is")

    # Filtro específico para "Fixing Commodity"
    label_fixing_commodity = ctk.CTkLabel(frame_filters, text="Fixing Commodity")
    label_fixing_commodity.grid(row=2, column=0, padx=5, pady=5)

    # Menu de opções para "Fixing Commodity"
    option_menu_fixing_commodity = ctk.CTkOptionMenu(frame_filters, variable=fixing_commodity_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_fixing_commodity_entries)
    option_menu_fixing_commodity.grid(row=2, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_fixing_commodity_entries("is")

    # Filtro específico para "Fixing CCY"
    label_fixing_ccy = ctk.CTkLabel(frame_filters, text="Fixing CCY")
    label_fixing_ccy.grid(row=3, column=0, padx=5, pady=5)

    # Menu de opções para "Fixing CCY"
    option_menu_fixing_ccy = ctk.CTkOptionMenu(frame_filters, variable=fixing_ccy_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_fixing_ccy_entries)
    option_menu_fixing_ccy.grid(row=3, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_fixing_ccy_entries("is")

    # Combobox para "Counterparty"
    label_counterparty = ctk.CTkLabel(frame_filters, text="Counterparty")
    label_counterparty.grid(row=4, column=0, padx=5, pady=5)
    cntpy_accronym = carregar_counterparty_combobox_opcao()
    counterparty_combobox_opcao = ctk.CTkComboBox(frame_filters, width=180, dropdown_fg_color ="white", button_color="#5A5368")
    CTkScrollableDropdown(counterparty_combobox_opcao, values=cntpy_accronym, justify="left", autocomplete=True)
    counterparty_combobox_opcao.grid(row=4, column=1, padx=5, pady=5)
    counterparty_combobox_opcao.set("")

    # Entry para "Notional"
    label_notional = ctk.CTkLabel(frame_filters, text="Notional")
    label_notional.grid(row=5, column=0, padx=5, pady=5)
    notional_entry = ctk.CTkEntry(frame_filters)
    notional_entry.grid(row=5, column=1, padx=5, pady=5)

    # Entry para "Strike"
    label_strike = ctk.CTkLabel(frame_filters, text="Strike")
    label_strike.grid(row=6, column=0, padx=5, pady=5)
    strike_entry = ctk.CTkEntry(frame_filters)
    strike_entry.grid(row=6, column=1, padx=5, pady=5)

    # Entry para "Athena ID"
    label_athena_id = ctk.CTkLabel(frame_filters, text="Athena ID")
    label_athena_id.grid(row=7, column=0, padx=5, pady=5)
    athena_id_entry = ctk.CTkEntry(frame_filters)
    athena_id_entry.grid(row=7, column=1, padx=5, pady=5)

    # OptionMenu para "Status"
    label_status = ctk.CTkLabel(frame_filters, text="Status")
    label_status.grid(row=8, column=0, padx=5, pady=5)
    status_option_menu = ctk.CTkOptionMenu(frame_filters, variable=status_option, values=["Pending", "Ok", "All"], width=50)
    status_option_menu.grid(row=8, column=1, padx=5, pady=5)

    # Frame para os botões de ação
    frame_actions = ctk.CTkFrame(query_janela)
    frame_actions.pack(side="bottom", fill="x", padx=10, pady=5)
    treeviews = {
        'tabela_opcao_cliente': tabela_opcao_cliente,
        'tabela_opcao_b2b': tabela_opcao_b2b,
        'tabela_arquivoopcao_cliente': tabela_arquivoopcao_cliente,
        'tabela_arquivoopcao_b2b': tabela_arquivoopcao_b2b,
        'tabela_fixingsopcao_cliente': tabela_fixingsopcao_cliente,
        'tabela_fixingsopcao_b2b': tabela_fixingsopcao_b2b
    }

    fonte_botao = ctk.CTkFont(family="League Spartan", size=13, weight="bold")
    # Botões de ação ,  width=100, height=26, corner_radius=8, fg_color="#5A5368", , font=fonte_botao
    ctk.CTkButton(frame_actions,  width=100, height=26, corner_radius=8, fg_color="#5A5368", text="Clear", font=fonte_botao, command=clear_entries).pack(side="left", padx=5)
    ctk.CTkButton(frame_actions,  width=100, height=26, corner_radius=8, fg_color="#5A5368", text="Run (F9)",font=fonte_botao, command= lambda: on_run_button_option_click(query_janela,    
    option_menu_trade_date,
    date_entry1_trade_date,
    date_entry2_trade_date,
    option_menu_settlement_date,
    date_entry1_settlement_date,
    date_entry2_settlement_date,
    option_menu_fixing_commodity,
    date_entry1_fixing_commodity,
    date_entry2_fixing_commodity,
    fixing_ccy_option,
    date_entry1_fixing_ccy,
    date_entry2_fixing_ccy,
    counterparty_combobox_opcao,
    notional_entry,
    strike_entry,
    athena_id_entry,
    status_option,
    treeviews
    )).pack(side="left", padx=5)    
    #run_button.configure(command=on_run_button_click)
    ctk.CTkButton(frame_actions,  width=100, height=26, corner_radius=8, fg_color="#5A5368", text="Cancel", font=fonte_botao, command=cancel).pack(side="left", padx=5)  
    
def create_deal_query_window():
    global janela    
    global deal_query_window_active
    global deal_option_query_window_active
    deal_query_window_active = True
    deal_option_query_window_active = False
    # Cria um Toplevel para a janela de consulta     
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")

    query_janela = ctk.CTkToplevel(janela)    
    query_janela.title("Deal Query")
    query_janela.iconbitmap(os.path.join(r"I:\Confirmation\Derivativos\Movimento\Liquidações do Dia\X_icone.ico"))

    # Traz a janela Toplevel para o primeiro plano
    query_janela.lift()
    query_janela.focus_set()
    query_janela.grab_set()  # Opcional: impede interação com a janela principal até que a Toplevel seja fechada

    # Frame para os filtros de consulta
    frame_filters = ctk.CTkFrame(query_janela)
    frame_filters.pack(side="top", fill="both", expand=True, padx=10, pady=5)

    # Variáveis para armazenar as seleções dos OptionMenus
    trade_date_option = tk.StringVar(value="is")
    settlement_date_option = tk.StringVar(value="is")
    fixing_commodity_option = tk.StringVar(value="is")
    fixing_ccy_option = tk.StringVar(value="is")
    status_option = tk.StringVar(value="All")  # Variável para o novo OptionMenu
    
    # Inicialização das entradas de data
    date_entry1_trade_date = None
    date_entry2_trade_date = None 
    date_entry1_settlement_date = None
    date_entry2_settlement_date = None
    date_entry1_fixing_commodity = None
    date_entry2_fixing_commodity = None
    date_entry1_fixing_ccy = None
    date_entry2_fixing_ccy = None

    # Função para limpar todas as entradas e comboboxes
    def clear_entries():
        for widget in frame_filters.winfo_children():
            if isinstance(widget, ctk.CTkEntry) or isinstance(widget, ttk.Combobox):
                widget.delete(0, tk.END)
            elif isinstance(widget, DateEntry):
                widget.delete(0, tk.END)  # Limpa a entrada de data
            elif isinstance(widget, ctk.CTkOptionMenu):
                widget.set('')  # Reseta o OptionMenu

    # Função para fechar a janela Toplevel
    def cancel():
        query_janela.destroy()

    # Função para atualizar as entradas de data para Trade Date
    def update_trade_date_entries(option):
        nonlocal option_menu_trade_date, date_entry1_trade_date, date_entry2_trade_date
        for widget in frame_filters.grid_slaves(row=0, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=0, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=0, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=0, column=4):
            widget.destroy()

        option_menu_trade_date = ctk.CTkOptionMenu(frame_filters, variable=trade_date_option, values=["is", "after", "before", "is between"], width=50, command=update_trade_date_entries)
        option_menu_trade_date.grid(row=0, column=1, padx=5, pady=5)

        date_entry1_trade_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_trade_date.grid(row=0, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=0, column=3, padx=5, pady=5)
            date_entry2_trade_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_trade_date.grid(row=0, column=4, padx=5, pady=5)

    # Função para atualizar as entradas de data para Settlement Date
    def update_settlement_date_entries(option):
        nonlocal option_menu_settlement_date, date_entry1_settlement_date, date_entry2_settlement_date
        for widget in frame_filters.grid_slaves(row=1, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=1, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=1, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=1, column=4):
            widget.destroy()

        option_menu_settlement_date = ctk.CTkOptionMenu(frame_filters, variable=settlement_date_option, values=["is", "after", "before", "is between"], width=50, command=update_settlement_date_entries)
        option_menu_settlement_date.grid(row=1, column=1, padx=5, pady=5)

        date_entry1_settlement_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_settlement_date.delete(0, "end")
        date_entry1_settlement_date.grid(row=1, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=1, column=3, padx=5, pady=5)
            date_entry2_settlement_date = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_settlement_date.delete(0, "end")
            date_entry2_settlement_date.grid(row=1, column=4, padx=5, pady=5)

    # Função para atualizar as entradas de data para Fixing Commodity
    def update_fixing_commodity_entries(option):
        nonlocal option_menu_fixing_commodity, date_entry1_fixing_commodity, date_entry2_fixing_commodity
        for widget in frame_filters.grid_slaves(row=2, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=2, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=2, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=2, column=4):
            widget.destroy()

        option_menu_fixing_commodity = ctk.CTkOptionMenu(frame_filters, variable=fixing_commodity_option, values=["is", "after", "before", "is between"], width=50, command=update_fixing_commodity_entries)
        option_menu_fixing_commodity.grid(row=2, column=1, padx=5, pady=5)

        date_entry1_fixing_commodity = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_fixing_commodity.delete(0, "end")
        date_entry1_fixing_commodity.grid(row=2, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=2, column=3, padx=5, pady=5)
            date_entry2_fixing_commodity = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_fixing_commodity.delete(0, "end")
            date_entry2_fixing_commodity.grid(row=2, column=4, padx=5, pady=5)

    # Função para atualizar as entradas de data para Fixing CCY
    def update_fixing_ccy_entries(option):
        nonlocal option_menu_fixing_ccy, date_entry1_fixing_ccy, date_entry2_fixing_ccy
        for widget in frame_filters.grid_slaves(row=3, column=1):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=3, column=2):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=3, column=3):
            widget.destroy()
        for widget in frame_filters.grid_slaves(row=3, column=4):
            widget.destroy()

        option_menu_fixing_ccy = ctk.CTkOptionMenu(frame_filters, variable=fixing_ccy_option, values=["is", "after", "before", "is between"], width=50, command=update_fixing_ccy_entries)
        option_menu_fixing_ccy.grid(row=3, column=1, padx=5, pady=5)

        date_entry1_fixing_ccy = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
        date_entry1_fixing_ccy.delete(0, "end")
        date_entry1_fixing_ccy.grid(row=3, column=2, padx=5, pady=5)

        if option == "is between":
            ctk.CTkLabel(frame_filters, text="and").grid(row=3, column=3, padx=5, pady=5)
            date_entry2_fixing_ccy = DateEntry(frame_filters, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd-MM-yyyy')
            date_entry2_fixing_ccy.delete(0, "end")
            date_entry2_fixing_ccy.grid(row=3, column=4, padx=5, pady=5)

    # Filtro específico para "Trade Date"
    label_trade_date = ctk.CTkLabel(frame_filters, text="Trade Date")
    label_trade_date.grid(row=0, column=0, padx=5, pady=5)

    # Menu de opções para "Trade Date"
    option_menu_trade_date = ctk.CTkOptionMenu(frame_filters, variable=trade_date_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_trade_date_entries)
    option_menu_trade_date.grid(row=0, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_trade_date_entries("is")

    # Filtro específico para "Settlement Date"
    label_settlement_date = ctk.CTkLabel(frame_filters, text="Settlement Date")
    label_settlement_date.grid(row=1, column=0, padx=5, pady=5)

    # Menu de opções para "Settlement Date"
    option_menu_settlement_date = ctk.CTkOptionMenu(frame_filters, variable=settlement_date_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_settlement_date_entries)
    option_menu_settlement_date.grid(row=1, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_settlement_date_entries("is")

    # Filtro específico para "Fixing Commodity"
    label_fixing_commodity = ctk.CTkLabel(frame_filters, text="Fixing Commodity")
    label_fixing_commodity.grid(row=2, column=0, padx=5, pady=5)

    # Menu de opções para "Fixing Commodity"
    option_menu_fixing_commodity = ctk.CTkOptionMenu(frame_filters, variable=fixing_commodity_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_fixing_commodity_entries)
    option_menu_fixing_commodity.grid(row=2, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_fixing_commodity_entries("is")

    # Filtro específico para "Fixing CCY"
    label_fixing_ccy = ctk.CTkLabel(frame_filters, text="Fixing CCY")
    label_fixing_ccy.grid(row=3, column=0, padx=5, pady=5)

    # Menu de opções para "Fixing CCY"
    option_menu_fixing_ccy = ctk.CTkOptionMenu(frame_filters, variable=fixing_ccy_option, values=["is", "after", "before", "is between"], button_color="#5A5368", width=50, command=update_fixing_ccy_entries)
    option_menu_fixing_ccy.grid(row=3, column=1, padx=5, pady=5)

    # Inicializa com a opção padrão
    update_fixing_ccy_entries("is")

    # Combobox para "Counterparty"
    label_counterparty = ctk.CTkLabel(frame_filters, text="Counterparty")
    label_counterparty.grid(row=4, column=0, padx=5, pady=5)
    cntpy_accronym = carregar_counterparty_combobox_termo()
    counterparty_combobox_termo = ctk.CTkComboBox(frame_filters, width=180, dropdown_fg_color ="white", button_color="#5A5368")
    CTkScrollableDropdown(counterparty_combobox_termo, values=cntpy_accronym, justify="left", autocomplete=True)
    counterparty_combobox_termo.grid(row=4, column=1, padx=5, pady=5)
    counterparty_combobox_termo.set("")

    # Entry para "Notional"
    label_notional = ctk.CTkLabel(frame_filters, text="Notional")
    label_notional.grid(row=5, column=0, padx=5, pady=5)
    notional_entry = ctk.CTkEntry(frame_filters)
    notional_entry.grid(row=5, column=1, padx=5, pady=5)

    # Entry para "Strike"
    label_strike = ctk.CTkLabel(frame_filters, text="Strike")
    label_strike.grid(row=6, column=0, padx=5, pady=5)
    strike_entry = ctk.CTkEntry(frame_filters)
    strike_entry.grid(row=6, column=1, padx=5, pady=5)

    # Entry para "Athena ID"
    label_athena_id = ctk.CTkLabel(frame_filters, text="Athena ID")
    label_athena_id.grid(row=7, column=0, padx=5, pady=5)
    athena_id_entry = ctk.CTkEntry(frame_filters)
    athena_id_entry.grid(row=7, column=1, padx=5, pady=5)

    # OptionMenu para "Status"
    label_status = ctk.CTkLabel(frame_filters, text="Status")
    label_status.grid(row=8, column=0, padx=5, pady=5)
    status_option_menu = ctk.CTkOptionMenu(frame_filters, variable=status_option, values=["Pending", "Ok", "All"], width=50)
    status_option_menu.grid(row=8, column=1, padx=5, pady=5)

    # Frame para os botões de ação
    frame_actions = ctk.CTkFrame(query_janela)
    frame_actions.pack(side="bottom", fill="x", padx=10, pady=5)
    treeviews = {
        'tabela_termo_cliente': tabela_termo_cliente,
        'tabela_termo_b2b': tabela_termo_b2b,
        'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
        'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
        'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
        'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b
    }

    fonte_botao = ctk.CTkFont(family="League Spartan", size=13, weight="bold")
    # Botões de ação ,  width=100, height=26, corner_radius=8, fg_color="#5A5368", , font=fonte_botao
    ctk.CTkButton(frame_actions,  width=100, height=26, corner_radius=8, fg_color="#5A5368", text="Clear", font=fonte_botao, command=clear_entries).pack(side="left", padx=5)
    ctk.CTkButton(frame_actions,  width=100, height=26, corner_radius=8, fg_color="#5A5368", text="Run (F9)",font=fonte_botao, command= lambda: on_run_button_click(query_janela,    
    option_menu_trade_date,
    date_entry1_trade_date,
    date_entry2_trade_date,
    option_menu_settlement_date,
    date_entry1_settlement_date,
    date_entry2_settlement_date,
    option_menu_fixing_commodity,
    date_entry1_fixing_commodity,
    date_entry2_fixing_commodity,
    fixing_ccy_option,
    date_entry1_fixing_ccy,
    date_entry2_fixing_ccy,
    counterparty_combobox_termo,
    notional_entry,
    strike_entry,
    athena_id_entry,
    status_option,
    treeviews
    )).pack(side="left", padx=5)    
   
    
    #run_button.configure(command=on_run_button_click)
    ctk.CTkButton(frame_actions,  width=100, height=26, corner_radius=8, fg_color="#5A5368", text="Cancel", font=fonte_botao, command=cancel).pack(side="left", padx=5)
# Iniciar o listener para eventos de teclado

def load_query_entry(entry_query_termo, treeviews):
    # Obter o texto da entrada
    query_text = entry_query_termo.get()
    
    results = execute_load_query_sqlite(query_text)
    
    if treeviews:
        update_treeviews(results, treeviews)
       

        
def execute_load_query_sqlite( query_text: str = None) -> Dict[str, List[List[Any]]]:   
    
    """
    Executa uma query no banco de dados SQLite com base nos filtros fornecidos
    e distribui os resultados em múltiplas tabelas, incluindo dados relacionados de outras tabelas.

    Args:
        conn: Caminho para o arquivo do banco de dados SQLite
        trade_date_filter: Dicionário com tipo de filtro e valor(es) para Trade Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        settlement_date_filter: Dicionário com tipo de filtro e valor(es) para Settlement Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        fixing_commodity_filter: Dicionário com tipo de filtro e valor para Fixing Commodity
            Formato: {'type': 'is', 'value': str}
        fixing_ccy_filter: Dicionário com tipo de filtro e valor para Fixing CCY
            Formato: {'type': 'is', 'value': str}
        counterparty_filter: Valor para filtrar Counterparty
        notional_filter: Valor para filtrar Notional
        strike_filter: Valor para filtrar Strike
        athena_id_filter: Valor para filtrar Athena ID
        status_filter: Tipo de filtro para Status ('Pending', 'OK', 'All')

    Returns:
        Dicionário contendo seis listas:
        - 'tabela_termo_cliente': Registros de termo_base_deals onde Counterparty != 'Lawton'
        - 'tabela_termo_b2b': Registros de termo_base_deals onde Counterparty == 'Lawton'
        - 'tabela_arquivotermo_cliente': Registros de termo_base_file relacionados com tabela_termo_cliente
        - 'tabela_arquivotermo_b2b': Registros de termo_base_file relacionados com tabela_termo_b2b
        - 'tabela_fixingstermo_cliente': Registros de termo_base_fixings relacionados com tabela_termo_cliente
        - 'tabela_fixingstermo_b2b': Registros de termo_base_fixings relacionados com tabela_termo_b2b
    """
    # Separar a consulta SQL e os parâmetros        
 
    
    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
    
    # Construir a consulta SQL base    
    params = []    
    query, params = query_text.split(" , ")
    params = eval(params)  # Converte a string de parâmetros de volta para uma lista/tupla
    
    # Executar a consulta
    cursor.execute(query, params)
    filtered_data = cursor.fetchall()
    
    # Distribuir dados filtrados para as tabelas principais
    tabela_termo_cliente = []
    tabela_termo_b2b = []
    
    # Coletar identificadores para busca em tabelas relacionadas
    cliente_identifiers = []
    b2b_identifiers = []
    id_cliente = []
    id_b2b = []
    contraparte = "00041007"
    for row in filtered_data:
        # Obter o valor do campo "Identifier" (índice -4)
        identifier = row[-4] if len(row) >= 4 else None
        id = row[0]
            
        if row[18] == 'Lawton':
            tabela_termo_b2b.append(row)
            if identifier:
                b2b_identifiers.append(identifier)
                id_b2b.append(id)
        else:
            tabela_termo_cliente.append(row)
            if identifier:
                cliente_identifiers.append(identifier)
                id_cliente.append(id)
    
    # Buscar dados relacionados na tabela termo_base_file
    tabela_arquivotermo_cliente = []
    tabela_arquivotermo_b2b = []
    
    if cliente_identifiers:
        placeholders = ', '.join(['?' for _ in cliente_identifiers])        
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND Contraparte <> ?"
        cursor.execute(query, cliente_identifiers + [contraparte])
        tabela_arquivotermo_cliente = cursor.fetchall()
    
    if b2b_identifiers:
        placeholders = ', '.join(['?' for _ in b2b_identifiers])
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND (Contraparte = ? OR Observaçao = ?)"
        cursor.execute(query, b2b_identifiers + [contraparte, contraparte])
        tabela_arquivotermo_b2b = cursor.fetchall()
    
    # Buscar dados relacionados na tabela termo_base_fixings
    tabela_fixingstermo_cliente = []
    tabela_fixingstermo_b2b = []
    
    if cliente_identifiers:
        placeholders_identifiers = ', '.join(['?' for _ in cliente_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_cliente])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, cliente_identifiers + id_cliente)
        tabela_fixingstermo_cliente = cursor.fetchall()
    
    if b2b_identifiers:        
        placeholders_identifiers = ', '.join(['?' for _ in b2b_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_b2b])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"        
        cursor.execute(query, b2b_identifiers + id_b2b)
        tabela_fixingstermo_b2b = cursor.fetchall()
    
    # Fechar a conexão com o banco de dados
    conn.commit()
    conn.close()
    
    # Retornar todas as tabelas em um dicionário
    return {
        'tabela_termo_cliente': tabela_termo_cliente,
        'tabela_termo_b2b': tabela_termo_b2b,
        'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
        'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
        'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
        'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b
    }

    
# Chamar a função para executar a consulta
def execute_athenaid_search_query_sqlite(
    search_type: str = None,    
    ids_list: Union[str, List[str]] = None
) -> Dict[str, List[List[Any]]]:
    """
    Executes a query on the SQLite database based on the provided filters
    and distributes the results into multiple tables.

    Args:
        search_type: Type of search ('Swap', 'Option', 'All')
        include_dead_deals_var: Whether to include dead deals
        search_options: Specific instruments to search for
        economic_info_var: Whether to include only economic info
        ids_list: List of Athena IDs to filter

    Returns:
        Dictionary containing lists for each table.
    """
    
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
    
    # Initialize result containers
    tabela_termo_cliente = []
    tabela_termo_b2b = []
    tabela_opcao_cliente = []
    tabela_opcao_b2b = []
    cliente_identifiers_termo = []
    b2b_identifiers_termo = []
    cliente_identifiers_opcao = []
    b2b_identifiers_opcao = []
    id_cliente_termo = []
    id_b2b_termo = []
    id_cliente_opcao = []
    id_b2b_opcao = []
    contraparte = "00041007"
    contraparte_option = "00041.00-7"

    # Function to execute query and process results
    def execute_and_process(base_query, is_option=False):
        nonlocal tabela_termo_cliente, tabela_termo_b2b, tabela_opcao_cliente, tabela_opcao_b2b
        nonlocal tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, tabela_arquivotermo_b2b
        nonlocal tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b
        nonlocal cliente_identifiers_termo, b2b_identifiers_termo, cliente_identifiers_opcao, b2b_identifiers_opcao
        nonlocal id_cliente_termo, id_b2b_termo, id_cliente_opcao, id_b2b_opcao

        params = []       
            
        query, params = _build_query_athenaid_with_filters(
            base_query,
            params,                                           
            ids_list
        )
        cursor.execute(query, params)
        filtered_data = cursor.fetchall()
        
        for row in filtered_data:
            identifier = row[-4] if len(row) >= 4 else None
            id = row[0]
                
            if row[18] == 'Lawton':
                if is_option:
                    tabela_opcao_b2b.append(row)
                    if identifier:
                        b2b_identifiers_opcao.append(identifier)
                        id_b2b_opcao.append(id)
                else:
                    tabela_termo_b2b.append(row)
                    if identifier:
                        b2b_identifiers_termo.append(identifier)
                        id_b2b_termo.append(id)
            else:
                if is_option:
                    tabela_opcao_cliente.append(row)
                    if identifier:
                        cliente_identifiers_opcao.append(identifier)
                        id_cliente_opcao.append(id)
                else:
                    tabela_termo_cliente.append(row)
                    if identifier:
                        cliente_identifiers_termo.append(identifier)
                        id_cliente_termo.append(id)
                        
            #return id_cliente_opcao, id_cliente_termo, b2b_identifiers_opcao, b2b_identifiers_termo, id_cliente_termo, id_cliente_opcao, cliente_identifiers_opcao, cliente_identifiers_termo
    
    # Execute queries based on search type
    if search_type == {'type':'All'} or search_type == {'type':'Swap'}:
        execute_and_process("SELECT * FROM termo_base_deals WHERE 1=1")
    if search_type == {'type':'All'} or search_type == {'type':'Option'}:    
        execute_and_process("SELECT * FROM opcao_base_deals WHERE 1=1", is_option=True)

    # Search related data in the appropriate tables
    tabela_arquivotermo_cliente = []
    tabela_arquivotermo_b2b = []
    tabela_fixingstermo_cliente = []
    tabela_fixingstermo_b2b = []
    tabela_arquivoopcao_cliente = []
    tabela_arquivoopcao_b2b = []
    tabela_fixingsopcao_cliente = []
    tabela_fixingsopcao_b2b = []
    
    # Termo related data
    if cliente_identifiers_termo:
        placeholders = ', '.join(['?' for _ in cliente_identifiers_termo])
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND Contraparte <> ?"
        cursor.execute(query, cliente_identifiers_termo + [contraparte])
        tabela_arquivotermo_cliente.extend(cursor.fetchall())
    
    if b2b_identifiers_termo:
        placeholders = ', '.join(['?' for _ in b2b_identifiers_termo])
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND (Contraparte = ? OR Observaçao = ?)"
        cursor.execute(query, b2b_identifiers_termo + [contraparte, contraparte])
        tabela_arquivotermo_b2b.extend(cursor.fetchall())
    
    if cliente_identifiers_termo:
        placeholders_identifiers = ', '.join(['?' for _ in cliente_identifiers_termo])
        placeholders_ids = ', '.join(['?' for _ in id_cliente_termo])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, cliente_identifiers_termo + id_cliente_termo)
        tabela_fixingstermo_cliente.extend(cursor.fetchall())
    
    if b2b_identifiers_termo:
        placeholders_identifiers = ', '.join(['?' for _ in b2b_identifiers_termo])
        placeholders_ids = ', '.join(['?' for _ in id_b2b_termo])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, b2b_identifiers_termo + id_b2b_termo)
        tabela_fixingstermo_b2b.extend(cursor.fetchall())

    # Opcao related data
    if cliente_identifiers_opcao:
        placeholders = ', '.join(['?' for _ in cliente_identifiers_opcao])
        query = f"SELECT * FROM base_opcao_file WHERE Identifier IN ({placeholders}) AND Conta_Contraparte <> ?"
        cursor.execute(query, cliente_identifiers_opcao + [contraparte_option])
        tabela_arquivoopcao_cliente.extend(cursor.fetchall())
    
    if b2b_identifiers_opcao:
        placeholders = ', '.join(['?' for _ in b2b_identifiers_opcao])
        query = f"SELECT * FROM base_opcao_file WHERE Identifier IN ({placeholders}) AND Conta_Contraparte = ?"
        cursor.execute(query, b2b_identifiers_opcao + [contraparte_option])
        tabela_arquivoopcao_b2b.extend(cursor.fetchall())
    
    if cliente_identifiers_opcao:
        placeholders_identifiers = ', '.join(['?' for _ in cliente_identifiers_opcao])
        placeholders_ids = ', '.join(['?' for _ in id_cliente_opcao])
        query = f"SELECT * FROM base_opcao_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, cliente_identifiers_opcao + id_cliente_opcao)
        tabela_fixingsopcao_cliente.extend(cursor.fetchall())
    
    if b2b_identifiers_opcao:
        placeholders_identifiers = ', '.join(['?' for _ in b2b_identifiers_opcao])
        placeholders_ids = ', '.join(['?' for _ in id_b2b_opcao])
        query = f"SELECT * FROM base_opcao_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, b2b_identifiers_opcao + id_b2b_opcao)
        tabela_fixingsopcao_b2b.extend(cursor.fetchall())
    
    # Close the database connection
    conn.commit()
    conn.close()

    # Return all tables in a dictionary
    return {
        'tabela_termo_cliente': tabela_termo_cliente,
        'tabela_termo_b2b': tabela_termo_b2b,
        'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
        'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
        'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
        'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b,
        'tabela_opcao_cliente': tabela_opcao_cliente,
        'tabela_opcao_b2b': tabela_opcao_b2b,
        'tabela_arquivoopcao_cliente': tabela_arquivoopcao_cliente,
        'tabela_arquivoopcao_b2b': tabela_arquivoopcao_b2b,
        'tabela_fixingsopcao_cliente': tabela_fixingsopcao_cliente,
        'tabela_fixingsopcao_b2b': tabela_fixingsopcao_b2b
    }
    
def _build_query_athenaid_with_filters(
    base_query: str,
    params: List[Any],     
    ids_list: Union[str, List[str]] = None
) -> Tuple[str, List[Any]]:
    """
    Constructs the SQL query with the provided filters.

    Args:
        base_query: The base SQL query string.
        params: The list of parameters to be used in the query.
        search_type: Type of search ('Swap', 'Option', 'All')
        include_dead_deals_var: Whether to include dead deals
        search_options: Specific instruments to search for
        economic_info_var: Whether to include only economic info
        ids_list: List of Athena IDs to filter

    Returns:
        A tuple containing the constructed SQL query and the list of parameters.
    """
    query = base_query   
 
    
    # Athena ID Filter
    if ids_list:
        query += " AND [DealName] IN (" + ", ".join(["?"] * len(ids_list)) + ")"
        params.extend(ids_list)
    
    return query, params




def create_query_from_athenaid_search(
    search_type: str = None,        
    ids_list: Union[str, List[str]] = None    
) -> Dict[str, Any]:
    """
    Cria um dicionário de parâmetros para a função execute_query_sqlite a partir dos valores da interface.
    
    Args:
        trade_date_type: Tipo de filtro para Trade Date ('is', 'is_between', 'after', 'before')
        trade_date_value: Valor(es) para filtro de Trade Date
        settlement_date_type: Tipo de filtro para Settlement Date ('is', 'is_between', 'after', 'before')
        settlement_date_value: Valor(es) para filtro de Settlement Date
        fixing_commodity_type: Tipo de filtro para Fixing Commodity ('is')
        fixing_commodity_value: Valor para filtro de Fixing Commodity
        fixing_ccy_type: Tipo de filtro para Fixing CCY ('is')
        fixing_ccy_value: Valor para filtro de Fixing CCY
        counterparty: Valor para filtro de Counterparty
        notional: Valor para filtro de Notional
        strike: Valor para filtro de Strike
        athena_id: Valor para filtro de Athena ID
        status: Tipo de filtro para Status ('Pending', 'OK', 'All')
        
    Returns:
        Dicionário com parâmetros para a função execute_query_sqlite
    """
    params = {}
    
    # Filtro de Trade Date
    if search_type:
        params['search_type'] = {
            'type': search_type,            
        }   
        
    if ids_list:
        params['ids_list'] = ids_list
    
    
    return params

def run_query_from_athenaid_sqlite(    
    search_options,    
    ids_text
) -> Dict[str, List[List[Any]]]:   
   
    search_type = search_options if search_options else None
    
    # Obter valores dos widgets de entrada para tipo de Operação    
    ids_list = ids_text.split("\n") 

    # Criar parâmetros para a query
    params = create_query_from_athenaid_search(   
        search_type=search_type,                  
        ids_list=ids_list        
    )
    
    # Executar a query
    return execute_athenaid_search_query_sqlite(**params)

# Configurar o botão "Run (F9)" para chamar a função
def on_run_athenaid_search_click(window,   
    search_combobox,    
    results_textbox,    
    treeviews
):
    # Obter valores dos widgets
    values = {}
    
     
    # Obter valores dos tipos de filtro
    values['search_options'] = search_combobox.get() if search_combobox else None    
    values['ids_text'] = results_textbox.get("1.0", "end-1c").strip() if results_textbox else None
    
    
    # Executar a query
    results = run_query_from_athenaid_sqlite(        
        values['search_options'],        
        values['ids_text']       
    )

    # Atualizar as Treeviews com os resultados
    if treeviews:
        update_treeviews(results, treeviews)
     
    
    window.destroy()
    
def execute_query_sqlite(    
    trade_date_filter: Dict[str, Any] = None,
    settlement_date_filter: Dict[str, Any] = None,
    fixing_commodity_filter: Dict[str, Any] = None,
    fixing_ccy_filter: Dict[str, Any] = None,
    counterparty_filter: str = None,
    notional_filter: str = None,
    strike_filter: str = None,
    athena_id_filter: str = None,
    status_filter: str = None
) -> Dict[str, List[List[Any]]]:
    """
    Executa uma query no banco de dados SQLite com base nos filtros fornecidos
    e distribui os resultados em múltiplas tabelas, incluindo dados relacionados de outras tabelas.

    Args:        
        trade_date_filter: Dicionário com tipo de filtro e valor(es) para Trade Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        settlement_date_filter: Dicionário com tipo de filtro e valor(es) para Settlement Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        fixing_commodity_filter: Dicionário com tipo de filtro e valor para Fixing Commodity
            Formato: {'type': 'is', 'value': str}
        fixing_ccy_filter: Dicionário com tipo de filtro e valor para Fixing CCY
            Formato: {'type': 'is', 'value': str}
        counterparty_filter: Valor para filtrar Counterparty
        notional_filter: Valor para filtrar Notional
        strike_filter: Valor para filtrar Strike
        athena_id_filter: Valor para filtrar Athena ID
        status_filter: Tipo de filtro para Status ('Pending', 'OK', 'All')

    Returns:
        Dicionário contendo seis listas:
        - 'tabela_termo_cliente': Registros de termo_base_deals onde Counterparty != 'Lawton'
        - 'tabela_termo_b2b': Registros de termo_base_deals onde Counterparty == 'Lawton'
        - 'tabela_arquivotermo_cliente': Registros de termo_base_file relacionados com tabela_termo_cliente
        - 'tabela_arquivotermo_b2b': Registros de termo_base_file relacionados com tabela_termo_b2b
        - 'tabela_fixingstermo_cliente': Registros de termo_base_fixings relacionados com tabela_termo_cliente
        - 'tabela_fixingstermo_b2b': Registros de termo_base_fixings relacionados com tabela_termo_b2b
    """
    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
    
    # Construir a consulta SQL base
    base_query = "SELECT * FROM termo_base_deals WHERE 1=1"
    params = []
    
    # Adicionar filtros à consulta
    query, params = _build_query_with_filters(
        base_query,
        params,
        trade_date_filter,
        settlement_date_filter,
        fixing_commodity_filter,
        fixing_ccy_filter,
        counterparty_filter,
        notional_filter,
        strike_filter,
        athena_id_filter,
        status_filter
    )
    
    global entry_query_termo_cliente
    global entry_query_termo_b2b
    global entry_query_arquivotermo_cliente
    global entry_query_arquivotermo_b2b
    
    entry_query_termo_cliente.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_termo_cliente.insert(0, f"{query} , {params}")
    entry_query_termo_b2b.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_termo_b2b.insert(0, f"{query} , {params}")
    
    entry_query_arquivotermo_cliente.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_arquivotermo_cliente.insert(0, f"{query} , {params}")
    entry_query_arquivotermo_b2b.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_arquivotermo_b2b.insert(0, f"{query} , {params}")
    
    
    # Executar a consulta
    cursor.execute(query, params)
    filtered_data = cursor.fetchall()
    
    # Distribuir dados filtrados para as tabelas principais
    tabela_termo_cliente = []
    tabela_termo_b2b = []
    
    # Coletar identificadores para busca em tabelas relacionadas
    cliente_identifiers = []
    b2b_identifiers = []
    id_cliente = []
    id_b2b = []
    contraparte = "00041007"
    for row in filtered_data:
        # Obter o valor do campo "Identifier" (índice -4)
        identifier = row[-4] if len(row) >= 4 else None
        id = row[0]
            
        if row[18] == 'Lawton':
            tabela_termo_b2b.append(row)
            if identifier:
                b2b_identifiers.append(identifier)
                id_b2b.append(id)
        else:
            tabela_termo_cliente.append(row)
            if identifier:
                cliente_identifiers.append(identifier)
                id_cliente.append(id)
    
    # Buscar dados relacionados na tabela termo_base_file
    tabela_arquivotermo_cliente = []
    tabela_arquivotermo_b2b = []
    
    if cliente_identifiers:
        placeholders = ', '.join(['?' for _ in cliente_identifiers])        
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND (Contraparte <> ? AND Observaçao <> ?)"
        cursor.execute(query, cliente_identifiers + [contraparte, contraparte])
        tabela_arquivotermo_cliente = cursor.fetchall()
    
    if b2b_identifiers:
        placeholders = ', '.join(['?' for _ in b2b_identifiers])
        query = f"SELECT * FROM termo_base_file WHERE Identifier IN ({placeholders}) AND (Contraparte = ? OR Observaçao = ?)"
        cursor.execute(query, b2b_identifiers + [contraparte, contraparte])
        tabela_arquivotermo_b2b = cursor.fetchall()
    
    # Buscar dados relacionados na tabela termo_base_fixings
    tabela_fixingstermo_cliente = []
    tabela_fixingstermo_b2b = []
    
    if cliente_identifiers:
        placeholders_identifiers = ', '.join(['?' for _ in cliente_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_cliente])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, cliente_identifiers + id_cliente)
        tabela_fixingstermo_cliente = cursor.fetchall()
    
    if b2b_identifiers:        
        placeholders_identifiers = ', '.join(['?' for _ in b2b_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_b2b])
        query = f"SELECT * FROM termo_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"        
        cursor.execute(query, b2b_identifiers + id_b2b)
        tabela_fixingstermo_b2b = cursor.fetchall()
    
    # Fechar a conexão com o banco de dados
    conn.commit()
    conn.close()

    # Retornar todas as tabelas em um dicionário
    return {
        'tabela_termo_cliente': tabela_termo_cliente,
        'tabela_termo_b2b': tabela_termo_b2b,
        'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
        'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
        'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
        'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b
    }

def _build_query_with_filters(
    base_query: str,
    params: List[Any],
    trade_date_filter: Dict[str, Any] = None,
    settlement_date_filter: Dict[str, Any] = None,
    fixing_commodity_filter: Dict[str, Any] = None,
    fixing_ccy_filter: Dict[str, Any] = None,
    counterparty_filter: str = None,
    notional_filter: float = None,
    strike_filter: float = None,
    athena_id_filter: str = None,
    status_filter: str = None
) -> Tuple[str, List[Any]]:
    """
    Constrói a consulta SQL com os filtros fornecidos.
    
    Returns:
        Tupla contendo a consulta SQL e a lista de parâmetros.
    """
    query = base_query
    
    # Filtro de Trade Date (índice 1)
    if trade_date_filter:
        filter_type = trade_date_filter.get('type')
        filter_value = trade_date_filter.get('value')
        
        if filter_type and filter_value:
            if filter_type == 'is':
                query += " AND [TradeDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [TradeDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [TradeDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [TradeDate] < ?"
                params.append(_convert_date_format(filter_value))
    
    # Filtro de Settlement Date (índice 10)
    if settlement_date_filter:
        filter_type = settlement_date_filter.get('type')
        filter_value = settlement_date_filter.get('value')
        
        if filter_type and filter_value:
            if filter_type == 'is':
                query += " AND [SettlementDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [SettlementDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [SettlementDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [SettlementDate] < ?"
                params.append(_convert_date_format(filter_value))
                
    # Filtro de Fixing Commodity (índice 17)
    if fixing_commodity_filter:
        filter_type = fixing_commodity_filter.get('type')
        filter_value = fixing_commodity_filter.get('value')   
        
        if filter_type and filter_value:            
            if filter_type == 'is':
                query += " AND [FixingEndDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [FixingEndDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [FixingEndDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [FixingEndDate] < ?"
                params.append(_convert_date_format(filter_value))    
    
    
    # Filtro de Fixing CCY (índice 15)   
    if fixing_ccy_filter:
        filter_type = fixing_ccy_filter.get('type')
        filter_value = fixing_ccy_filter.get('value')   
                
        if filter_type and filter_value:            
            if filter_type == 'is':
                query += " AND [FXConvDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [FXConvDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [FXConvDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [FXConvDate] < ?"
                params.append(_convert_date_format(filter_value))  
    
    # Filtro de Counterparty (índice 18)
    if counterparty_filter:
        query += " AND Counterparty = ?"
        params.append(counterparty_filter)
    
    # Filtro de Notional (índice 9)
    if notional_filter is not None:
        query += " AND [TotalNotional] = ?"
        params.append(notional_filter)
    
    # Filtro de Strike (índice 6)
    if strike_filter is not None:
        query += " AND Strike = ?"
        params.append(strike_filter)
    
    # Filtro de Athena ID (índice 0)
    if athena_id_filter:
        query += " AND [DealName] = ?"
        params.append(athena_id_filter)
    
    # Filtro de Status (índice -2)
    if status_filter:
        if status_filter == 'Pending':
            query += " AND Status <> 'Concluded'"
        elif status_filter == 'OK':
            query += " AND Status = 'Concluded'"
        # Para 'All', não aplicamos filtro
   
    return query, params


def _convert_date_format(date_str: str, input_format: str = "%d-%m-%Y") -> str:
    """
    Converte uma string de data do formato de entrada para o formato do banco de dados.
    
    Args:
        date_str: String de data no formato de entrada (ex: "21-04-2025")
        input_format: Formato da data de entrada (padrão: "%d-%m-%Y")
        
    Returns:
        String de data no formato do banco de dados (ex: "21-Apr-2025")
    """
    try:
        # Converter a string para objeto datetime
        date_obj = datetime.strptime(date_str, input_format)
        # Converter de volta para string no formato do banco
        return date_obj.strftime("%d-%b-%Y")
    except ValueError:
        # Se não conseguir converter, retornar a string original
        return date_str


def create_query_from_interface(
    trade_date_type: str = None,
    trade_date_value: Union[str, List[str]] = None,
    settlement_date_type: str = None,
    settlement_date_value: Union[str, List[str]] = None,
    fixing_commodity_type: str = None,
    fixing_commodity_value: str = None,
    fixing_ccy_type: str = None,
    fixing_ccy_value: str = None,
    counterparty: str = None,
    notional: float = None,
    strike: float = None,
    athena_id: str = None,
    status: str = None
) -> Dict[str, Any]:
    """
    Cria um dicionário de parâmetros para a função execute_query_sqlite a partir dos valores da interface.
    
    Args:
        trade_date_type: Tipo de filtro para Trade Date ('is', 'is_between', 'after', 'before')
        trade_date_value: Valor(es) para filtro de Trade Date
        settlement_date_type: Tipo de filtro para Settlement Date ('is', 'is_between', 'after', 'before')
        settlement_date_value: Valor(es) para filtro de Settlement Date
        fixing_commodity_type: Tipo de filtro para Fixing Commodity ('is')
        fixing_commodity_value: Valor para filtro de Fixing Commodity
        fixing_ccy_type: Tipo de filtro para Fixing CCY ('is')
        fixing_ccy_value: Valor para filtro de Fixing CCY
        counterparty: Valor para filtro de Counterparty
        notional: Valor para filtro de Notional
        strike: Valor para filtro de Strike
        athena_id: Valor para filtro de Athena ID
        status: Tipo de filtro para Status ('Pending', 'OK', 'All')
        
    Returns:
        Dicionário com parâmetros para a função execute_query_sqlite
    """
    params = {}
    
    # Filtro de Trade Date
    if trade_date_type and trade_date_value:
        params['trade_date_filter'] = {
            'type': trade_date_type,
            'value': trade_date_value
        }
    
    # Filtro de Settlement Date
    if settlement_date_type and settlement_date_value:
        params['settlement_date_filter'] = {
            'type': settlement_date_type,
            'value': settlement_date_value
        }
    
    # Filtro de Fixing Commodity
    if fixing_commodity_type and fixing_commodity_value:
        params['fixing_commodity_filter'] = {
            'type': fixing_commodity_type,
            'value': fixing_commodity_value
        }
    
    # Filtro de Fixing CCY
    if fixing_ccy_type and fixing_ccy_value:
        params['fixing_ccy_filter'] = {
            'type': fixing_ccy_type,
            'value': fixing_ccy_value
        }
    
    # Outros filtros
    if counterparty:
        params['counterparty_filter'] = counterparty
    
    if notional is not None:
        params['notional_filter'] = notional
    
    if strike is not None:
        params['strike_filter'] = strike
    
    if athena_id:
        params['athena_id_filter'] = athena_id
    
    if status:
        params['status_filter'] = status
    
    return params

def run_query_from_interface_sqlite(    
    trade_date_option,
    date_entry1_trade_date,
    date_entry2_trade_date,
    settlement_date_option,
    date_entry1_settlement_date,
    date_entry2_settlement_date,
    fixing_commodity_option,
    date_entry1_fixing_commodity,
    date_entry2_fixing_commodity,
    fixing_ccy_option,
    date_entry1_fixing_ccy,
    date_entry2_fixing_ccy,
    counterparty_combobox_termo,
    notional_entry,
    strike_entry,
    athena_id_entry,
    status_option
) -> Dict[str, List[List[Any]]]:
    # Obter valores dos tipos de filtro
    trade_date_type = trade_date_option if trade_date_option else None
    settlement_date_type = settlement_date_option if settlement_date_option else None
    fixing_commodity_type = fixing_commodity_option if fixing_commodity_option else None
    fixing_ccy_type = fixing_ccy_option if fixing_ccy_option else None
    status = status_option if status_option else None
    
    # Obter valores dos widgets de entrada para Trade Date
    trade_date_value = None
    if trade_date_type:
        if trade_date_type == "is between" and date_entry2_trade_date:
            trade_date_value = [
                date_entry1_trade_date,
                date_entry2_trade_date
            ]
        else:
            trade_date_value = date_entry1_trade_date
    
    # Obter valores dos widgets de entrada para Settlement Date
    settlement_date_value = None
    if settlement_date_type and date_entry1_settlement_date:
        if settlement_date_type == "is between" and date_entry2_settlement_date:
            settlement_date_value = [
                date_entry1_settlement_date,
                date_entry2_settlement_date
            ]
        else:
            settlement_date_value = date_entry1_settlement_date
    
    # Obter valores dos widgets de entrada para Fixing Commodity
    fixing_commodity_value = None
    if fixing_commodity_type and date_entry1_fixing_commodity:
        if fixing_commodity_type == "is between" and date_entry2_fixing_commodity:
            fixing_commodity_value = [
                date_entry1_fixing_commodity,
                date_entry2_fixing_commodity
            ]
        else:
            fixing_commodity_value = date_entry1_fixing_commodity
    
    # Obter valores dos widgets de entrada para Fixing CCY
    fixing_ccy_value = None
    if fixing_ccy_type and date_entry1_fixing_ccy:
        if fixing_ccy_type == "is between" and date_entry2_fixing_ccy:
            fixing_ccy_value = [
                date_entry1_fixing_ccy,
                date_entry2_fixing_ccy
            ]
        else:
            fixing_ccy_value = date_entry1_fixing_ccy
    
    # Obter valores dos outros campos
    counterparty = counterparty_combobox_termo if counterparty_combobox_termo else None
    
    notional = None
    if notional_entry:
        try:
            notional = float(notional_entry)
        except ValueError:
            pass
    
    strike = None
    if strike_entry:
        try:
            strike = float(strike_entry)
        except ValueError:
            pass
    
    athena_id = athena_id_entry if athena_id_entry else None
    
    # Criar parâmetros para a query
    params = create_query_from_interface(
        trade_date_type=trade_date_type,
        trade_date_value=trade_date_value,
        settlement_date_type=settlement_date_type,
        settlement_date_value=settlement_date_value,
        fixing_commodity_type=fixing_commodity_type,
        fixing_commodity_value=fixing_commodity_value,
        fixing_ccy_type=fixing_ccy_type,
        fixing_ccy_value=fixing_ccy_value,
        counterparty=counterparty,
        notional=notional,
        strike=strike,
        athena_id=athena_id,
        status=status
    )
    
    # Executar a query
    return execute_query_sqlite(**params)

# Configurar o botão "Run (F9)" para chamar a função
def on_run_button_click(window,   
    option_menu_trade_date,
    date_entry1_trade_date,
    date_entry2_trade_date,
    option_menu_settlement_date,
    date_entry1_settlement_date,
    date_entry2_settlement_date,
    option_menu_fixing_commodity,
    date_entry1_fixing_commodity,
    date_entry2_fixing_commodity,
    option_menu_fixing_ccy,
    date_entry1_fixing_ccy,
    date_entry2_fixing_ccy,
    counterparty_combobox_termo,
    notional_entry,
    strike_entry,
    athena_id_entry,
    status_option_menu,
    treeviews
):
    # Obter valores dos widgets
    values = {}
    
     
    # Obter valores dos tipos de filtro
    values['trade_date_option'] = option_menu_trade_date.get() if option_menu_trade_date else None
    values['settlement_date_option'] = option_menu_settlement_date.get() if option_menu_settlement_date else None
    values['fixing_commodity_option'] = option_menu_fixing_commodity.get() if option_menu_fixing_commodity else None
    values['fixing_ccy_option'] = option_menu_fixing_ccy.get() if option_menu_fixing_ccy else None
    values['status_option'] = status_option_menu.get() if status_option_menu else None
    
    # Obter valores dos widgets de entrada para Trade Date
    values['date_entry1_trade_date'] = date_entry1_trade_date.get() if date_entry1_trade_date else None
    values['date_entry2_trade_date'] = date_entry2_trade_date.get() if date_entry2_trade_date else None
    
    # Obter valores dos widgets de entrada para Settlement Date
    values['date_entry1_settlement_date'] = date_entry1_settlement_date.get() if date_entry1_settlement_date else None
    values['date_entry2_settlement_date'] = date_entry2_settlement_date.get() if date_entry2_settlement_date else None
    
    # Obter valores dos widgets de entrada para Fixing Commodity
    values['date_entry1_fixing_commodity'] = date_entry1_fixing_commodity.get() if date_entry1_fixing_commodity else None
    values['date_entry2_fixing_commodity'] = date_entry2_fixing_commodity.get() if date_entry2_fixing_commodity else None
    
    # Obter valores dos widgets de entrada para Fixing CCY
    values['date_entry1_fixing_ccy'] = date_entry1_fixing_ccy.get() if date_entry1_fixing_ccy else None
    values['date_entry2_fixing_ccy'] = date_entry2_fixing_ccy.get() if date_entry2_fixing_ccy else None
    
    # Obter valores dos outros campos
    values['counterparty_combobox_termo'] = counterparty_combobox_termo.get() if counterparty_combobox_termo else None
    values['notional_entry'] = notional_entry.get() if notional_entry else None
    values['strike_entry'] = strike_entry.get() if strike_entry else None
    values['athena_id_entry'] = athena_id_entry.get() if athena_id_entry else None
    
    # Executar a query
    results = run_query_from_interface_sqlite(        
        values['trade_date_option'],
        values['date_entry1_trade_date'],
        values['date_entry2_trade_date'],
        values['settlement_date_option'],
        values['date_entry1_settlement_date'],
        values['date_entry2_settlement_date'],
        values['fixing_commodity_option'],
        values['date_entry1_fixing_commodity'],
        values['date_entry2_fixing_commodity'],
        values['fixing_ccy_option'],
        values['date_entry1_fixing_ccy'],
        values['date_entry2_fixing_ccy'],
        values['counterparty_combobox_termo'],
        values['notional_entry'],
        values['strike_entry'],
        values['athena_id_entry'],
        values['status_option'],
    )

    # Atualizar as Treeviews com os resultados
    if treeviews:
        update_treeviews(results, treeviews)
     
    
    window.destroy()
def load_query_entry_option(entry_query_opcao, treeviews):
    # Obter o texto da entrada
    query_text = entry_query_opcao.get()
    
    results = execute_load_query_option_sqlite(query_text)
    
    if treeviews:
        update_treeviews(results, treeviews)
        
def execute_load_query_option_sqlite( query_text: str = None) -> Dict[str, List[List[Any]]]:   
    
    """
    Executa uma query no banco de dados SQLite com base nos filtros fornecidos
    e distribui os resultados em múltiplas tabelas, incluindo dados relacionados de outras tabelas.

    Args:
        conn: Caminho para o arquivo do banco de dados SQLite
        trade_date_filter: Dicionário com tipo de filtro e valor(es) para Trade Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        settlement_date_filter: Dicionário com tipo de filtro e valor(es) para Settlement Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        fixing_commodity_filter: Dicionário com tipo de filtro e valor para Fixing Commodity
            Formato: {'type': 'is', 'value': str}
        fixing_ccy_filter: Dicionário com tipo de filtro e valor para Fixing CCY
            Formato: {'type': 'is', 'value': str}
        counterparty_filter: Valor para filtrar Counterparty
        notional_filter: Valor para filtrar Notional
        strike_filter: Valor para filtrar Strike
        athena_id_filter: Valor para filtrar Athena ID
        status_filter: Tipo de filtro para Status ('Pending', 'OK', 'All')

    Returns:
        Dicionário contendo seis listas:
        - 'tabela_opcao_cliente': Registros de opcao_base_deals onde Counterparty != 'Lawton'
        - 'tabela_opcao_b2b': Registros de opcao_base_deals onde Counterparty == 'Lawton'
        - 'tabela_arquivoopcao_cliente': Registros de opcao_base_file relacionados com tabela_opcao_cliente
        - 'tabela_arquivoopcao_b2b': Registros de opcao_base_file relacionados com tabela_opcao_b2b
        - 'tabela_fixingsopcao_cliente': Registros de opcao_base_fixings relacionados com tabela_opcao_cliente
        - 'tabela_fixingsopcao_b2b': Registros de opcao_base_fixings relacionados com tabela_opcao_b2b
    """
    # Separar a consulta SQL e os parâmetros        
 
    
    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
    
    # Construir a consulta SQL base    
    params = []    
    query, params = query_text.split(" , ")
    params = eval(params)  # Converte a string de parâmetros de volta para uma lista/tupla
    
    # Executar a consulta
    cursor.execute(query, params)
    filtered_data = cursor.fetchall()
    
    # Distribuir dados filtrados para as tabelas principais
    tabela_opcao_cliente = []
    tabela_opcao_b2b = []
    
    # Coletar identificadores para busca em tabelas relacionadas
    cliente_identifiers = []
    b2b_identifiers = []
    id_cliente = []
    id_b2b = []
    contraparte = "00041.00-7"
    for row in filtered_data:
        # Obter o valor do campo "Identifier" (índice -4)
        identifier = row[-4] if len(row) >= 4 else None
        id = row[0]
            
        if row[18] == 'Lawton':
            tabela_opcao_b2b.append(row)
            if identifier:
                b2b_identifiers.append(identifier)
                id_b2b.append(id)
        else:
            tabela_opcao_cliente.append(row)
            if identifier:
                cliente_identifiers.append(identifier)
                id_cliente.append(id)
    
    # Buscar dados relacionados na tabela opcao_base_file
    tabela_arquivoopcao_cliente = []
    tabela_arquivoopcao_b2b = []
    
    if cliente_identifiers:
        placeholders = ', '.join(['?' for _ in cliente_identifiers])        
        query = f"SELECT * FROM opcao_base_file WHERE Identifier IN ({placeholders}) AND Conta_Contraparte <> ?"
        cursor.execute(query, cliente_identifiers + [contraparte])
        tabela_arquivoopcao_cliente = cursor.fetchall()
    
    if b2b_identifiers:
        placeholders = ', '.join(['?' for _ in b2b_identifiers])
        query = f"SELECT * FROM opcao_base_file WHERE Identifier IN ({placeholders}) AND Conta_Contraparte = ?"
        cursor.execute(query, b2b_identifiers + [contraparte])
        tabela_arquivoopcao_b2b = cursor.fetchall()
    
    # Buscar dados relacionados na tabela opcao_base_fixings
    tabela_fixingsopcao_cliente = []
    tabela_fixingsopcao_b2b = []
    
    if cliente_identifiers:
        placeholders_identifiers = ', '.join(['?' for _ in cliente_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_cliente])
        query = f"SELECT * FROM opcao_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, cliente_identifiers + id_cliente)
        tabela_fixingsopcao_cliente = cursor.fetchall()
    
    if b2b_identifiers:        
        placeholders_identifiers = ', '.join(['?' for _ in b2b_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_b2b])
        query = f"SELECT * FROM opcao_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"        
        cursor.execute(query, b2b_identifiers + id_b2b)
        tabela_fixingsopcao_b2b = cursor.fetchall()
    
    # Fechar a conexão com o banco de dados
    conn.commit()
    conn.close()
    
    # Retornar todas as tabelas em um dicionário
    return {
        'tabela_opcao_cliente': tabela_opcao_cliente,
        'tabela_opcao_b2b': tabela_opcao_b2b,
        'tabela_arquivoopcao_cliente': tabela_arquivoopcao_cliente,
        'tabela_arquivoopcao_b2b': tabela_arquivoopcao_b2b,
        'tabela_fixingsopcao_cliente': tabela_fixingsopcao_cliente,
        'tabela_fixingsopcao_b2b': tabela_fixingsopcao_b2b
    }
    
def execute_query_option_sqlite(    
    trade_date_filter: Dict[str, Any] = None,
    settlement_date_filter: Dict[str, Any] = None,
    fixing_commodity_filter: Dict[str, Any] = None,
    fixing_ccy_filter: Dict[str, Any] = None,
    counterparty_filter: str = None,
    notional_filter: str = None,
    strike_filter: str = None,
    athena_id_filter: str = None,
    status_filter: str = None
) -> Dict[str, List[List[Any]]]:
    """
    Executa uma query no banco de dados SQLite com base nos filtros fornecidos
    e distribui os resultados em múltiplas tabelas, incluindo dados relacionados de outras tabelas.

    Args:        
        trade_date_filter: Dicionário com tipo de filtro e valor(es) para Trade Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        settlement_date_filter: Dicionário com tipo de filtro e valor(es) para Settlement Date
            Formato: {'type': 'is'|'is_between'|'after'|'before', 'value': str|List[str]}
        fixing_commodity_filter: Dicionário com tipo de filtro e valor para Fixing Commodity
            Formato: {'type': 'is', 'value': str}
        fixing_ccy_filter: Dicionário com tipo de filtro e valor para Fixing CCY
            Formato: {'type': 'is', 'value': str}
        counterparty_filter: Valor para filtrar Counterparty
        notional_filter: Valor para filtrar Notional
        strike_filter: Valor para filtrar Strike
        athena_id_filter: Valor para filtrar Athena ID
        status_filter: Tipo de filtro para Status ('Pending', 'OK', 'All')

    Returns:
        Dicionário contendo seis listas:
        - 'tabela_opcao_cliente': Registros de opcao_base_deals onde Counterparty != 'Lawton'
        - 'tabela_opcao_b2b': Registros de opcao_base_deals onde Counterparty == 'Lawton'
        - 'tabela_arquivoopcao_cliente': Registros de opcao_base_file relacionados com tabela_opcao_cliente
        - 'tabela_arquivoopcao_b2b': Registros de opcao_base_file relacionados com tabela_opcao_b2b
        - 'tabela_fixingsopcao_cliente': Registros de opcao_base_fixings relacionados com tabela_opcao_cliente
        - 'tabela_fixingsopcao_b2b': Registros de opcao_base_fixings relacionados com tabela_opcao_b2b
    """
    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
    cursor = conn.cursor()
    cursor.execute("BEGIN TRANSACTION;")
    
    # Construir a consulta SQL base
    base_query = "SELECT * FROM opcao_base_deals WHERE 1=1"
    params = []
    
    # Adicionar filtros à consulta
    query, params = _build_query_option_with_filters(
        base_query,
        params,
        trade_date_filter,
        settlement_date_filter,
        fixing_commodity_filter,
        fixing_ccy_filter,
        counterparty_filter,
        notional_filter,
        strike_filter,
        athena_id_filter,
        status_filter
    )
    
    global entry_query_opcao_cliente
    global entry_query_opcao_b2b
    global entry_query_arquivoopcao_cliente
    global entry_query_arquivoopcao_b2b
    entry_query_opcao_cliente.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_opcao_cliente.insert(0, f"{query} , {params}")
    entry_query_opcao_b2b.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_opcao_b2b.insert(0, f"{query} , {params}")
    entry_query_arquivoopcao_cliente.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_arquivoopcao_cliente.insert(0, f"{query} , {params}")
    entry_query_arquivoopcao_b2b.delete(0, tk.END)
    # Inserir a consulta e os parâmetros formatados como string
    entry_query_arquivoopcao_b2b.insert(0, f"{query} , {params}")
    
    
    # Executar a consulta
    cursor.execute(query, params)
    filtered_data = cursor.fetchall()
    
    # Distribuir dados filtrados para as tabelas principais
    tabela_opcao_cliente = []
    tabela_opcao_b2b = []
    
    # Coletar identificadores para busca em tabelas relacionadas
    cliente_identifiers = []
    b2b_identifiers = []
    id_cliente = []
    id_b2b = []
    contraparte = "00041.00-7"
    for row in filtered_data:
        # Obter o valor do campo "Identifier" (índice -4)
        identifier = row[-4] if len(row) >= 4 else None
        id = row[0]
            
        if row[18] == 'Lawton':
            tabela_opcao_b2b.append(row)
            if identifier:
                b2b_identifiers.append(identifier)
                id_b2b.append(id)
        else:
            tabela_opcao_cliente.append(row)
            if identifier:
                cliente_identifiers.append(identifier)
                id_cliente.append(id)
    
    # Buscar dados relacionados na tabela opcao_base_file
    tabela_arquivoopcao_cliente = []
    tabela_arquivoopcao_b2b = []
    
    if cliente_identifiers:
        placeholders = ', '.join(['?' for _ in cliente_identifiers])        
        query = f"SELECT * FROM opcao_base_file WHERE Identifier IN ({placeholders}) AND Conta_Contraparte <> ?"
        cursor.execute(query, cliente_identifiers + [contraparte])
        tabela_arquivoopcao_cliente = cursor.fetchall()
    
    if b2b_identifiers:
        placeholders = ', '.join(['?' for _ in b2b_identifiers])
        query = f"SELECT * FROM opcao_base_file WHERE Identifier IN ({placeholders}) AND Conta_Contraparte = ?"
        cursor.execute(query, b2b_identifiers + [contraparte])
        tabela_arquivoopcao_b2b = cursor.fetchall()
    
    # Buscar dados relacionados na tabela opcao_base_fixings
    tabela_fixingsopcao_cliente = []
    tabela_fixingsopcao_b2b = []
    
    if cliente_identifiers:
        placeholders_identifiers = ', '.join(['?' for _ in cliente_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_cliente])
        query = f"SELECT * FROM opcao_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"
        cursor.execute(query, cliente_identifiers + id_cliente)
        tabela_fixingsopcao_cliente = cursor.fetchall()
    
    if b2b_identifiers:        
        placeholders_identifiers = ', '.join(['?' for _ in b2b_identifiers])
        placeholders_ids  = ', '.join(['?' for _ in id_b2b])
        query = f"SELECT * FROM opcao_base_fixings WHERE Identifier IN ({placeholders_identifiers}) and AthenaID in ({placeholders_ids})"        
        cursor.execute(query, b2b_identifiers + id_b2b)
        tabela_fixingsopcao_b2b = cursor.fetchall()
    
    # Fechar a conexão com o banco de dados
    conn.commit()
    conn.close()

    # Retornar todas as tabelas em um dicionário
    return {
        'tabela_opcao_cliente': tabela_opcao_cliente,
        'tabela_opcao_b2b': tabela_opcao_b2b,
        'tabela_arquivoopcao_cliente': tabela_arquivoopcao_cliente,
        'tabela_arquivoopcao_b2b': tabela_arquivoopcao_b2b,
        'tabela_fixingsopcao_cliente': tabela_fixingsopcao_cliente,
        'tabela_fixingsopcao_b2b': tabela_fixingsopcao_b2b
    }

def _build_query_option_with_filters(
    base_query: str,
    params: List[Any],
    trade_date_filter: Dict[str, Any] = None,
    settlement_date_filter: Dict[str, Any] = None,
    fixing_commodity_filter: Dict[str, Any] = None,
    fixing_ccy_filter: Dict[str, Any] = None,
    counterparty_filter: str = None,
    notional_filter: float = None,
    strike_filter: float = None,
    athena_id_filter: str = None,
    status_filter: str = None
) -> Tuple[str, List[Any]]:
    """
    Constrói a consulta SQL com os filtros fornecidos.
    
    Returns:
        Tupla contendo a consulta SQL e a lista de parâmetros.
    """
    query = base_query
    
    # Filtro de Trade Date (índice 1)
    if trade_date_filter:
        filter_type = trade_date_filter.get('type')
        filter_value = trade_date_filter.get('value')
        
        if filter_type and filter_value:
            if filter_type == 'is':
                query += " AND [TradeDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [TradeDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [TradeDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [TradeDate] < ?"
                params.append(_convert_date_format(filter_value))
    
    # Filtro de Settlement Date (índice 10)
    if settlement_date_filter:
        filter_type = settlement_date_filter.get('type')
        filter_value = settlement_date_filter.get('value')
        
        if filter_type and filter_value:
            if filter_type == 'is':
                query += " AND [SettlementDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [SettlementDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [SettlementDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [SettlementDate] < ?"
                params.append(_convert_date_format(filter_value))
                
    # Filtro de Fixing Commodity (índice 17)
    if fixing_commodity_filter:
        filter_type = fixing_commodity_filter.get('type')
        filter_value = fixing_commodity_filter.get('value')   
        
        if filter_type and filter_value:            
            if filter_type == 'is':
                query += " AND [FixingEndDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [FixingEndDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [FixingEndDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [FixingEndDate] < ?"
                params.append(_convert_date_format(filter_value))    
    
    
    # Filtro de Fixing CCY (índice 15)   
    if fixing_ccy_filter:
        filter_type = fixing_ccy_filter.get('type')
        filter_value = fixing_ccy_filter.get('value')   
                
        if filter_type and filter_value:            
            if filter_type == 'is':
                query += " AND [FXConvDate] = ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'is between' and isinstance(filter_value, list) and len(filter_value) == 2:
                query += " AND [FXConvDate] BETWEEN ? AND ?"
                params.append(_convert_date_format(filter_value[0]))
                params.append(_convert_date_format(filter_value[1]))
            elif filter_type == 'after':
                query += " AND [FXConvDate] > ?"
                params.append(_convert_date_format(filter_value))
            elif filter_type == 'before':
                query += " AND [FXConvDate] < ?"
                params.append(_convert_date_format(filter_value))  
    
    # Filtro de Counterparty (índice 18)
    if counterparty_filter:
        query += " AND Counterparty = ?"
        params.append(counterparty_filter)
    
    # Filtro de Notional (índice 9)
    if notional_filter is not None:
        query += " AND [TotalNotional] = ?"
        params.append(notional_filter)
    
    # Filtro de Strike (índice 6)
    if strike_filter is not None:
        query += " AND Strike = ?"
        params.append(strike_filter)
    
    # Filtro de Athena ID (índice 0)
    if athena_id_filter:
        query += " AND [DealName] = ?"
        params.append(athena_id_filter)
    
    # Filtro de Status (índice -2)
    if status_filter:
        if status_filter == 'Pending':
            query += " AND Status <> 'Concluded'"
        elif status_filter == 'OK':
            query += " AND Status = 'Concluded'"
        # Para 'All', não aplicamos filtro
   
    return query, params



def create_query_option_from_interface(
    trade_date_type: str = None,
    trade_date_value: Union[str, List[str]] = None,
    settlement_date_type: str = None,
    settlement_date_value: Union[str, List[str]] = None,
    fixing_commodity_type: str = None,
    fixing_commodity_value: str = None,
    fixing_ccy_type: str = None,
    fixing_ccy_value: str = None,
    counterparty: str = None,
    notional: float = None,
    strike: float = None,
    athena_id: str = None,
    status: str = None
) -> Dict[str, Any]:
    """
    Cria um dicionário de parâmetros para a função execute_query_option_sqlite a partir dos valores da interface.
    
    Args:
        trade_date_type: Tipo de filtro para Trade Date ('is', 'is_between', 'after', 'before')
        trade_date_value: Valor(es) para filtro de Trade Date
        settlement_date_type: Tipo de filtro para Settlement Date ('is', 'is_between', 'after', 'before')
        settlement_date_value: Valor(es) para filtro de Settlement Date
        fixing_commodity_type: Tipo de filtro para Fixing Commodity ('is')
        fixing_commodity_value: Valor para filtro de Fixing Commodity
        fixing_ccy_type: Tipo de filtro para Fixing CCY ('is')
        fixing_ccy_value: Valor para filtro de Fixing CCY
        counterparty: Valor para filtro de Counterparty
        notional: Valor para filtro de Notional
        strike: Valor para filtro de Strike
        athena_id: Valor para filtro de Athena ID
        status: Tipo de filtro para Status ('Pending', 'OK', 'All')
        
    Returns:
        Dicionário com parâmetros para a função execute_query_option_sqlite
    """
    params = {}
    
    # Filtro de Trade Date
    if trade_date_type and trade_date_value:
        params['trade_date_filter'] = {
            'type': trade_date_type,
            'value': trade_date_value
        }
    
    # Filtro de Settlement Date
    if settlement_date_type and settlement_date_value:
        params['settlement_date_filter'] = {
            'type': settlement_date_type,
            'value': settlement_date_value
        }
    
    # Filtro de Fixing Commodity
    if fixing_commodity_type and fixing_commodity_value:
        params['fixing_commodity_filter'] = {
            'type': fixing_commodity_type,
            'value': fixing_commodity_value
        }
    
    # Filtro de Fixing CCY
    if fixing_ccy_type and fixing_ccy_value:
        params['fixing_ccy_filter'] = {
            'type': fixing_ccy_type,
            'value': fixing_ccy_value
        }
    
    # Outros filtros
    if counterparty:
        params['counterparty_filter'] = counterparty
    
    if notional is not None:
        params['notional_filter'] = notional
    
    if strike is not None:
        params['strike_filter'] = strike
    
    if athena_id:
        params['athena_id_filter'] = athena_id
    
    if status:
        params['status_filter'] = status
    
    return params

def run_query_option_from_interface_sqlite(    
    trade_date_option,
    date_entry1_trade_date,
    date_entry2_trade_date,
    settlement_date_option,
    date_entry1_settlement_date,
    date_entry2_settlement_date,
    fixing_commodity_option,
    date_entry1_fixing_commodity,
    date_entry2_fixing_commodity,
    fixing_ccy_option,
    date_entry1_fixing_ccy,
    date_entry2_fixing_ccy,
    counterparty_combobox_opcao,
    notional_entry,
    strike_entry,
    athena_id_entry,
    status_option
) -> Dict[str, List[List[Any]]]:
    # Obter valores dos tipos de filtro
    trade_date_type = trade_date_option if trade_date_option else None
    settlement_date_type = settlement_date_option if settlement_date_option else None
    fixing_commodity_type = fixing_commodity_option if fixing_commodity_option else None
    fixing_ccy_type = fixing_ccy_option if fixing_ccy_option else None
    status = status_option if status_option else None
    
    # Obter valores dos widgets de entrada para Trade Date
    trade_date_value = None
    if trade_date_type:
        if trade_date_type == "is between" and date_entry2_trade_date:
            trade_date_value = [
                date_entry1_trade_date,
                date_entry2_trade_date
            ]
        else:
            trade_date_value = date_entry1_trade_date
    
    # Obter valores dos widgets de entrada para Settlement Date
    settlement_date_value = None
    if settlement_date_type and date_entry1_settlement_date:
        if settlement_date_type == "is between" and date_entry2_settlement_date:
            settlement_date_value = [
                date_entry1_settlement_date,
                date_entry2_settlement_date
            ]
        else:
            settlement_date_value = date_entry1_settlement_date
    
    # Obter valores dos widgets de entrada para Fixing Commodity
    fixing_commodity_value = None
    if fixing_commodity_type and date_entry1_fixing_commodity:
        if fixing_commodity_type == "is between" and date_entry2_fixing_commodity:
            fixing_commodity_value = [
                date_entry1_fixing_commodity,
                date_entry2_fixing_commodity
            ]
        else:
            fixing_commodity_value = date_entry1_fixing_commodity
    
    # Obter valores dos widgets de entrada para Fixing CCY
    fixing_ccy_value = None
    if fixing_ccy_type and date_entry1_fixing_ccy:
        if fixing_ccy_type == "is between" and date_entry2_fixing_ccy:
            fixing_ccy_value = [
                date_entry1_fixing_ccy,
                date_entry2_fixing_ccy
            ]
        else:
            fixing_ccy_value = date_entry1_fixing_ccy
    
    # Obter valores dos outros campos
    counterparty = counterparty_combobox_opcao if counterparty_combobox_opcao else None
    
    notional = None
    if notional_entry:
        try:
            notional = float(notional_entry)
        except ValueError:
            pass
    
    strike = None
    if strike_entry:
        try:
            strike = float(strike_entry)
        except ValueError:
            pass
    
    athena_id = athena_id_entry if athena_id_entry else None
    
    # Criar parâmetros para a query
    params = create_query_option_from_interface(
        trade_date_type=trade_date_type,
        trade_date_value=trade_date_value,
        settlement_date_type=settlement_date_type,
        settlement_date_value=settlement_date_value,
        fixing_commodity_type=fixing_commodity_type,
        fixing_commodity_value=fixing_commodity_value,
        fixing_ccy_type=fixing_ccy_type,
        fixing_ccy_value=fixing_ccy_value,
        counterparty=counterparty,
        notional=notional,
        strike=strike,
        athena_id=athena_id,
        status=status
    )
    
    # Executar a query
    return execute_query_option_sqlite(**params)

# Configurar o botão "Run (F9)" para chamar a função
def on_run_button_option_click(window,   
    option_menu_trade_date,
    date_entry1_trade_date,
    date_entry2_trade_date,
    option_menu_settlement_date,
    date_entry1_settlement_date,
    date_entry2_settlement_date,
    option_menu_fixing_commodity,
    date_entry1_fixing_commodity,
    date_entry2_fixing_commodity,
    option_menu_fixing_ccy,
    date_entry1_fixing_ccy,
    date_entry2_fixing_ccy,
    counterparty_combobox_opcao,
    notional_entry,
    strike_entry,
    athena_id_entry,
    status_option_menu,
    treeviews
):
    # Obter valores dos widgets
    values = {}
    
     
    # Obter valores dos tipos de filtro
    values['trade_date_option'] = option_menu_trade_date.get() if option_menu_trade_date else None
    values['settlement_date_option'] = option_menu_settlement_date.get() if option_menu_settlement_date else None
    values['fixing_commodity_option'] = option_menu_fixing_commodity.get() if option_menu_fixing_commodity else None
    values['fixing_ccy_option'] = option_menu_fixing_ccy.get() if option_menu_fixing_ccy else None
    values['status_option'] = status_option_menu.get() if status_option_menu else None
    
    # Obter valores dos widgets de entrada para Trade Date
    values['date_entry1_trade_date'] = date_entry1_trade_date.get() if date_entry1_trade_date else None
    values['date_entry2_trade_date'] = date_entry2_trade_date.get() if date_entry2_trade_date else None
    
    # Obter valores dos widgets de entrada para Settlement Date
    values['date_entry1_settlement_date'] = date_entry1_settlement_date.get() if date_entry1_settlement_date else None
    values['date_entry2_settlement_date'] = date_entry2_settlement_date.get() if date_entry2_settlement_date else None
    
    # Obter valores dos widgets de entrada para Fixing Commodity
    values['date_entry1_fixing_commodity'] = date_entry1_fixing_commodity.get() if date_entry1_fixing_commodity else None
    values['date_entry2_fixing_commodity'] = date_entry2_fixing_commodity.get() if date_entry2_fixing_commodity else None
    
    # Obter valores dos widgets de entrada para Fixing CCY
    values['date_entry1_fixing_ccy'] = date_entry1_fixing_ccy.get() if date_entry1_fixing_ccy else None
    values['date_entry2_fixing_ccy'] = date_entry2_fixing_ccy.get() if date_entry2_fixing_ccy else None
    
    # Obter valores dos outros campos
    values['counterparty_combobox_opcao'] = counterparty_combobox_opcao.get() if counterparty_combobox_opcao else None
    values['notional_entry'] = notional_entry.get() if notional_entry else None
    values['strike_entry'] = strike_entry.get() if strike_entry else None
    values['athena_id_entry'] = athena_id_entry.get() if athena_id_entry else None
    
    # Executar a query
    results = run_query_option_from_interface_sqlite(        
        values['trade_date_option'],
        values['date_entry1_trade_date'],
        values['date_entry2_trade_date'],
        values['settlement_date_option'],
        values['date_entry1_settlement_date'],
        values['date_entry2_settlement_date'],
        values['fixing_commodity_option'],
        values['date_entry1_fixing_commodity'],
        values['date_entry2_fixing_commodity'],
        values['fixing_ccy_option'],
        values['date_entry1_fixing_ccy'],
        values['date_entry2_fixing_ccy'],
        values['counterparty_combobox_opcao'],
        values['notional_entry'],
        values['strike_entry'],
        values['athena_id_entry'],
        values['status_option'],
    )

    # Atualizar as Treeviews com os resultados
    if treeviews:
        update_treeviews(results, treeviews)
     
    
    window.destroy()
    
     
def update_treeviews(results: Dict[str, List[List[Any]]], treeviews: Dict[str, ttk.Treeview]):
    """
    Updates the treeviews with the filtered data.
    
    Args:
        results: Dictionary with query results
        treeviews: Dictionary with treeviews for each table
    """
    global tabela_termo_cliente, tabela_termo_b2b, tabela_opcao_cliente, tabela_opcao_b2b
    global tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b
    global tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b

    # Clear treeviews
    for treeview_name, treeview in treeviews.items():
        for item in treeview.get_children():
            treeview.delete(item)
    
    # Insert filtered data
    for table_name, data in results.items():
        if table_name in treeviews:
            for row in data:
                treeviews[table_name].insert('', 'end', values=row)
    
    # Adjust column widths
    ajustar_largura_colunas(tabela_arquivoopcao_cliente, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivoopcao_b2b, colunas_arquivo_opcao, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_cliente, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_arquivotermo_b2b, colunas_arquivo_termo, tabview)
    ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
    ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
    ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)
    
    # Highlight duplicates if there is data
    if tabela_opcao_cliente or tabela_opcao_b2b or tabela_arquivoopcao_cliente or tabela_arquivoopcao_b2b or tabela_fixingsopcao_cliente or tabela_fixingsopcao_b2b:
        highlight_duplicates(tabela_opcao_cliente, 'deals')
        highlight_duplicates(tabela_opcao_b2b, 'deals')
        highlight_duplicates(tabela_arquivoopcao_cliente, 'arquivo')
        highlight_duplicates(tabela_arquivoopcao_b2b, 'arquivo')
        highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
        highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
    
    if tabela_termo_cliente or tabela_termo_b2b or tabela_arquivotermo_cliente or tabela_arquivotermo_b2b or tabela_fixingstermo_cliente or tabela_fixingstermo_b2b:
        highlight_duplicates(tabela_termo_cliente, 'deals')
        highlight_duplicates(tabela_termo_b2b, 'deals')        
        highlight_duplicates(tabela_arquivotermo_cliente, 'arquivo')
        highlight_duplicates(tabela_arquivotermo_b2b, 'arquivo')        
        highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
        highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo')
    
    # Update labels with the number of deals
    global label_qty_deals_cliente_termo    
    global label_qty_deals_b2b_termo    
    qty_deals_termo = number_of_deals(tabela_termo_cliente)
    label_qty_deals_cliente_termo.configure(text=str(qty_deals_termo))      
    qty_deals_termo_b2b = number_of_deals(tabela_termo_b2b)
    label_qty_deals_b2b_termo.configure(text=str(qty_deals_termo_b2b))
    
def email_if(tabela_termo_cliente):
    # Extract necessary data from tables
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    
    # Get today's date
    today = datetime.today().strftime("%d-%b-%Y")
    
    accronyms = []
    mercadorias = []
    for row_id_comm in tabela_termo_cliente.get_children():
        row_comm = tabela_termo_cliente.item(row_id_comm)["values"]
        mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)
        accronym = row_comm[18]
        if mercadoria not in mercadorias:
            mercadorias.append(mercadoria)
        if accronym not in accronyms:
            accronyms.append(accronym)
    
    for accronym in accronyms:
        for mercadoria in mercadorias:
            # Filter items by acronym and commodity type
            filtered_items_participante = [
                row_id for row_id in tabela_termo_cliente.get_children()
                if (row := tabela_termo_cliente.item(row_id)["values"])[18] == accronym and row[1] == today and lookup(row[2], commodities_ric, commodities_mercadoria) == mercadoria and lookup(accronym, cntpy_accronym, cntpy_b3_account) != '73760.10-2'
            ]
            
            if not filtered_items_participante:
                messagebox.showwarning("Attention!", "Economic Affirmation em D0")
                return
            
            if filtered_items_participante:
                # Define items based on filtered_items
                items = [tabela_termo_cliente.item(row_id)["values"] for row_id in filtered_items_participante]
                
                # Calculate financial results                
                b3_account = lookup(accronym, cntpy_accronym, cntpy_b3_account)
                contraparte = lookup(accronym, cntpy_accronym, cntpy_name)
            
                # Construct email body in HTML
                email_body = f"""
                <html>
                <body style="font-family: 'Calibri'; font-size: 11pt;">
                <p>Prezados Senhores,</p>
                <p>Por gentileza, poderiam confirmar os dados da(s) operação(ões) abaixo?</p>

                <p>Conta CETIP {contraparte}: {b3_account}</p>
                <p>Conta CETIP BANCO JP MORGAN S.A: 73760.00-9</p>
                
                <table style="font-family: 'Arial'; font-size: 10pt; border-collapse: collapse; width: auto; border: 1px solid black; text-align: center;">
                    <tr style="font-weight: bold; border: 1px solid black;">
                        <td style="border: 1px solid black;">Posição JPMorgan</td>
                        <td style="border: 1px solid black;">Data Operação</td>
                        <td style="border: 1px solid black;">Valor Base/Quantidade</td>
                        <td style="border: 1px solid black;">Taxa Forward</td>
                        <td style="border: 1px solid black;">Moeda/Ativo</td>
                        <td style="border: 1px solid black;">Inicio Fixing Mercadoria</td>
                        <td style="border: 1px solid black;">Final Fixing Mercadoria</td>
                        <td style="border: 1px solid black;">Fixing Moeda</td>
                        <td style="border: 1px solid black;">Data Vencimento</td>
                    </tr>
                """
                
                def format_strike(value):
                    """Format the currency value, using parentheses for negative numbers."""
                    if value < 0:
                        return f"({abs(value):,.4f})"
                    else:
                        return f"{value:,.4f}"
                    
                for item in items:
                    posicao = "Vendedor" if item[3] == "Sell" else "Comprador"
                    data_operacao = datetime.strptime(item[1], "%d-%b-%Y").strftime("%d/%m/%Y")
                    data_vencimento = datetime.strptime(item[10], "%d-%b-%Y").strftime("%d/%m/%Y")
                    fixing_inicio = datetime.strptime(item[16], "%d-%b-%Y").strftime("%d/%m/%Y")
                    fixing_final = datetime.strptime(item[17], "%d-%b-%Y").strftime("%d/%m/%Y")
                    fixing_ccy = datetime.strptime(item[15], "%d-%b-%Y").strftime("%d/%m/%Y")
                    mercadoria = lookup(item[2], commodities_ric, commodities_mercadoria)
                    valor_base_str = str(item[9]) if isinstance(item[9], int) else item[9]
                    valor_base = abs(float(valor_base_str.replace(",", "")))  # Remove comma and convert to float
                    strike_str = str(item[6]) if isinstance(item[6], int) else item[6]
                    strike = abs(float(strike_str.replace(",", "")))  # Remove comma and convert to float
                    
                    email_body += f"""
                    <tr style="border: 1px solid black;">
                        <td style="border: 1px solid black;">{posicao}</td>
                        <td style="border: 1px solid black;">{data_operacao}</td>
                        <td style="border: 1px solid black;">{f"{valor_base:,.0f}".replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                        <td style="border: 1px solid black;">{format_strike(strike).replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                        <td style="border: 1px solid black;">{item[2]}({mercadoria})</td>
                        <td style="border: 1px solid black;">{fixing_inicio if fixing_inicio != fixing_final else "N/A"}</td>
                        <td style="border: 1px solid black;">{fixing_final}</td>
                        <td style="border: 1px solid black;">{fixing_ccy}</td>
                        <td style="border: 1px solid black;">{data_vencimento}</td>
                    </tr>
                    """
                
                # Close the table after all rows have been added
                email_body += "</table><br>"
                
                # Add footer
                email_body += """                
                <p>Atenciosamente,</p>
                <p>Banco J.P. Morgan S.A. | Av. Brigadeiro Faria Lima, 3729 - 15º andar - São Paulo - SP | T: 55 11 4950 6717 | F: 55 11 4950 3557 |<br>
                brsp_otc_derivatives_ops@jpmorgan.com | jpmorgan.com | Ouvidoria JPMorgan:  Tel.: 0800 – 7700847 / E-mail: ouvidoria.jp.morgan@jpmorgan.com</p>
                </body>
                </html>
                """
                subject = f"Confirmação da(s) Operação(ões) Fechada(s) em {datetime.today().strftime('%d/%m/%Y')} - {contraparte} - Termo de Mercadoria "
            
            # Create and display the email in Outlook
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = ""  # Replace with the actual recipient
            mail.CC = "brazil.otc.ops@jpmorgan.com"
            mail.Subject = subject
            mail.HTMLBody = email_body  # Use HTMLBody for HTML content
            mail.Display()  # Display the email for review
                

def email_Premio(tabela_opcao_cliente):
    # Extrair dados necessários das tabelas    
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    
    # Obter a data de hoje
    today = datetime.today().strftime("%d-%b-%Y")
    
    accronyms = []
    mercadorias = []
    for row_id_comm in tabela_opcao_cliente.get_children():
            row_comm = tabela_opcao_cliente.item(row_id_comm)["values"]                    
            mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)
            accronym = row_comm[18]
            if mercadoria not in mercadorias:
                mercadorias.append(mercadoria)      
            if accronym not in accronyms:    
                accronyms.append(accronym) 
        
    for accronym in accronyms:        
        for mercadoria in mercadorias:
        # Filtrar itens por acrônimo e tipo de mercadoria
            filtered_items_cliente1 = [
                row_id for row_id in tabela_opcao_cliente.get_children()
                if (row := tabela_opcao_cliente.item(row_id)["values"])[18] == accronym and row[22] == today and lookup(row[2], commodities_ric, commodities_mercadoria) == mercadoria and lookup(accronym, cntpy_accronym, cntpy_b3_account) == '73760.10-2']       
            filtered_items_participante = [
                row_id for row_id in tabela_opcao_cliente.get_children()
                if (row := tabela_opcao_cliente.item(row_id)["values"])[18] == accronym and row[22] == today and lookup(row[2], commodities_ric, commodities_mercadoria) == mercadoria and lookup(accronym, cntpy_accronym, cntpy_b3_account) != '73760.10-2']       
            
            if not filtered_items_cliente1 and not filtered_items_participante :
                messagebox.showwarning("Attention!", "Nenhum Prêmio com pagamento D0.")
                return  
            
            if filtered_items_cliente1:
                # Define items based on filtered_items
                items = [tabela_opcao_cliente.item(row_id)["values"] for row_id in filtered_items_cliente1]
                
                # Calcular o resultado apurado e IR
                resultado_apurado = sum(float(item[19].replace(",", "")) if item[3] == "Sell" else float(item[19].replace(",", "")) * -1 for item in items)
                ir = 0 if resultado_apurado >= 0 else resultado_apurado * 0.00005 * -1
                resultado_final = resultado_apurado + ir if resultado_apurado < 0 else resultado_apurado
                b3_account = lookup(accronym, cntpy_accronym, cntpy_b3_account)                
                contraparte = lookup(accronym, cntpy_accronym, cntpy_name)
            
                # Montar o corpo do e-mail em HTML
                email_body = f"""
                <html>
                <body style="font-family: 'Times New Roman'; font-size: 12pt;">
                <p>Prezados Senhores,</p>
                <p>Vimos confirmar a(s) liquidação(ões) da(s) operação(ões) de derivativos abaixo especificada(s):</p>
                
                <table style="font-family: 'Arial'; font-size: 10pt; border-collapse: collapse; width: auto; border: 1px solid black; text-align: center;">
                    <tr style="font-weight: bold; border: 1px solid black;">
                        <td style="border: 1px solid black;">Contrato</td>
                        <td style="border: 1px solid black;">Data Operação</td>
                        <td style="border: 1px solid black;">Data Vencimento</td>
                        <td style="border: 1px solid black;">Moeda/Ativo</td>
                        <td style="border: 1px solid black;">Valor Base/Quantidade</td>
                        <td style="border: 1px solid black;">Resultado Final</td>
                    </tr>      
                """            
            
                def format_currency(value):
                    """Format the currency value, using parentheses for negative numbers."""
                    if value < 0:
                        return f"({abs(value):,.2f})"
                    else:
                        return f"{value:,.2f}"
                    
                for item in items:
                    
                    contrato = item[23]
                    data_operacao = datetime.strptime(item[1], "%d-%b-%Y").strftime("%d/%m/%Y")
                    data_vencimento = datetime.strptime(item[10], "%d-%b-%Y").strftime("%d/%m/%Y")
                    mercadoria = lookup(item[2], commodities_ric, commodities_mercadoria)
                    # Ensure item[9] is a string before using replace
                    valor_base_str = str(item[9]) if isinstance(item[9], int) else item[9]
                    valor_base = abs(float(valor_base_str.replace(",", "")))  # Remove comma and convert to float
                    resultado = float(item[19].replace(",", "")) if item[3] == "Sell" else float(item[19].replace(",", "")) * -1
                
                    email_body += f"""
                    <tr style="border: 1px solid black;">
                        <td style="border: 1px solid black;">{contrato}</td>
                        <td style="border: 1px solid black;">{data_operacao}</td>
                        <td style="border: 1px solid black;">{data_vencimento}</td>
                        <td style="border: 1px solid black;">{item[2]}({mercadoria})</td>
                        <td style="border: 1px solid black;">{f"{valor_base:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                        <td style="border: 1px solid black;">{format_currency(resultado).replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                    </tr>
                    """
                
                # Close the table after all rows have been added
                email_body += "</table><br>"

                
                cnpj = formatar_cnpj(lookup(accronym, cntpy_accronym, cntpy_taxid))
                

                email_body += "</table>"
                

                email_body += f"""
                <table style="font-family: 'Times New Roman'; font-size: 12pt; border-collapse: collapse; width: auto;">
                    <tr>
                        <td style="font-weight: bold;">Resultado Apurado:</td>
                        <td style="font-weight: bold;">R$ {format_currency(resultado_apurado).replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">IR (0,005%):</td>
                        <td style="font-weight: bold;">R$ {f"{ir:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">Resultado Final:</td>
                        <td style="font-weight: bold;">R$ {format_currency(resultado_final).replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                    </tr>
                </table>
                """
                    

                # Adicionar instruções de pagamento
                if resultado_final < 0:
                    email_body += """
                    <p>Conforme entendimentos mantidos, informamos que providenciaremos nesta data a transferência financeira do montante correspondente ao Resultado Final Apurado em vosso favor,
                    conforme os dados a seguir, transmitidos por meio da Autorização Permanente para Liquidação Financeira e/ou confirmados por ligação telefônica:</p>
                    """
                elif resultado_final > 0:
                    email_body += """
                    <p>Sendo assim, informamos que debitaremos os valores descritos acima da conta corrente do Cliente junto ao Banco J.P.Morgan S.A., mediante confirmação de saldo e nos moldes da autorização de débito encaminhada 
                    pelos Srs. Caso não tenham encaminhado autorização de débito, solicitamos que o montante correspondente ao Resultado Final Apurado acima seja transferido em favor do Banco J.P Morgan S.A. nesta data, conforme os dados a seguir:</p>
                    """

            
                bank = lookup(accronym, cntpy_accronym, cntpy_bank)
                ag = lookup(accronym, cntpy_accronym, cntpy_ag)
                cc = lookup(accronym, cntpy_accronym, cntpy_cc)

                # Adicionar dados bancários
                email_body += "<table style='font-family: Times New Roman; font-size: 12pt; border-collapse: collapse; width: auto;'>"
                if resultado_final < 0:
                    email_body += f"""
                    <tr>
                        <td>Nome e nº do banco:</td>
                        <td style="font-weight: bold;">{bank}</td>
                    </tr>
                    <tr>
                        <td>Nº e nome da agência:</td>
                        <td style="font-weight: bold;">{ag}</td>
                    </tr>
                    <tr>
                        <td>Conta–corrente nº:</td>
                        <td style="font-weight: bold;">{cc}</td>
                    </tr>
                    <tr>
                        <td>CNPJ/MF nº:</td>
                        <td style="font-weight: bold;">{cnpj}</td>
                    </tr>
                    """
                else:
                    email_body += """
                    <tr>
                        <td>Nome e nº do banco:</td>
                        <td style="font-weight: bold;">BANCO JP MORGAN S/A - 376</td>
                    </tr>
                    <tr>
                        <td>Nº e nome da agência:</td>
                        <td style="font-weight: bold;">0011</td>
                    </tr>
                    <tr>
                        <td>Conta–corrente nº:</td>
                        <td style="font-weight: bold;">985116003</td>
                    </tr>
                    <tr>
                        <td>CNPJ/MF nº:</td>
                        <td style="font-weight: bold;">33.172.537/0001-98</td>
                    </tr>
                    """
                email_body += "</table>"

                # Adicionar rodapé
                email_body += """
                <p>A presente Ficha de Liquidação é parte integrante e inseparável do Contrato e/ou da Confirmação de Operação de Derivativo em referência.</p>
                <p>Atenciosamente,</p>
                <p>Banco J.P. Morgan S.A. | Av. Brigadeiro Faria Lima, 3729 - 15º andar - São Paulo - SP | T: 55 11 4950 6717 | F: 55 11 4950 3557 |<br>
                brsp_otc_derivatives_ops@jpmorgan.com | jpmorgan.com | Ouvidoria JPMorgan:  Tel.: 0800 – 7700847 / E-mail: ouvidoria.jp.morgan@jpmorgan.com</p>
                </body>
                </html>
                """
                subject = f"(Pagamento de Prêmio) Liquidação de Operação de Derivativo (Commodities) - {datetime.today().strftime('%d/%m/%Y')} - {contraparte}"

            if filtered_items_participante:
                 # Define items based on filtered_items
                items = [tabela_opcao_cliente.item(row_id)["values"] for row_id in filtered_items_participante]
                
                # Calcular o resultado apurado e IR
                resultado_apurado = sum(float(item[19].replace(",", "")) if item[3] == "Sell" else float(item[19].replace(",", "")) * -1 for item in items)
                ir = 0 if resultado_apurado >= 0 else resultado_apurado * 0.00005 * -1
                resultado_final = resultado_apurado + ir if resultado_apurado < 0 else resultado_apurado
                b3_account = lookup(accronym, cntpy_accronym, cntpy_b3_account)                
                contraparte = lookup(accronym, cntpy_accronym, cntpy_name)
                # Montar o corpo do e-mail em HTML
                email_body = f"""
                <html>
                <body style="font-family: 'Times New Roman'; font-size: 12pt;">
                <p>Prezados Senhores,</p>
                <p>Por gentileza, poderiam confirmar os dados da(s) operação(ões) abaixo?</p>

                <p>Conta CETIP {contraparte}: {b3_account}</p>
                <p>Conta CETIP BANCO JP MORGAN S.A: 73760.00-9</p>

                <table style="font-family: 'Arial'; font-size: 10pt; border-collapse: collapse; width: auto; border: 1px solid black; text-align: center;">
                    <tr style="font-weight: bold; border: 1px solid black;">
                        <td style="border: 1px solid black;">Contrato</td>
                        <td style="border: 1px solid black;">Tipo Opção</td>
                        <td style="border: 1px solid black;">Titular</td>
                        <td style="border: 1px solid black;">Data Operação</td>
                        <td style="border: 1px solid black;">Data Vencimento</td>
                        <td style="border: 1px solid black;">Exercício</td>
                        <td style="border: 1px solid black;">Moeda/Ativo</td>
                        <td style="border: 1px solid black;">Fixing Moeda</td>
                        <td style="border: 1px solid black;">Fixing Mercadoria</td>
                        <td style="border: 1px solid black;">Valor Base/Quantidade</td>
                        <td style="border: 1px solid black;">Resultado Final</td>
                        <td style="border: 1px solid black;">Data Pagamento Prêmio</td>
                    </tr>
                """

                def format_currency(value):
                    """Format the currency value, using parentheses for negative numbers."""
                    if value < 0:
                        return f"({abs(value):,.2f})"
                    else:
                        return f"{value:,.2f}"
                    
                for item in items:                    
                    contrato = item[23] if b3_account == 73760.10-2 or (b3_account != 73760.10-2 and item[3] == "Sell") else "Mnemonico Lançador"
                    titular = b3_account if item[3] == "Sell" else "73760.00-9"
                    option_type = "Put" if item[4] == "Option (Put)" else "Call"
                    data_operacao = datetime.strptime(item[1], "%d-%b-%Y").strftime("%d/%m/%Y")
                    data_vencimento = datetime.strptime(item[10], "%d-%b-%Y").strftime("%d/%m/%Y")
                    fixing_moeda = datetime.strptime(item[15], "%d-%b-%Y").strftime("%d/%m/%Y")
                    fixing_mercadoria = datetime.strptime(item[17], "%d-%b-%Y").strftime("%d/%m/%Y")
                    pagamento_premio = datetime.strptime(item[22], "%d-%b-%Y").strftime("%d/%m/%Y")
                    mercadoria = lookup(item[2], commodities_ric, commodities_mercadoria)
                    # Ensure item[9] is a string before using replace
                    valor_base_str = str(item[9]) if isinstance(item[9], int) else item[9]
                    valor_base = abs(float(valor_base_str.replace(",", "")))  # Remove comma and convert to float
                    resultado = float(item[19].replace(",", "")) if item[3] == "Sell" else float(item[19].replace(",", "")) * -1
                
                    email_body += f"""
                    <tr style="border: 1px solid black;">
                        <td style="border: 1px solid black;">{contrato}</td>
                        <td style="border: 1px solid black;">{option_type}</td>
                        <td style="border: 1px solid black;">{titular}</td>
                        <td style="border: 1px solid black;">{data_operacao}</td>
                        <td style="border: 1px solid black;">{data_vencimento}</td>
                        <td style="border: 1px solid black;">EUROPEIA</td>
                        <td style="border: 1px solid black;">{item[2]}({mercadoria})</td>
                        <td style="border: 1px solid black;">{fixing_moeda}</td>
                        <td style="border: 1px solid black;">{fixing_mercadoria}</td>
                        <td style="border: 1px solid black;">{f"{valor_base:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                        <td style="border: 1px solid black;">{format_currency(resultado).replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                        <td style="border: 1px solid black;">{pagamento_premio}</td>
                    </tr>
                    """
                
                # Close the table after all rows have been added
                email_body += "</table><br>"

                
                cnpj = formatar_cnpj(lookup(accronym, cntpy_accronym, cntpy_taxid))
                

                email_body += "</table>"
                

                email_body += f"""
                <table style="font-family: 'Times New Roman'; font-size: 12pt; border-collapse: collapse; width: auto;">
                    <tr>
                        <td style="font-weight: bold;">Resultado Apurado:</td>
                        <td style="font-weight: bold;">R$ {format_currency(resultado_apurado).replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">IR (0,005%):</td>
                        <td style="font-weight: bold;">R$ 0,00</td>
                    </tr>
                    <tr>
                        <td style="font-weight: bold;">Resultado Final:</td>
                        <td style="font-weight: bold;">R$ {format_currency(resultado_apurado).replace(",", " ").replace(".", ",").replace(" ", ".")}</td>
                    </tr>
                </table>
                """
                 

                # Adicionar rodapé
                email_body += """                
                <p>Atenciosamente,</p>
                <p>Banco J.P. Morgan S.A. | Av. Brigadeiro Faria Lima, 3729 - 15º andar - São Paulo - SP | T: 55 11 4950 6717 | F: 55 11 4950 3557 |<br>
                brsp_otc_derivatives_ops@jpmorgan.com | jpmorgan.com | Ouvidoria JPMorgan:  Tel.: 0800 – 7700847 / E-mail: ouvidoria.jp.morgan@jpmorgan.com</p>
                </body>
                </html>
                """
                subject = f" Confirmação das Operações Fechadas em {datetime.today().strftime('%d/%m/%Y')} - {contraparte} - Opção Mercadoria "
            
            # Criar e exibir o e-mail no Outlook
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = ""  # Substitua pelo destinatário real
            mail.CC = "Liquidação"
            mail.Subject = subject
            mail.HTMLBody = email_body  # Use HTMLBody for HTML content
            mail.Display()  # Exibe o e-mail para revisão
            
def formatar_cnpj(cnpj):
    # Remove qualquer caractere que não seja dígito
    cnpj = ''.join(filter(str.isdigit, cnpj))
    # Verifica se o CNPJ tem 14 dígitos
    if len(cnpj) != 14:
        raise ValueError("CNPJ deve ter 14 dígitos")
    # Formata o CNPJ
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
      
def importar_operacoes(tabview, abas_existentes, tree):    
    global janela
    # Perguntar ao usuário se deseja seguir com o cadastro
    resposta = messagebox.askyesno("Import Deals", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")        
        shared_mailbox = outlook.Folders['brazil.otc.ops@jpmorgan.com']
        inbox = shared_mailbox.Folders["Inbox"]
        new_deals_folder = inbox.Folders["New deals"]
        b2bs_automatic_folder = new_deals_folder.Folders["B2Bs Automatic"]
        

        messages_option = inbox.Items.Restrict("@SQL=\"http://schemas.microsoft.com/mapi/proptag/0x0E1D001F\" LIKE '%Brazil Booking Recap %'")
        messages_swap = inbox.Items.Restrict("@SQL=\"http://schemas.microsoft.com/mapi/proptag/0x0E1D001F\" LIKE '%Brazil Booking Recap %Swap %'")
        filtered_messages = list(messages_option) + list(messages_swap)
        SID = getpass.getuser()
        SID = SID[0].upper() + SID[1:]

        if not filtered_messages:
            messagebox.showwarning("Attention!", "Nenhum e-mail encontrado.")
            return   
    
        # Verificar se a aba "Monitor" já existe
        if "Monitor" in abas_existentes:
            pass
        else:
            # Inicializar o monitor de operações e obter os tabviews
            monitor_operacoes(tabview, aba_inicio, abas_existentes, tree)
            abas_existentes.append("Monitor")              

        for message in filtered_messages:
            if message.Class == 43:
                subject = message.Subject
                body = message.HTMLBody
                start_pos = body.find("<table")
                end_pos = body.find("</table>") + len("</table>")
                if start_pos > 0 and end_pos > 0:
                    table_html = body[start_pos:end_pos]
                    soup = BeautifulSoup(table_html, "html.parser")
                    table = soup.find("table")
                    rows = table.find_all("tr")                    
                    if "CANCEL" in subject:
                         message.Delete()
                    else:
                        if "Option" in subject:
                            colunas = colunas_opcao
                            treeview_cliente = tabela_opcao_cliente
                            treeview_b2b = tabela_opcao_b2b
                        elif "Swap" in subject:
                            colunas = colunas_termo
                            treeview_cliente = tabela_termo_cliente
                            treeview_b2b = tabela_termo_b2b
                    

                        cliente_data = [cell.get_text() for cell in rows[2].find_all("td")]
                        while len(cliente_data) < len(colunas):
                            cliente_data.append("")
                        if "Option" in subject:
                            # Gerar e inserir o código Mnemonico
                            type_value = cliente_data[3]  # Supondo que 'Type' está na posição 3
                            mnemonico = generate_mnemonic_cliente(type_value)
                            cliente_data[23] = mnemonico  # Inserir o Mnemonico na posição correta
                            identificador = str(random.randint(1,9)) + ''.join([str(random.randint(0,9)) for _ in range(10)])                        
                            cliente_data[24] = identificador
                            cliente_data[25] = '1'
                            cliente_data[26] = 'New'
                            cliente_data[27] =  SID
                            treeview_cliente.insert("", "end", values=cliente_data)
                        elif "Swap" in subject:
                            identificador = str(random.randint(1,9)) + ''.join([str(random.randint(0,9)) for _ in range(10)])      
                            cliente_data[19] = identificador
                            cliente_data[20] = '1'
                            cliente_data[21] = 'New'
                            cliente_data[22] =  SID
                            treeview_cliente.insert("", "end", values=cliente_data)

                        b2b_data = [cell.get_text() for cell in rows[3].find_all("td")]
                        while len(b2b_data) < len(colunas):
                            b2b_data.append("")

                        if "Option" in subject:
                            # Gerar e inserir o código Mnemonico
                            type_value = b2b_data[3]  # Supondo que 'Type' está na posição 3
                            mnemonico = generate_mnemonic_b2b(type_value)
                            b2b_data[18] = 'Lawton'
                            b2b_data[23] = mnemonico  # Inserir o Mnemonico na posição correta
                            b2b_data[24] = identificador
                            b2b_data[25] = '1'
                            b2b_data[26] = 'New'
                            b2b_data[27] =  SID
                            treeview_b2b.insert("", "end", values=b2b_data)
                        elif "Swap" in subject:
                            b2b_data[18] = 'Lawton'
                            b2b_data[19] = identificador
                            b2b_data[20] = '1'
                            b2b_data[21] = 'New'
                            b2b_data[22] =  SID
                            treeview_b2b.insert("", "end", values=b2b_data) 
                        # Ajustar a largura das colunas
                        ajustar_largura_colunas(treeview_cliente, colunas, tabview)
                        ajustar_largura_colunas(treeview_b2b, colunas, tabview)           

                        # Vincular eventos de duplo clique e navegação                    
                        vincular_evento_duplo_clique_status(treeview_cliente, colunas)
                        vincular_evento_duplo_clique_status(treeview_b2b, colunas)

                        vincular_navegacao_setas(treeview_cliente)
                        vincular_navegacao_setas(treeview_b2b) 
                        ajustar_operacoes_opcao()
                        ajustar_operacoes_termo()
                        # Supondo que você já tenha um Treeview com dados inseridos
                        highlight_duplicates(treeview_cliente, 'deals')
                        highlight_duplicates(treeview_b2b, 'deals')                       
                        message.Move(b2bs_automatic_folder)           

        # Ordenar as tabelas pela coluna "SettlementDate"        
        ordenar_por(tabela_termo_cliente, ["SettlementDate"], False)
        ordenar_por(tabela_termo_b2b, ["SettlementDate"], False)
        ordenar_por(tabela_opcao_cliente, ["SettlementDate"], False)
        ordenar_por(tabela_opcao_b2b, ["SettlementDate"], False)         
        # Atualiza o label com a quantidade de deals  
        global label_qty_deals_cliente_termo
        global label_qty_deals_cliente_opcao
        global label_qty_deals_b2b_termo
        global label_qty_deals_b2b_opcao
        qty_deals_termo = number_of_deals(tabela_termo_cliente)
        label_qty_deals_cliente_termo.configure(text=str(qty_deals_termo))     
        qty_deals_opcao = number_of_deals(tabela_opcao_cliente)
        label_qty_deals_cliente_opcao.configure(text=str(qty_deals_opcao))   
        qty_deals_termo_b2b = number_of_deals(tabela_termo_b2b)
        label_qty_deals_b2b_termo.configure(text=str(qty_deals_termo_b2b))   
        qty_deals_opcao_b2b = number_of_deals(tabela_opcao_b2b)
        label_qty_deals_b2b_opcao.configure(text=str(qty_deals_opcao_b2b))        

        # Chame a função após o preenchimento das tabelas de opção
        tabela_feriados = calendarios_bolsas(tabview, abas_existentes)  # Inicializa e obtém tabela_feriados
        # Chame a função após o preenchimento das tabelas de opção
        preencher_fixings(tabview, abas_existentes)

        messagebox.showinfo("Success!", "Deals imported!")
    
    except Exception as e:
        print(f"Erro ao importar operações: {e}")
        traceback.print_exc()


   
def formatar_data(data_str):
    # Converte a string de data no formato dd/mm/yyyy para o formato desejado
    data = datetime.strptime(data_str, "%d/%m/%Y")
    meses_portugues = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]
    dia = f"{data.day:02}"
    mes = meses_portugues[data.month - 1]
    ano = data.year
    return f"{dia} de {mes} de {ano}"  
    
    
def convert_pdf_to_images(pdf_path, downloads_path, scale=1):
    doc = fitz.open(pdf_path)
    image_paths = []
    for page_number in range(len(doc)):
        page = doc.load_page(page_number)
        mat = fitz.Matrix(scale, scale)
        pix = page.get_pixmap(matrix=mat)
        image_path = os.path.join(downloads_path, f"page_{page_number + 1}.png")
        pix.save(image_path)
        image_paths.append(image_path)
    doc.close()
    return image_paths

def carregar_e_exibir(pdf_path, downloads_path):
    global image_paths  # Make image_paths global to access it in delete_confirmation
    image_paths = convert_pdf_to_images(pdf_path, downloads_path, scale=1)
    for image_path in image_paths:
        image = Image.open(image_path)
        photo = ctk.CTkImage(image, size=(1024, 768))
        label = ctk.CTkLabel(inner_frame, image=photo)
        label.configure(text='')
        label.image = photo
        label.pack(pady=0)
    canvas.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))

def get_email_signature():
    # Path to the Outlook signatures folder
    sig_path = os.path.join(os.environ["APPDATA"], "Microsoft", "Signatures")
    sig_file = "INTERNO.htm"  # Name of the signature file

    # Check if the signature file exists
    if not os.path.exists(os.path.join(sig_path, sig_file)):
        sig_file = "(INTERNAL) Luiza.htm"  # Alternative signature file name
    if not os.path.exists(os.path.join(sig_path, sig_file)):
        sig_file = "Internal.htm"  # Another alternative signature file name

    # Read the signature file if it exists
    signature_path = os.path.join(sig_path, sig_file)
    if os.path.exists(signature_path):
        try:
            with open(signature_path, 'r', encoding='utf-8') as file:
                signature = file.read()
        except UnicodeDecodeError:
            # Fallback to cp1252 encoding if utf-8 fails
            with open(signature_path, 'r', encoding='cp1252') as file:
                signature = file.read()
    else:
        signature = ""

    return signature

def create_outlook_email(subject, to_email, cc_email, body, attachments):
    try:
        # Create an Outlook application instance
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)  # 0: olMailItem

        # Set email properties
        mail.Subject = subject
        mail.To = to_email
        mail.CC = cc_email

        # Get the email signature
        signature = get_email_signature()

        # Set the email body with the signature as HTML
        mail.HTMLBody = f"{body}<br><br>{signature}"

        # Attach files
        for attachment in attachments:
            mail.Attachments.Add(attachment)

        # Display the email (opens in Outlook for review and sending)
        mail.Display()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to create email in Outlook: {e}") 
        
def update_bookmark(doc, bookmark_name, text):
    """Update a bookmark in a Word document using COM."""
    if doc.Bookmarks.Exists(bookmark_name):
        bookmark = doc.Bookmarks(bookmark_name)
        range_ = bookmark.Range
        range_.Text = text
        doc.Bookmarks.Add(bookmark_name, range_)

def update_table(doc, bookmark_name, dados_linhas):
    """Update a table in a Word document using COM."""
    if doc.Bookmarks.Exists(bookmark_name):
        bookmark = doc.Bookmarks(bookmark_name)
        table = bookmark.Range.Tables(1)  # Assume there's only one table in the bookmark range

        # Clear existing rows except the header
        while table.Rows.Count > 2:
            table.Rows(3).Delete()

        # Add new rows and fill them with data
        for i, row_data in enumerate(dados_linhas, start=1):
            if i > 1:
                 table.Rows.Add()
            for j, cell_data in enumerate(row_data, start=1):
                table.Cell(i + 1, j).Range.Text = str(cell_data)
                
# Function to open a new window with the progress bar
def open_progress_window_opcao():
    global janela, progress_window, label_counterparty, label_mercadoria, progressbar_confirmation
    progress_window = ctk.CTkToplevel(janela)
    progress_window.title("Option Confirmation Progress")
    progress_window.lift()
    progress_window.focus_set()

    label_counterparty = ctk.CTkLabel(progress_window, text="Cliente: Initializing")
    label_counterparty.grid(column=0, row=0, pady=2, padx=5)

    label_mercadoria = ctk.CTkLabel(progress_window, text="Mercadoria: Initializing")
    label_mercadoria.grid(column=1, row=0, pady=2, padx=5)
    
    progressbar_confirmation = ctk.CTkProgressBar(
        master=progress_window,
        orientation='horizontal',
        width=300,
        height=20,
        corner_radius=1,
        mode="indeterminate",
        fg_color="lightgray",
        progress_color="blue"
    )
    progressbar_confirmation.grid(column=0, columnspan=2, row=1, pady=5, padx=5)

def update_progress_window(nome_cliente, mercadoria):
    label_counterparty.configure(text=f"Cliente: {nome_cliente}")
    label_mercadoria.configure(text=f"Mercadoria: {mercadoria}")

def process_documents(tabview, accronyms, indexes_conf, mercadorias):
    # Initialize COM library
    pythoncom.CoInitialize()
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    
    try:
        # Start the progress bar
        progressbar_confirmation.start()

        for index_conf in indexes_conf:
            for accronym in accronyms:
                for mercadoria in mercadorias:
                    update_progress_window(accronym, mercadoria)

                    # Filter items for option contract types
                    filtered_items = [
                        row_id for row_id in tabela_opcao_cliente.get_children()
                        if (row := tabela_opcao_cliente.item(row_id)["values"])[18] == accronym and lookup(row[2], commodities_ric, commodities_mercadoria) == mercadoria and str(row[-3]) == str(index_conf) and row[-2] != "New" and row[-2] != "Pending Review"
                    ]

                    opcao_asiatico = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] != tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_mercadoria) != "OLEO DE PALMA EM USD" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) != "BLOOMBGERG" and tabela_opcao_cliente.item(item)["values"][2] != "CO1-2"]
                    opcao_bullet = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] == tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_mercadoria) != "OLEO DE PALMA EM USD" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) != "BLOOMBGERG"]
                    opcao_platts_asiatico_BRL = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] != tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] == "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG" ]
                    opcao_platts_bullet_BRL = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] == tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] == "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG" ]
                    opcao_platts_asiatico = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] != tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG" ]
                    opcao_platts_bullet = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] == tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG" ]
                    opcao_co1_2 = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] != tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] != "BRR" and tabela_opcao_cliente.item(item)["values"][2] == "CO1-2" ]            
                    opcao_palmoil_asiatico = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] != tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_mercadoria) == "OLEO DE PALMA EM USD"]
                    opcao_palmoil_bullet = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] == tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_opcao_cliente.item(item)["values"][2], commodities_ric, commodities_mercadoria) == "OLEO DE PALMA EM USD"]
                    opcao_asiatico_BRL = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] != tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] == "BRR"]
                    opcao_bullet_BRL = [item for item in filtered_items if tabela_opcao_cliente.item(item)["values"][16] == tabela_opcao_cliente.item(item)["values"][17] and tabela_opcao_cliente.item(item)["values"][8] == "BRR"]
                    nome_empresa = lookup(accronym, cntpy_accronym, cntpy_name) if accronym != "" else ""
                    # Assuming data_atual is a string in the format 'YYYY-MM-DD' or similar
                    data_atual_str = row[1]  # Replace with the actual format of your date string

                    # Convert the string to a datetime object
                    data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Adjust the format as needed
                    

                    # Now you can use strftime
                    mes2 = data_atual.strftime("%m")

                    # Dicionário para os meses em português
                    meses_portugues = {
                        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
                        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
                        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
                    }

                    # Obter o nome do mês em português
                    mes = meses_portugues[mes2]
                    ano = data_atual.strftime("%Y")
                    dia = data_atual.strftime("%d")
                        
                    if str(index_conf) == 1:
                        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
                        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "OPÇÃO", f"{mercadoria}")                        
                    else:     
                        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
                        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "OPÇÃO",  f"#{index_conf} {mercadoria}")
                    
                    # Check for existing files
                    docx_files = glob.glob(os.path.join(downloads_path, "*CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS*.doc"))
                    pdf_files = glob.glob(os.path.join(downloads_path, "*CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS*.pdf"))
                    
                    if docx_files and pdf_files:
                        docx_path = docx_files[0]  # Assuming you want the first match
                        pdf_path = pdf_files[0]  # Assuming you want the first match
                        criar_preview(downloads_path, nome_empresa, pdf_path, docx_path, data_atual, mercadoria, index_conf)
                    else:   
                        # Create documents for lines with different values
                        if opcao_asiatico:
                            criar_documento(opcao_asiatico, "diferentes", accronym, index_conf, mercadoria)
                        
                        if opcao_co1_2:
                            criar_documento(opcao_co1_2, "diferentes", accronym, index_conf, mercadoria, tipo_co1_2=True)

                        if opcao_asiatico_BRL:
                            criar_documento(opcao_asiatico_BRL, "diferentes", accronym, index_conf, mercadoria, tipo_brl=True)

                        if opcao_platts_asiatico_BRL:
                            criar_documento(opcao_platts_asiatico_BRL, "diferentes", accronym, index_conf, mercadoria, tipo_brl=True, tipo_platts=True)
                        
                        if opcao_platts_asiatico:
                            criar_documento(opcao_platts_asiatico, "diferentes", accronym, index_conf, mercadoria, tipo_brl=False, tipo_platts=True)
                        
                        if opcao_palmoil_asiatico:
                            criar_documento(opcao_palmoil_asiatico, "diferentes", accronym, index_conf, mercadoria, tipo_brl=False, tipo_palmoil=True)
                        
                        # Create documents for lines with equal values
                        if opcao_bullet:
                            criar_documento(opcao_bullet, "iguais", accronym, index_conf, mercadoria)   
                                                    
                        if opcao_platts_bullet:
                            criar_documento(opcao_platts_bullet, "iguais", accronym, index_conf, mercadoria, tipo_brl=False, tipo_platts=True)

                        if opcao_platts_bullet_BRL:
                            criar_documento(opcao_platts_bullet_BRL, "iguais", accronym, index_conf, mercadoria, tipo_brl=True, tipo_platts=True)

                        if opcao_bullet_BRL:
                            criar_documento(opcao_bullet_BRL, "iguais", accronym, index_conf, mercadoria, tipo_brl=True, tipo_platts=False, tipo_palmoil=False)
                        
                        if opcao_palmoil_bullet:
                            criar_documento(opcao_palmoil_bullet, "iguais", accronym, index_conf, mercadoria, tipo_brl=False, tipo_palmoil=True)

    finally:
        # Uninitialize COM library
        pythoncom.CoUninitialize()

    # Stop the progress bar and close the window after processing is complete
    progressbar_confirmation.stop()
    progress_window.destroy()    

def criar_documento(linhas, tipo_linha, accronym, index_conf, mercadoria, tipo_brl=False, tipo_platts=False, tipo_palmoil=False, tipo_co1_2=False):
    # Lists to accumulate data for database update
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    Counterparties = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []

    xlookup_dict = {
        "JAN": "F", "FEB": "G", "MAR": "H", "APR": "J", "MAY": "K",
        "JUN": "M", "JUL": "N", "AUG": "Q", "SEP": "U", "OCT": "V",
        "NOV": "X", "DEC": "Z"
    }

    def previous_workday(start_date, workdays):
        workdays = int(workdays)
        previous_day = start_date - timedelta(days=workdays)
        while previous_day.weekday() >= 5:  # 5 = Saturday, 6 = Sunday
            previous_day -= timedelta(days=1)
        return previous_day

    def XLookup(value):
        return xlookup_dict.get(value.upper(), "")
    
    global entry_filtro_commodities
    entry_filtro_commodities.delete(0, tk.END)
    global tabela_commodities_data
    for item in tabela_commodities_data:            
            tabela_commodities.insert("", "end", values=item)
    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    first_item_values = tabela_opcao_cliente.item(linhas[0], 'values')
    market = first_item_values[2]                            
    cntpy = first_item_values[18]
    data_formatada_cgd = formatar_data(lookup(cntpy, cntpy_accronym, cntpy_cgd))
    nome_empresa = lookup(cntpy, cntpy_accronym, cntpy_name) if cntpy != "" else ""
    cnpj_formatado = formatar_cnpj(lookup(cntpy, cntpy_accronym, cntpy_taxid))
    trade_date = first_item_values[1]
    trade_date_pg_2 = safe_date_conversion_dd_MM(trade_date)
    data_por_extenso_assinatura = data_por_extenso(trade_date)

    # Check for missing values and show warnings
    missing_messages = []
    if cnpj_formatado is None:
        missing_messages.append(f"Missing TAX ID for {accronym}")
    if data_formatada_cgd is None:
        missing_messages.append(f"Missing CGD for {accronym}")
    if nome_empresa is None:
        missing_messages.append(f"Missing Counterparty for {accronym}")

    if missing_messages:
        messagebox.showwarning("Missing Information", "\n".join(missing_messages))
        return  # Interrupt the function if any value is missing

   
        # Determine the template path based on the type
    if tipo_brl:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\OPÇÃO COMMODITY - BRL.doc"
    elif tipo_co1_2:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\OPÇÃO COMMODITY-CO1-2.doc"
    elif tipo_platts and not tipo_brl:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\OPÇÃO COMMODITY - PLATTS.doc"
    elif tipo_platts and tipo_brl:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\OPÇÃO COMMODITY - BRL - PLATTS.doc"   
    elif tipo_palmoil:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\OPÇÃO COMMODITY - PALM OIL.doc"
    else:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\OPÇÃO COMMODITY.doc"

    # Open the Word document using COM
    word = win32.Dispatch('Word.Application')
    word.Visible = False
    doc = word.Documents.Open(template_path)                        
  
    # Update bookmarks
    update_bookmark(doc, "CGD", data_formatada_cgd)
    update_bookmark(doc, "CNPJ_CONTRAPARTE", cnpj_formatado)
    update_bookmark(doc, "CONTRAPARTE", nome_empresa)
    update_bookmark(doc, "CONTRAPARTE_ASSINATURA", nome_empresa)
    update_bookmark(doc, "DATA_NEGOCIACAO_EXTENSO", data_por_extenso_assinatura)
    update_bookmark(doc, "DATA_NEGOCIACAO_PG2", trade_date_pg_2)

    # Definir os dados das linhas
    dados_linhas = []
    for index, item in enumerate(linhas, start=1):  # Começa a contagem em 1
        values = tabela_opcao_cliente.item(item, 'values')        
        TradeDates.append(values[1])
        Counterparties.append(values[18])
        AthenaIDs.append(values[0])                
        Confirmations.append("OK")        
        Identifiers.append(str(values[-4]))
        Indexes.append(str(values[-3]))
        Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        
        market = values[2]
        accronym = values[18]
        mercadoria = lookup(values[2], commodities_ric, commodities_mercadoria)
        strike = strike_confirmation(market, values[6], values[8], commodities_ric, commodities_factor, tabview).replace(",", " ").replace(".", ",").replace(" ", ".")
        trade_date = values[1]
        valor_base_str = values[9].replace(",", ".").replace("-", "")
        ptax = safe_date_conversion_dd_MM(values[15])
        exchange = lookup(market, commodities_ric, commodities_exchange)
        premium = values[19].replace(",", "")
        valor_numerico = float(premium)
        # Formate o número
        premium = f"{valor_numerico:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")
        
        
        settlement_date = safe_date_conversion_dd_MM(values[10])
        
        # Check for CO1-2 market
        if market == "CO1-2":
            # Extract necessary date components
            settlement_date = safe_date_conversion_dd_MM(first_item_values[10])
            reference_date = safe_date_conversion_dd_MM(first_item_values[17])
            
            # Convert dates to datetime objects
            settlement_date_obj = datetime.strptime(settlement_date, "%d/%m/%Y")
            reference_date_obj = datetime.strptime(reference_date, "%d/%m/%Y")
            
            # Extract month and year
            settlement_month = settlement_date_obj.strftime("%b").upper()
            settlement_year = settlement_date_obj.strftime("%Y")[-1]
            
            # Determine the market code based on the reference date
            if reference_date_obj.month == 12:
                # December logic
                two_days_before = previous_workday(reference_date_obj, 2)  
                one_day_before = previous_workday(reference_date_obj, 1)                      
                market_1 = f"COG{settlement_year}"            
                market_2 = f"COH{settlement_year}"
                market = (f"Para as Datas de Verificação entre a Data Inicial de Verificação de Mercadoria e {two_days_before.strftime('%d/%m/%Y')} "
                        f"significa {market_1} e para as Datas de Verificação em {one_day_before.strftime('%d/%m/%Y')} e {reference_date}, significa {market_2}")
            else:
                # Non-December logic
                one_day_before = previous_workday(reference_date_obj, 1) 
                if settlement_date_obj.month == 12:
                    market_1 = f"CO{XLookup('JAN')}{int(settlement_year) + 1}"
                else:
                    market_1 = f"CO{XLookup(settlement_month)}{settlement_year}"
                if settlement_date_obj.month == 11:
                    market_2 = f"CO{XLookup('JAN')}{int(settlement_year) + 1}"
                else:
                    market_2 = f"CO{XLookup(settlement_month)}{settlement_year}"  
                market = (f"Para as Datas de Verificação entre a Data Inicial de Verificação de Mercadoria e {one_day_before.strftime('%d/%m/%Y')} "
                        f"significa {market_1} e para a Data de Verificação {reference_date}, significa {market_2}")
        elif market == "NACX0005":
            market = "VLSFO (AMFSA00)"
        elif market == "PEURNPHY":
            market = "PAAL00"
        elif market == "PCRUDTB1":
            market = "PCAAS00"                           

        if tipo_linha == "diferentes":
            if tipo_palmoil:
                dados_linhas.append([
                    index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, exchange, 'Não Aplicável', valor_base_str, 'MYR USD', safe_date_conversion_dd_MM(values[17]),
                    ptax, 'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])                                    
            elif tipo_brl and tipo_platts:
                dados_linhas.append([
                    index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, "PLATTS", valor_base_str, ptax,
                    'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), 'Não Aplicável', safe_date_conversion_dd_MM(values[10])
                ])
            elif tipo_brl and not tipo_platts:
                dados_linhas.append([
                        index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, exchange, valor_base_str, ptax,
                        'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  safe_date_conversion_dd_MM(values[16]),
                        safe_date_conversion_dd_MM(values[17]), 'Não Aplicável', safe_date_conversion_dd_MM(values[10])
                    ])
            elif tipo_platts and not tipo_brl:
                dados_linhas.append([
                        index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, "PLATTS", valor_base_str, ptax,
                        'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  safe_date_conversion_dd_MM(values[16]),
                        safe_date_conversion_dd_MM(values[17]), 'Não Aplicável', safe_date_conversion_dd_MM(values[10])
                    ])
            elif tipo_co1_2:
                    dados_linhas.append([
                    index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, exchange, valor_base_str, ptax,
                    'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), 'Não Aplicável', safe_date_conversion_dd_MM(values[10])
                ])    
                    
            else:
                dados_linhas.append([
                    index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, exchange, valor_base_str, ptax,
                    'Parte B' if values[3] == "Sell" else 'Parte A','R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), 'Não Aplicável', safe_date_conversion_dd_MM(values[10])
                ])
        else:
            if tipo_palmoil:
                dados_linhas.append([
                    index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, exchange, 'Não Aplicável', valor_base_str, 'MYR USD', safe_date_conversion_dd_MM(values[17]),
                    ptax, 'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  'Não Aplicável',
                    'Não Aplicável', safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])   
            elif tipo_brl and tipo_platts:
                dados_linhas.append([
                    index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, "PLATTS", valor_base_str, ptax,
                    'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  'Não Aplicável', 'Não Aplicável', safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
            elif tipo_brl and not tipo_platts:
                dados_linhas.append([
                        index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, exchange, valor_base_str, ptax,
                        'Parte B' if values[3] == "Sell" else 'Parte A','R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  'Não Aplicável', 'Não Aplicável', safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                    ])
            elif tipo_platts and not tipo_brl:
                dados_linhas.append([
                        index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, "PLATTS", valor_base_str, ptax,
                        'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  'Não Aplicável', 'Não Aplicável', safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                    ])
            else:
                dados_linhas.append([
                    index, values[23], "Venda" if values[4] == "Option (Put)" else "Compra", 'Europeia', market, exchange, valor_base_str, ptax,
                    'Parte B' if values[3] == "Sell" else 'Parte A', 'R$' + str(premium), safe_date_conversion_dd_MM(values[22]), strike,  'Não Aplicável', 'Não Aplicável', safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
    
 # Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += ["OK"] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
    # Update the table with the data
    update_table(doc, "TABELA_OPERACOES", dados_linhas)   
  
    # Assuming data_atual is a string in the format 'YYYY-MM-DD' or similar
    data_atual_str = values[1]  # Replace with the actual format of your date string

    # Convert the string to a datetime object
    data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Adjust the format as needed
    

    # Now you can use strftime
    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    if index == 1:
        update_bookmark(doc, "TITULO", f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS nº {values[23]}")
        docx_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS nº {values[23]}.doc"
        pdf_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS nº {values[23]}.pdf"                            
    else:
        docx_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS - {safe_date_conversion_confirma(values[1])}.doc"
        pdf_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS - {safe_date_conversion_confirma(values[1])}.pdf"
        
    if str(index_conf) == '1':
        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
        if diretorio_raiz_confirmation is None:
            missing_messages.append(f"Missing confirmation path mapping for {accronym}")
        if missing_messages:
            messagebox.showwarning("Missing Information", "\n".join(missing_messages))
            return  # Interrupt the function if any value is missing
        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "OPÇÃO", f"{mercadoria}")
        if not os.path.exists(downloads_path):
                    os.makedirs(downloads_path)
    else:     
        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
        if diretorio_raiz_confirmation is None:
            missing_messages.append(f"Missing confirmation path mapping for {accronym}")
        if missing_messages:
            messagebox.showwarning("Missing Information", "\n".join(missing_messages))
            return  # Interrupt the function if any value is missing
        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "OPÇÃO",  f"#{index_conf} {mercadoria}")
        if not os.path.exists(downloads_path):
            os.makedirs(downloads_path)   
                        
    docx_path = os.path.join(downloads_path, docx_filename)
    pdf_path = os.path.join(downloads_path, pdf_filename)                           
    
    # Save the document using SaveAs                                                
    doc.SaveAs2(pdf_path, FileFormat=17)    
    doc.SaveAs2(docx_path)    
    word.Quit()                          
    # Release COM objects
    del doc
    del word                       
    criar_preview(downloads_path, nome_empresa, pdf_path, docx_path, data_atual, mercadoria,  index_conf)          
               
def generate_confirmation_opcao(tabview):
    global tabela_opcao_cliente
    resposta = messagebox.askyesno("Confirmation generate", "Wish to proceed?")
    if not resposta:
        return

    # Open the progress window immediately
    open_progress_window_opcao()

    # Extract necessary data
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()

    # Extract unique acronyms from the "Cliente" table
    accronyms = []
    indexes_conf = []
    mercadorias = []
    for row_id_comm in tabela_opcao_cliente.get_children():
        row_comm = tabela_opcao_cliente.item(row_id_comm)["values"]
        mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)
        accronym = row_comm[18]
        index_conf = row_comm[-3]
        if accronym not in accronyms:
            accronyms.append(accronym)
        if index_conf not in indexes_conf:
            indexes_conf.append(index_conf)
        if mercadoria not in mercadorias:
            mercadorias.append(mercadoria)

    # Start the document processing in a separate thread
    threading.Thread(target=process_documents, args=(tabview, accronyms, indexes_conf, mercadorias)).start()
 
# Function to open a new window with the progress bar
def open_progress_window_termo():
    global janela, progress_window, label_counterparty, label_mercadoria, progressbar_confirmation
    progress_window = ctk.CTkToplevel(janela)
    progress_window.title("Termo Confirmation Progress")
    progress_window.lift()
    progress_window.focus_set()

    label_counterparty = ctk.CTkLabel(progress_window, text="Cliente: Initializing")
    label_counterparty.grid(column=0, row=0, pady=2, padx=5)

    label_mercadoria = ctk.CTkLabel(progress_window, text="Mercadoria: Initializing")
    label_mercadoria.grid(column=1, row=0, pady=2, padx=5)

    progressbar_confirmation = ctk.CTkProgressBar(
        master=progress_window,
        orientation='horizontal',
        width=300,
        height=20,
        corner_radius=1,
        mode="indeterminate",
        fg_color="lightgray",
        progress_color="blue"
    )
    progressbar_confirmation.grid(column=0, columnspan=2, row=1, pady=5, padx=5)

def update_progress_window_termo(nome_cliente, mercadoria):
    label_counterparty.configure(text=f"Cliente: {nome_cliente}")
    label_mercadoria.configure(text=f"Mercadoria: {mercadoria}")

def process_documents_termo(tabview, accronyms, indexes_conf, mercadorias):
    # Initialize COM library
    pythoncom.CoInitialize()
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    
    try:
        # Start the progress bar
        progressbar_confirmation.start()

        for index_conf in indexes_conf:
            for accronym in accronyms:
                for mercadoria in mercadorias:
                    update_progress_window_termo(accronym, mercadoria)

                    filtered_items = [
                        row_id for row_id in tabela_termo_cliente.get_children()
                        if (row := tabela_termo_cliente.item(row_id)["values"])[18] == accronym and lookup(row[2], commodities_ric, commodities_mercadoria) == mercadoria and str(row[-3]) == str(index_conf) and row[-2] != "New" and row[-2] != "Pending Review"
                    ]

                    termo_asiatico_BRL = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] != tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] == "BRR"]
                    termo_bullet_BRL = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] == tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] == "BRR"]
                    termo_asiatico = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] != tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] != "BRR"]
                    termo_bullet = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] == tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] != "BRR"  and lookup(tabela_termo_cliente.item(item)["values"][2], commodities_ric, commodities_mercadoria) != "OLEO DE PALMA EM USD"]
                    termo_platts_asiatico_BRL = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] != tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] == "BRR" and lookup(tabela_termo_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG"]
                    termo_platts_bullet_BRL = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] == tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] == "BRR" and lookup(tabela_termo_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG"]
                    termo_platts_asiatico = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] != tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_termo_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG"]
                    termo_platts_bullet = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] == tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_termo_cliente.item(item)["values"][2], commodities_ric, commodities_exchange) == "BLOOMBGERG"]
                    termo_palmoil_asiatico = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] != tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_termo_cliente.item(item)["values"][2], commodities_ric, commodities_mercadoria) == "OLEO DE PALMA EM USD"]
                    termo_palmoil_bullet = [item for item in filtered_items if tabela_termo_cliente.item(item)["values"][16] == tabela_termo_cliente.item(item)["values"][17] and tabela_termo_cliente.item(item)["values"][8] != "BRR" and lookup(tabela_termo_cliente.item(item)["values"][2], commodities_ric, commodities_mercadoria) == "OLEO DE PALMA EM USD"]
                    
                    nome_empresa = lookup(accronym, cntpy_accronym, cntpy_name) if accronym != "" else ""
                    # Assuming data_atual is a string in the format 'YYYY-MM-DD' or similar
                    data_atual_str = row[1]  # Replace with the actual format of your date string

                    # Convert the string to a datetime object
                    data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Adjust the format as needed
                    

                    # Now you can use strftime
                    mes2 = data_atual.strftime("%m")

                    # Dicionário para os meses em português
                    meses_portugues = {
                        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
                        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
                        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
                    }

                    # Obter o nome do mês em português
                    mes = meses_portugues[mes2]
                    ano = data_atual.strftime("%Y")
                    dia = data_atual.strftime("%d")
                        
                    if index_conf == 1:
                        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
                        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "TERMO", f"{mercadoria}")                        
                    else:     
                        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
                        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "TERMO",  f"#{index_conf} {mercadoria}")
                    
                    # Check for existing files
                    docx_files = glob.glob(os.path.join(downloads_path, "*CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS*.doc"))
                    pdf_files = glob.glob(os.path.join(downloads_path, "*CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS*.pdf"))
                    
                    if docx_files and pdf_files:
                        docx_path = docx_files[0]  # Assuming you want the first match
                        pdf_path = pdf_files[0]  # Assuming you want the first match
                        criar_preview(downloads_path, nome_empresa, pdf_path, docx_path, data_atual, mercadoria,  index_conf)
                    else:
                        # Create documents for lines with different values
                        if termo_asiatico:
                            criar_documento_termo(termo_asiatico, "diferentes", accronym, index_conf, mercadoria)

                        if termo_platts_asiatico:
                            criar_documento_termo(termo_platts_asiatico, "diferentes", accronym, index_conf, mercadoria, tipo_platts=True)

                        if termo_asiatico_BRL:
                            criar_documento_termo(termo_asiatico_BRL, "diferentes", accronym, index_conf, mercadoria, tipo_brl=True)

                        if termo_platts_asiatico_BRL:
                            criar_documento_termo(termo_asiatico_BRL, "diferentes", accronym, index_conf, mercadoria, tipo_brl=True, tipo_platts=True)

                        if termo_bullet:
                            criar_documento_termo(termo_bullet, "iguais", accronym, index_conf, mercadoria)

                        if termo_platts_bullet:
                            criar_documento_termo(termo_platts_bullet, "iguais", accronym, index_conf, mercadoria, tipo_platts=True)

                        if termo_platts_bullet_BRL:
                            criar_documento_termo(termo_platts_bullet_BRL, "iguais", accronym, index_conf, mercadoria, tipo_brl=True, tipo_platts=True)

                        if termo_bullet_BRL:
                            criar_documento_termo(termo_bullet_BRL, "iguais", accronym, index_conf, mercadoria, tipo_brl=True)

                        if termo_palmoil_asiatico:
                            criar_documento_termo(termo_palmoil_asiatico, "diferentes", accronym, index_conf, mercadoria, tipo_palmoil=True)

                        if termo_palmoil_bullet:
                            criar_documento_termo(termo_palmoil_bullet, "iguais", accronym, index_conf, mercadoria, tipo_palmoil=True)
                        
    finally:
        # Uninitialize COM library
        pythoncom.CoUninitialize()

    # Stop the progress bar and close the window after processing is complete
    progressbar_confirmation.stop()
    progress_window.destroy()
    #messagebox.showinfo("Sucesso!", "Confirmações geradas com sucesso!")
    
def criar_documento_termo(linhas, tipo_linha, accronym, index_conf, mercadoria, tipo_brl=False, tipo_platts=False, tipo_palmoil=False):    
    # Lists to accumulate data for database update
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    Counterparties = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []   
    global entry_filtro_commodities
    entry_filtro_commodities.delete(0, tk.END)
    global tabela_commodities_data
    for item in tabela_commodities_data:            
            tabela_commodities.insert("", "end", values=item)
    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()               
    first_item_values = tabela_termo_cliente.item(linhas[0], 'values')
    market = first_item_values[2]                        
    cntpy = first_item_values[18]
    data_formatada_cgd = formatar_data(lookup(cntpy, cntpy_accronym, cntpy_cgd))
    nome_empresa = lookup(cntpy, cntpy_accronym, cntpy_name) if cntpy != "" else ""
    cnpj_formatado = formatar_cnpj(lookup(cntpy, cntpy_accronym, cntpy_taxid))
    trade_date = first_item_values[1]
    trade_date_pg_2 = safe_date_conversion_dd_MM(trade_date)
    data_por_extenso_assinatura = data_por_extenso(trade_date)
    
    # Check for missing values and show warnings
    missing_messages = []
    if cnpj_formatado is None:
        missing_messages.append(f"Missing TAX ID for {accronym}")
    if data_formatada_cgd is None:
        missing_messages.append(f"Missing CGD for {accronym}")
    if nome_empresa is None:
        missing_messages.append(f"Missing Counterparty for {accronym}")

    if missing_messages:
        messagebox.showwarning("Missing Information", "\n".join(missing_messages))
        return  # Interrupt the function if any value is missing

    # Determine the template path based on the type
    if tipo_brl:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\TERMO - BRL.doc"
    elif tipo_platts and not tipo_brl:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\TERMO - PLATTS.doc"
    elif tipo_platts and tipo_brl:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\TERMO - BRL - PLATTS.doc"   
    elif tipo_palmoil:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\TERMO - PALM OIL.doc"
    else:
        template_path = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia\\CommodiXchange\\Modelo Confirmações\\TERMO.doc"

    # Open the Word document using COM
    word = win32.Dispatch('Word.Application')
    word.Visible = False
    word.DisplayAlerts = 0  # Suppress alerts
    doc = word.Documents.Open(template_path)
        
    # Update bookmarks
    update_bookmark(doc, "CGD", data_formatada_cgd)
    update_bookmark(doc, "CNPJ_CONTRAPARTE", cnpj_formatado)
    update_bookmark(doc, "CONTRAPARTE", nome_empresa)
    update_bookmark(doc, "CONTRAPARTE_ASSINATURA", nome_empresa)
    update_bookmark(doc, "DATA_NEGOCIACAO_EXTENSO", data_por_extenso_assinatura)
    update_bookmark(doc, "DATA_NEGOCIACAO_PG2", trade_date_pg_2)

    # Define os dados das linhas
    dados_linhas = []
    for index, item in enumerate(linhas, start=1):
        values = tabela_termo_cliente.item(item, 'values')
        TradeDates.append(values[1])
        Counterparties.append(values[18])
        AthenaIDs.append(values[0])                
        Confirmations.append("OK")        
        Identifiers.append(str(values[-4]))
        Indexes.append(str(values[-3]))
        Time_Stamps.append(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        market = values[2]                            
        strike = strike_confirmation(market, values[6], values[8], commodities_ric, commodities_factor, tabview).replace(",", " ").replace(".", ",").replace(" ", ".")
        trade_date = values[1]
        valor_base_str = values[9].replace(",", ".").replace("-", "")
        ptax = safe_date_conversion_dd_MM(values[15])
        exchange = lookup(market, commodities_ric, commodities_exchange)
        if market == "NACX0005":
            market = "VLSFO (AMFSA00)"
        elif market == "PEURNPHY":
            market = "PAAL00"
        elif market == "PCRUDTB1":
            market = "PCAAS00"

        if tipo_linha == "diferentes":
            if tipo_palmoil:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, "MDE-BURSA MALAYSIA", 'Não Aplicável', valor_base_str,
                    'MYR USD', safe_date_conversion_dd_MM(values[17]), 'Não Aplicável', 'Não Aplicável', 'Não Aplicável', ptax, strike, 
                        safe_date_conversion_dd_MM(values[16]), safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
            elif tipo_brl and tipo_platts:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, "PLATTS", valor_base_str,
                    'Não Aplicável', 'Não Aplicável', 'Não Aplicável', safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), strike, safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
            elif tipo_brl and not tipo_platts:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, exchange, valor_base_str,
                    'Não Aplicável', 'Não Aplicável', 'Não Aplicável', safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), strike, safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
            else:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, exchange, valor_base_str,
                    'Não Aplicável', 'Não Aplicável', 'Não Aplicável', ptax, strike, safe_date_conversion_dd_MM(values[16]),
                    safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
        else:
            if tipo_palmoil:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, "MDE-BURSA MALAYSIA", 'Não Aplicável', valor_base_str,
                    'MYR USD', safe_date_conversion_dd_MM(values[17]), 'Não Aplicável', 'Não Aplicável', 'Não Aplicável', ptax, strike, 
                    'Não Aplicável', safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
            elif tipo_brl and tipo_platts:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, "PLATTS", valor_base_str,
                    'Não Aplicável', 'Não Aplicável', 'Não Aplicável', 'Não Aplicável',
                    safe_date_conversion_dd_MM(values[17]), strike, 'Não Aplicável',
                    safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ]) 
            elif tipo_brl and not tipo_platts:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, exchange, valor_base_str,
                    'Não Aplicável', 'Não Aplicável', 'Não Aplicável', 'Não Aplicável',
                    safe_date_conversion_dd_MM(values[17]), strike, 'Não Aplicável',
                    safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
            else:
                dados_linhas.append([
                    index, values[0], 'Parte B' if values[3] == "Sell" else 'Parte A', market, exchange, valor_base_str,
                    'Não Aplicável', 'Não Aplicável', 'Não Aplicável', ptax, strike, 'Não Aplicável',
                    safe_date_conversion_dd_MM(values[17]), safe_date_conversion_dd_MM(values[10])
                ])
    
# Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += ["OK"] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
    # Update the table with the data
    update_table(doc, "TABELA_OPERACOES", dados_linhas)   
    
    # Assuming data_atual is a string in the format 'YYYY-MM-DD' or similar
    data_atual_str = values[1]  # Replace with the actual format of your date string

    # Convert the string to a datetime object
    data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Adjust the format as needed
    

    # Now you can use strftime
    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    if index == 1:
        update_bookmark(doc, "TITULO", f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS nº {values[0]}")
        docx_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS nº {values[0]}.doc"
        pdf_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS nº {values[0]}.pdf"                            
    else:
        docx_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS - {safe_date_conversion_confirma(values[1])}.doc"
        pdf_filename = f"CONFIRMAÇÃO DE OPERAÇÕES DE DERIVATIVOS - {safe_date_conversion_confirma(values[1])}.pdf"
        
    if str(index_conf) == '1':
        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
        if diretorio_raiz_confirmation is None:
            missing_messages.append(f"Missing confirmation path mapping for {accronym}")
        if missing_messages:
            messagebox.showwarning("Missing Information", "\n".join(missing_messages))
            return  # Interrupt the function if any value is missing
        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "TERMO", f"{mercadoria}")
        if not os.path.exists(downloads_path):
                    os.makedirs(downloads_path)
    else:     
        diretorio_raiz_confirmation = lookup(accronym, cntpy_accronym, cntpy_confirmacoes)
        if diretorio_raiz_confirmation is None:
            missing_messages.append(f"Missing confirmation path mapping for {accronym}")
        if missing_messages:
            messagebox.showwarning("Missing Information", "\n".join(missing_messages))
            return  # Interrupt the function if any value is missing
        downloads_path = os.path.join(diretorio_raiz_confirmation, ano, f"{mes2}. {mes}", dia, "TERMO",  f"#{index_conf} {mercadoria}")
        if not os.path.exists(downloads_path):
            os.makedirs(downloads_path)   
                            

    docx_path = os.path.join(downloads_path, docx_filename)
    pdf_path = os.path.join(downloads_path, pdf_filename)                      
    
    # Save the document using SaveAs
    doc.SaveAs(pdf_path, FileFormat=17)     
    doc.SaveAs(docx_path)       
    word.Quit()           
     # Release COM objects
    del doc
    del word                   
    criar_preview(downloads_path, nome_empresa, pdf_path, docx_path, data_atual, mercadoria,  index_conf)
               
def generate_confirmation_termo(tabview):
    global tabela_termo_cliente
    resposta = messagebox.askyesno("Confirmation generate", "Wish to proceed?")
    if not resposta:
        return   
    
    # Open the progress window immediately
    open_progress_window_termo()
    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    
    # Extrai os acrônimos únicos da tabela "Cliente"
    accronyms = []    
    indexes_conf = []
    mercadorias = []      
    for row_id_comm in tabela_termo_cliente.get_children():
        row_comm = tabela_termo_cliente.item(row_id_comm)["values"]
        mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)
        accronym = row_comm[18]
        index_conf = row_comm[-3]        
        if accronym not in accronyms:
            accronyms.append(accronym)
        if index_conf not in indexes_conf:
            indexes_conf.append(index_conf)
        if mercadoria not in mercadorias:
                mercadorias.append(mercadoria)    


    # Start the document processing in a separate thread
    threading.Thread(target=process_documents_termo, args=(tabview, accronyms, indexes_conf, mercadorias)).start()


def criar_preview(downloads_path, nome_empresa, pdf_path, docx_path, data_atual, mercadoria,  index_conf):
    global inner_frame, canvas, janela, toplevel_windows 
    # Tkinter setup
    ctk.set_appearance_mode("System")
    ctk.set_default_color_theme("dark-blue")

    nova_janela = ctk.CTkToplevel(janela)
    if index_conf == 1:
        nova_janela.title(f"Confirmation Preview - {nome_empresa} - {mercadoria}")
    else:
        nova_janela.title(f"Confirmation Preview - {nome_empresa} - #{index_conf} {mercadoria}") 
        
    nova_janela.geometry("1280x800")    
    nova_janela.iconbitmap(os.path.join(r"I:\Confirmation\Derivativos\Movimento\Liquidações do Dia\X_icone.ico"))
    # Traz a janela Toplevel para o primeiro plano
    nova_janela.lift()
    nova_janela.focus_set()
    
    # Frame for checkboxes on the left
    frame_check_confirmation = ctk.CTkFrame(nova_janela, height=668, width=100)
    frame_check_confirmation.grid(row=0, column=0, sticky='ns') 

    # Initialize checkboxes and labels
    checkboxes = []

    # CGD
    label_cgd = ctk.CTkLabel(frame_check_confirmation, text="CGD")
    label_cgd.grid(row=0, column=0, padx=2, pady=1, sticky='w')
    checkbox_cgd = ctk.CTkCheckBox(frame_check_confirmation, text="")
    checkbox_cgd.grid(row=0, column=1, padx=2, pady=1, sticky='e')
    checkboxes.append((checkbox_cgd, label_cgd))

    # Contraparte
    label_contraparte = ctk.CTkLabel(frame_check_confirmation, text="Contraparte")
    label_contraparte.grid(row=1, column=0, padx=2, pady=1, sticky='w')
    checkbox_contraparte = ctk.CTkCheckBox(frame_check_confirmation, text="")
    checkbox_contraparte.grid(row=1, column=1, padx=2, pady=1, sticky='e')
    checkboxes.append((checkbox_contraparte, label_contraparte))

    # CNPJ
    label_cnpj = ctk.CTkLabel(frame_check_confirmation, text="CNPJ")
    label_cnpj.grid(row=2, column=0, padx=2, pady=1, sticky='w')
    checkbox_cnpj = ctk.CTkCheckBox(frame_check_confirmation, text="")
    checkbox_cnpj.grid(row=2, column=1, padx=2, pady=1, sticky='e')
    checkboxes.append((checkbox_cnpj, label_cnpj))

    # Trade Date
    label_trade_date = ctk.CTkLabel(frame_check_confirmation, text="Trade Date")
    label_trade_date.grid(row=3, column=0, padx=2, pady=1, sticky='w')
    checkbox_trade_date = ctk.CTkCheckBox(frame_check_confirmation, text="")
    checkbox_trade_date.grid(row=3, column=1, padx=2, pady=1, sticky='e')
    checkboxes.append((checkbox_trade_date, label_trade_date))

    # Trade Date pg.6
    label_trade_date_pg6 = ctk.CTkLabel(frame_check_confirmation, text="Trade Date pg.6")
    label_trade_date_pg6.grid(row=4, column=0, padx=2, pady=1, sticky='w')
    checkbox_trade_date_pg6 = ctk.CTkCheckBox(frame_check_confirmation, text="")
    checkbox_trade_date_pg6.grid(row=4, column=1, padx=2, pady=1, sticky='e')
    checkboxes.append((checkbox_trade_date_pg6, label_trade_date_pg6))

    # Contraparte pg.6
    label_contraparte_pg6 = ctk.CTkLabel(frame_check_confirmation, text="Contraparte pg.6")
    label_contraparte_pg6.grid(row=5, column=0, padx=2, pady=1, sticky='w')
    checkbox_contraparte_pg6 = ctk.CTkCheckBox(frame_check_confirmation, text="")
    checkbox_contraparte_pg6.grid(row=5, column=1, padx=2, pady=1, sticky='e')
    checkboxes.append((checkbox_contraparte_pg6, label_contraparte_pg6))

    # Anexo I
    label_anexo_i = ctk.CTkLabel(frame_check_confirmation, text="Anexo I")
    label_anexo_i.grid(row=6, column=0, padx=2, pady=1, sticky='w')
    checkbox_anexo_i = ctk.CTkCheckBox(frame_check_confirmation, text="")
    checkbox_anexo_i.grid(row=6, column=1, padx=2, pady=1, sticky='e')
    checkboxes.append((checkbox_anexo_i, label_anexo_i))

    # Frame for preview on the right
    frame_preview = ctk.CTkFrame(nova_janela)
    frame_preview.grid(row=0, column=1, sticky='nsew')
    frame_preview.grid_rowconfigure(0, weight=1)
    frame_preview.grid_columnconfigure(0, weight=1)
    frame_preview.grid_columnconfigure(1, weight=1)

    # Canvas for scrolling within the preview frame
    canvas = ctk.CTkCanvas(
        frame_preview,
        width=1024,
        height=768       
    )
    canvas.pack(side='left', fill='both', expand=True)

    # Scrollbars
    scrollbar_x_preview = ctk.CTkScrollbar(frame_preview, orientation='horizontal', command=canvas.xview)
    scrollbar_y_preview = ctk.CTkScrollbar(frame_preview, orientation='vertical', command=canvas.yview)

    scrollbar_x_preview.pack(side='bottom', fill='x')
    scrollbar_y_preview.pack(side='right', fill='y')
    scrollbar_x_preview.configure(height=25)
    scrollbar_y_preview.configure(width=25)

    # Configure the canvas to use the scrollbars
    canvas.configure(xscrollcommand=scrollbar_x_preview.set, yscrollcommand=scrollbar_y_preview.set)
    
    # Inner frame containing scrollable content
    inner_frame = ctk.CTkFrame(canvas, height=768, width=1124)
    canvas.create_window((0, 0), window=inner_frame, anchor='nw')

    # Update the scroll region when the inner frame changes size
    inner_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Frame for buttons below
    frame_botao = ctk.CTkFrame(nova_janela, height=150)
    frame_botao.grid(row=2, column=0, columnspan=3, sticky='ew', pady=5)
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    botao_gerar_confirmation = ctk.CTkButton(frame_botao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRM", font=fonte_botao, command=lambda: confirm_confirmation(checkboxes, nome_empresa, mercadoria, pdf_path, downloads_path, data_atual, nova_janela, index_conf))
    botao_gerar_confirmation.grid(column=0, row=0, padx=10, pady=10)
    botao_delete_confirmation = ctk.CTkButton(frame_botao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", font=fonte_botao, command=lambda: delete_confirmation(docx_path, pdf_path, nova_janela, downloads_path))
    botao_delete_confirmation.grid(column=1, row=0, padx=10, pady=10)

    # Configure expansion
    nova_janela.grid_rowconfigure(0, weight=1)
    nova_janela.grid_columnconfigure(1, weight=1)    
    carregar_e_exibir(pdf_path, downloads_path)



    def delete_confirmation(docx_path, pdf_path, toplevel_window, downloads_path):
        resposta = messagebox.askyesno("Delete", "Wish to proceed?")
        if resposta:
            try:
                # Delete the DOCX file
                if os.path.exists(docx_path):
                    os.remove(docx_path)
                
                # Delete the PDF file
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                
                # Delete all .png files in the downloads_path
                for file in glob.glob(os.path.join(downloads_path, "*.png")):
                    if os.path.exists(file):
                        os.remove(file)
                        
                # Close the Toplevel window
                toplevel_window.destroy()
                
                
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while deleting files: {e}")

    def fechar_toplevel(docx_path, pdf_path, toplevel_window):    
        # Remover a janela da lista de janelas abertas
        if toplevel_window in toplevel_windows:
            toplevel_windows.remove(toplevel_window)

        delete_confirmation(docx_path, pdf_path, toplevel_window, downloads_path)
        
    def on_closing():
        # global toplevel_windows
        if toplevel_windows:
            messagebox.showwarning("Aviso", "Preview aberto, favor verificar")
        else:
            janela.destroy()

    


    def confirm_confirmation(checkboxes, nome_empresa, mercadoria, pdf_path, downloads_path, data_atual, toplevel_window, index_conf):
        # Check if all checkboxes are selected
        all_checked = all(checkbox.get() for checkbox, label in checkboxes)
        if not all_checked:
            # Find which checkboxes are not checked
            unchecked_labels = [label.cget("text") for checkbox, label in checkboxes if not checkbox.get()]
            messagebox.showwarning("Incomplete", f"The following items are not checked: {', '.join(unchecked_labels)}")
        else:
            tipo = "Opção" if "OPÇÃO" in pdf_path else "Termo"
            # Get the current date in dd/mm/yyyy format
            formatted_date = data_atual.strftime("%d/%m/%Y")
            SID = getpass.getuser()
            # First email details
            if index_conf == 1:
                subject1 = f"Contrato Validado: {tipo} Commodities - {nome_empresa} - {formatted_date} - {mercadoria}"
            else:
                subject1 = f"Contrato Validado: {tipo} Commodities - {nome_empresa} - {formatted_date} - #{index_conf} {mercadoria}"
            to_email1 = "brazil.otc.ops@jpmorgan.com"
            body1 = f"A(s) confirmação(ões) anexa(s) foi(ram) gerada(s) pela ferramenta CommodiXchange e foi(ram) validada(s) por {SID} ."

            # Find the .msg file containing "INTERNAL" in the name
            msg_files = glob.glob(os.path.join(downloads_path, "*INTERNAL*.msg"))
            if msg_files:
                msg_file = msg_files[0]  # Take the first match
            else:
                messagebox.showerror("Error", "No INTERNAL .msg file found.")
                return
            

            # Create the first email in Outlook
            create_outlook_email(subject1, to_email1, "", body1, [pdf_path, msg_file])

            # Determine the greeting based on the current time
            current_hour = datetime.now().hour
            if current_hour < 12:
                greeting = "Bom dia"
            elif 12 <= current_hour < 18:
                greeting = "Boa tarde"
            else:
                greeting = "Boa noite"

            # Second email details
            if index_conf == 1:
                subject2 = f"Validação Contrato: {tipo} Commodities - {nome_empresa} - {formatted_date} - {mercadoria}"
            else:
                subject2 = f"Validação Contrato: {tipo} Commodities - {nome_empresa} - {formatted_date} - #{index_conf} {mercadoria}"
            to_email2 = "brazil_sales_support_mo@jpmchase.com"
            cc_email2 = "brazil.otc.ops@jpmorgan.com"
            # Use the smaller filled circle character
            body2 = (
                f"{greeting},<br>"
                "Tudo bem?<br><br>"
                "Por gentileza, poderiam validar o(s) item(ns) abaixo:<br><br>"
                "&#9679; Contrato<br><br>"  # Smaller filled circle
                "Obrigado(a),"
            )
            # Create the second email in Outlook
            create_outlook_email(subject2, to_email2, cc_email2, body2, [pdf_path, msg_file])
          
                # Close the Toplevel window
            toplevel_window.destroy()

def extrair_dados_tabelas():
    # Extrair dados da tabela_base_comitentes
    cntpy_taxid = []
    cntpy_accronym = []
    cntpy_name = []
    cntpy_b3_account = []
    cntpy_cgd = []
    cntpy_confirmacoes = []
    cntpy_bank = []
    cntpy_cc = []
    cntpy_ag = []
    cntpy_status = []
    for item in tabela_base_comitentes.get_children():
        values = tabela_base_comitentes.item(item, 'values')
        cntpy_taxid.append(values[0])  # Supondo que o CNPJ está na primeira coluna
        cntpy_accronym.append(values[5])   # Supondo que o nome do cliente está na sexta coluna
        cntpy_name.append(values[4])   # Supondo que o nome do cliente está na quinta coluna
        cntpy_b3_account.append(values[6]) # Supondo que a conta cetip do cliente está na sétima coluna
        cntpy_cgd.append(values[7]) # Supondo que a conta cetip do cliente está na sétima coluna
        cntpy_confirmacoes.append(values[8])        
        cntpy_bank.append(values[9])
        cntpy_ag.append(values[10])
        cntpy_cc.append(values[11])
        cntpy_status.append(values[12])

    # Extrair dados da tabela_commodities
    commodities_ric = []
    commodities_factor = []
    commodities_exchange = []
    commodities_type = []
    commodities_mercadoria = []
    commodities_YYYY = []
    commodities_MM = []
    commodities_unity = []
    commodities_status = []
    for item in tabela_commodities.get_children():
        values = tabela_commodities.item(item, 'values')
        commodities_ric.append(values[0])  # Supondo que o RIC está na primeira coluna
        commodities_factor.append(values[8])  # Supondo que o fator de conversão está na nona coluna
        commodities_exchange.append(values[1])  # Supondo que a bolsa está na segunda coluna
        commodities_type.append(values[5])  # Supondo que o tipo está na sexta coluna
        commodities_mercadoria.append(values[9])  # Supondo que o tipo está na décima coluna
        commodities_MM.append(values[3])  # Supondo que o tipo está na décima coluna
        commodities_YYYY.append(values[4])  # Supondo que o tipo está na décima coluna
        commodities_unity.append(values[6])  # Supondo que o tipo está na décima coluna
        commodities_status.append(values[10])
    return commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status


        

    
    
def create_email_validation_opcao(caminho_completo, base_path, mercadoria, accronym, cliente, formatted_date):   
 # Determine the greeting based on the current time
    current_hour = datetime.now().hour
    if current_hour < 12:
        greeting = "bom dia"
    elif 12 <= current_hour < 18:
        greeting = "boa tarde"
    else:
        greeting = "boa noite"   
    

    # Verificar se existe um folder "#2 TERMO"
    if f"#2 {mercadoria}" in os.listdir(base_path):
        # Se existir, perguntar ao usuário qual pasta usar
        chosen_folder = askdirectory(title="Select a folder to attach", initialdir=base_path)
        if not chosen_folder:
            messagebox.showerror("Erro", "Escolha inválida. Operação cancelada.")
            exit()
    else:
        # Caso contrário, definir chosen_folder como o caminho padrão
        chosen_folder = os.path.join(caminho_completo, f"{cliente}", 'OPÇÃO', f"{mercadoria}")
    
    downloads_path = os.path.join(base_path, chosen_folder)

    # Coletar arquivos para anexar apenas uma vez por acrônimo
    msg_files = glob.glob(os.path.join(downloads_path, "*.msg"))
    xlsx_files = glob.glob(os.path.join(downloads_path, "*.xlsx"))
    lawton_files = glob.glob(os.path.join(downloads_path, "*_LAWTON*.txt"))

    # Adiciona arquivos à lista de anexos
    attachments = msg_files + xlsx_files + lawton_files
    

    # Detalhes do email
    sufixo = f" - {chosen_folder}" if chosen_folder != "OPÇÃO" else ""
    subject = f"Validar Planilha + Lançar B2B: Opção Commodities - {cliente} - {formatted_date} - {mercadoria}"
    to_email = "OTeamC"
    cc_email = ""

    body = (
        f"Pessoal, {greeting},<br>"
        "Tudo bem?<br><br>"
        "Por gentileza, poderiam validar o(s) item(ns) anexo(s) e se de acordo, registrar o(s) arquivo(s) ponta Lawton na B3?<br><br>"
        "Obrigado(a),"
    )

    # Criar o email no Outlook
    create_outlook_email(subject, to_email, cc_email, body, attachments)


def create_email_validation_termo(downloads_path, mercadoria, accronym, cliente, formatted_date, index):   
    # Determine the greeting based on the current time
    current_hour = datetime.now().hour
    if current_hour < 12:
        greeting = "bom dia"
    elif 12 <= current_hour < 18:
        greeting = "boa tarde"
    else:
        greeting = "boa noite"   
    
    # Coletar arquivos para anexar para cada combinação
    msg_files = glob.glob(os.path.join(downloads_path, "*.msg"))
    xlsx_files = glob.glob(os.path.join(downloads_path, "*.xlsx"))
    lawton_files = glob.glob(os.path.join(downloads_path, "*_LAWTON*.txt"))

    # Adiciona arquivos à lista de anexos
    attachments = msg_files + xlsx_files + lawton_files
    
    # Detalhes do email    
    subject = f"Validar Planilha + Lançar B2B: Termo Commodities - {cliente} - {formatted_date} - {mercadoria}" if index == 1 else f"Validar Planilha + Lançar B2B: Termo Commodities - {cliente} - {formatted_date} - #{index} {mercadoria}"
    to_email = "OTeamC"
    cc_email = ""

    body = (
        f"Pessoal, {greeting},<br>"
        "Tudo bem?<br><br>"
        "Por gentileza, poderiam validar o(s) item(ns) anexo(s) e se de acordo, registrar o(s) arquivo(s) ponta Lawton na B3?<br><br>"
        "Obrigado(a),"
    )

    # Criar o email no Outlook
    create_outlook_email(subject, to_email, cc_email, body, attachments)


def validation_email_termo(tabela_termo_cliente):
    resposta = messagebox.askyesno("Validation E-mail - Termo", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função

    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"

    # Data atual
    data_atual = datetime.now()
    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    # Caminho completo para o diretório
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")

    # Criar o diretório se não existir
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)

    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()

    # Extrai os acrônimos únicos da tabela "Cliente"
    accronyms = []
    mercadorias = []
    indexes = []
    for row_id_comm in tabela_termo_cliente.get_children():
        row_comm = tabela_termo_cliente.item(row_id_comm)["values"]                    
        mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)
        accronym = row_comm[18]
        index = row_comm[20]
        if mercadoria not in mercadorias:
            mercadorias.append(mercadoria) 
        if accronym not in accronyms:
            accronyms.append(accronym)      
        if index not in indexes:
            indexes.append(index)
    
    for index in indexes:
        for accronym in accronyms:
            for mercadoria in mercadorias:
                formatted_date = None
                identifiers_termo = []
                cliente = lookup(accronym, cntpy_acronym, cntpy_name)
                for row_id_tco in tabela_termo_cliente.get_children():
                    row_comm_tco = tabela_termo_cliente.item(row_id_tco)["values"]                    
                    identifier_termo = str(row_comm_tco[-1])
                    index_tco = row_comm_tco[20]
                    if lookup(row_comm_tco[2], commodities_ric, commodities_mercadoria) == mercadoria and row_comm_tco[18] == accronym and index_tco == index:
                        identifiers_termo.append(identifier_termo)
                        formatted_date = datetime.strptime(row_comm_tco[1], "%d-%b-%Y").strftime("%d/%m/%Y")
                        
                if identifiers_termo:
                    downloads_path = os.path.join(caminho_completo, f"{cliente}", "TERMO", f"{mercadoria}" if index == 1 else f"#{index} {mercadoria}")
                    create_email_validation_termo(downloads_path, mercadoria, accronym, cliente, formatted_date, index)
               
def validation_email_opcao(tabela_opcao_cliente):
    resposta = messagebox.askyesno("Validation E-mail - Opção", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função

    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"

    # Data atual
    data_atual = datetime.now()
    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    # Caminho completo para o diretório
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")

    # Criar o diretório se não existir
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)

    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()

    # Extrai os acrônimos únicos da tabela "Cliente"
    accronyms = []
    mercadorias = []
    for row_id_comm in tabela_opcao_cliente.get_children():
        row_comm = tabela_opcao_cliente.item(row_id_comm)["values"]                    
        mercadoria = lookup(row_comm[2], commodities_ric, commodities_mercadoria)
        accronym = row_comm[18]
        if mercadoria not in mercadorias:
            mercadorias.append(mercadoria) 
        if accronym not in accronyms:
            accronyms.append(accronym)     
    

    for accronym in accronyms:
        # Itera sobre cada acrônimo único
        for mercadoria in mercadorias:
            formatted_date = None
            identifiers_opcao = []
            # Inicializa variáveis para coletar informações e anexos
            cliente = lookup(accronym, cntpy_accronym, cntpy_name)
            for row_id_opt in tabela_opcao_cliente.get_children():
                row_comm_opt = tabela_opcao_cliente.item(row_id_opt)["values"]                    
                identifier_termo = str(row_comm_opt[-1])
                if lookup(row_comm_opt[2], commodities_ric, commodities_mercadoria) == mercadoria and row_comm_opt[18] == accronym:
                    identifiers_opcao.append(identifier_termo)
                    formatted_date = datetime.strptime(row_comm_opt[1], "%d-%b-%Y").strftime("%d/%m/%Y")
                       
    # Inicializa variáveis para tipo e data formatada                                   
            if identifiers_opcao:
                # Supondo que caminho_completo e cliente já estejam definidos
                base_path = os.path.join(caminho_completo, f"{cliente}", "OPÇÃO")
                create_email_validation_opcao(caminho_completo, base_path, mercadoria, accronym, cliente, formatted_date)

def validation_email_opcao_(tabela_opcao_cliente):
    resposta = messagebox.askyesno("Validation E-mail - Opção", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função

    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"

    # Data atual
    data_atual = datetime.now()
    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    # Caminho completo para o diretório
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")

    # Criar o diretório se não existir
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)

    # Extrair dados necessários
    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()

    # Extrai os acrônimos únicos da tabela "Cliente"
    accronyms = {tabela_opcao_cliente.item(row_id)["values"][18] for row_id in tabela_opcao_cliente.get_children()}

    # Itera sobre cada acrônimo único
    for accronym in accronyms:
        # Determine the greeting based on the current time
        current_hour = datetime.now().hour
        if current_hour < 12:
            greeting = "bom dia"
        elif 12 <= current_hour < 18:
            greeting = "boa tarde"
        else:
            greeting = "boa noite"

        # Inicializa variáveis para coletar informações e anexos
        cliente = lookup(accronym, cntpy_accronym, cntpy_name)
          # Supondo que caminho_completo e cliente já estejam definidos
        base_path = os.path.join(caminho_completo, f"{cliente}")

        # Verificar se existe um folder "#2 TERMO"
        if "#2 OPÇÃO" in os.listdir(base_path):
            # Se existir, perguntar ao usuário qual pasta usar
            chosen_folder = askdirectory(title="Select a folder to attach", initialdir=base_path)
            if not chosen_folder:
                messagebox.showerror("Erro", "Escolha inválida. Operação cancelada.")
                exit()
        else:
            # Caso contrário, definir chosen_folder como o caminho padrão
            chosen_folder = os.path.join(caminho_completo, f"{cliente}", 'OPÇÃO')

        downloads_path = os.path.join(base_path, chosen_folder)

        if not os.path.exists(downloads_path):
            os.makedirs(downloads_path)

        # Coletar arquivos para anexar apenas uma vez por acrônimo
        msg_files = glob.glob(os.path.join(downloads_path, "*.msg"))
        xlsx_files = glob.glob(os.path.join(downloads_path, "*.xlsx"))
        lawton_files = glob.glob(os.path.join(downloads_path, "*_Lawton*.txt"))

        # Adiciona arquivos à lista de anexos
        attachments = msg_files + xlsx_files + lawton_files

        # Inicializa variáveis para mercadoria e data formatada
        mercadoria = None
        formatted_date = None

        for row_id in tabela_opcao_cliente.get_children():
            row = tabela_opcao_cliente.item(row_id)["values"]
            if row[18] == accronym:
                mercadoria = lookup(row[2], commodities_ric, commodities_mercadoria)
                date_object = datetime.strptime(row[1], "%d-%b-%Y")  # Ajuste o formato conforme necessário
                formatted_date = date_object.strftime("%d/%m/%Y")
                break  # Saia do loop após encontrar a primeira correspondência

        # Detalhes do email
        sufixo = f" - {chosen_folder}" if chosen_folder != "OPÇÃO" else ""
        subject = f"Validar Planilha + Lançar B2B: Opção Commodities - {cliente} - {formatted_date} - {mercadoria}"
        to_email = "OTeamC"
        cc_email = ""

        body = (
            f"Pessoal, {greeting},<br>"
            "Tudo bem?<br><br>"
            "Por gentileza, poderiam validar o(s) item(ns) anexo(s) e se de acordo, registrar o(s) arquivo(s) ponta Lawton na B3?<br><br>"
            "Obrigado(a),"
        )

        # Criar o email no Outlook
        create_outlook_email(subject, to_email, cc_email, body, attachments)


def highlight_deals_not_mapped(treeview, tipo):
    # Verifica o tipo e percorre os itens do Treeview
    for item in treeview.get_children():
        row = treeview.item(item, 'values')
        
        if tipo == "TERMO":
            # Verifica o valor no índice 1
            if row[1] == "Codigo_Cetip":
                # Pinta a célula do índice 1 de lightcoral
                treeview.tag_configure('highlight', background='lightcoral')
                treeview.item(item, tags=('highlight',))
            else:
                # Deixa a célula do índice 1 em branco
                treeview.tag_configure('normal', background='white')
                treeview.item(item, tags=('normal',))
        
        elif tipo == "OPÇÃO":
            # Verifica o valor no índice 6
            if row[6] == "Codigo_Cetip":
                # Pinta a célula do índice 6 de lightcoral
                treeview.tag_configure('highlight', background='lightcoral')
                treeview.item(item, tags=('highlight',))
            else:
                # Deixa a célula do índice 6 em branco
                treeview.tag_configure('normal', background='white')
                treeview.item(item, tags=('normal',))
                
def mapping_deals(tabela_termo_b2b, tabela_boletatermo):
    deals_cliente, td_cliente, market_client, type_cliente, instrument_cliente, strike_cliente, intCCY_cliente, tn_cliente, sd_cliente, fxd_cliente, fsd_cliente, fed_cliente, accronym, identifier_cliente, status_cliente, sid_cliente, deals_b2b, td_b2b, market_client, type_b2b, instrument_b2b, strike_b2b, intCCY_b2b, tn_b2b, sd_b2b, fxd_b2b, fsd_b2b, fed_b2b, accronym_b2b, identifier_b2b, status_b2b, sid_b2b = extrair_dados_deals(tabela_termo_cliente, tabela_termo_b2b)

    # Lists to accumulate data for database update
    Makers = []
    Checkers = []
    Instruments = []
    TradeDates = []
    Counterparties = []
    AthenaIDs = []
    B3_IDs = []    
    B2B_AthenaIDs = []        
    B2B_B3_IDs = []            
    Confirmations = []
    SS_Validations = []
    Identifiers = []
    Indexes = []
    Time_Stamps = []
    Statuses = []

    # Diretório raiz
    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"

    # Data atual
    data_atual = datetime.now()
    mes2 = data_atual.strftime("%m")

    # Dicionário para os meses em português
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }

    # Obter o nome do mês em português
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")

    # Caminho completo para o diretório
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")

    # Iterar sobre os arquivos no diretório
    for arquivo in os.listdir(caminho_completo):
        if arquivo.endswith(".CETIP21"):
            caminho_arquivo = os.path.join(caminho_completo, arquivo)
            with open(caminho_arquivo, 'r', encoding='utf-8') as f:
                # Ler o conteúdo do arquivo linha por linha
                linhas = f.readlines()
                for linha in linhas[1:]:  # Começar a partir da linha 2
                    # Variável 1: linha_retorno
                    linha_retorno = linha.strip()

                    # Variável 2: athena_id
                    athena_id = linha_retorno[690:700].strip()
                    B2B_AthenaIDs.append(athena_id)
                    
                    # Iterar sobre a tabela_termo_b2b e comparar                   
                    for values_b2b in tabela_termo_b2b.get_children():
                        values = tabela_termo_b2b.item(values_b2b)['values']
                        if athena_id == str(values[0]):                           
                            # Variável 3: id_boleta
                            id_boleta = str(values[-4])
                                                       
                            # Variável 4: cetip_id
                            cetip_id = linha_retorno[15:26].strip()
                            if cetip_id != "":
                                B2B_B3_IDs.append(cetip_id)
                                AthenaIDs.append(lookup_approve(id_boleta, identifier_cliente, deals_cliente))                            
                            # Iterar sobre a tabela_boletatermo
                            for values_boleta_id in tabela_boletatermo.get_children():
                                values_boleta = tabela_boletatermo.item(values_boleta_id)['values']                                
                                if str(values_boleta[-1]) == id_boleta:
                                    values_boleta[1] = str(cetip_id)
                                    tabela_boletatermo.item(values_boleta_id, values=values_boleta)                                                                     
                            break  # Sai do loop se encontrar um match na tabela_termo_b2b

 # Ensure all lists have the same length and fill with empty strings if necessary
    max_length = max(len(AthenaIDs), len(B2B_AthenaIDs), len(B3_IDs), len(Instruments), len(Statuses), len(Makers), len(Checkers), len(Time_Stamps), len(Confirmations), len(SS_Validations), len(Identifiers), len(Indexes))
    TradeDates += [""] * (max_length - len(TradeDates))
    Counterparties += [""] * (max_length - len(Counterparties))
    AthenaIDs += [""] * (max_length - len(AthenaIDs))
    B3_IDs += [""] * (max_length - len(B3_IDs))
    B2B_AthenaIDs += [""] * (max_length - len(B2B_AthenaIDs))
    B2B_B3_IDs += [""] * (max_length - len(B2B_B3_IDs))
    Instruments += [""] * (max_length - len(Instruments))
    Statuses += [""] * (max_length - len(Statuses))
    Makers += [""] * (max_length - len(Makers))
    Checkers += [""] * (max_length - len(Checkers))
    Time_Stamps += [""] * (max_length - len(Time_Stamps))
    Confirmations += [""] * (max_length - len(Confirmations))
    SS_Validations += [""] * (max_length - len(SS_Validations))
    Identifiers += [""] * (max_length - len(Identifiers))
    Indexes += [""] * (max_length - len(Indexes))
    
    # Pass the lists to the insert_or_update_base_deals function
    insert_or_update_base_deals(
        TradeDates, Counterparties, AthenaIDs, B3_IDs, B2B_AthenaIDs, B2B_B3_IDs,
        Instruments, Statuses, Makers, Checkers, Time_Stamps, Confirmations,
        SS_Validations, Identifiers, Indexes
    )
    
    highlight_deals_not_mapped(tabela_boletatermo, "TERMO")


def popular_boleta(tabview, abas_existentes, tabela_anbima=None):
# Perguntar ao usuário se deseja seguir com o cadastro
    resposta = messagebox.askyesno("Intrag File", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função
    try:
        
        global entry_filtro_commodities
        entry_filtro_commodities.delete(0, tk.END)
        global tabela_commodities_data
        for item in tabela_commodities_data:            
                tabela_commodities.insert("", "end", values=item)

        if "Boleta Dinâmica - Intrag" not in abas_existentes:
            # Inicializar o monitor de operações e obter os tabviews
            intrag(tabview,abas_existentes)
            abas_existentes.append("Boleta Dinâmica - Intrag") 
        else:
            tabela_boletaopcao.delete(*tabela_boletaopcao.get_children())
            tabela_boletatermo.delete(*tabela_boletatermo.get_children())          

        # Extrair dados necessários das tabelas
        commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()

        def rept(char, count):
            return char * count

        # Processar dados para tabela_opcao_cliente
        for i, item in enumerate(tabela_opcao_b2b.get_children(), start=3):
            try:
                values = tabela_opcao_b2b.item(item, 'values')
                status = values[-2]
                if status != "New" and status != "Pending Review":
                    market = values[2]                                
                    strike = math.trunc(float(values[6].replace(",", "").replace("-", ""))   * float(lookup(market, commodities_ric, commodities_factor).replace(",", ".")) * 10**8) /10**8
                    exchange = lookup(market, commodities_ric, commodities_exchange)               
                    premium = float(values[19].replace(",", "").replace("-", ""))
                    valor_base = int(values[9].replace(",", "").replace("-", ""))
                    premiumperunit = math.trunc(float(values[20].replace(",", "").replace("-", ""))   * float(lookup(market, commodities_ric, commodities_factor).replace(",", ".")) * 10**8) /10**8                                                      
                    simbolo_moeda = lookup(market, commodities_ric, commodities_mercadoria)[:3]

                    
                    linha = [
                        "INTRAGJP552",	# Id. Sist.
                        "OPCAO",	# ID Tipo de Linha
                        "1",	# Codigo da Operação
                        safe_date_conversion_dd_MM(values[1]), #Data de Registro
                        "73760.00-9" if values[3] == "Sell" else "00041.00-7", #Conta Titular
                        "BANCO J.P. MORGAN S/A" if values[3] == "Sell" else "LAWTON MULTIMERCADO-FI", #Nome do Titular
                        "OFVC" if values[4] == "Option (Put)" else "OFCC",	# Tipo Contrato
                        values[23], #Contrato Cetip
                        ''.join([str(random.randint(1,9)) for _ in range(5)]),	# Meu Número
                        "002", #Tipo da Operaçao
                        "00041.00-7" if values[3] == "Sell" else "73760.00-9",
                        "LAWTON MULTIMERCADO-FI" if values[3] == "Sell" else "BANCO J.P. MORGAN S/A",
                        safe_date_conversion_dd_MM(values[1]),	# Data de Inicio
                        safe_date_conversion_dd_MM(values[10]),	# Data de Vencimento
                        ''.join([str(random.randint(1,9)) for _ in range(16)]), #Numero Cetip
                        "COM", #Código SISBACEN da Moeda
                        simbolo_moeda, #Símbolo da Moeda
                        str(premium), #Valor da Aplicação
                        str(valor_base) + ".00", #Valor Base Moeda Estrangeira
                        "", #Valor Antecipado
                        "", #Preço Unitário da Antecipação
                        "0.00", #Valor Resgate
                        "" if values[4] == "Option (Put)" else str(strike), #Preço de Exercício de Call
                        "" if values[4] != "Option (Put)" else str(strike), #Preço de Exercício de Put
                        "" if values[4] == "Option (Put)" else str(premiumperunit), #Prêmio Unitário de Call
                        "" if values[4] != "Option (Put)" else str(premiumperunit), #Prêmio Unitário de Put
                        "",  #Cotação Barreira
                        "EUROPEIA", #Tipo de Exercício
                        "COMMODITIES", #Fonte de Informação
                        "9", #Boletim   
                        "", #Horário do Boletim
                        str(networkdays(values[17], values[10])),  # Cotação para o Vencimento
                        "D-" + str(networkdays(values[17], values[10])),  # Descrição de Cotação para Vencimento
                        exchange, #Fonte de Consulta
                        market, #TICKER
                        str(valor_base) + ".00", #Quantidade
                        safe_date_conversion_dd_MM(values[22]),	# DATA DE PAGAMENTO DO PRÊMIO
                        "APLICÁVEL" if values[16] != values[17] else "Não Aplicável",   # MÉDIA OPÇÃO ASIÁTICA
                        values[-4],
                    ]                

                    tabela_boletaopcao.insert("", "end", values=linha)
                    ajustar_largura_colunas(tabela_boletaopcao, colunas_boleta_opcao, tabview)
                    vincular_evento_duplo_clique(tabela_boletaopcao, colunas_boleta_opcao)
                    vincular_navegacao_setas(tabela_boletaopcao)
            except Exception as e:
                print(f"Erro ao processar item na tabela_boletaopcao: {e}")         
  
     # Processar dados para tabela_termo_b2b
        for i, item in enumerate(tabela_termo_b2b.get_children(), start=3):
            try:
                values = tabela_termo_b2b.item(item, 'values')
                status = values[-2]
                if status != "New" and status != "Pending Review":
                    market = values[2]
                    strike = values[6]          
                    exchange = lookup(market, commodities_ric, commodities_exchange)              
                    mercadoria = lookup(market, commodities_ric, commodities_mercadoria)                                             
                    type = lookup(market, commodities_ric, commodities_type)
                    notional = round(int(values[9].replace(",", "").replace("-", "")) * float(values[6].replace(",", "").replace("-", "")) * float(lookup(market, commodities_ric, commodities_factor).replace(",", ".")) ,2)
                    premium = float(values[19].replace(",", "").replace("-", ""))
                    valor_base = int(values[9].replace(",", "").replace("-", ""))              
                    strike = math.trunc(float(values[6].replace(",", "").replace("-", ""))   * float(lookup(market, commodities_ric, commodities_factor).replace(",", ".")) * 10**8) /10**8
                    mm_yyyy = str(lookup(market, commodities_ric, commodities_MM)) + "-" + str(lookup(market, commodities_ric, commodities_YYYY))             
                    simbolo_moeda = lookup(market, commodities_ric, commodities_mercadoria)[:3]
                    unity = lookup(market, commodities_ric, commodities_unity)
                    # Inicialize `datas_fixing` com um valor padrão
                    datas_fixing = ""

                    if values[16] != values[17]:                    
                        for fix_row in tabela_fixingstermo_b2b.get_children():
                            fix_values = tabela_fixingstermo_b2b.item(fix_row, 'values')
                            if fix_values[-1] == str(values[-1]):   
                                datas_fixing = ""                    
                                for j in range(5, 85):  # Iterar sobre os valores de fixação
                                    if fix_values[j].strip():                                                                        
                                        # Converta `next_date` para string antes de concatenar
                                        datas_fixing += fix_values[j] + ","

                                    else:
                                        break  # Stop the loop if an empty field is found

                    linha = [
                            "NDF -  TERMO MERCADORIA",  # TIPO DO CONTRATO A TERMO
                            "Codigo_Cetip",             # CÓDIGO CONTRATO CETIP
                            "INTRAGJP552",              # CÓDIGO DA CARTEIRA
                            "VENDEDOR" if values[3] == "Sell" else "COMPRADOR",  # POSIÇÃO DO PARTICIPANTE
                            "",                         # CPF/CNPJ CLIENTE PARTE
                            "JPM",                      # CONTRAPARTE
                            "",                         # CPF/CNPJ CLIENTE CONTRAPARTE
                            "NÃO",                      # CESTA GARANTIAS CONTRAPARTE
                            "NÃO",                      # CESTA GARANTIAS PARTE
                            str(notional),              # VALOR BASE (NOTIONAL)
                            safe_date_conversion_y_m_d(values[1]),  # DATA DE OPERAÇÃO
                            safe_date_conversion_y_m_d(values[1]),  # DATA DE REGISTRO
                            safe_date_conversion_y_m_d(values[10]), # DATA DE VENCIMENTO
                            "N/A",                      # MOEDA
                            exchange,                   # BOLSA REFERÊNCIA
                            mercadoria,                 # COMMODITY
                            type,                       # TIPO (FORMA)
                            str(valor_base),            # QUANTIDADE
                            unity,                      # UNIDADE DE NEGOCIAÇÃO
                            strike,                     # PREÇO DA OPERAÇÃO
                            values[8],                  # PARIDADE PARA LIQUIDAÇÃO
                            mm_yyyy,                    # MÊS E ANO DO VENCIMENTO
                            "D-" + str(networkdays(values[15], values[10], tabela_anbima)),  # COTAÇÃO PARA AJUSTE/SPOT
                            "0",                        # TAXA A TERMO (R$/MOEDA)
                            "N/A",                      # TAXA MÉDIA PARA TERMO ASIÁTICO
                            "N/A",                      # FONTE DE INFORMAÇÃO
                            "D-" + str(networkdays(values[17], values[10])),  # COTAÇÃO PARA O VENCIMENTO
                            "FINAL" if values[16] == values[17] else "ASIATCO", # TIPO DE AJUSTE
                            "Strike em BRL," + datas_fixing[:-1] if values[8] =="BRR" else datas_fixing[:-1],  # OBSERVAÇÃO
                            "N/A",                      # Fator de Desconto
                            values[-4],                 # Identifier
                        ]    

                    tabela_boletatermo.insert("", "end", values=linha)
                    ajustar_largura_colunas(tabela_boletatermo, colunas_boleta_termo, tabview)
                    vincular_evento_duplo_clique(tabela_boletatermo, colunas_boleta_termo)
                    vincular_navegacao_setas(tabela_boletatermo)
                    highlight_deals_not_mapped(tabela_boletatermo, "TERMO")
            except Exception as e:
                print(f"Erro ao processar item na tabela_boletatermo: {e}")         
    except Exception as e:
                print(f"Erro ao processar item nas tabelas: {e}")  
                
def insert_or_update_all_tables(data_and_tables):
    """
    Insere ou atualiza múltiplas linhas em múltiplas tabelas no banco de dados em uma única operação.

    :param conn: Conexão com o banco de dados.
    :param data_and_tables: Lista de tuplas, onde cada tupla contém (rows_data, table_name).
    """
    try:
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")
        ids_opcao_edit = []
        ids_termo_edit = []
        ids_opcao = []
        ids_termo = []

        # Verificar e atualizar/inserir para opcao
        for rows_data, table_name in data_and_tables:
            if table_name == "opcao_base_deals":
                for row in rows_data:
                    key_value = row[0]
                    id = row[-4]
                    if id != "":
                       ids_opcao.append(id) 
                    # Verificar se a operação já existe na opcao_base_deals
                    cursor.execute('SELECT COUNT(*) FROM opcao_base_deals WHERE DealName = ?', (key_value,))
                    exists = cursor.fetchone()[0] > 0
                    if exists:
                        ids_opcao_edit.append(id)
        if ids_opcao:
            if ids_opcao_edit:
                # Deletar as linhas existentes para opcao
                cursor.execute('DELETE FROM opcao_base_deals WHERE Identifier IN ({})'.format(','.join('?' for _ in ids_opcao)), ids_opcao)
                cursor.execute('DELETE FROM opcao_base_file WHERE Identifier IN ({})'.format(','.join('?' for _ in ids_opcao)), ids_opcao)
                cursor.execute('DELETE FROM opcao_base_fixings WHERE Identifier IN ({})'.format(','.join('?' for _ in ids_opcao)), ids_opcao)
            
            # Inserir nas tabelas correspondentes para opcao
            for rows_data, table_name in data_and_tables:
                if "opcao" in table_name and rows_data:
                    placeholders = ", ".join(["?" for _ in range(len(rows_data[0]))])
                    sql_insert = f"INSERT INTO {table_name} VALUES ({placeholders})"
                    cursor.executemany(sql_insert, rows_data)

        # Verificar e atualizar/inserir para termo
        for rows_data, table_name in data_and_tables:
            if table_name == "termo_base_deals":
                for row in rows_data:
                    key_value = row[0]
                    id = row[-4]
                    if id != "":
                        ids_termo.append(id)
                    # Verificar se a operação já existe na termo_base_deals
                    cursor.execute('SELECT COUNT(*) FROM termo_base_deals WHERE DealName = ?', (key_value,))
                    exists = cursor.fetchone()[0] > 0
                    if exists:
                        ids_termo_edit.append(id)
        if ids_termo:
            if ids_termo_edit:
                # Atualizar as tabelas correspondentes para termo
                cursor.execute('DELETE FROM termo_base_deals WHERE Identifier IN ({})'.format(','.join('?' for _ in ids_termo)), ids_termo)
                cursor.execute('DELETE FROM termo_base_file WHERE Identifier IN ({})'.format(','.join('?' for _ in ids_termo)), ids_termo)
                cursor.execute('DELETE FROM termo_base_fixings WHERE Identifier IN ({})'.format(','.join('?' for _ in ids_termo)), ids_termo)
            
            # Inserir nas tabelas correspondentes para termo
            for rows_data, table_name in data_and_tables:
                if "termo" in table_name and rows_data:
                    placeholders = ", ".join(["?" for _ in range(len(rows_data[0]))])
                    sql_insert = f"INSERT INTO {table_name} VALUES ({placeholders})"
                    cursor.executemany(sql_insert, rows_data)

        conn.commit()
        conn.close()
    except Exception as e:
        conn.close()
        messagebox.showerror("Database Error", f"An error occurred: {e}")
        
def get_columns_for_table(table_name):
    """
    Retorna as colunas para a tabela especificada.

    :param table_name: Nome da tabela.
    :return: Lista de colunas.
    """
    # Supondo que as colunas estejam declaradas globalmente
    global termo_deals_columns, opcao_fixings_columns, termo_fixings_columns, termo_file_columns, opcao_file_columns, opcao_deals_columns

    columns_map = {
        "opcao_base_deals": opcao_deals_columns,
        "opcao_base_file": opcao_file_columns,
        "opcao_base_fixings": opcao_fixings_columns,
        "termo_base_deals": termo_deals_columns,
        "termo_base_file": termo_file_columns,
        "termo_base_fixings": termo_fixings_columns
    }

    return columns_map[table_name]



        
def popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None):
    # Perguntar ao usuário se deseja seguir com o cadastro
    resposta = messagebox.askyesno("B3 File", "Wish to proceed?")
    if not resposta:
        return  # Se o usuário clicar em "No", encerra a função
    try:        
        # Verificar se a aba "Monitor" já existe
        global entry_filtro_commodities
        entry_filtro_commodities.delete(0, tk.END)
        global tabela_commodities_data
        for item in tabela_commodities_data:            
                tabela_commodities.insert("", "end", values=item)
        if "Arquivo B3" not in abas_existentes:
            # Inicializar o monitor de operações e obter os tabviews
            # arquivo_b3(tabview,abas_existentes)
            abas_existentes.append("Arquivo B3")               

        # Extrair dados necessários das tabelas
        commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()

        def rept(char, count):
            return char * count
        
         # Listas para acumular linhas a serem inseridas
        linhas_opcao_cliente = []
        linhas_opcao_b2b = []
        linhas_opcao_arquivo_cliente = []
        linhas_opcao_arquivo_b2b = []
        linhas_opcao_fixings_cliente = []
        linhas_opcao_fixings_b2b = []
        linhas_termo_cliente = []
        linhas_termo_b2b = []
        linhas_termo_arquivo_cliente = []
        linhas_termo_arquivo_b2b = []
        linhas_termo_fixings_cliente = []
        linhas_termo_fixings_b2b = []        
        DealNames = []
        Markets = []
        Types = []
        Strikes = []
        IntermediatesCCY = []
        TotalNotionals = []
        SettlementDates = []
        FxConvDates = []
        FirstFixingDates = []
        LastFixingDates = []
        SIDS = []    
        Instruments = []
        TradeDates_update = [] 
        Counterparties_update = [] 
        Statuses_update = []
        Status_update = "Pending Review"
        # Processar dados para tabela_opcao_cliente
        for i, item in enumerate(tabela_opcao_cliente.get_children(), start=3):
            try:
                values = list(tabela_opcao_cliente.item(item, 'values'))  # Converte a tupla em uma lista
                market = values[2]
                strike = values[6]
                intermediate_ccy = values[8]
                cntpy = values[18]
                premiumperunit = values[20]
                athena_id = values[0]
                start_date = values[17]
                end_date = values[10]
                SID = getpass.getuser()
                SID = SID[0].upper() + SID[1:]
                status = values[26]

                strike_result = strike_formula_opcao(market, strike, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                cnpj_cliente = lookup(cntpy, cntpy_acronym, cntpy_taxid)
                decimal_strike_result = decimal_formula_opcao(market, strike, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                valor_base = int(values[9].replace(",", "").replace("-", ""))
                pu_result = pu_formula_opcao(market, premiumperunit, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                pu_decimal_result = pudecimal_formula_opcao(market, premiumperunit, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                cetip_account = lookup(cntpy, cntpy_acronym, cntpy_b3_account)

               # Inicialize `datas_fixing` com um valor padrão
                datas_fixing = rept(" ", 8)
                if status == "New":
                    values[26] = "Pending Review"  # Modifica a lista
                    values[27] = SID
                    tabela_opcao_cliente.item(item, values=values)  # Atualiza o Treeview
                    linhas_opcao_cliente.append(values)
                    Instruments.append(values[4])
                    TradeDates_update.append(values[1])                                        
                    Counterparties_update.append(values[18])                       
                    DealNames.append(values[0])                    
                    Markets.append(values[2])
                    Types.append(values[3])
                    Strikes.append(values[6])
                    IntermediatesCCY.append(values[8])
                    TotalNotionals.append(values[9])
                    SettlementDates.append(values[10])
                    FxConvDates.append(values[15])
                    FirstFixingDates.append(values[16])
                    LastFixingDates.append(values[17])
                    Statuses_update.append(values[-2])
                    if values[16] != values[17]:                    
                        for fix_row in tabela_fixingsopcao_cliente.get_children():
                            fix_values = list(tabela_fixingsopcao_cliente.item(fix_row, 'values'))
                            if fix_values[-4] == str(values[-4]): 
                                fix_values[-2] = "Pending Review"
                                fix_values[-1] = SID
                                tabela_fixingsopcao_cliente.item(fix_row, values=fix_values)  # Atualiza o Treeview
                                linhas_opcao_fixings_cliente.append(fix_values)
                                datas_fixing = ""                    
                                for j in range(5, 50):  # Iterar sobre os valores de fixação
                                    if fix_values[j].strip():  
                                        start_date = datetime.strptime(fix_values[j], '%d/%m/%Y')                                      
                                        next_date = next_workday_opcao(start_date, networkdays(values[17], values[10]))
                                        # Convert `next_date` to a string in the format 'dd/mm/yyyy'
                                        next_date_str = next_date.strftime("%d/%m/%Y")                                    
                                        # Converta `next_date` para string antes de concatenar
                                        datas_fixing += safe_date_conversion_yyyymmaa(next_date_str) + rept("0", 9 - len(str(valor_base))) + str(valor_base) + rept("0", 8)

                                    else:
                                        break  # Stop the loop if an empty field is found
                    
                    linha = [
                        "OPCAO",	# Id. Sist.
                        "1",	# ID Tipo de Linha
                        "0002",	# Codigo da Operação
                        "OFVC" if values[4] == "Option (Put)" else "OFCC",	# Tipo Contrato
                        values[23],	# Código
                        "73760.00-9",	# Conta Parte
                        cetip_account,	# Conta Contraparte
                        "P2" if values[3] == "Sell" else "P1",	# Papel Parte
                        rept(" ", 3),	# Moeda Base/Índice/Ações
                        safe_date_conversion(values[1]),	# Data Inicio
                        safe_date_conversion(values[10]),	# Data de Vencimento
                        strike_result,	# Preço de Exercício
                        decimal_strike_result,	# Casas Decimais Preço Exercicio
                        pu_result,	# Prêmio Unitário
                        pu_decimal_result,	# Casas Decimais do Prêmio Unitário
                        rept("0", 14 - len(str(valor_base))) + str(valor_base),	# Valor Base em Moeda Estrangeira / Quantidade
                        rept("0", 8),	# Casas Decimais do Valor Base em Moeda Estrangeira ou Quantidade.
                        rept(" ", 10),	# Cotação / Índice Limite
                        rept(" ", 8),	# Casas Decimais da Cotação / Índice Limite
                        "2",	# Tipo de Exercício
                        rept(" ", 10),	# Banco Liquidante
                        "0" if cetip_account == "73760.10-2" or (datetime.strptime(values[22], "%d-%b-%Y").date() != datetime.today().date() and cetip_account != "73760.10-2") else "2",	# Modalidade
                        "1",	# Adesão a Contrato
                        ''.join([str(random.randint(0,9)) for _ in range(10)]),	# Meu Número
                        rept(" ", 10),	# Conta Intermediador
                        rept(" ", 10),	# Comissão paga pelo Titular
                        rept(" ", 2),	# Casas Decimais da Comissão paga pelo Titular
                        rept(" ", 10),	# Comissão paga pelo Lançador
                        rept(" ", 2),	# Casas Decimais da Comissão paga pelo Lançador
                        "N",	# Cross-Rate na Avaliação
                        "23",	# Fonte de Informação
                        str(networkdays(values[17], values[10])),  # Cotação para o Vencimento
                        rept(" ", 1),	# Boletim
                        rept(" ", 1),	# Horário do Boletim
                        rept(" ", 1),	# Fonte de Consulta
                        rept(" ", 20),	# Outra Fonte de Consulta
                        rept(" ", 20),	# Tela ou Função de Consulta
                        rept(" ", 20),	# Praça de Negociação
                        rept(" ", 5),	# Horário de Consulta
                        rept(" ", 1),	# Cotação – Taxa de Câmbio
                        rept(" ", 1),	# Cotação – Paridade
                        rept(" ", 8),	# Data de Avaliação
                        rept(" ", 14),	# CPF / CNPJ da Parte
                        cnpj_cliente if cetip_account == "73760.10-2" else rept(" ", 14),	# CPF / CNPJ da Contraparte
                        rept(" ", 3),	# Moeda Cotada
                        rept(" ", 1),	# Barreiras
                        rept(" ", 10),	# Trigger In
                        rept(" ", 8),	# Casas Decimais do Trigger In
                        rept(" ", 10),	# Trigger Out
                        rept(" ", 8),	# Casas Decimais do Trigger Out
                        "N",	# Cesta de Garantias - Lançador
                        rept(" ", 1),	# Forma de Verificação
                        rept(" ", 1),	# Rebate
                        rept(" ", 10),	# Valor do Rebate
                        rept(" ", 8),	# Casas decimais do Valor do Rebate
                        rept(" ", 1),	# Liquidação do Rebate
                        rept(" ", 10),	# Código da Ação / Indice Internacional
                        rept(" ", 1),	# Ajuste de Proventos pelas
                        rept(" ", 1),	# Proteção contra Provento em Dinheiro
                        rept(" ", 1),	# Trigger – Proporção
                        rept(" ", 1),	# Trigger – Forma de Disparo
                        rept(" ", 1),	# Trigger – Tipo de Disparo
                        "S" if values[8] == "BRR" else "N",	# Preço de Exercício em Reais
                        "N",	# Opção Quanto
                        rept(" ", 3),	# Cotação para Opção Quanto
                        rept(" ", 8),	# Casas decimais do Cotação para Opção Quanto
                        safe_date_conversion(values[22]),	# Data de Liquidação do Prêmio
                        rept(" ", 15 - len(values[2])) + values[2],	# Mercadoria
                        str(networkdays(values[15], values[10], tabela_anbima)),	# Cotação para Moeda
                        rept(" ", 280),	# Observação
                        rept(" ", 1) if values[16] == values[17] else "1",	# Média para Opção Asiática
                        datas_fixing,	# Data de Verificação
                        rept(" ", 9) if values[16] == values[17] else "",	# Valor/Quantidade de Referência
                        rept(" ", 8) if values[16] == values[17] else "",	# Casas Decimais do Valor/Quantidade de Referência
                        rept(" ", 8) if values[16] == values[17] else "",	# Data de Verificação
                        rept(" ", 9) if values[16] == values[17] else "",	# Valor/Quantidade de Referência
                        rept(" ", 8) if values[16] == values[17] else "",	# Casas Decimais do Valor/Quantidade de Referência
                        values[24],       # Identifier   
                        values[25],        # Index      
                        "Pending Review",        # Status      
                        SID,        # SID      
                    ]                                                         
                    linhas_opcao_arquivo_cliente.append(linha)
                    tabela_arquivoopcao_cliente.insert("", "end", values=linha) 
                    SIDS.append(values[-1])                                      
            except Exception as e:
                print(f"Erro ao processar item na tabela_opcao_cliente: {e}")

        # Processar dados para tabela_opcao_b2b
        for i, item in enumerate(tabela_opcao_b2b.get_children(), start=3):
            try:
                values = list(tabela_opcao_b2b.item(item, 'values'))  # Converte a tupla em uma lista
                market = values[2]
                strike = values[6]
                intermediate_ccy = values[8]
                cntpy = values[17]
                premiumperunit = values[20]               
                SID = getpass.getuser()
                SID = SID[0].upper() + SID[1:]
                status = values[26]

                strike_result = strike_formula_opcao(market, strike, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                cnpj_cliente = lookup(cntpy, cntpy_acronym, cntpy_taxid)
                decimal_strike_result = decimal_formula_opcao(market, strike, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                valor_base = int(values[9].replace(",", "").replace("-", ""))
                pu_result = pu_formula_opcao(market, premiumperunit, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                pu_decimal_result = pudecimal_formula_opcao(market, premiumperunit, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                cetip_account = lookup(cntpy, cntpy_acronym, cntpy_b3_account)

               # Inicialize `datas_fixing` com um valor padrão
                datas_fixing = rept(" ", 8)
                if status == "New":
                    values[26] = "Pending Review"  # Modifica a lista
                    values[-1] = SID
                    tabela_opcao_b2b.item(item, values=values)  # Atualiza o Treeview
                    linhas_opcao_b2b.append(values)
                    Instruments.append(values[4])
                    TradeDates_update.append(values[1])                                        
                    Counterparties_update.append(values[18])                       
                    DealNames.append(values[0])                    
                    Markets.append(values[2])
                    Types.append(values[3])
                    Strikes.append(values[6])
                    IntermediatesCCY.append(values[8])
                    TotalNotionals.append(values[9])
                    SettlementDates.append(values[10])
                    FxConvDates.append(values[15])
                    FirstFixingDates.append(values[16])
                    LastFixingDates.append(values[17])
                    Statuses_update.append(values[-2])
                    if values[16] != values[17]:                    
                        for fix_row in tabela_fixingsopcao_b2b.get_children():
                            fix_values = list(tabela_fixingsopcao_b2b.item(fix_row, 'values'))
                            if fix_values[-4] == str(values[-4]):
                                fix_values[-2] = "Pending Review"    
                                fix_values[-1] = SID                            
                                tabela_fixingsopcao_b2b.item(fix_row, values=fix_values)  # Atualiza o Treeview
                                linhas_opcao_fixings_b2b.append(fix_values)
                                datas_fixing = ""                    
                                for j in range(5, 50):  # Iterar sobre os valores de fixação
                                    if fix_values[j].strip():   
                                        start_date = datetime.strptime(fix_values[j], '%d/%m/%Y')                                      
                                        next_date = next_workday_opcao(start_date, networkdays(values[17], values[10]) - 1)
                                        # Convert `next_date` to a string in the format 'dd/mm/yyyy'
                                        next_date_str = next_date.strftime("%d/%m/%Y")                                    
                                        # Converta `next_date` para string antes de concatenar
                                        datas_fixing += safe_date_conversion_yyyymmaa(next_date_str) + rept("0", 9 - len(str(valor_base))) + str(valor_base) + rept("0", 8)                                                                    
                                    else:
                                        break  # Stop the loop if an empty field is found

                    linha = [
                        "OPCAO",	# Id. Sist.
                        "1",	# ID Tipo de Linha
                        "0002",	# Codigo da Operação
                        "OFVC" if values[4] == "Option (Put)" else "OFCC",	# Tipo Contrato
                        values[23],	# Código
                        "73760.00-9",	# Conta Parte
                        "00041.00-7",	# Conta Contraparte
                        "P1" if values[3] == "Sell" else "P2",	# Papel Parte
                        rept(" ", 3),	# Moeda Base/Índice/Ações
                        safe_date_conversion(values[1]),	# Data Inicio
                        safe_date_conversion(values[10]),	# Data de Vencimento
                        strike_result,	# Preço de Exercício
                        decimal_strike_result,	# Casas Decimais Preço Exercicio
                        pu_result,	# Prêmio Unitário
                        pu_decimal_result,	# Casas Decimais do Prêmio Unitário
                        rept("0", 14 - len(str(valor_base))) + str(valor_base),	# Valor Base em Moeda Estrangeira / Quantidade
                        rept("0", 8),	# Casas Decimais do Valor Base em Moeda Estrangeira ou Quantidade.
                        rept(" ", 10),	# Cotação / Índice Limite
                        rept(" ", 8),	# Casas Decimais da Cotação / Índice Limite
                        "2",	# Tipo de Exercício
                        rept(" ", 10),	# Banco Liquidante
                        "0" if datetime.strptime(values[22], "%d-%b-%Y").date() != datetime.today().date() else "2",# Modalidade
                        "1",	# Adesão a Contrato
                        ''.join([str(random.randint(0,9)) for _ in range(10)]),	# Meu Número
                        rept(" ", 10),	# Conta Intermediador
                        rept(" ", 10),	# Comissão paga pelo Titular
                        rept(" ", 2),	# Casas Decimais da Comissão paga pelo Titular
                        rept(" ", 10),	# Comissão paga pelo Lançador
                        rept(" ", 2),	# Casas Decimais da Comissão paga pelo Lançador
                        "N",	# Cross-Rate na Avaliação
                        "23",	# Fonte de Informação
                        str(networkdays(values[17], values[10])),  # Cotação para o Vencimento
                        rept(" ", 1),	# Boletim
                        rept(" ", 1),	# Horário do Boletim
                        rept(" ", 1),	# Fonte de Consulta
                        rept(" ", 20),	# Outra Fonte de Consulta
                        rept(" ", 20),	# Tela ou Função de Consulta
                        rept(" ", 20),	# Praça de Negociação
                        rept(" ", 5),	# Horário de Consulta
                        rept(" ", 1),	# Cotação – Taxa de Câmbio
                        rept(" ", 1),	# Cotação – Paridade
                        rept(" ", 8),	# Data de Avaliação
                        rept(" ", 14),	# CPF / CNPJ da Parte
                        rept(" ", 14),	# CPF / CNPJ da Contraparte
                        rept(" ", 3),	# Moeda Cotada
                        rept(" ", 1),	# Barreiras
                        rept(" ", 10),	# Trigger In
                        rept(" ", 8),	# Casas Decimais do Trigger In
                        rept(" ", 10),	# Trigger Out
                        rept(" ", 8),	# Casas Decimais do Trigger Out
                        "N",	# Cesta de Garantias - Lançador
                        rept(" ", 1),	# Forma de Verificação
                        rept(" ", 1),	# Rebate
                        rept(" ", 10),	# Valor do Rebate
                        rept(" ", 8),	# Casas decimais do Valor do Rebate
                        rept(" ", 1),	# Liquidação do Rebate
                        rept(" ", 10),	# Código da Ação / Indice Internacional
                        rept(" ", 1),	# Ajuste de Proventos pelas
                        rept(" ", 1),	# Proteção contra Provento em Dinheiro
                        rept(" ", 1),	# Trigger – Proporção
                        rept(" ", 1),	# Trigger – Forma de Disparo
                        rept(" ", 1),	# Trigger – Tipo de Disparo
                        "S" if values[8] == "BRR" else "N",	# Preço de Exercício em Reais
                        "N",	# Opção Quanto
                        rept(" ", 3),	# Cotação para Opção Quanto
                        rept(" ", 8),	# Casas decimais do Cotação para Opção Quanto
                        safe_date_conversion(values[22]),	# Data de Liquidação do Prêmio
                        rept(" ", 15 - len(values[2])) + values[2],	# Mercadoria
                        str(networkdays(values[15], values[10], tabela_anbima)),	# Cotação para Moeda
                        rept(" ", 280),	# Observação
                        rept(" ", 1) if values[16] == values[17] else "1",	# Média para Opção Asiática
                        datas_fixing,	# Data de Verificação
                        rept(" ", 9) if values[16] == values[17] else "",	# Valor/Quantidade de Referência
                        rept(" ", 8) if values[16] == values[17] else "",	# Casas Decimais do Valor/Quantidade de Referência
                        rept(" ", 8) if values[16] == values[17] else "",	# Data de Verificação
                        rept(" ", 9) if values[16] == values[17] else "",	# Valor/Quantidade de Referência
                        rept(" ", 8) if values[16] == values[17] else "",	# Casas Decimais do Valor/Quantidade de Referência
                        values[24],       # Identifier   
                        values[25],        # Index      
                        "Pending Review",        # Status      
                        SID,        # SID         
                    ]                                   


                    tabela_arquivoopcao_b2b.insert("", "end", values=linha)                    
                    linhas_opcao_arquivo_b2b.append(linha)   
                    SIDS.append(values[-1])        
            except Exception as e:
                print(f"Erro ao processar item na tabela_opcao_b2b: {e}")
        
                

            # Processar dados para tabela_termo_cliente
        for i, item in enumerate(tabela_termo_cliente.get_children(), start=3):
            try:
                values = list(tabela_termo_cliente.item(item, 'values'))  # Converte para lista
                market = values[2]
                strike = values[6]
                intermediate_ccy = values[8]
                cntpy = values[18]
                SID = getpass.getuser()
                SID = SID[0].upper() + SID[1:]
                status = values[21]

                strike_result = strike_formula(market, strike, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                cnpj_cliente = lookup(cntpy, cntpy_acronym, cntpy_taxid)
                tas_result = tas_formula(market, strike, intermediate_ccy, commodities_ric, commodities_factor)
                cetip_account = lookup(cntpy, cntpy_acronym, cntpy_b3_account)
                valor_base = int(values[9].replace(",", "").replace("-", ""))
                quantidade_datas_fixing = rept("0", 3)

                if status == "New":
                    values[21] = "Pending Review"  # Modifica a lista
                    values[22] = SID
                    tabela_termo_cliente.item(item, values=values)  # Atualiza o Treeview
                    linhas_termo_cliente.append(values)
                    Instruments.append(values[4])
                    TradeDates_update.append(values[1])                                        
                    Counterparties_update.append(values[18])                       
                    DealNames.append(values[0])                    
                    Markets.append(values[2])
                    Types.append(values[3])
                    Strikes.append(values[6])
                    IntermediatesCCY.append(values[8])
                    TotalNotionals.append(values[9])
                    SettlementDates.append(values[10])
                    FxConvDates.append(values[15])
                    FirstFixingDates.append(values[16])
                    LastFixingDates.append(values[17])
                    Statuses_update.append(values[-2])
                    if values[16] != values[17]:
                        for fix_row in tabela_fixingstermo_cliente.get_children():
                            fix_values = list(tabela_fixingstermo_cliente.item(fix_row, 'values'))
                            if fix_values[-4] == str(values[-4]):
                                fix_values[87] = "Pending Review"
                                fix_values[88] = SID
                                tabela_fixingstermo_cliente.item(fix_row, values=fix_values)  # Atualiza o Treeview
                                linhas_termo_fixings_cliente.append(fix_values)
                                quantidade_datas_fixing = ""
                                quantidade_datas_fixing += "0" + fix_values[4]

                    linha = [
                        "TER" + rept(" ", 2),  # ID do Sistema
                        "1",  # ID Tipo de Linha
                        "0001",  # Código operação
                        ''.join([str(random.randint(0, 9)) for _ in range(10)]),  # Meu Número
                        "73760009",  # Lançamento do Participante (Conta)
                        0 if values[3] == "Buy" else 1,  # Papel (Posição do participante)
                        rept(" ", 14),  # CPF/CNPJ Cliente Parte
                        cetip_account.replace(".", "").replace("-", ""),  # Contraparte
                        cnpj_cliente if cetip_account == "73760.10-2" else rept(" ", 14),  # CPF/CNPJ Cliente Contraparte
                        "S",  # Contrato Global
                        rept(" ", 19) + "4",  # Classe do Ativo Subjacente
                        " 340" if values[2] in ["NACX0005", "PTS005", "PTS002", "PTS006", "PTS003"] else " 358",  # Fonte Informação
                        rept(" ", 3),  # Moeda de Referência
                        rept(" ", 3),  # Moeda Cotada
                        rept(" ", 1),  # Cotação para o Vencimento
                        rept("0", 14 - len(str(valor_base))) + str(valor_base) + "00",  # Valor Base / Quantidade
                        rept(" ", 10 - len(values[2])) + values[2],  # Código do Ativo Subjacente
                        strike_result,  # Taxa a Termo (R$/Moeda)
                        safe_date_conversion(values[17]) if values[16] == values[17] else rept(" ", 8),  # Data de fixing do Ativo Subjacente
                        safe_date_conversion(values[1]),  # Data de Operação
                        safe_date_conversion(values[10]),  # Data vencimento
                        rept(" ", 1),  # Boletim
                        "F" if values[2] in ["NACX0005", "PTS005", "PTS002", "PTS006", "PTS003"] else "A",  # Tipo de Cotação
                        safe_date_conversion(values[15]) if values[8] != "BRR" else rept(" ", 8),  # Data de Fixing da Moeda
                        rept(" ", 1),  # Cross Rate na Avaliação?
                        rept(" ", 1),  # Fonte de Consulta
                        rept(" ", 8),  # Tela ou Função de Consulta
                        rept(" ", 8),  # Praça de Negociação
                        rept(" ", 8),  # Horário de Consulta
                        rept(" ", 1),  # Cotação - Taxa de Câmbio R$/USD
                        rept(" ", 1),  # Cotação - Paridade (Moeda/USD ou USD/ Moeda)
                        rept(" ", 8),  # Data de Avaliação
                        rept(" ", 10),  # Código da paridade cross
                        rept(" ", 8),  # Data de fixing da paridade cross
                        "S" if values[6].upper().startswith("TAS") else "N",  # Termo a Termo
                        safe_date_conversion(values[1]) if values[6].upper().startswith("TAS") else rept(" ", 8),  # Data de Fixação
                        "V" if values[6].upper().startswith("TAS") else rept(" ", 1),  # Forma de Atualização
                        tas_result,  # Valor / Percentual Negociado
                        rept(" ", 1),  # Cotação para fixing
                        "N",  # Atualizar Valor Base?
                        rept(" ", 12),  # Cotação Inicial
                        "N",  # Ajustar Taxa
                        rept(" ", 1),  # Responsável pelo Ajuste da Taxa
                        rept(" ", 8),  # Data Inicial para Ajuste da Taxa
                        rept(" ", 8),  # Data Final para Ajuste de taxa
                        rept(" ", 1),  # Limites
                        rept(" ", 14),  # Superior (Paridade)
                        rept(" ", 14),  # Inferior (Paridade)
                        rept(" ", 8),  # Data de Liquidação do Prêmio
                        rept(" ", 1),  # Prêmio a ser pago pelo
                        rept(" ", 16),  # Valor do Prêmio
                        rept(" ", 1),  # Modalidade de Liquidação
                        rept(" ", 1),  # Prêmio em Moeda Estrangeira
                        rept(" ", 8),  # Data de fixing da moeda do prêmio
                        "S" if values[8] == "BRR" else rept(" ", 1),  # Taxa a Termo em Reais
                        rept(" ", 280),  # Observação
                        rept(" ", 14 - len(values[0])) + values[0],  # Código Identificador
                        "N" if values[16] == values[17] else "A",  # Tipo Média Asiático
                        quantidade_datas_fixing,  # Quantidade de Datas de Verificação
                        values[19],  # Identifier
                        values[20],  # Index
                        "Pending Review",  # Status
                        SID,  # SID
                    ]

                    tabela_arquivotermo_cliente.insert("", "end", values=linha)
                    linhas_termo_arquivo_cliente.append(linha)
                    SIDS.append(values[-1])                    

                    if values[16] != values[17]:
                        for fix_row in tabela_fixingstermo_cliente.get_children():
                            fix_values = list(tabela_fixingstermo_cliente.item(fix_row, 'values'))
                            if fix_values[85] == str(values[-4]):                               
                                for j in range(5, len(fix_values)):  # Adjusted to iterate over values_fixing
                                    if fix_values[j].strip():
                                        linha = [
                                            "TER" + rept(" ", 2),
                                            "2",
                                            "0001",
                                            safe_date_conversion_yyyymmaa(fix_values[j]),
                                            rept("0", 16),
                                            safe_date_conversion_yyyymmaa(fix_values[j]) if values[8] == "BRR" else rept(" ", 8),
                                            rept(" ", 8) if values[8] == "BRR" else rept(" ", 9),
                                            rept(" ", 1),
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",  
                                            "",
                                            "",
                                            "", 
                                            "",
                                            "",
                                            "",
                                            "",                                     
                                            values[19],  # Identifier
                                            values[20],  # Index
                                            "Pending Review",  # Status
                                            SID,  # SID                             
                                        ]
                                        tabela_arquivotermo_cliente.insert("", "end", values=linha)  
                                        linhas_termo_arquivo_cliente.append(linha)                                  
                                    else:
                                        break  # Stop the loop if an empty field is found
                    
            except Exception as e:
                print(f"Erro ao processar item na tabela_termo_cliente: {e}")

        # Processar dados para tabela_termo_b2b
        for i, item in enumerate(tabela_termo_b2b.get_children(), start=3):
            try:
                values = list(tabela_termo_b2b.item(item, 'values'))  # Converte para lista
                market = values[2]
                strike = values[6]
                intermediate_ccy = values[8]
                cntpy = values[18]
                SID = getpass.getuser()
                SID = SID[0].upper() + SID[1:]
                status = values[21]

                strike_result = strike_formula(market, strike, intermediate_ccy, commodities_ric, commodities_factor, tabview)
                cnpj_cliente = lookup(cntpy, cntpy_acronym, cntpy_taxid)
                tas_result = tas_formula(market, strike, intermediate_ccy, commodities_ric, commodities_factor)
                cetip_account = lookup(cntpy, cntpy_acronym, cntpy_b3_account)
                valor_base = int(values[9].replace(",", "").replace("-", ""))
                quantidade_datas_fixing = rept("0", 3)

                if status == "New":
                    values[21] = "Pending Review"  # Modifica a lista
                    values[22] = SID
                    tabela_termo_b2b.item(item, values=values)  # Atualiza o Treeview
                    linhas_termo_b2b.append(values)
                    Instruments.append(values[4])
                    TradeDates_update.append(values[1])                                        
                    Counterparties_update.append(values[18])                       
                    DealNames.append(values[0])                    
                    Markets.append(values[2])
                    Types.append(values[3])
                    Strikes.append(values[6])
                    IntermediatesCCY.append(values[8])
                    TotalNotionals.append(values[9])
                    SettlementDates.append(values[10])
                    FxConvDates.append(values[15])
                    FirstFixingDates.append(values[16])
                    LastFixingDates.append(values[17])
                    Statuses_update.append(values[-2])
                    if values[16] != values[17]:
                        for fix_row in tabela_fixingstermo_b2b.get_children():
                            fix_values = list(tabela_fixingstermo_b2b.item(fix_row, 'values'))
                            if fix_values[-4] == str(values[-4]):     
                                fix_values[87] = "Pending Review"
                                fix_values[88] = SID
                                tabela_fixingstermo_b2b.item(fix_row, values=fix_values)  # Atualiza o Treeview                                                           
                                linhas_termo_fixings_b2b.append(fix_values)                                
                                quantidade_datas_fixing = ""
                                quantidade_datas_fixing += "0" + fix_values[4]            

                    linha = [			
                            "TER" + rept(" ", 2),	#ID do Sistema
                            "1",				    #ID Tipo de Linha
                            "0001",				    #Código operação
                            ''.join([str(random.randint(0,9)) for _ in range(10)]),				#Meu Número
                            "73760009",				#Lançamento do Participante (Conta)
                            1 if values[3] == "Buy" else 0,				#Papel (Posição do participante)
                            rept(" ", 14),			#CPF/CNPJ Cliente Parte
                            "00041007", 			#Contraparte
                            rept(" ", 14),			#CPF/CNPJ Cliente Contraparte
                            "S",				    #Contrato Global
                            rept(" ", 19) + "4",				#Classe do Ativo Subjacente
                            " 340" if values[2] in ["NACX0005", "PTS005", "PTS002", "PTS006", "PTS003"] else " 358",				#Fonte Informação
                            rept(" ", 3),			#Moeda de Referência
                            rept(" ", 3),			#Moeda Cotada
                            rept(" ", 1),			#Cotação para o Vencimento
                            rept("0", 14 - len(str(valor_base))) + str(valor_base) + "00",				#Valor Base / Quantidade
                            rept(" ", 10 - len(values[2])) + values[2],				#Código do Ativo Subjacente
                            strike_result,			#Taxa a Termo (R$/Moeda)
                            safe_date_conversion(values[17]) if values[16] == values[17] else rept(" ", 8),				#Data de fixing do Ativo Subjacente
                            safe_date_conversion(values[1]),				#Data de Operação
                            safe_date_conversion(values[10]),				#Data vencimento
                            rept(" ", 1),			#Boletim
                            "F" if values[2] in ["NACX0005", "PTS005", "PTS002", "PTS006", "PTS003"] else "A",				#Tipo de Cotação
                            safe_date_conversion(values[15]) if values[8] != "BRR" else rept(" ", 8),				#Data de Fixing da Moeda
                            rept(" ", 1),			#Cross Rate na Avaliação?
                            rept(" ", 1),			#Fonte de Consulta
                            rept(" ", 8),			#Tela ou Função de Consulta
                            rept(" ", 8),			#Praça de Negociação
                            rept(" ", 8),			#Horário de Consulta
                            rept(" ", 1),			#Cotação - Taxa de Câmbio R$/USD
                            rept(" ", 1),			#Cotação - Paridade (Moeda/USD ou USD/ Moeda)
                            rept(" ", 8),			#Data de Avaliação
                            rept(" ", 10),			#Código da paridade cross
                            rept(" ", 8),			#Data de fixing da paridade cross
                            "S" if values[6].upper().startswith("TAS") else "N",				#Termo a Termo
                            safe_date_conversion(values[1]) if values[6].upper().startswith("TAS") else rept(" ", 8),				#Data de Fixação
                            "V" if values[6].upper().startswith("TAS") else rept(" ", 1),				#Forma de Atualização
                            tas_result,				#Valor / Percentual Negociado
                            rept(" ", 1),			#Cotação para fixing
                            "N",				    #Atualizar Valor Base?
                            rept(" ", 12),			#Cotação Inicial
                            "N",				    #Ajustar Taxa
                            rept(" ", 1),			#Responsável pelo Ajuste da Taxa
                            rept(" ", 8),			#Data Inicial para Ajuste da Taxa
                            rept(" ", 8),			#Data Final para Ajuste de taxa
                            rept(" ", 1),			#Limites
                            rept(" ", 14),			#Superior (Paridade)
                            rept(" ", 14),			#Inferior (Paridade)
                            rept(" ", 8),			#Data de Liquidação do Prêmio
                            rept(" ", 1),			#Prêmio a ser pago pelo
                            rept(" ", 16),			#Valor do Prêmio
                            rept(" ", 1),			#Modalidade de Liquidação
                            rept(" ", 1),			#Prêmio em Moeda Estrangeira
                            rept(" ", 8),			#Data de fixing da moeda do prêmio
                            "S" if values[8] == "BRR" else rept(" ", 1),			#Taxa a Termo em Reais
                            rept(" ", 280),			#Observação
                            rept(" ", 14 - len(values[0])) + values[0],				#Código Identificador
                            "N" if values[16] == values[17] else "A",				#Tipo Média Asiático
                            quantidade_datas_fixing,			#Quantidade de Datas de Verificação
                            values[19],  # Identifier
                            values[20],  # Index
                            "Pending Review",  # Status
                            SID,  # SID

                    ]	
                    tabela_arquivotermo_b2b.insert("", "end", values=linha)
                    linhas_termo_arquivo_b2b.append(linha)
                    SIDS.append(values[-1])                  
                                        
                    if values[16] != values[17]:                    
                        for fix_row in tabela_fixingstermo_b2b.get_children():
                            fix_values = list(tabela_fixingstermo_b2b.item(fix_row, 'values'))
                            if fix_values[-4] == str(values[-4]):                                                                                 
                                for j in range(5, len(fix_values)):  # Adjusted to iterate over values_fixing
                                    if fix_values[j].strip():                            
                                        linha = [
                                            "TER" + rept(" ", 2),
                                            "2",
                                            "0001",
                                            safe_date_conversion_yyyymmaa(fix_values[j]),
                                            rept("0", 16),
                                            safe_date_conversion_yyyymmaa(fix_values[j]) if values[8] == "BRR" else rept(" ", 8),
                                            rept(" ", 8) if values[8] == "BRR" else rept(" ", 9),    
                                            rept(" ", 1),
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",  
                                            "",
                                            "",
                                            "", 
                                            "00041007",
                                            "",
                                            "",
                                            "",                                                                           
                                            values[19],  # Identifier
                                            values[20],  # Index
                                            "Pending Review",  # Status
                                            SID,  # SID
                                        ]                                  

                                        tabela_arquivotermo_b2b.insert("", "end", values=linha)     
                                        linhas_termo_arquivo_b2b.append(linha)                               
                                    else:
                                        break  # Stop the loop if an empty field is found                  
               
            except Exception as e:
                print(f"Erro ao processar item na tabela_termo_b2b: {e}")
       
        # Chamada única para inserir dados em todas as tabelas
        data_and_tables = [
            (linhas_opcao_cliente, "opcao_base_deals"),
            (linhas_opcao_b2b, "opcao_base_deals"),
            (linhas_opcao_arquivo_cliente, "opcao_base_file"),
            (linhas_opcao_arquivo_b2b, "opcao_base_file"),
            (linhas_opcao_fixings_cliente, "opcao_base_fixings"),
            (linhas_opcao_fixings_b2b, "opcao_base_fixings"),
            (linhas_termo_cliente, "termo_base_deals"),
            (linhas_termo_b2b, "termo_base_deals"),
            (linhas_termo_arquivo_cliente, "termo_base_file"),
            (linhas_termo_arquivo_b2b, "termo_base_file"),
            (linhas_termo_fixings_cliente, "termo_base_fixings"),
            (linhas_termo_fixings_b2b, "termo_base_fixings")
        ]
        
        insert_or_update_all_tables(data_and_tables)
       # Ajustar colunas e vincular eventos
        ajustar_largura_colunas(tabela_arquivoopcao_cliente, colunas_arquivo_opcao, tabview)
        ajustar_largura_colunas(tabela_arquivoopcao_b2b, colunas_arquivo_opcao, tabview)
        ajustar_largura_colunas(tabela_arquivotermo_cliente, colunas_arquivo_termo, tabview)
        ajustar_largura_colunas(tabela_arquivotermo_b2b, colunas_arquivo_termo, tabview)
        ajustar_largura_colunas(tabela_termo_cliente, colunas_termo, tabview)
        ajustar_largura_colunas(tabela_termo_b2b, colunas_termo, tabview)
        ajustar_largura_colunas(tabela_opcao_cliente, colunas_opcao, tabview)
        ajustar_largura_colunas(tabela_opcao_b2b, colunas_opcao, tabview)
        ajustar_largura_colunas(tabela_fixingstermo_cliente, colunas_fixings_termo, tabview)
        ajustar_largura_colunas(tabela_fixingstermo_b2b, colunas_fixings_termo, tabview)
        ajustar_largura_colunas(tabela_fixingsopcao_cliente, colunas_fixings_opcao, tabview)
        ajustar_largura_colunas(tabela_fixingsopcao_b2b, colunas_fixings_opcao, tabview)       
        highlight_duplicates(tabela_opcao_cliente, 'deals')
        highlight_duplicates(tabela_opcao_b2b, 'deals')
        highlight_duplicates(tabela_termo_cliente, 'deals')
        highlight_duplicates(tabela_termo_b2b, 'deals')
        highlight_duplicates(tabela_arquivoopcao_cliente, 'arquivo')
        highlight_duplicates(tabela_arquivoopcao_b2b, 'arquivo')
        highlight_duplicates(tabela_arquivotermo_cliente, 'arquivo')
        highlight_duplicates(tabela_arquivotermo_b2b, 'arquivo')
        highlight_duplicates(tabela_fixingsopcao_cliente, 'arquivo')
        highlight_duplicates(tabela_fixingsopcao_b2b, 'arquivo')
        highlight_duplicates(tabela_fixingstermo_cliente, 'arquivo')
        highlight_duplicates(tabela_fixingstermo_b2b, 'arquivo')         
        
        vincular_evento_duplo_clique_status(tabela_arquivoopcao_cliente, colunas_arquivo_opcao)
        vincular_evento_duplo_clique_status(tabela_arquivoopcao_b2b, colunas_arquivo_opcao)
        vincular_evento_duplo_clique_status(tabela_arquivotermo_cliente, colunas_arquivo_termo)
        vincular_evento_duplo_clique_status(tabela_arquivotermo_b2b, colunas_arquivo_termo)
        vincular_navegacao_setas(tabela_arquivoopcao_cliente)
        vincular_navegacao_setas(tabela_arquivoopcao_b2b)
        vincular_navegacao_setas(tabela_arquivotermo_cliente)
        vincular_navegacao_setas(tabela_arquivotermo_b2b) 
        status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update)                         
        
    except Exception as e:
        print(f"Erro ao popular arquivos B3: {e}")

def gerar_arquivos_tco_asian(caminho_completo, accronym, tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, identifiers, data_atual, mercadoria, index):
    # Caminhos para os arquivos TCO Asian
    if index == '1':
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"TCOAsian_Cliente_{accronym}_{mercadoria}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"TCOAsian_Banco_{accronym}_{mercadoria}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"TCOAsian_Lawton_{accronym}_{mercadoria}.txt")
    else:
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"TCOAsian_Cliente_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"TCOAsian_Banco_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"TCOAsian_Lawton_{accronym}_{mercadoria}_#{index}.txt")

    # Primeira linha do arquivo
    primeira_linha = f"TER  0{'0001'}JPMORGANBM{' ' * 10}{data_atual.strftime('%Y%m%d')}00003"
    primeira_linha_lawton = f"TER  0{'0001'}INTRAGLAWTONFDO{' ' * 5}{data_atual.strftime('%Y%m%d')}00003"

    # Gerar arquivos para cliente, banco e lawton
    gerar_arquivo(caminho_arquivo_cliente, primeira_linha, tabela_arquivotermo_cliente, identifiers)
    gerar_arquivo(caminho_arquivo_banco, primeira_linha, tabela_arquivotermo_b2b, identifiers)
    gerar_arquivo_lawton(caminho_arquivo_lawton, primeira_linha_lawton, tabela_arquivotermo_b2b, identifiers)

def gerar_arquivos_tco(caminho_completo, accronym, tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, identifiers, data_atual, mercadoria, index):
    # Caminhos para os arquivos TCO
    if index == '1':
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"TCO_Cliente_{accronym}_{mercadoria}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"TCO_Banco_{accronym}_{mercadoria}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"TCO_Lawton_{accronym}_{mercadoria}.txt")
    else:
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"TCO_Cliente_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"TCO_Banco_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"TCO_Lawton_{accronym}_{mercadoria}_#{index}.txt")

    # Primeira linha do arquivo
    primeira_linha = f"TER  0{'0001'}JPMORGANBM{' ' * 10}{data_atual.strftime('%Y%m%d')}00003"
    primeira_linha_lawton = f"TER  0{'0001'}INTRAGLAWTONFDO{' ' * 5}{data_atual.strftime('%Y%m%d')}00003"

    # Gerar arquivos para cliente, banco e lawton
    gerar_arquivo(caminho_arquivo_cliente, primeira_linha, tabela_arquivotermo_cliente, identifiers)
    gerar_arquivo(caminho_arquivo_banco, primeira_linha, tabela_arquivotermo_b2b, identifiers)
    gerar_arquivo_lawton(caminho_arquivo_lawton, primeira_linha_lawton, tabela_arquivotermo_b2b, identifiers)

def gerar_arquivo(caminho_arquivo, primeira_linha, tabela, identifiers):    
    with open(caminho_arquivo, 'w', encoding='utf-8') as arquivo:
        arquivo.write(primeira_linha + '\n')
        for item in tabela.get_children():
            valores_completos = tabela.item(item, 'values')
            # Verifique se o identificador está na lista
            if str(valores_completos[-4]) in identifiers: 
                linha = list(valores_completos[:-4])
                # Certifique-se de que o índice 17 existe e é uma string
                if len(linha) > 17:
                    linha[17] = linha[17].replace(".", "")
                linha_concatenada = ''.join(linha)
                arquivo.write(linha_concatenada + '\n')

def gerar_arquivo_lawton(caminho_arquivo, primeira_linha_lawton, tabela, identifiers):    
    with open(caminho_arquivo, 'w', encoding='utf-8') as arquivo:
        arquivo.write(primeira_linha_lawton + '\n')
        for item in tabela.get_children():
            valores_completos = tabela.item(item, 'values')
            if str(valores_completos[-4]) in identifiers:
                linha = list(valores_completos[:-4])
                if str(valores_completos[1]) == '1':
                    if len(linha) > 17:
                        linha[17] = linha[17].replace(".", "")
                    parte1 = ''.join(linha[0:3])
                    parte2 = ''.join([str(random.randint(0,9)) for _ in range(10)])
                    parte3 = linha[7]
                    parte4 = '1' if linha[5] == '0' else '0'                    
                    parte5 = linha[6]
                    parte6 = linha[4]
                    parte7 = ''.join(linha[8:])
                    linha_concatenada = parte1 + parte2 + parte3 + parte4 + parte5 + parte6 + parte7
                elif str(valores_completos[1]) == '2':
                    linha_concatenada = ''.join(linha)
                arquivo.write(linha_concatenada + '\n')
                
def gerar_arquivos_b3_termo(tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, switch_cliente_arquivotermo):
    resposta = messagebox.askyesno("Generate B3 File", "Wish to proceed?")
    if not resposta:
        return

    # Extract data from tables
    deals_cliente, td_cliente, market_client, type_cliente, instrument_cliente, strike_cliente, intCCY_cliente, tn_cliente, sd_cliente, fxd_cliente, fsd_cliente, fed_cliente, accronym, identifier_cliente, status_cliente, sid_cliente, deals_b2b, td_b2b, market_client, type_b2b, instrument_b2b, strike_b2b, intCCY_b2b, tn_b2b, sd_b2b, fxd_b2b, fsd_b2b, fed_b2b, accronym_b2b, identifier_b2b, status_b2b, sid_b2b = extrair_dados_deals(tabela_termo_cliente, tabela_termo_b2b)
    global entry_filtro_commodities
    entry_filtro_commodities.delete(0, tk.END)
    global tabela_commodities_data
    for item in tabela_commodities_data:            
            tabela_commodities.insert("", "end", values=item)
  
    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"
    for item in tabela_termo_cliente.get_children():
        data_atual_str = tabela_termo_cliente.item(item)["values"][1]  # Adjust as needed

        # Convert the string to a datetime object
        try:
            data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Adjust the format as needed
        except ValueError as e:
            messagebox.showerror("Erro", f"Formato de data inválido: {data_atual_str}")
            continue
        
        mes2 = data_atual.strftime("%m")
        meses_portugues = {
            "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
            "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
            "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
        }
        mes = meses_portugues[mes2]
        ano = data_atual.strftime("%Y")
        dia = data_atual.strftime("%d")

    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    
    if switch_cliente_arquivotermo.get() == "on":
        mercadorias = []
        cnpjs = []
        indexes = []
        SID = getpass.getuser()
        SID = SID[0].upper() + SID[1:]
        identifiers = []
                
        rows = tabela_arquivotermo_cliente.get_children()
        
        for row_id_file in rows:
            row_comm_file = tabela_arquivotermo_cliente.item(row_id_file)["values"]              

            if str(row_comm_file[1]) == '1':             
                mercadoria = lookup(row_comm_file[16].strip(), commodities_ric, commodities_mercadoria)
                cnpj = row_comm_file[8] if row_comm_file[7] == 73760102 else row_comm_file[7]
                index = str(row_comm_file[-3])
                if mercadoria not in mercadorias:
                    mercadorias.append(mercadoria)
                if cnpj not in cnpjs:
                    cnpjs.append(cnpj)            
                if index not in indexes:
                    indexes.append(index)

        for index in indexes:          
            for mercadoria in mercadorias: 
                for cnpj in cnpjs:
                    cnpj_str = str(cnpj).zfill(14) if len(str(cnpj)) != 8 else f"{str(cnpj)[:5]}.{str(cnpj)[5:7]}-{str(cnpj)[7:]}"
                    cliente = lookup(cnpj_str, cntpy_taxid, cntpy_name) if len(str(cnpj)) != 8 else lookup(cnpj_str, cntpy_b3_account, cntpy_name)
                    accronym = lookup(cnpj_str, cntpy_taxid, cntpy_acronym) if len(str(cnpj)) != 8 else lookup(cnpj_str, cntpy_b3_account, cntpy_acronym)                      
                    identifiers_asian = []
                    identifiers_bullet = []

                    for row_id_file in rows:
                        row_file = tabela_arquivotermo_cliente.item(row_id_file)["values"]
                        status = row_file[61]  # Ensure this is the correct index for status                        
                        if status == "Approved" or status == "Generated":
                            if len(row_file) > 57 and (row_file[8] == cnpj or row_file[7] == cnpj) and lookup(row_file[16].strip(), commodities_ric, commodities_mercadoria) == mercadoria and str(row_file[-3]) == index:
                                identifier = str(row_file[-4])                                                                                  
                                if row_file[57] == 1:
                                    identifiers_asian.append(identifier)
                                    identifiers.append(identifier)
                                else:
                                    identifiers_bullet.append(identifier)
                                    identifiers.append(identifier)                                

                    # Generate files for TCO Asian
                    if identifiers_asian:
                        if index == '1':
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "TERMO", f"{mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)
                        else:
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "TERMO", f"#{index} {mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)                        
                        gerar_arquivos_tco_asian(caminho_completo, accronym, tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, identifiers_asian, data_atual, mercadoria, index)

                    # Generate files for TCO
                    if identifiers_bullet:
                        if index == '1':
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "TERMO", f"{mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)
                        else:
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "TERMO", f"#{index} {mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)                        
                        gerar_arquivos_tco(caminho_completo, accronym, tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, identifiers_bullet, data_atual, mercadoria, index)           

        # Call update_status once after all files are generated
        update_status(tabela_termo_cliente, tabela_termo_b2b, tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, identifiers)
    else:
        caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")
        if not os.path.exists(caminho_completo):
            os.makedirs(caminho_completo)
        gerar_arquivos_tco(caminho_completo, "", tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, [], data_atual)

    messagebox.showinfo("Sucesso!", "Arquivos gerados com sucesso!")

# Certifique-se de que todos os identificadores esperados estão sendo coletados
def gerar_arquivos_opt_asian(caminho_completo, accronym, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, identifiers, data_atual, mercadoria, index):
    # Caminhos para os arquivos OPT Asian
    if index == '1':
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"OPTAsian_Cliente_{accronym}_{mercadoria}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"OPTAsian_Banco_{accronym}_{mercadoria}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"OPTAsian_Lawton_{accronym}_{mercadoria}.txt")
    else:
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"OPTAsian_Cliente_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"OPTAsian_Banco_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"OPTAsian_Lawton_{accronym}_{mercadoria}_#{index}.txt")
    
    # Primeira linha do arquivo
    primeira_linha_opt = f"OPCAO0{'0002'}JPMORGANBM{' ' * 10}{data_atual.strftime('%Y%m%d')}{' ' * 666}"
    primeira_linha_lawton_opt = f"OPCAO0{'0002'}INTRAGLAWTONFDO{' ' * 5}{data_atual.strftime('%Y%m%d')}{' ' * 666}"

    # Gerar arquivos para cliente, banco e lawton
    gerar_arquivoopcao(caminho_arquivo_cliente, primeira_linha_opt, tabela_arquivoopcao_cliente, identifiers)
    gerar_arquivoopcao(caminho_arquivo_banco, primeira_linha_opt, tabela_arquivoopcao_b2b, identifiers)
    gerar_arquivoopcao_lawton(caminho_arquivo_lawton, primeira_linha_lawton_opt, tabela_arquivoopcao_b2b, identifiers)

def gerar_arquivos_opt(caminho_completo, accronym, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, identifiers, data_atual, mercadoria, index):
    # Caminhos para os arquivos OPT
    if index == '1':
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"OPT_Cliente_{accronym}_{mercadoria}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"OPT_Banco_{accronym}_{mercadoria}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"OPT_Lawton_{accronym}_{mercadoria}.txt")
    else:
        caminho_arquivo_cliente = os.path.join(caminho_completo, f"OPT_Cliente_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_banco = os.path.join(caminho_completo, f"OPT_Banco_{accronym}_{mercadoria}_#{index}.txt")
        caminho_arquivo_lawton = os.path.join(caminho_completo, f"OPT_Lawton_{accronym}_{mercadoria}_#{index}.txt")

    # Primeira linha do arquivo
    primeira_linha_opt = f"OPCAO0{'0002'}JPMORGANBM{' ' * 10}{data_atual.strftime('%Y%m%d')}{' ' * 666}"
    primeira_linha_lawton_opt = f"OPCAO0{'0002'}INTRAGLAWTONFDO{' ' * 5}{data_atual.strftime('%Y%m%d')}{' ' * 666}"

    # Gerar arquivos para cliente, banco e lawton
    gerar_arquivoopcao(caminho_arquivo_cliente, primeira_linha_opt, tabela_arquivoopcao_cliente, identifiers)
    gerar_arquivoopcao(caminho_arquivo_banco, primeira_linha_opt, tabela_arquivoopcao_b2b, identifiers)
    gerar_arquivoopcao_lawton(caminho_arquivo_lawton, primeira_linha_lawton_opt, tabela_arquivoopcao_b2b, identifiers)

def gerar_arquivoopcao(caminho_arquivo, primeira_linha, tabela, identifiers):
    with open(caminho_arquivo, 'w', encoding='utf-8') as arquivo:
        arquivo.write(primeira_linha + '\n')
        quantidade_operacoes = 0
        for item in tabela.get_children():
            valores_completos = tabela.item(item, 'values')
            # Verifique se o identificador está na lista
            if str(valores_completos[-4]) in identifiers:
                linha = list(valores_completos[:-4])              
                linha_concatenada = ''.join(linha)
                arquivo.write(linha_concatenada + '\n')
                quantidade_operacoes += 1

        # Adicionar a última linha com a quantidade de operações
        quantidade_str = str(quantidade_operacoes)
        zeros_necessarios = 10 - len(quantidade_str)
        ultima_linha = f"OPCAO9{'0' * zeros_necessarios}{quantidade_str}{' ' * 688}"
        arquivo.write(ultima_linha)  # Remova a quebra de linha aqui




def gerar_arquivoopcao_lawton(caminho_arquivo, primeira_linha_lawton, tabela, identifiers):
    with open(caminho_arquivo, 'w', encoding='utf-8') as arquivo:
        arquivo.write(primeira_linha_lawton + '\n')
        quantidade_operacoes = 0
        for item in tabela.get_children():
            valores_completos = tabela.item(item, 'values')
            if str(valores_completos[-4]) in identifiers:
                linha = list(valores_completos[:-4])                               
                parte1 = ''.join(linha[0:5])
                parte2 = linha[6]
                parte3 = linha[5]
                parte4 = 'P1' if linha[7] == 'P2' else 'P2'                                        
                parte5 = ''.join(linha[8:23])
                parte6 = ''.join([str(random.randint(0,9)) for _ in range(10)])
                parte7 = ''.join(linha[24:])
                
                linha_concatenada = parte1 + parte2 + parte3 + parte4 + parte5 + parte6 + parte7             
                arquivo.write(linha_concatenada + '\n')
                quantidade_operacoes += 1
                
        # Adicionar a última linha com a quantidade de operações
        quantidade_str = str(quantidade_operacoes)
        zeros_necessarios = 10 - len(quantidade_str)
        ultima_linha = f"OPCAO9{'0' * zeros_necessarios}{quantidade_str}{' ' * 688}"
        arquivo.write(ultima_linha)  # Remova a quebra de linha aqui
        
def update_status(treeview_deals, treeview_deals_b2b, treeview_file, treeview_file_b2b, identifiers):  
    try:
        # Determina quais Treeviews de fixings usar com base no tipo de treeview_deals
        if treeview_deals == tabela_opcao_cliente:
            treeview_fixings = tabela_fixingsopcao_cliente
            treeview_fixings_b2b = tabela_fixingsopcao_b2b
        else:
            treeview_fixings = tabela_fixingstermo_cliente
            treeview_fixings_b2b = tabela_fixingstermo_b2b

        # Conecta ao banco de dados SQLite
        conn = sqlite3.connect(db_path, timeout=20, isolation_level="EXCLUSIVE")
        cursor = conn.cursor()
        cursor.execute("BEGIN TRANSACTION;")  # Inicia uma transação no banco de dados
        
        # Inicializa listas para armazenar IDs e identificadores
        ids_opcao = []
        ids_termo = []
        identifiers_opcao = []
        identifiers_termo = []        
        identifiers_fixings = []
        DealNames = []
        Markets = []
        Types = []
        Strikes = []
        IntermediatesCCY = []
        TotalNotionals = []
        SettlementDates = []
        FxConvDates = []
        FirstFixingDates = []
        LastFixingDates = []
        SIDS = []    
        Instruments = []
        TradeDates_update = [] 
        Counterparties_update = [] 
        Statuses_update = []
        Status_update = "Generated"
        SID = getpass.getuser()
        SID = SID[0].upper() + SID[1:]

        # Atualiza o status dos itens no Treeview de arquivos para "Generated"
        rows = treeview_file.get_children()    
        for item in rows:
            row_data_file = list(treeview_file.item(item, 'values'))  # Converte a tupla em uma lista
            identifier = str(row_data_file[-4])
            status = row_data_file[-2]
            if status == "Approved" and identifier in identifiers:
                row_data_file[-2] = "Generated"       
                row_data_file[-1] = SID
                treeview_file.item(item, values=row_data_file)  # Atualiza o Treeview

        # Atualiza o status dos itens no Treeview de arquivos B2B para "Generated"
        rows = treeview_file_b2b.get_children()    
        for item in rows:
            row_data_file_b2b = list(treeview_file_b2b.item(item, 'values'))  # Converte a tupla em uma lista
            identifier = str(row_data_file_b2b[-4])
            status = row_data_file_b2b[-2]
            if status == "Approved" and identifier in identifiers:
                row_data_file_b2b[-2] = "Generated"     
                row_data_file_b2b[-1] = SID               
                treeview_file_b2b.item(item, values=row_data_file_b2b)  # Atualiza o Treeview

        # Destaca duplicatas nos Treeviews de arquivos
        highlight_duplicates(treeview_file, 'arquivo')
        highlight_duplicates(treeview_file_b2b, 'arquivo')
        
        # Atualiza o status dos itens no Treeview de fixings para "Generated"
        rows = treeview_fixings.get_children()    
        for item in rows:
            row_data_fixings = list(treeview_fixings.item(item, 'values'))  # Converte a tupla em uma lista
            identifier = str(row_data_fixings[-4])
            status = row_data_fixings[-2]
            if status == "Approved" and identifier in identifiers:
                row_data_fixings[-2] = "Generated"  
                row_data_fixings[-1] = SID                     
                treeview_fixings.item(item, values=row_data_fixings)  # Atualiza o Treeview
                identifiers_fixings.append(identifier)

        # Atualiza o status dos itens no Treeview de fixings B2B para "Generated"
        rows = treeview_fixings_b2b.get_children()    
        for item in rows:
            row_data_fixings_b2b = list(treeview_fixings_b2b.item(item, 'values'))  # Converte a tupla em uma lista
            identifier = str(row_data_fixings_b2b[-4])
            status = row_data_fixings_b2b[-2]
            if status == "Approved" and identifier in identifiers:
                row_data_fixings_b2b[-2] = "Generated"      
                row_data_fixings_b2b[-1] = SID                    
                treeview_fixings_b2b.item(item, values=row_data_fixings_b2b)  # Atualiza o Treeview

        # Destaca duplicatas nos Treeviews de fixings
        highlight_duplicates(treeview_fixings, 'deals')
        highlight_duplicates(treeview_fixings_b2b, 'deals')
                
        # Atualiza o status dos itens no Treeview de transações para "Generated"
        rows = treeview_deals.get_children()    
        for item in rows:
            row_data_deals = list(treeview_deals.item(item, 'values'))  # Converte a tupla em uma lista
            identifier = str(row_data_deals[-4])
            status = row_data_deals[-2]
            id = row_data_deals[0]
            if status == "Approved" and identifier in identifiers:
                row_data_deals[-2] = "Generated"   
                row_data_deals[-1] = SID  
                treeview_deals.item(item, values=row_data_deals)  # Atualiza o Treeview
                Instruments.append(row_data_deals[4])
                TradeDates_update.append(row_data_deals[1])                
                Counterparties_update.append(row_data_deals[18])                
                DealNames.append(row_data_deals[0])                
                Markets.append(row_data_deals[2])
                Types.append(row_data_deals[3])
                Strikes.append(row_data_deals[6])
                IntermediatesCCY.append(row_data_deals[8])
                TotalNotionals.append(row_data_deals[9])
                SettlementDates.append(row_data_deals[10])
                FxConvDates.append(row_data_deals[15])
                FirstFixingDates.append(row_data_deals[16])
                LastFixingDates.append(row_data_deals[17])
                Statuses_update.append(row_data_deals[-2])
                SIDS.append(row_data_deals[-1])
                if treeview_deals == tabela_opcao_cliente:
                    identifiers_opcao.append(identifier)
                    ids_opcao.append(id)
                else:
                    identifiers_termo.append(identifier)
                    ids_termo.append(id)

        # Atualiza o status dos itens no Treeview de transações B2B para "Generated"
        rows = treeview_deals_b2b.get_children()    
        for item in rows:
            row_data_deals_b2b = list(treeview_deals_b2b.item(item, 'values'))  # Converte a tupla em uma lista
            identifier = str(row_data_deals_b2b[-4])
            status = row_data_deals_b2b[-2]
            id = row_data_deals_b2b[0]
            if status == "Approved" and identifier in identifiers:
                row_data_deals_b2b[-2] = "Generated"  
                row_data_deals_b2b[-1] = SID   
                treeview_deals_b2b.item(item, values=row_data_deals_b2b)  # Atualiza o Treeview
                Instruments.append(row_data_deals[4])
                TradeDates_update.append(row_data_deals_b2b[1])                
                Counterparties_update.append(row_data_deals_b2b[18])                
                DealNames.append(row_data_deals_b2b[0])                
                Markets.append(row_data_deals_b2b[2])
                Types.append(row_data_deals_b2b[3])
                Strikes.append(row_data_deals_b2b[6])
                IntermediatesCCY.append(row_data_deals_b2b[8])
                TotalNotionals.append(row_data_deals_b2b[9])
                SettlementDates.append(row_data_deals_b2b[10])
                FxConvDates.append(row_data_deals_b2b[15])
                FirstFixingDates.append(row_data_deals_b2b[16])
                LastFixingDates.append(row_data_deals_b2b[17])
                Statuses_update.append(row_data_deals_b2b[-2])
                SIDS.append(row_data_deals_b2b[-1])
                if treeview_deals == tabela_opcao_cliente:
                    ids_opcao.append(id)
                else:                    
                    ids_termo.append(id)

        # Destaca duplicatas nos Treeviews de transações
        highlight_duplicates(treeview_deals, 'deals')
        highlight_duplicates(treeview_deals_b2b, 'deals')
        
        # Atualiza o banco de dados com os IDs e identificadores processados
        if ids_opcao:           
            placeholders = ", ".join(["?" for _ in range(len(ids_opcao))])
            sql_update = f"UPDATE opcao_base_deals SET Status = 'Generated' WHERE DealName IN ({placeholders})"
            cursor.execute(sql_update, ids_opcao)
            
            # Verifica se há dados para fixings antes de atualizar
            if identifiers_fixings:
                placeholders = ", ".join(["?" for _ in range(len(ids_opcao))])
                sql_update = f"UPDATE opcao_base_fixings SET Status = 'Generated' WHERE AthenaID IN ({placeholders})"
                cursor.execute(sql_update, ids_opcao)
            
            placeholders = ", ".join(["?" for _ in range(len(identifiers_opcao))])
            sql_update = f"UPDATE opcao_base_file SET Status = 'Generated' WHERE Identifier IN ({placeholders})"
            cursor.execute(sql_update, identifiers_opcao)

        if ids_termo:
            placeholders = ", ".join(["?" for _ in range(len(ids_termo))])
            sql_update = f"UPDATE termo_base_deals SET Status = 'Generated' WHERE DealName IN ({placeholders})"
            cursor.execute(sql_update, ids_termo)
            
            # Verifica se há dados para fixings antes de atualizar
            if identifiers_fixings:
                placeholders = ", ".join(["?" for _ in range(len(ids_termo))])
                sql_update = f"UPDATE termo_base_fixings SET Status = 'Generated' WHERE AthenaID IN ({placeholders})"
                cursor.execute(sql_update, ids_termo)
            
            placeholders = ", ".join(["?" for _ in range(len(identifiers_termo))])
            sql_update = f"UPDATE termo_base_file SET Status = 'Generated' WHERE Identifier IN ({placeholders})"
            cursor.execute(sql_update, identifiers_termo)
        
        # Confirma as alterações no banco de dados e fecha a conexão
        conn.commit()
        conn.close()
        status_change_email(DealNames, TradeDates_update, Markets, Types, Instruments, Strikes, IntermediatesCCY, TotalNotionals, SettlementDates, FxConvDates, FirstFixingDates, LastFixingDates, Counterparties_update, Statuses_update, SIDS, Status_update)                
    except sqlite3.Error as e:
        # Trata erros de banco de dados
        print(f"Erro ao acessar o banco de dados: {e}")
        conn.rollback()  # Reverte a transação em caso de erro
        conn.close()

            
def gerar_arquivos_b3_opcao(tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, switch_cliente_arquivoopcao):
    resposta = messagebox.askyesno("Generate B3 File", "Wish to proceed?")
    if not resposta:
        return

    # Extract data from tables
    deals_cliente, td_cliente, market_client, type_cliente, instrument_cliente, strike_cliente, intCCY_cliente, tn_cliente, sd_cliente, fxd_cliente, fsd_cliente, fed_cliente, accronym, identifier_cliente, status_cliente, sid_cliente, deals_b2b, td_b2b, market_client, type_b2b, instrument_b2b, strike_b2b, intCCY_b2b, tn_b2b, sd_b2b, fxd_b2b, fsd_b2b, fed_b2b, accronym_b2b, identifier_b2b, status_b2b, sid_b2b = extrair_dados_deals(tabela_termo_cliente, tabela_termo_b2b)
    global entry_filtro_commodities
    entry_filtro_commodities.delete(0, tk.END)
    global tabela_commodities_data
    for item in tabela_commodities_data:            
            tabela_commodities.insert("", "end", values=item)
    
  
    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Liquidações do Dia"
    for item in tabela_opcao_cliente.get_children():
        data_atual_str = tabela_opcao_cliente.item(item)["values"][1]  # Adjust as needed

        # Convert the string to a datetime object
        try:
            data_atual = datetime.strptime(data_atual_str, "%d-%b-%Y")  # Adjust the format as needed
        except ValueError as e:
            messagebox.showerror("Erro", f"Formato de data inválido: {data_atual_str}")
            continue
        
        mes2 = data_atual.strftime("%m")
        meses_portugues = {
            "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
            "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
            "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
        }
        mes = meses_portugues[mes2]
        ano = data_atual.strftime("%Y")
        dia = data_atual.strftime("%d")

    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_acronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    
    if switch_cliente_arquivoopcao.get() == "on":
        mercadorias = []
        cnpjs = []
        indexes = []
        SID = getpass.getuser()
        SID = SID[0].upper() + SID[1:]
        indentifiers = []
                
        rows = tabela_arquivoopcao_cliente.get_children()
        
        for row_id_file in rows:
            row_comm_file = tabela_arquivoopcao_cliente.item(row_id_file)["values"]              

            if str(row_comm_file[1]) == '1':             
                mercadoria = lookup(row_comm_file[67].strip(), commodities_ric, commodities_mercadoria)
                cnpj = row_comm_file[43] if row_comm_file[6] == "73760.10-2" else row_comm_file[6]
                index = str(row_comm_file[-3])
                if mercadoria not in mercadorias:
                    mercadorias.append(mercadoria)
                if cnpj not in cnpjs:
                    cnpjs.append(cnpj)            
                if index not in indexes:
                    indexes.append(index)

        for index in indexes:          
            for mercadoria in mercadorias: 
                for cnpj in cnpjs:
                    cnpj_str = str(cnpj).zfill(14) if len(str(cnpj)) != 10 else cnpj #f"{str(cnpj)[:5]}.{str(cnpj)[5:7]}-{str(cnpj)[7:]}"                  
                    cliente = lookup(cnpj_str, cntpy_taxid, cntpy_name) if len(str(cnpj)) != 10 else lookup(cnpj, cntpy_b3_account, cntpy_name)
                    accronym = lookup(cnpj_str, cntpy_taxid, cntpy_acronym) if len(str(cnpj)) != 10 else lookup(cnpj, cntpy_b3_account, cntpy_acronym)                    
                    identifiers_asian = []
                    identifiers_bullet = []

                    for row_id_file in rows:
                        row_file = tabela_arquivoopcao_cliente.item(row_id_file)["values"]
                        status = row_file[-2]  # Ensure this is the correct index for status                        
                        if status == "Approved" or status == "Generated":
                            if len(row_file) > 70 and (row_file[43] == cnpj or row_file[6] == cnpj) and lookup(row_file[67].strip(), commodities_ric, commodities_mercadoria) == mercadoria and str(row_file[-3]) == index:
                                identifier = str(row_file[-4])                                                                                  
                                if row_file[70] == 1:
                                    identifiers_asian.append(identifier)
                                    indentifiers.append(identifier)
                                else:
                                    identifiers_bullet.append(identifier)   
                                    indentifiers.append(identifier)                             

                    # Generate files for TCO Asian
                    if identifiers_asian:
                        if index == '1':
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "OPÇÃO", f"{mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)
                        else:
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "OPÇÃO", f"#{index} {mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)                        
                        gerar_arquivos_opt_asian(caminho_completo, accronym, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, identifiers_asian, data_atual, mercadoria, index)
                        

                    # Generate files for TCO
                    if identifiers_bullet:
                        if index == '1':
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "OPÇÃO", f"{mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)
                        else:
                            caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro", f'{cliente}', "OPÇÃO", f"#{index} {mercadoria}")
                            if not os.path.exists(caminho_completo):
                                os.makedirs(caminho_completo)                        
                        gerar_arquivos_opt(caminho_completo, accronym, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, identifiers_bullet, data_atual, mercadoria, index)     
                              
        # Call update_status once after all files are generated
        update_status(tabela_opcao_cliente, tabela_opcao_b2b, tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, indentifiers)
    else:
        caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia, "Registro")
        if not os.path.exists(caminho_completo):
            os.makedirs(caminho_completo)
        gerar_arquivos_opt(caminho_completo, "", tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, [], data_atual, mercadoria)

    messagebox.showinfo("Sucesso!", "Arquivos gerados com sucesso!")

def gerar_boleta_termo(tabela_boletatermo):
    resposta = messagebox.askyesno("Generate Intrag File", "Wish to proceed?")
    if not resposta:
        return
    
    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Itau"
    data_atual = datetime.now()
    mes2 = data_atual.strftime("%m")
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia)

    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)
    caminho_arquivo = os.path.join(caminho_completo, "Fdo_Tes_NDF_CETIP.txt")

    with open(caminho_arquivo, 'w', encoding='utf-8') as arquivo:
        for item in tabela_boletatermo.get_children():
            valores_completos = tabela_boletatermo.item(item, 'values')
            # Concatenar os valores com ';' como separador
            if valores_completos[2] != "Codigo_Cetip":
                linha = list(valores_completos[:-1])
                linha_concatenada = ';'.join(linha)
                arquivo.write(linha_concatenada + '\n')            
            

    messagebox.showinfo("Sucesso!", "Arquivos gerados com sucesso!")

def gerar_boleta_opcao(tabela_boletaopcao):
    resposta = messagebox.askyesno("Generate Intrag File", "Wish to proceed?")    
    if not resposta:
        return
    
    diretorio_raiz = "I:\\Confirmation\\Derivativos\\Movimento\\Itau"
    data_atual = datetime.now()
    mes2 = data_atual.strftime("%m")
    meses_portugues = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }
    mes = meses_portugues[mes2]
    ano = data_atual.strftime("%Y")
    dia = data_atual.strftime("%d")
    caminho_completo = os.path.join(diretorio_raiz, ano, f"{mes2}. {mes}", dia)

    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)
    caminho_arquivo = os.path.join(caminho_completo, "Fdo_Tes_Opções_Moeda.txt")

    with open(caminho_arquivo, 'w', encoding='utf-8') as arquivo:
        for item in tabela_boletaopcao.get_children():
            valores_completos = tabela_boletaopcao.item(item, 'values')
           # Concatenar os valores com ';' como separador
            linha = list(valores_completos[:-1])
            linha_concatenada = ';'.join(linha)
            arquivo.write(linha_concatenada + '\n') 

    
    messagebox.showinfo("Sucesso!", "Arquivos gerados com sucesso!")
    
# Certifique-se de que todos os identificadores esperados estão sendo coletados 
def arquivo_b3(tabview, abas_existentes):
    global sub_notebook_arquivo_b3, sub_sub_notebook_arquivotermo, sub_sub_notebook_arquivoopcao
    tree = []
    # Adicionar a aba "Arquivo B3" ao tabview
    tabview.insert(2, "Arquivo B3")
    aba_arquivo_b3 = tabview.tab("Arquivo B3")

    sub_notebook_arquivo_b3 = ctk.CTkTabview(aba_arquivo_b3)
    sub_notebook_arquivo_b3.pack(expand=True, fill='both')

    # Configuração para a sub-aba "Termo"
    sub_notebook_arquivo_b3.add("Termo")
    aba_arquivotermo = sub_notebook_arquivo_b3.tab("Termo")

    sub_sub_notebook_arquivotermo = ctk.CTkTabview(aba_arquivotermo)
    sub_sub_notebook_arquivotermo.pack(expand=True, fill='both')

    # Configuração para a sub-sub-aba "Cliente" em "Termo"
    sub_sub_notebook_arquivotermo.add("Cliente")
    aba_cliente_arquivtermo = sub_sub_notebook_arquivotermo.tab("Cliente")
    
    frame_query_cliente_arquivotermo = ctk.CTkFrame(aba_cliente_arquivtermo, height=1, fg_color="#D3D3D3")
    frame_query_cliente_arquivotermo.pack(fill='x')   
    
    frame_approval_cliente_arquivotermo = ctk.CTkFrame(aba_cliente_arquivtermo, width=220, fg_color="#D3D3D3")
    frame_approval_cliente_arquivotermo.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_cliente_arquivotermo)


    frame_treeview_cliente_arquivotermo = ctk.CTkFrame(aba_cliente_arquivtermo)
    frame_treeview_cliente_arquivotermo.pack(expand=True, fill='both', side='left')

    frame_scrollbary_cliente_arquivotermo = ctk.CTkFrame(frame_treeview_cliente_arquivotermo, width=4)
    frame_scrollbary_cliente_arquivotermo.pack(fill='y', side='right')

    scrollbar_x_cliente_arquivotermo = ctk.CTkScrollbar(frame_treeview_cliente_arquivotermo, orientation='horizontal')
    scrollbar_y_cliente_arquivotermo = ctk.CTkScrollbar(frame_scrollbary_cliente_arquivotermo, orientation='vertical')

    global tabela_arquivotermo_cliente
    tabela_arquivotermo_cliente = ttk.Treeview(frame_treeview_cliente_arquivotermo, columns=colunas_arquivo_termo, show='headings', xscrollcommand=scrollbar_x_cliente_arquivotermo.set, yscrollcommand=scrollbar_y_cliente_arquivotermo.set)
    tabela_arquivotermo_cliente.pack(expand=True, fill='both')
    scrollbar_x_cliente_arquivotermo.configure(command=tabela_arquivotermo_cliente.xview, height=25)
    scrollbar_y_cliente_arquivotermo.configure(command=tabela_arquivotermo_cliente.yview, width=25)
    
    # Configurar cabeçalhos das colunas
    for coluna in colunas_arquivo_termo:
        tabela_arquivotermo_cliente.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_arquivotermo_cliente, _col, False))
    vincular_evento_duplo_clique_status(tabela_arquivotermo_cliente, colunas_arquivo_termo)
    frame_botoes_cliente_arquivotermo = ctk.CTkFrame(frame_treeview_cliente_arquivotermo, height=150)
    frame_botoes_cliente_arquivotermo.pack(fill='x', side='bottom')

    

    # Configuração para a sub-sub-aba "B2B" em "Termo"
    sub_sub_notebook_arquivotermo.add("B2B")
    aba_b2b_arquivotermo = sub_sub_notebook_arquivotermo.tab("B2B")
    
    frame_query_b2b_arquivotermo = ctk.CTkFrame(aba_b2b_arquivotermo, height=1, fg_color="#D3D3D3")
    frame_query_b2b_arquivotermo.pack(fill='x')  
    
    frame_approval_b2b_arquivotermo = ctk.CTkFrame(aba_b2b_arquivotermo, width=220, fg_color="#D3D3D3")
    frame_approval_b2b_arquivotermo.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_b2b_arquivotermo)

    frame_treeview_b2b_arquivotermo = ctk.CTkFrame(aba_b2b_arquivotermo)
    frame_treeview_b2b_arquivotermo.pack(expand=True, fill='both', side='left')

    frame_scrollbary_b2b_arquivotermo = ctk.CTkFrame(frame_treeview_b2b_arquivotermo, width=4)
    frame_scrollbary_b2b_arquivotermo.pack(fill='y', side='right')

    scrollbar_x_b2b_arquivotermo = ctk.CTkScrollbar(frame_treeview_b2b_arquivotermo, orientation='horizontal')
    scrollbar_y_b2b_arquivotermo = ctk.CTkScrollbar(frame_scrollbary_b2b_arquivotermo, orientation='vertical')

    global tabela_arquivotermo_b2b
    tabela_arquivotermo_b2b = ttk.Treeview(frame_treeview_b2b_arquivotermo, columns=colunas_arquivo_termo, show='headings', xscrollcommand=scrollbar_x_b2b_arquivotermo.set, yscrollcommand=scrollbar_y_b2b_arquivotermo.set)
    tabela_arquivotermo_b2b.pack(expand=True, fill='both')
    scrollbar_x_b2b_arquivotermo.configure(command=tabela_arquivotermo_b2b.xview, height=25)
    scrollbar_y_b2b_arquivotermo.configure(command=tabela_arquivotermo_b2b.yview, width=25)
    

    # Configurar cabeçalhos das colunas
    for coluna in colunas_arquivo_termo:
        tabela_arquivotermo_b2b.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_arquivotermo_b2b, _col, False))
    vincular_evento_duplo_clique_status(tabela_arquivotermo_b2b, colunas_arquivo_termo)
    frame_botoes_b2b_arquivotermo = ctk.CTkFrame(frame_treeview_b2b_arquivotermo, height=150)
    frame_botoes_b2b_arquivotermo.pack(fill='x', side='bottom')

    

    # Configuração para a sub-aba "Opção"
    sub_notebook_arquivo_b3.add("Opção")
    aba_arquivoopcao = sub_notebook_arquivo_b3.tab("Opção")

    sub_sub_notebook_arquivoopcao = ctk.CTkTabview(aba_arquivoopcao)
    sub_sub_notebook_arquivoopcao.pack(expand=True, fill='both')

    # Configuração para a sub-sub-aba "Cliente" em "Opção"
    sub_sub_notebook_arquivoopcao.add("Cliente")
    aba_cliente_arquivopcao = sub_sub_notebook_arquivoopcao.tab("Cliente")
    
    frame_query_cliente_arquivopcao = ctk.CTkFrame(aba_cliente_arquivopcao, height=1, fg_color="#D3D3D3")
    frame_query_cliente_arquivopcao.pack(fill='x') 
    
    frame_approval_cliente_arquivopcao = ctk.CTkFrame(aba_cliente_arquivopcao, width=220, fg_color="#D3D3D3")
    frame_approval_cliente_arquivopcao.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_cliente_arquivopcao)

    frame_treeview_cliente_arquivoopcao = ctk.CTkFrame(aba_cliente_arquivopcao)
    frame_treeview_cliente_arquivoopcao.pack(expand=True, fill='both', side='left')

    frame_scrollbary_cliente_arquivoopcao = ctk.CTkFrame(frame_treeview_cliente_arquivoopcao, width=4)
    frame_scrollbary_cliente_arquivoopcao.pack(fill='y', side='right')

    scrollbar_x_cliente_arquivoopcao = ctk.CTkScrollbar(frame_treeview_cliente_arquivoopcao, orientation='horizontal')
    scrollbar_y_cliente_arquivoopcao = ctk.CTkScrollbar(frame_scrollbary_cliente_arquivoopcao, orientation='vertical')

    global tabela_arquivoopcao_cliente
    tabela_arquivoopcao_cliente = ttk.Treeview(frame_treeview_cliente_arquivoopcao, columns=colunas_arquivo_opcao, show='headings', xscrollcommand=scrollbar_x_cliente_arquivoopcao.set, yscrollcommand=scrollbar_y_cliente_arquivoopcao.set)
    tabela_arquivoopcao_cliente.pack(expand=True, fill='both')
    scrollbar_x_cliente_arquivoopcao.configure(command=tabela_arquivoopcao_cliente.xview, height=25)
    scrollbar_y_cliente_arquivoopcao.configure(command=tabela_arquivoopcao_cliente.yview, width=25)

    # Configurar cabeçalhos das colunas
    for coluna in colunas_arquivo_opcao:
        tabela_arquivoopcao_cliente.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_arquivoopcao_cliente, _col, False))
    vincular_evento_duplo_clique_status(tabela_arquivoopcao_cliente, colunas_arquivo_opcao)
    frame_botoes_cliente_arquivoopcao = ctk.CTkFrame(frame_treeview_cliente_arquivoopcao, height=150)
    frame_botoes_cliente_arquivoopcao.pack(fill='x', side='bottom')

    

    # Configuração para a sub-sub-aba "B2B" em "Opção"
    sub_sub_notebook_arquivoopcao.add("B2B")
    aba_b2b_arquivoopcao = sub_sub_notebook_arquivoopcao.tab("B2B")
    
    frame_query_b2b_arquivopcao = ctk.CTkFrame(aba_b2b_arquivoopcao, height=1, fg_color="#D3D3D3")
    frame_query_b2b_arquivopcao.pack(fill='x')
    
    frame_approval_b2b_arquivopcao = ctk.CTkFrame(aba_b2b_arquivoopcao, width=220, fg_color="#D3D3D3")
    frame_approval_b2b_arquivopcao.pack(expand=True, fill='y', side='left') 

    # Cria a legenda de cores
    create_color_legend(frame_approval_b2b_arquivopcao)

    frame_treeview_b2b_arquivoopcao = ctk.CTkFrame(aba_b2b_arquivoopcao)
    frame_treeview_b2b_arquivoopcao.pack(expand=True, fill='both', side='left')

    frame_scrollbary_b2b_arquivoopcao = ctk.CTkFrame(frame_treeview_b2b_arquivoopcao, width=4)
    frame_scrollbary_b2b_arquivoopcao.pack(fill='y', side='right')

    scrollbar_x_b2b_arquivoopcao = ctk.CTkScrollbar(frame_treeview_b2b_arquivoopcao, orientation='horizontal')
    scrollbar_y_b2b_arquivoopcao = ctk.CTkScrollbar(frame_scrollbary_b2b_arquivoopcao, orientation='vertical')

    global tabela_arquivoopcao_b2b
    tabela_arquivoopcao_b2b = ttk.Treeview(frame_treeview_b2b_arquivoopcao, columns=colunas_arquivo_opcao, show='headings', xscrollcommand=scrollbar_x_b2b_arquivoopcao.set, yscrollcommand=scrollbar_y_b2b_arquivoopcao.set)
    tabela_arquivoopcao_b2b.pack(expand=True, fill='both')
    scrollbar_x_b2b_arquivoopcao.configure(command=tabela_arquivoopcao_b2b.xview, height=25)
    scrollbar_y_b2b_arquivoopcao.configure(command=tabela_arquivoopcao_b2b.yview, width=25)

    # Configurar cabeçalhos das colunas
    for coluna in colunas_arquivo_opcao:
        tabela_arquivoopcao_b2b.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_arquivoopcao_b2b, _col, False))
    vincular_evento_duplo_clique_status(tabela_arquivoopcao_b2b, colunas_arquivo_opcao)
    frame_botoes_b2b_arquivoopcao = ctk.CTkFrame(frame_treeview_b2b_arquivoopcao, height=150)
    frame_botoes_b2b_arquivoopcao.pack(fill='x', side='bottom')

    #Botoes Arquivo B3 Termo Cliente
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")
    #botao_email_cliente_arquivotermo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="E-MAIL", font=fonte_botao, command=lambda: validation_email_termo(tabela_termo_cliente))
    #botao_email_cliente_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_importar_arquivotermo_cliente = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_arquivotermo_cliente.pack(side='left', padx=1, pady=10)

    botao_limpar_cliente_arquivotermo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_arquivotermo())
    botao_limpar_cliente_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_populardatas_arquivotermo_cliente = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_arquivotermo_cliente.pack(side='left', padx=1, pady=10)

    botao_popular_cliente_arquivotermo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_cliente_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_gerar_arquivocliente_termo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="GENERATE B3", font=fonte_botao, command=lambda: gerar_arquivos_b3_termo(tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, switch_cliente_arquivotermo))
    botao_gerar_arquivocliente_termo.pack(side='left', padx=1, pady=10)

    botao_popularboleta_cliente_arquivotermo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_cliente_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_delete_cliente_arquivotermo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_arquivotermo_cliente, []))
    botao_delete_cliente_arquivotermo.pack(side='left', padx=1, pady=10)   
                                                                                                                                                                                                    
    #botao_excel_cliente_arquivotermo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_termo(abas_existentes, tabela_termo_cliente, tabela_termo_b2b, tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, switch_cliente_arquivotermo))
    #botao_excel_cliente_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_confirmation_cliente_arquivotermo = ctk.CTkButton(frame_botoes_cliente_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_cliente_arquivotermo.pack(side='left', padx=1, pady=10)

    switch_var_cliente_arquivotermo = ctk.StringVar(value="on")

    def alterna_switch_cliente_arquivotermo():
        switch_cliente_arquivotermo.configure(text="Per Client" if switch_var_cliente_arquivotermo.get() == "on" else "All")

    switch_cliente_arquivotermo = ctk.CTkSwitch(frame_botoes_cliente_arquivotermo, text="Per Client", variable=switch_var_cliente_arquivotermo, onvalue="on", offvalue="off", command=lambda: alterna_switch_cliente_arquivotermo())
    switch_cliente_arquivotermo.pack(side='left', padx=1, pady=10)

    scrollbar_x_cliente_arquivotermo.pack(side='bottom', fill='x')
    scrollbar_y_cliente_arquivotermo.pack(side='right', fill='y')
     #Botoes Arquivo B3 Termo B2b
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    #botao_email_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="E-MAIL", font=fonte_botao, command=lambda: validation_email_termo(tabela_termo_cliente))
    #botao_email_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_importar_arquivotermo_b2b = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_arquivotermo_b2b.pack(side='left', padx=1, pady=10)

    botao_limpar_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_arquivotermo())
    botao_limpar_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_populardatas_arquivotermo_b2b = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_arquivotermo_b2b.pack(side='left', padx=1, pady=10)

    botao_popular_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_gerar_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="GENERATE B3", font=fonte_botao, command=lambda: gerar_arquivos_b3_termo(tabela_arquivotermo_cliente, tabela_arquivotermo_b2b, switch_b2b_arquivotermo))
    botao_gerar_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_popularboleta_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_delete_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_arquivotermo_b2b, []))
    botao_delete_b2b_arquivotermo.pack(side='left', padx=1, pady=10)       

    #botao_excel_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_termo(abas_existentes, tabela_termo_cliente, tabela_termo_b2b, tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, switch_b2b_arquivotermo))
    #botao_excel_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    botao_confirmation_b2b_arquivotermo = ctk.CTkButton(frame_botoes_b2b_arquivotermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    switch_var_b2b_arquivotermo = ctk.StringVar(value="on")

    def alterna_switch_b2b_arquivotermo():
        switch_b2b_arquivotermo.configure(text="Per Client" if switch_var_b2b_arquivotermo.get() == "on" else "All")

    switch_b2b_arquivotermo = ctk.CTkSwitch(frame_botoes_b2b_arquivotermo, text="Per Client", variable=switch_var_b2b_arquivotermo, onvalue="on", offvalue="off", command=lambda: alterna_switch_b2b_arquivotermo())
    switch_b2b_arquivotermo.pack(side='left', padx=1, pady=10)

    scrollbar_x_b2b_arquivotermo.pack(side='bottom', fill='x')
    scrollbar_y_b2b_arquivotermo.pack(side='right', fill='y')

    #Botoes Arquivo B3 Opcao Cliente
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    #botao_email_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="E-MAIL", font=fonte_botao, command=lambda: validation_email_opcao(tabela_opcao_cliente))
    #botao_email_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_importar_arquivoopcao_cliente = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_arquivoopcao_cliente.pack(side='left', padx=1, pady=10)    

    botao_limpar_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_arquivoopcao())
    botao_limpar_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)
    
    botao_populardatas_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_popular_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_gerar_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="GENERATE B3", font=fonte_botao, command=lambda: gerar_arquivos_b3_opcao(tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, switch_cliente_arquivoopcao))
    botao_gerar_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_popularboleta_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_delete_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_arquivoopcao_cliente, []))
    botao_delete_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)

    #botao_excel_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_opcao(abas_existentes, tabela_opcao_cliente, tabela_opcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b, switch_cliente_arquivoopcao))
    #botao_excel_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)
   
    botao_confirmation_cliente_arquivoopcao = ctk.CTkButton(frame_botoes_cliente_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_opcao(tabview))
    botao_confirmation_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)

    switch_var_cliente_arquivoopcao = ctk.StringVar(value="on")

    def alterna_switch_cliente_arquivoopcao():
        switch_cliente_arquivoopcao.configure(text="Per Client" if switch_var_cliente_arquivoopcao.get() == "on" else "All")

    switch_cliente_arquivoopcao = ctk.CTkSwitch(frame_botoes_cliente_arquivoopcao, text="Per Client", variable=switch_var_cliente_arquivoopcao, onvalue="on", offvalue="off", command=lambda: alterna_switch_cliente_arquivoopcao())
    switch_cliente_arquivoopcao.pack(side='left', padx=1, pady=10)

    scrollbar_x_cliente_arquivoopcao.pack(side='bottom', fill='x')
    scrollbar_y_cliente_arquivoopcao.pack(side='right', fill='y')


    #Botoes Arquivo B3 Opcao B2b
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")

    #botao_email_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="E-MAIL", font=fonte_botao, command=lambda: validation_email_opcao(tabela_opcao_cliente))
    #botao_email_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_importar_arquivoopcao_cliente = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_arquivoopcao_cliente.pack(side='left', padx=1, pady=10)    

    botao_limpar_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_arquivoopcao())
    botao_limpar_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_popular_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_gerar_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="GENERATE B3", font=fonte_botao, command=lambda: gerar_arquivos_b3_opcao(tabela_arquivoopcao_cliente, tabela_arquivoopcao_b2b, switch_b2b_arquivoopcao))
    botao_gerar_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)  

    botao_popularboleta_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_delete_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_arquivoopcao_b2b, []))
    botao_delete_b2b_arquivoopcao.pack(side='left', padx=1, pady=10) 

    #botao_excel_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_opcao(abas_existentes, tabela_opcao_cliente, tabela_opcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b, switch_b2b_arquivoopcao))
    #botao_excel_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_populardatas_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)

    botao_confirmation_b2b_arquivoopcao = ctk.CTkButton(frame_botoes_b2b_arquivoopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)

    switch_var_b2b_arquivoopcao = ctk.StringVar(value="on")

    def alterna_switch_b2b_arquivoopcao():
        switch_b2b_arquivoopcao.configure(text="Per Client" if switch_var_b2b_arquivoopcao.get() == "on" else "All")

    switch_b2b_arquivoopcao = ctk.CTkSwitch(frame_botoes_b2b_arquivoopcao, text="Per Client", variable=switch_var_b2b_arquivoopcao, onvalue="on", offvalue="off", command=lambda: alterna_switch_b2b_arquivoopcao())
    switch_b2b_arquivoopcao.pack(side='left', padx=1, pady=10)

    scrollbar_x_b2b_arquivoopcao.pack(side='bottom', fill='x')
    scrollbar_y_b2b_arquivoopcao.pack(side='right', fill='y')  
    
    treeviews = {
    'tabela_termo_cliente': tabela_termo_cliente,
    'tabela_termo_b2b': tabela_termo_b2b,
    'tabela_arquivotermo_cliente': tabela_arquivotermo_cliente,
    'tabela_arquivotermo_b2b': tabela_arquivotermo_b2b,
    'tabela_fixingstermo_cliente': tabela_fixingstermo_cliente,
    'tabela_fixingstermo_b2b': tabela_fixingstermo_b2b,
    'tabela_opcao_cliente': tabela_opcao_cliente,
    'tabela_opcao_b2b': tabela_opcao_b2b,
    'tabela_arquivoopcao_cliente': tabela_arquivoopcao_cliente,
    'tabela_arquivoopcao_b2b': tabela_arquivoopcao_b2b,
    'tabela_fixingsopcao_cliente': tabela_fixingsopcao_cliente,
    'tabela_fixingsopcao_b2b': tabela_fixingsopcao_b2b
    }  
    
    # Query Termo Cliente
    botao_query_arquivotermo_cliente = tk.Button(frame_query_cliente_arquivotermo, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_query_window)
    botao_query_arquivotermo_cliente.pack(side='left', padx=2, pady=0)  
    global entry_query_arquivotermo_cliente
    entry_query_arquivotermo_cliente = ctk.CTkEntry(frame_query_cliente_arquivotermo, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_arquivotermo_cliente.pack(side='left', padx=2, pady=0)  
    botao_load_arquivotermo_cliente = tk.Button(frame_query_cliente_arquivotermo, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry(entry_query_arquivotermo_cliente, treeviews))                               
    botao_load_arquivotermo_cliente.pack(side='left', padx=2, pady=0)  
    # Query arquivotermo B2B
    botao_query_arquivotermo_b2b = tk.Button(frame_query_b2b_arquivotermo, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_query_window)
    botao_query_arquivotermo_b2b.pack(side='left', padx=2, pady=0)  
    global entry_query_arquivotermo_b2b
    entry_query_arquivotermo_b2b = ctk.CTkEntry(frame_query_b2b_arquivotermo, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_arquivotermo_b2b.pack(side='left', padx=2, pady=0)  
    botao_load_arquivotermo_b2b = tk.Button(frame_query_b2b_arquivotermo, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry(entry_query_arquivotermo_b2b, treeviews))                               
    botao_load_arquivotermo_b2b.pack(side='left', padx=2, pady=0)  
    # Query Opção Cliente
    botao_query_arquivoopcao_cliente = tk.Button(frame_query_cliente_arquivopcao, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_option_query_window)
    botao_query_arquivoopcao_cliente.pack(side='left', padx=2, pady=0)  
    global entry_query_arquivoopcao_cliente
    entry_query_arquivoopcao_cliente = ctk.CTkEntry(frame_query_cliente_arquivopcao, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_arquivoopcao_cliente.pack(side='left', padx=2, pady=0)  
    botao_load_arquivoopcao_cliente = tk.Button(frame_query_cliente_arquivopcao, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry_option(entry_query_arquivoopcao_cliente, treeviews))                               
    botao_load_arquivoopcao_cliente.pack(side='left', padx=2, pady=0)  
    # Query Opção B2B
    botao_query_arquivoopcao_b2b = tk.Button(frame_query_b2b_arquivopcao, height=1, width=8, text="QUERY", relief='raised', font=fonte_botao, command= create_deal_option_query_window)
    botao_query_arquivoopcao_b2b.pack(side='left', padx=2, pady=0)  
    global entry_query_arquivoopcao_b2b
    entry_query_arquivoopcao_b2b = ctk.CTkEntry(frame_query_b2b_arquivopcao, width=220, height=7, corner_radius=1, border_width=1, border_color='black')
    entry_query_arquivoopcao_b2b.pack(side='left', padx=2, pady=0)  
    botao_load_arquivoopcao_b2b = tk.Button(frame_query_b2b_arquivopcao, height=1, width=8, text="LOAD", relief='raised', font=fonte_botao, command= lambda: load_query_entry_option(entry_query_arquivoopcao_b2b, treeviews))                               
    botao_load_arquivoopcao_b2b.pack(side='left', padx=2, pady=0)  
    
    

def intrag(tabview, abas_existentes):
    global sub_notebook_boleta_dinamica
    if "Arquivo B3" not in abas_existentes:      
        tabview.insert(2, "Boleta Dinâmica - Intrag")
    else:
        tabview.insert(3, "Boleta Dinâmica - Intrag")
            
    # Adicionar a aba "Boleta Dinâmica - Intrag" ao tabview    
    aba_boleta_dinamica = tabview.tab("Boleta Dinâmica - Intrag")

    sub_notebook_boleta_dinamica = ctk.CTkTabview(aba_boleta_dinamica)
    sub_notebook_boleta_dinamica.pack(expand=True, fill='both')

    # Configuração para a sub-aba "Termo"
    sub_notebook_boleta_dinamica.add("Termo")
    aba_boletatermo = sub_notebook_boleta_dinamica.tab("Termo") 

    frame_treeview_boletatermo = ctk.CTkFrame(aba_boletatermo)
    frame_treeview_boletatermo.pack(expand=True, fill='both', side='left')

    frame_scrollbary_boletatermo = ctk.CTkFrame(frame_treeview_boletatermo, width=4)
    frame_scrollbary_boletatermo.pack(fill='y', side='right')

    scrollbar_x_boletatermo = ctk.CTkScrollbar(frame_treeview_boletatermo, orientation='horizontal')
    scrollbar_y_boletatermo = ctk.CTkScrollbar(frame_scrollbary_boletatermo, orientation='vertical')

    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    global tabela_boletatermo
    tabela_boletatermo = ttk.Treeview(frame_treeview_boletatermo, columns=colunas_boleta_termo, show='headings', xscrollcommand=scrollbar_x_boletatermo.set, yscrollcommand=scrollbar_y_boletatermo.set)
    tabela_boletatermo.pack(expand=True, fill='both')
    scrollbar_x_boletatermo.configure(command=tabela_boletatermo.xview, height=25)
    scrollbar_y_boletatermo.configure(command=tabela_boletatermo.yview, width=25)

      # Configurar cabeçalhos das colunas
    for coluna in colunas_boleta_termo:
        tabela_boletatermo.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_boletatermo, _col, False))

    frame_botoes_boletatermo = ctk.CTkFrame(frame_treeview_boletatermo, height=150)
    frame_botoes_boletatermo.pack(fill='x', side='bottom')
    

    # Configuração para a sub-aba "Opção"
    sub_notebook_boleta_dinamica.add("Opção")
    aba_boletaopcao = sub_notebook_boleta_dinamica.tab("Opção")

    frame_treeview_boletaopcao = ctk.CTkFrame(aba_boletaopcao)
    frame_treeview_boletaopcao.pack(expand=True, fill='both', side='left')

    frame_scrollbary_boletaopcao = ctk.CTkFrame(frame_treeview_boletaopcao, width=4)
    frame_scrollbary_boletaopcao.pack(fill='y', side='right')

    scrollbar_x_boletaopcao = ctk.CTkScrollbar(frame_treeview_boletaopcao, orientation='horizontal')
    scrollbar_y_boletaopcao = ctk.CTkScrollbar(frame_scrollbary_boletaopcao, orientation='vertical')

    commodities_ric, commodities_factor, commodities_exchange, commodities_mercadoria, commodities_type, commodities_MM, commodities_YYYY, commodities_unity, commodities_status, cntpy_taxid, cntpy_accronym, cntpy_name, cntpy_b3_account, cntpy_cgd, cntpy_confirmacoes, cntpy_bank, cntpy_cc, cntpy_ag, cntpy_status = extrair_dados_tabelas()
    global tabela_boletaopcao
    tabela_boletaopcao = ttk.Treeview(frame_treeview_boletaopcao, columns=colunas_boleta_opcao, show='headings', xscrollcommand=scrollbar_x_boletaopcao.set, yscrollcommand=scrollbar_y_boletaopcao.set)
    tabela_boletaopcao.pack(expand=True, fill='both')
    scrollbar_x_boletaopcao.configure(command=tabela_boletaopcao.xview, height=25)
    scrollbar_y_boletaopcao.configure(command=tabela_boletaopcao.yview, width=25)

      # Configurar cabeçalhos das colunas
    for coluna in colunas_boleta_opcao:
        tabela_boletaopcao.heading(coluna, text=coluna, command=lambda _col=coluna: ordenar_por(tabela_boletaopcao, _col, False))

    frame_botoes_boletaopcao = ctk.CTkFrame(frame_treeview_boletaopcao, height=150)
    frame_botoes_boletaopcao.pack(fill='x', side='bottom')

    #Botoes Boleta Termo 
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")
    botao_mapping_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="MAPPING DEALS", font=fonte_botao, command=lambda: mapping_deals(tabela_termo_b2b, tabela_boletatermo))
    botao_mapping_boletatermo.pack(side='left', padx=1, pady=10)

    botao_importar_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_boletatermo.pack(side='left', padx=1, pady=10)

    botao_limpar_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_boletatermo())
    botao_limpar_boletatermo.pack(side='left', padx=1, pady=10)

    botao_popular_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_boletatermo.pack(side='left', padx=1, pady=10)

    botao_delete_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_boletatermo, []))
    botao_delete_boletatermo.pack(side='left', padx=1, pady=10) 

    botao_populardatas_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_boletatermo.pack(side='left', padx=1, pady=10)

    botao_popularboleta_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_boletatermo.pack(side='left', padx=1, pady=10)

    botao_gerar_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="GENERATE INTRAG", font=fonte_botao, command=lambda: gerar_boleta_termo(tabela_boletatermo))
    botao_gerar_boletatermo.pack(side='left', padx=1, pady=10)    

    botao_excel_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_termo(abas_existentes, tabela_termo_cliente, tabela_termo_b2b, tabela_fixingstermo_cliente, tabela_fixingstermo_b2b, switch_boletatermo))
    botao_excel_boletatermo.pack(side='left', padx=1, pady=10)

    botao_confirmation_boletatermo = ctk.CTkButton(frame_botoes_boletatermo, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_termo(tabview))
    botao_confirmation_boletatermo.pack(side='left', padx=1, pady=10)

    switch_var_boletatermo = ctk.StringVar(value="on")

    def alterna_switch_boletatermo():
        switch_boletatermo.configure(text="Per Client" if switch_var_boletatermo.get() == "on" else "All")

    switch_boletatermo = ctk.CTkSwitch(frame_botoes_boletatermo, text="Per Client", variable=switch_var_boletatermo, onvalue="on", offvalue="off", command=lambda: alterna_switch_boletatermo())
    switch_boletatermo.pack(side='left', padx=1, pady=10)

    scrollbar_x_boletatermo.pack(side='bottom', fill='x')
    scrollbar_y_boletatermo.pack(side='right', fill='y')

    #Botoes Boleta Opção 
    fonte_botao = ctk.CTkFont(family="League Spartan", size=12, weight="bold")
    botao_email_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="E-MAIL", font=fonte_botao, command=lambda: validation_email_opcao(tabela_opcao_cliente))
    botao_email_boletaopcao.pack(side='left', padx=1, pady=10)

    botao_importar_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="IMPORT DEALS", font=fonte_botao, command=lambda: importar_operacoes(tabview, abas_existentes, tree))
    botao_importar_boletaopcao.pack(side='left', padx=1, pady=10)

    botao_limpar_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CLEAR", font=fonte_botao, command=lambda:limpar_dados_boletaopcao())
    botao_limpar_boletaopcao.pack(side='left', padx=1, pady=10)

    botao_popular_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE B3", font=fonte_botao, command=lambda: popular_arquivos_b3(tabview, abas_existentes, tabela_anbima=None))
    botao_popular_boletaopcao.pack(side='left', padx=1, pady=10)

    botao_delete_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="DELETE", command=lambda: delete_selected_item(tabela_boletaopcao, []))
    botao_delete_boletaopcao.pack(side='left', padx=1, pady=10) 

    botao_populardatas_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="FIXING DATES", font=fonte_botao, command=lambda: chamar_preencher_fixings(tabview, abas_existentes))
    botao_populardatas_boletaopcao.pack(side='left', padx=1, pady=10)

    botao_popularboleta_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="POPULATE INTRAG", font=fonte_botao, command=lambda: popular_boleta(tabview,abas_existentes, tabela_anbima=None))
    botao_popularboleta_boletaopcao.pack(side='left', padx=1, pady=10)

    botao_gerar_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="GENERATE INTRAG", font=fonte_botao, command=lambda: gerar_boleta_opcao(tabela_boletaopcao))
    botao_gerar_boletaopcao.pack(side='left', padx=1, pady=10)    

    botao_excel_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="EXCEL", font=fonte_botao, command=lambda: export_to_excel_opcao(abas_existentes, tabela_opcao_cliente, tabela_opcao_b2b, tabela_fixingsopcao_cliente, tabela_fixingsopcao_b2b, switch_boletaopcao))
    botao_excel_boletaopcao.pack(side='left', padx=1, pady=10)

    botao_confirmation_boletaopcao = ctk.CTkButton(frame_botoes_boletaopcao, width=122, height=26, corner_radius=8, fg_color="#5A5368", text="CONFIRMATION", font=fonte_botao, command=lambda: generate_confirmation_opcao(tabview))
    botao_confirmation_boletaopcao.pack(side='left', padx=1)

    switch_var_boletaopcao = ctk.StringVar(value="on")

    def alterna_switch_boletaopcao():
        switch_boletaopcao.configure(text="Per Client" if switch_var_boletaopcao.get() == "on" else "All")

    switch_boletaopcao = ctk.CTkSwitch(frame_botoes_boletaopcao, text="Per Client", variable=switch_var_boletaopcao, onvalue="on", offvalue="off", command=lambda: alterna_switch_boletaopcao())
    switch_boletaopcao.pack(side='left', padx=1)

    scrollbar_x_boletaopcao.pack(side='bottom', fill='x')
    scrollbar_y_boletaopcao.pack(side='right', fill='y')
    
    
    
    

def ajustar_operacoes_opcao():
    xlookup_dict = {
        "JAN": "F", "FEB": "G", "MAR": "H", "APR": "J", "MAY": "K",
        "JUN": "M", "JUL": "N", "AUG": "Q", "SEP": "U", "OCT": "V",
        "NOV": "X", "DEC": "Z"
    }

    def XLookup(value):
        return xlookup_dict.get(value.upper(), "")

    for item in tabela_opcao_cliente.get_children():
        values = tabela_opcao_cliente.item(item, 'values')
        market_value = values[2]
        contract_value = values[5]
        if market_value == "MPB_LME":
            tabela_opcao_cliente.set(item, "Market", "LOPBDY")
        elif market_value == "MCU_LME":
            tabela_opcao_cliente.set(item, "Market", "LOCADY")
        elif market_value == "MAL_LME":
            tabela_opcao_cliente.set(item, "Market", "LOAHDY")
        elif market_value == "HU_RBOB_NYMEX":
            tabela_opcao_cliente.set(item, "Market", "XB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "HO_NYMEX":
            tabela_opcao_cliente.set(item, "Market", "HO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MZN_LME":
            tabela_opcao_cliente.set(item, "Market", "LOZSDY")
        elif market_value == "MSN_LME":
            tabela_opcao_cliente.set(item, "Market", "LOSNDY")
        elif market_value == "FO_0.5%_SING_FOB":
            tabela_opcao_cliente.set(item, "Market", "NACX0005")
        elif market_value == "SB_ICE":
            tabela_opcao_cliente.set(item, "Market", "SB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "FCPO_BURSA_MYR":
            tabela_opcao_cliente.set(item, "Market", ".KO" + XLookup(contract_value[:3]) + contract_value[-1] + "BNMK F")
        elif market_value == "C_CBOT":
            tabela_opcao_cliente.set(item, "Market", "C " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "S_CBOT":
            tabela_opcao_cliente.set(item, "Market", "S " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BO_CBOT":
            tabela_opcao_cliente.set(item, "Market", "BO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "CC_ICE":
            tabela_opcao_cliente.set(item, "Market", "CC" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "W_CBOT":
            tabela_opcao_cliente.set(item, "Market", "W " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BRT_IPE":
            tabela_opcao_cliente.set(item, "Market", "CO" + XLookup(contract_value[:3]) + contract_value[-1]) if values[16] == values[17] else tabela_opcao_cliente.set(item, "Market", "CO1-2")  
        elif market_value == "SM_CBOT":
            tabela_opcao_cliente.set(item, "Market", "SM" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MAL_MW_PREMIUM":
            tabela_opcao_cliente.set(item, "Market", "PMMUAKE0")
        elif market_value == "BRT_DTD":
            tabela_opcao_cliente.set(item, "Market", "PCRUDTB1")
        elif market_value == "CT_ICE":
            tabela_opcao_cliente.set(item, "Market", "CT" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "KC_ICE":
            tabela_opcao_cliente.set(item, "Market", "KC" + XLookup(contract_value[:3]) + contract_value[-1])

    for item in tabela_opcao_b2b.get_children():
        values = tabela_opcao_b2b.item(item, 'values')
        market_value = values[2]
        contract_value = values[5]
        if market_value == "MPB_LME":
            tabela_opcao_b2b.set(item, "Market", "LOPBDY")
        elif market_value == "MCU_LME":
            tabela_opcao_b2b.set(item, "Market", "LOCADY")
        elif market_value == "MAL_LME":
            tabela_opcao_b2b.set(item, "Market", "LOAHDY")
        elif market_value == "HU_RBOB_NYMEX":
            tabela_opcao_b2b.set(item, "Market", "XB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "HO_NYMEX":
            tabela_opcao_b2b.set(item, "Market", "HO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MZN_LME":
            tabela_opcao_b2b.set(item, "Market", "LOZSDY")
        elif market_value == "MSN_LME":
            tabela_opcao_b2b.set(item, "Market", "LOSNDY")
        elif market_value == "FO_0.5%_SING_FOB":
            tabela_opcao_b2b.set(item, "Market", "NACX0005")
        elif market_value == "SB_ICE":
            tabela_opcao_b2b.set(item, "Market", "SB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "FCPO_BURSA_MYR":
            tabela_opcao_b2b.set(item, "Market", ".KO" + XLookup(contract_value[:3]) + contract_value[-1] + "BNMK F")
        elif market_value == "C_CBOT":
            tabela_opcao_b2b.set(item, "Market", "C " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "S_CBOT":
            tabela_opcao_b2b.set(item, "Market", "S " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BO_CBOT":
            tabela_opcao_b2b.set(item, "Market", "BO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "CC_ICE":
            tabela_opcao_b2b.set(item, "Market", "CC" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "W_CBOT":
            tabela_opcao_b2b.set(item, "Market", "W " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BRT_IPE":
            tabela_opcao_b2b.set(item, "Market", "CO" + XLookup(contract_value[:3]) + contract_value[-1]) if values[16] == values[17] else tabela_opcao_b2b.set(item, "Market", "CO1-2")             
        elif market_value == "SM_CBOT":
            tabela_opcao_b2b.set(item, "Market", "SM" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MAL_MW_PREMIUM":
            tabela_opcao_b2b.set(item, "Market", "PMMUAKE0")
        elif market_value == "BRT_DTD":
            tabela_opcao_b2b.set(item, "Market", "PCRUDTB1")
        elif market_value == "CT_ICE":
            tabela_opcao_b2b.set(item, "Market", "CT" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "KC_ICE":
            tabela_opcao_b2b.set(item, "Market", "KC" + XLookup(contract_value[:3]) + contract_value[-1])

def ajustar_operacoes_termo():
    # Dicionário para mapear os meses para os caracteres do contrato
    xlookup_dict = {
        "JAN": "F", "FEB": "G", "MAR": "H", "APR": "J", "MAY": "K",
        "JUN": "M", "JUL": "N", "AUG": "Q", "SEP": "U", "OCT": "V",
        "NOV": "X", "DEC": "Z"
    }

    # Função para simular o XLookup
    def XLookup(value):
        result = xlookup_dict.get(value.upper(), "")
        return result

    # Iterar sobre as linhas da tabela_termo_cliente
    for item in tabela_termo_cliente.get_children():
        values = tabela_termo_cliente.item(item, 'values')
        market_value = values[2]  # Supondo que "Market" é a terceira coluna
        contract_value = values[5]  # Supondo que "Contract" é a sexta coluna
        if market_value == "MPB_LME":
            tabela_termo_cliente.set(item, "Market", "LOPBDY")
        elif market_value == "MCU_LME":
            tabela_termo_cliente.set(item, "Market", "LOCADY")
        elif market_value == "MAL_LME":
            tabela_termo_cliente.set(item, "Market", "LOAHDY")
        elif market_value == "HU_RBOB_NYMEX":
            tabela_termo_cliente.set(item, "Market", "XB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "HO_NYMEX":
            tabela_termo_cliente.set(item, "Market", "HO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MZN_LME":
            tabela_termo_cliente.set(item, "Market", "LOZSDY")
        elif market_value == "MSN_LME":
            tabela_termo_cliente.set(item, "Market", "LOSNDY")
        elif market_value == "FO_0.5%_SING_FOB":
            tabela_termo_cliente.set(item, "Market", "NACX0005")
        elif market_value == "SB_ICE":
            tabela_termo_cliente.set(item, "Market", "SB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "FCPO_BURSA_MYR":
            tabela_termo_cliente.set(item, "Market", "KO" + XLookup(contract_value[:3]) + contract_value[-1] + "BNMK")
        elif market_value == "C_CBOT":
            tabela_termo_cliente.set(item, "Market", "C " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "S_CBOT":
            tabela_termo_cliente.set(item, "Market", "S " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BO_CBOT":
            tabela_termo_cliente.set(item, "Market", "BO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "CC_ICE":
            tabela_termo_cliente.set(item, "Market", "CC" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "W_CBOT":
            tabela_termo_cliente.set(item, "Market", "W " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BRT_IPE":
            tabela_termo_cliente.set(item, "Market", "CO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "SM_CBOT":
            tabela_termo_cliente.set(item, "Market", "SM" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MAL_MW_PREMIUM":
            tabela_termo_cliente.set(item, "Market", "PMMUAKE0")
        elif market_value == "BRT_DTD":
            tabela_termo_cliente.set(item, "Market", "PCRUDTB1")
        elif market_value == "CT_ICE":
            tabela_termo_cliente.set(item, "Market", "CT" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "KC_ICE":
            tabela_termo_cliente.set(item, "Market", "KC" + XLookup(contract_value[:3]) + contract_value[-1])

    # Repetir o mesmo processo para tabela_termo_b2b
    for item in tabela_termo_b2b.get_children():
        values = tabela_termo_b2b.item(item, 'values')
        market_value = values[2]  # Supondo que "Market" é a terceira coluna
        contract_value = values[5]  # Supondo que "Contract" é a sexta coluna
    
        if market_value == "MPB_LME":
            tabela_termo_b2b.set(item, "Market", "LOPBDY")
        elif market_value == "MCU_LME":
            tabela_termo_b2b.set(item, "Market", "LOCADY")
        elif market_value == "MAL_LME":
            tabela_termo_b2b.set(item, "Market", "LOAHDY")
        elif market_value == "HU_RBOB_NYMEX":
            tabela_termo_b2b.set(item, "Market", "XB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "HO_NYMEX":
            tabela_termo_b2b.set(item, "Market", "HO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MZN_LME":
            tabela_termo_b2b.set(item, "Market", "LOZSDY")
        elif market_value == "MSN_LME":
            tabela_termo_b2b.set(item, "Market", "LOSNDY")
        elif market_value == "FO_0.5%_SING_FOB":
            tabela_termo_b2b.set(item, "Market", "NACX0005")
        elif market_value == "SB_ICE":
            tabela_termo_b2b.set(item, "Market", "SB" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "FCPO_BURSA_MYR":
            tabela_termo_b2b.set(item, "Market", "KO" + XLookup(contract_value[:3]) + contract_value[-1] + "BNMK")
        elif market_value == "C_CBOT":
            tabela_termo_b2b.set(item, "Market", "C " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "S_CBOT":
            tabela_termo_b2b.set(item, "Market", "S " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BO_CBOT":
            tabela_termo_b2b.set(item, "Market", "BO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "CC_ICE":
            tabela_termo_b2b.set(item, "Market", "CC" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "W_CBOT":
            tabela_termo_b2b.set(item, "Market", "W " + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "BRT_IPE":
            tabela_termo_b2b.set(item, "Market", "CO" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "SM_CBOT":
            tabela_termo_b2b.set(item, "Market", "SM" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "MAL_MW_PREMIUM":
            tabela_termo_b2b.set(item, "Market", "PMMUAKE0")
        elif market_value == "BRT_DTD":
            tabela_termo_b2b.set(item, "Market", "PCRUDTB1")
        elif market_value == "CT_ICE":
            tabela_termo_b2b.set(item, "Market", "CT" + XLookup(contract_value[:3]) + contract_value[-1])
        elif market_value == "KC_ICE":
            tabela_termo_b2b.set(item, "Market", "KC" + XLookup(contract_value[:3]) + contract_value[-1]) 

def limpar_dados_fixingstermo():
    for item in tabela_fixingstermo_cliente.get_children():
        tabela_fixingstermo_cliente.delete(item)
    for item in tabela_fixingstermo_b2b.get_children():
        tabela_fixingstermo_b2b.delete(item)
    

def limpar_dados_boletatermo():
    for item in tabela_boletatermo.get_children():
        tabela_boletatermo.delete(item)

def limpar_dados_boletaopcao():
    for item in tabela_boletaopcao.get_children():
        tabela_boletaopcao.delete(item) 

def limpar_dados_fixingsopcao():
    for item in tabela_fixingsopcao_cliente.get_children():
        tabela_fixingsopcao_cliente.delete(item)
    for item in tabela_fixingsopcao_b2b.get_children():
        tabela_fixingsopcao_b2b.delete(item)

def limpar_dados_arquivotermo():
    for item in tabela_arquivotermo_cliente.get_children():
        tabela_arquivotermo_cliente.delete(item)
    for item in tabela_arquivotermo_b2b.get_children():
        tabela_arquivotermo_b2b.delete(item)
    
    messagebox.showinfo("Atualização", "Dados do Arquivo de Termo limpos!")
    

def limpar_dados_arquivoopcao():
    for item in tabela_arquivoopcao_cliente.get_children():
        tabela_arquivoopcao_cliente.delete(item)
    for item in tabela_arquivoopcao_b2b.get_children():
        tabela_arquivoopcao_b2b.delete(item)
    messagebox.showinfo("Atualização", "Dados do Arquivo de Opção limpos!")

def limpar_dados_termo(label_qty_deals):
    for item in tabela_termo_cliente.get_children():
        tabela_termo_cliente.delete(item)
    for item in tabela_termo_b2b.get_children():
        tabela_termo_b2b.delete(item)
    for item in tabela_fixingstermo_cliente.get_children():
        tabela_fixingstermo_cliente.delete(item)
    for item in tabela_fixingstermo_b2b.get_children():
        tabela_fixingstermo_b2b.delete(item)
    # Atualiza o label com a quantidade de deals                        
        qty_deals = number_of_deals(tabela_termo_cliente)
        label_qty_deals.configure(text=str(qty_deals))
    messagebox.showinfo("Atualização", "Dados de Termo limpos!")

def limpar_dados_opcao(label_qty_deals):
    for item in tabela_opcao_cliente.get_children():
        tabela_opcao_cliente.delete(item)
    for item in tabela_opcao_b2b.get_children():
        tabela_opcao_b2b.delete(item)
    for item in tabela_fixingsopcao_cliente.get_children():
        tabela_fixingsopcao_cliente.delete(item)
    for item in tabela_fixingsopcao_b2b.get_children():
        tabela_fixingsopcao_b2b.delete(item)
    messagebox.showinfo("Atualização", "Dados de Opção limpos!")
import tkinter as tk

def on_double_click_status(event, tree, colunas):
    # Get the region where the double-click occurred
    region = tree.identify("region", event.x, event.y)
    if region == "cell":
        # Get the column index
        column_index = int(tree.identify_column(event.x)[1:]) - 1  # Convert from '#n' to index
        # Check if the column is not one of the last three
        if column_index < len(colunas) - 2: #inserir - 2 (para nao deixar Status e SID serem alterados manualmente)
            # Get the item ID and column name
            row_id = tree.identify_row(event.y)
            column_id = tree.identify_column(event.x)
            col_idx = int(column_id.replace("#", "")) - 1
            if col_idx < 0 or col_idx >= len(colunas):
                return
            if not row_id:
                return
            current_values = tree.item(row_id, "values")
            if col_idx >= len(current_values):
                return
            current_value = current_values[col_idx] if current_values[col_idx] is not None else ""

            # Check the value of the second-to-last column
            if current_values[-2] == "New":
                # Create Entry to edit the cell
                entry = tk.Entry(tree)
                entry.insert(0, current_value)
                bbox = tree.bbox(row_id, column_id)
                if bbox:
                    entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])

                def save_entry(event=None):
                    new_value = entry.get()
                    current_values = list(tree.item(row_id, "values"))
                    current_values[col_idx] = new_value
                    tree.item(row_id, values=current_values)
                    entry.destroy()

                entry.bind("<Return>", save_entry)
                entry.bind("<FocusOut>", save_entry)  # Save on focus out
                entry.focus_set()
            else:
                # Call the approve_window_status function
                approve_window_status()

def vincular_evento_duplo_clique_status(tree, colunas):
    # Bind the double-click event to the Treeview
    tree.bind("<Double-1>", lambda event: on_double_click_status(event, tree, colunas))
    
def on_double_click(event, treeview, colunas):
    region = treeview.identify("region", event.x, event.y)
    if region == "cell":
        row_id = treeview.identify_row(event.y)
        column_id = treeview.identify_column(event.x)
        col_idx = int(column_id.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(colunas):
            return
        if not row_id:
            return
        current_values = treeview.item(row_id, "values")
        if col_idx >= len(current_values):
            return
        current_value = current_values[col_idx] if current_values[col_idx] is not None else ""

        # Criar Entry para editar a célula
        entry = tk.Entry(treeview)
        entry.insert(0, current_value)
        bbox = treeview.bbox(row_id, column_id)
        if bbox:
            entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])

        def save_entry(event=None):
            new_value = entry.get()
            current_values = list(treeview.item(row_id, "values"))
            current_values[col_idx] = new_value
            treeview.item(row_id, values=current_values)
            entry.destroy()

        entry.bind("<Return>", save_entry)
        entry.bind("<FocusOut>", save_entry)  # Salva ao perder o foco
        entry.focus_set()    

def vincular_evento_duplo_clique_base(tree, colunas):
    # Bind the double-click event to the Treeview
    tree.bind("<Double-1>", lambda event: on_double_click_base(event, tree, colunas))

def on_double_click_base(event, tree, colunas):
    # Get the region where the double-click occurred
    region = tree.identify("region", event.x, event.y)
    if region == "cell":
        # Get the column index
        column_index = int(tree.identify_column(event.x)[1:]) - 1  # Convert from '#n' to index
        # Check if the column is not one of the last three
        if column_index < len(colunas) - 3:
            # Get the item ID and column name
            row_id = tree.identify_row(event.y)
            column_id = tree.identify_column(event.x)
            col_idx = int(column_id.replace("#", "")) - 1
            if col_idx < 0 or col_idx >= len(colunas):
                return
            if not row_id:
                return
            current_values = tree.item(row_id, "values")
            if col_idx >= len(current_values):
                return
            current_value = current_values[col_idx] if current_values[col_idx] is not None else ""

            # Create Entry to edit the cell
            entry = tk.Entry(tree)
            entry.insert(0, current_value)
            bbox = tree.bbox(row_id, column_id)
            if bbox:
                entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])

            def save_entry(event=None):
                new_value = entry.get()
                current_values = list(tree.item(row_id, "values"))
                current_values[col_idx] = new_value
                tree.item(row_id, values=current_values)
                entry.destroy()

            entry.bind("<Return>", save_entry)
            entry.bind("<FocusOut>", save_entry)  # Save on focus out
            entry.focus_set()

def vincular_evento_duplo_clique(treeview, colunas):
    treeview.bind("<Double-1>", lambda event, tv=treeview, cols=colunas: on_double_click(event, tv, cols))

def vincular_navegacao_setas(tree):
    def on_up_key(event):
        selected = tree.selection()
        if selected:
            prev_item = tree.prev(selected[0])
            if prev_item:
                tree.selection_set(prev_item)
                tree.focus(prev_item)
                tree.see(prev_item)
        return "break"  # Impede o comportamento padrão

# Executar a função para criar a interface
criar_interface()

